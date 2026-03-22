"""
Routes API pour le Plan de Surveillance
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from pathlib import Path
import uuid
import logging
import os

from models import (
    SurveillanceItem,
    SurveillanceItemCreate,
    SurveillanceItemUpdate,
    SurveillanceItemStatus,
    SurveillanceResponsible,
    AIAnalysisHistory,
    ActionType,
    EntityType,
    SuccessResponse
)
from dependencies import get_current_user, get_current_admin_user
from audit_service import AuditService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/surveillance", tags=["surveillance"])

# Variables globales (seront injectées depuis server.py)
db = None
audit_service = None
realtime_manager = None

def init_surveillance_routes(database, audit_svc, realtime_mgr=None):
    """Initialise les routes avec la connexion DB et audit service"""
    global db, audit_service, realtime_manager
    db = database
    audit_service = audit_svc
    realtime_manager = realtime_mgr


# ==================== Utilitaires récurrence ====================

def parse_periodicite_to_months(periodicite: str) -> int:
    """Convertit une périodicité texte en nombre de mois.
    Ex: '1 an' -> 12, '6 mois' -> 6, '3 mois' -> 3, '1 mois' -> 1, '2 ans' -> 24
    """
    p = periodicite.lower().strip()
    import re
    # Patterns: "X an(s)", "X mois", "trimestriel", "semestriel", "annuel", "mensuel"
    if "trimestriel" in p:
        return 3
    if "semestriel" in p:
        return 6
    if "annuel" in p:
        return 12
    if "mensuel" in p:
        return 1
    
    m = re.match(r'(\d+)\s*(an|ans|année|années)', p)
    if m:
        return int(m.group(1)) * 12
    
    m = re.match(r'(\d+)\s*(mois)', p)
    if m:
        return int(m.group(1))
    
    # Fallback: essayer de trouver un nombre
    m = re.search(r'(\d+)', p)
    if m:
        num = int(m.group(1))
        if "an" in p:
            return num * 12
        return num
    
    return 12  # Défaut: annuel


def add_months_to_date(date: datetime, months: int) -> datetime:
    """Ajoute N mois à une date."""
    month = date.month - 1 + months
    year = date.year + month // 12
    month = month % 12 + 1
    day = min(date.day, [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return date.replace(year=year, month=month, day=day)


def get_year_from_date_str(date_str: str) -> Optional[int]:
    """Extrait l'année d'une date ISO string."""
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str.replace('Z', '+00:00')).year
    except:
        try:
            return int(date_str[:4])
        except:
            return None


def periodicite_to_days(periodicite: str) -> int:
    """Convertit une périodicité texte en nombre de jours approximatif."""
    months = parse_periodicite_to_months(periodicite)
    return int(months * 30.44)  # 1 mois ≈ 30.44 jours


def calculate_tolerance_days(periodicite: str) -> int:
    """
    Calcule la tolérance en jours = floor(periodicite_en_jours * 0.08)
    Exemples:
        3 mois (90j) → 7 jours
        6 mois (180j) → 14 jours
        1 an (365j) → 29 jours
        2 ans (730j) → 58 jours
    """
    import math
    days = periodicite_to_days(periodicite)
    return int(math.floor(days * 0.08))


async def find_matching_occurrence(ctrl: dict, document_info: dict, db_instance) -> Optional[dict]:
    """
    Cherche une occurrence PLANIFIER existante qui correspond au contrôle extrait par l'IA.
    
    Critères de correspondance:
    - Même catégorie
    - Classe/type similaire (contenance de mots-clés communs)
    - Même exécutant ou organisme
    - Même bâtiment (si renseigné)
    - Date dans la tolérance ±8% de la périodicité
    
    Retourne l'occurrence trouvée ou None.
    """
    category = ctrl.get("category", "AUTRE")
    classe_type = (ctrl.get("classe_type") or "").strip().lower()
    executant = (ctrl.get("executant") or document_info.get("organisme_controle") or "").strip().lower()
    batiment = (ctrl.get("batiment") or "").strip().lower()
    derniere_visite = ctrl.get("derniere_visite") or document_info.get("date_intervention")
    
    if not derniere_visite or not classe_type:
        return None
    
    try:
        date_realisation = datetime.fromisoformat(derniere_visite).date()
    except Exception:
        return None
    
    # Chercher les occurrences non réalisées de la même catégorie
    query = {
        "category": category,
        "status": {"$in": [SurveillanceItemStatus.PLANIFIER.value, SurveillanceItemStatus.PLANIFIE.value]},
        "prochain_controle": {"$ne": None, "$exists": True}
    }
    
    candidates = await db_instance.surveillance_items.find(query, {"_id": 0}).to_list(length=500)
    
    if not candidates:
        return None
    
    best_match = None
    best_score = 0
    
    for candidate in candidates:
        score = 0
        
        # 1. Correspondance de classe/type (mots-clés communs)
        cand_classe = (candidate.get("classe_type") or "").strip().lower()
        if not cand_classe:
            continue
        
        # Calculer la similarité par mots communs
        words_ctrl = set(w for w in classe_type.split() if len(w) > 2)
        words_cand = set(w for w in cand_classe.split() if len(w) > 2)
        if words_ctrl and words_cand:
            common = words_ctrl & words_cand
            similarity = len(common) / max(len(words_ctrl), len(words_cand))
            if similarity < 0.3:
                continue  # Trop différent
            score += similarity * 40  # Max 40 points
        elif classe_type == cand_classe:
            score += 40
        else:
            continue
        
        # 2. Correspondance exécutant/organisme
        cand_executant = (candidate.get("executant") or "").strip().lower()
        cand_organisme = (candidate.get("organisme_controle") or "").strip().lower()
        if executant and (executant in cand_executant or executant in cand_organisme or 
                          cand_executant in executant or cand_organisme in executant):
            score += 20
        
        # 3. Correspondance bâtiment
        cand_batiment = (candidate.get("batiment") or "").strip().lower()
        if batiment and cand_batiment:
            if batiment == cand_batiment:
                score += 20
            elif batiment in cand_batiment or cand_batiment in batiment:
                score += 10
        elif not batiment and not cand_batiment:
            score += 10  # Les deux sans bâtiment = ok
        
        # 4. Correspondance de date (dans la tolérance ±8%)
        try:
            cand_date = datetime.fromisoformat(candidate["prochain_controle"]).date()
            cand_periodicite = candidate.get("periodicite", "1 an")
            tolerance = calculate_tolerance_days(cand_periodicite)
            ecart = abs((date_realisation - cand_date).days)
            
            if ecart <= tolerance:
                score += 20  # Dans la tolérance parfaite
            elif ecart <= tolerance * 2:
                score += 10  # Proche mais hors tolérance stricte
            else:
                continue  # Trop loin en date
        except Exception:
            continue
        
        if score > best_score:
            best_score = score
            best_match = candidate
    
    # Seuil minimum de confiance: 60/100
    if best_match and best_score >= 60:
        best_match["_match_score"] = best_score
        best_match["_match_confidence"] = "high" if best_score >= 80 else "medium"
        return best_match
    
    return None


def generate_recurring_controls(base_item_dict: dict, start_date_str: str, periodicite: str) -> list:
    """Génère tous les contrôles récurrents de start_date jusqu'à fin N+1.
    Retourne une liste de dicts prêts à être insérés en DB.
    """
    months = parse_periodicite_to_months(periodicite)
    if months <= 0:
        return []
    
    try:
        start_date = datetime.fromisoformat(start_date_str.replace('Z', '+00:00'))
    except:
        try:
            start_date = datetime.strptime(start_date_str[:10], "%Y-%m-%d")
        except:
            return []
    
    current_year = datetime.now().year
    end_year = current_year + 1
    end_date = datetime(end_year, 12, 31, 23, 59, 59)
    
    groupe_id = base_item_dict.get("groupe_controle_id") or str(uuid.uuid4())
    
    recurring = []
    next_date = add_months_to_date(start_date, months)
    
    while next_date <= end_date:
        new_item = {
            **base_item_dict,
            "id": str(uuid.uuid4()),
            "prochain_controle": next_date.strftime("%Y-%m-%d"),
            "annee": next_date.year,
            "groupe_controle_id": groupe_id,
            "status": "PLANIFIER",
            "date_realisation": None,
            "derniere_visite": None,
            "ecart_jours": None,
            "alerte_envoyee": False,
            "email_rappel_envoye": False,
            "numero_rapport": None,
            "resultat_controle": None,
            "commentaire": None,
            "attachments": [],
            # Reset des mois
            "janvier": False, "fevrier": False, "mars": False,
            "avril": False, "mai": False, "juin": False,
            "juillet": False, "aout": False, "septembre": False,
            "octobre": False, "novembre": False, "decembre": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        # Nettoyer _id si présent
        new_item.pop("_id", None)
        recurring.append(new_item)
        next_date = add_months_to_date(next_date, months)
    
    return recurring


# ==================== CRUD Routes ====================

@router.get("/items", response_model=List[dict])
async def get_surveillance_items(
    category: Optional[str] = None,
    responsable: Optional[str] = None,
    batiment: Optional[str] = None,
    status: Optional[str] = None,
    annee: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer tous les items du plan de surveillance avec filtres"""
    try:
        query = {"deleted_at": {"$exists": False}}
        
        if category:
            query["category"] = category
        if responsable:
            query["responsable"] = responsable
        if batiment:
            query["batiment"] = batiment
        if status:
            query["status"] = status
        if annee:
            query["annee"] = annee
        
        items = await db.surveillance_items.find(query).to_list(length=None)
        
        # Convertir _id en string
        for item in items:
            if "_id" in item:
                del item["_id"]
        
        return items
    except Exception as e:
        logger.error(f"Erreur récupération items surveillance: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/items/{item_id}")
async def get_surveillance_item(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer un item spécifique"""
    try:
        item = await db.surveillance_items.find_one({"id": item_id})
        
        if not item:
            raise HTTPException(status_code=404, detail="Item non trouvé")
        
        if "_id" in item:
            del item["_id"]
        
        return item
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur récupération item {item_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/items")
async def create_surveillance_item(
    item_data: SurveillanceItemCreate,
    current_user: dict = Depends(get_current_user)
):
    """Créer un nouvel item de surveillance + contrôles récurrents jusqu'à N+1"""
    try:
        # Déterminer l'année du contrôle
        annee = get_year_from_date_str(item_data.prochain_controle) if item_data.prochain_controle else datetime.now().year
        groupe_id = item_data.groupe_controle_id if hasattr(item_data, 'groupe_controle_id') and item_data.groupe_controle_id else str(uuid.uuid4())
        
        item = SurveillanceItem(
            **item_data.model_dump(),
            annee=annee,
            groupe_controle_id=groupe_id,
            created_by=current_user.get("id"),
            updated_by=current_user.get("id")
        )
        
        item_dict = item.model_dump()
        await db.surveillance_items.insert_one(item_dict)
        
        # Générer les contrôles récurrents futurs jusqu'à fin N+1
        recurring_count = 0
        if item_data.prochain_controle and item_data.periodicite:
            recurring = generate_recurring_controls(item_dict, item_data.prochain_controle, item_data.periodicite)
            if recurring:
                for r in recurring:
                    r["created_by"] = current_user.get("id")
                    r["updated_by"] = current_user.get("id")
                await db.surveillance_items.insert_many(recurring)
                recurring_count = len(recurring)
                logger.info(f"✅ {recurring_count} contrôle(s) récurrent(s) créé(s) pour {item.classe_type}")
        
        # Audit
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.CREATE,
            entity_type=EntityType.SURVEILLANCE,
            entity_id=item.id,
            entity_name=f"Plan surveillance: {item.classe_type}"
        )
        
        if "_id" in item_dict:
            del item_dict["_id"]
        
        item_dict["recurring_count"] = recurring_count
        
        # Broadcast WebSocket pour la synchronisation temps réel
        if realtime_manager:
            await realtime_manager.emit_event(
                "surveillance_plans",
                "created",
                item_dict,
                user_id=current_user["id"]
            )
        
        return item_dict
    except Exception as e:
        logger.error(f"Erreur création item surveillance: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/items/{item_id}")
async def update_surveillance_item(
    item_id: str,
    item_update: SurveillanceItemUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour un item de surveillance"""
    try:
        logger.info(f"🔍 UPDATE REQUEST - Item ID: {item_id}")
        logger.info(f"📦 Données reçues: {item_update.model_dump()}")
        
        # Vérifier que l'item existe
        existing = await db.surveillance_items.find_one({"id": item_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Item non trouvé")
        
        # Préparer les mises à jour
        update_data = {
            k: v for k, v in item_update.model_dump(exclude_unset=True).items()
            if v is not None
        }
        
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        update_data["updated_by"] = current_user.get("id")
        
        # Mettre à jour
        await db.surveillance_items.update_one(
            {"id": item_id},
            {"$set": update_data}
        )
        
        # Récupérer l'item mis à jour
        updated_item = await db.surveillance_items.find_one({"id": item_id})
        
        # Audit
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType.SURVEILLANCE,
            entity_id=item_id,
            entity_name=f"Plan surveillance: {existing.get('classe_type')}"
        )
        
        if "_id" in updated_item:
            del updated_item["_id"]
        
        # Broadcast WebSocket pour la synchronisation temps réel
        if realtime_manager:
            await realtime_manager.emit_event(
                "surveillance_plans",
                "updated",
                updated_item,
                user_id=current_user["id"]
            )
        
        return updated_item
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur mise à jour item {item_id}: {str(e)}")
        logger.error(f"❌ Type erreur: {type(e).__name__}")
        import traceback
        logger.error(f"❌ Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/items/{item_id}", response_model=SuccessResponse)
async def delete_surveillance_item(
    item_id: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """Supprimer un item de surveillance (Admin uniquement)"""
    try:
        item = await db.surveillance_items.find_one({"id": item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Item non trouvé")
        
        from datetime import datetime, timezone
        await db.surveillance_items.update_one(
            {"id": item_id},
            {"$set": {
                "deleted_at": datetime.now(timezone.utc),
                "deleted_by": current_user["id"],
                "deleted_by_name": f"{current_user['prenom']} {current_user['nom']}"
            }}
        )
        
        # Audit
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.DELETE,
            entity_type=EntityType.SURVEILLANCE,
            entity_id=item_id,
            entity_name=f"Plan surveillance: {item.get('classe_type')}"
        )
        
        # Broadcast WebSocket pour la synchronisation temps réel
        if realtime_manager:
            await realtime_manager.emit_event(
                "surveillance_plans",
                "deleted",
                {"id": item_id},
                user_id=current_user["id"]
            )
        
        return {"success": True, "message": "Item supprimé"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur suppression item {item_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/occurrences/{groupe_controle_id}")
async def get_control_occurrences(
    groupe_controle_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer toutes les occurrences d'un contrôle récurrent par son groupe_controle_id"""
    try:
        items = await db.surveillance_items.find(
            {"groupe_controle_id": groupe_controle_id},
            {"_id": 0, "id": 1, "annee": 1, "prochain_controle": 1, "status": 1,
             "classe_type": 1, "periodicite": 1, "date_realisation": 1}
        ).sort("annee", 1).to_list(length=50)

        return {"success": True, "occurrences": items, "total": len(items)}
    except Exception as e:
        logger.error(f"Erreur récupération occurrences {groupe_controle_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/batch-trends")
async def get_batch_trends(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Calcul des tendances de conformité pour un lot de contrôles récurrents"""
    try:
        ids = data.get("groupe_controle_ids", [])
        current_year = data.get("current_year", datetime.now(timezone.utc).year)

        if not ids:
            return {"success": True, "trends": {}}

        pipeline = [
            {"$match": {"groupe_controle_id": {"$in": ids}}},
            {"$group": {
                "_id": "$groupe_controle_id",
                "occurrences": {"$push": {
                    "annee": "$annee",
                    "status": "$status",
                    "prochain_controle": "$prochain_controle",
                    "date_realisation": "$date_realisation"
                }}
            }}
        ]
        results = await db.surveillance_items.aggregate(pipeline).to_list(length=500)

        trends = {}
        for group in results:
            gid = group["_id"]
            occs = group["occurrences"]
            past = [o for o in occs if (o.get("annee") or 9999) < current_year]

            if not past:
                trends[gid] = {"trend": "none", "realized": 0, "total": 0}
                continue

            realized = sum(1 for o in past if o.get("status") == "REALISE")
            total_past = len(past)
            ratio = realized / total_past

            if ratio >= 0.8:
                trend = "up"
            elif ratio >= 0.5:
                trend = "stable"
            else:
                trend = "down"

            trends[gid] = {"trend": trend, "realized": realized, "total": total_past}

        return {"success": True, "trends": trends}
    except Exception as e:
        logger.error(f"Erreur calcul tendances: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Années et Migration ====================

@router.get("/available-years")
async def get_available_years(current_user: dict = Depends(get_current_user)):
    """Récupérer la liste des années ayant des contrôles"""
    try:
        pipeline = [
            {"$match": {"annee": {"$ne": None}}},
            {"$group": {"_id": "$annee"}},
            {"$sort": {"_id": 1}}
        ]
        result = await db.surveillance_items.aggregate(pipeline).to_list(length=None)
        years = [r["_id"] for r in result if r["_id"]]
        
        current_year = datetime.now().year
        # S'assurer que l'année courante et N+1 sont incluses
        if current_year not in years:
            years.append(current_year)
        if (current_year + 1) not in years:
            years.append(current_year + 1)
        
        years = sorted(set(years))
        return {"years": years, "current_year": current_year}
    except Exception as e:
        logger.error(f"Erreur récupération années: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/migrate-years")
async def migrate_surveillance_years(current_user: dict = Depends(get_current_user)):
    """Migration: assigner annee aux contrôles existants et générer les récurrences manquantes"""
    try:
        items = await db.surveillance_items.find({"$or": [{"annee": None}, {"annee": {"$exists": False}}]}).to_list(length=None)
        
        updated = 0
        generated = 0
        
        for item in items:
            # 1. Assigner l'année
            annee = get_year_from_date_str(item.get("prochain_controle"))
            if not annee:
                annee = get_year_from_date_str(item.get("created_at")) or datetime.now().year
            
            update_fields = {"annee": annee}
            
            # Assigner un groupe_controle_id si manquant
            if not item.get("groupe_controle_id"):
                update_fields["groupe_controle_id"] = str(uuid.uuid4())
            
            await db.surveillance_items.update_one(
                {"id": item["id"]},
                {"$set": update_fields}
            )
            updated += 1
            
            # 2. Générer les contrôles récurrents futurs si pas déjà fait
            groupe_id = item.get("groupe_controle_id") or update_fields.get("groupe_controle_id")
            if item.get("prochain_controle") and item.get("periodicite"):
                # Vérifier qu'il n'y a pas déjà des récurrences pour ce groupe
                existing_count = await db.surveillance_items.count_documents({
                    "groupe_controle_id": groupe_id,
                    "id": {"$ne": item["id"]}
                })
                if existing_count == 0:
                    item_copy = {**item}
                    item_copy.pop("_id", None)
                    item_copy["groupe_controle_id"] = groupe_id
                    
                    recurring = generate_recurring_controls(
                        item_copy,
                        item["prochain_controle"],
                        item["periodicite"]
                    )
                    if recurring:
                        for r in recurring:
                            r["created_by"] = item.get("created_by")
                            r["updated_by"] = item.get("updated_by")
                        await db.surveillance_items.insert_many(recurring)
                        generated += len(recurring)
        
        return {
            "success": True,
            "updated": updated,
            "generated": generated,
            "message": f"{updated} contrôle(s) mis à jour, {generated} contrôle(s) récurrent(s) générés"
        }
    except Exception as e:
        logger.error(f"Erreur migration années: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Statistiques et Indicateurs ====================

@router.get("/stats")
async def get_surveillance_stats(
    annee: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer les statistiques du plan de surveillance (filtrables par année)"""
    try:
        query = {}
        if annee:
            query["annee"] = annee
        
        items = await db.surveillance_items.find(query).to_list(length=None)
        
        total = len(items)
        realises = len([i for i in items if i.get("status") == SurveillanceItemStatus.REALISE.value])
        planifies = len([i for i in items if i.get("status") == SurveillanceItemStatus.PLANIFIE.value])
        a_planifier = len([i for i in items if i.get("status") == SurveillanceItemStatus.PLANIFIER.value])
        
        # Par catégorie (dynamique - récupère toutes les catégories existantes)
        by_category = {}
        categories = list(set([i.get("category") for i in items if i.get("category")]))
        for cat in categories:
            cat_items = [i for i in items if i.get("category") == cat]
            cat_realises = len([i for i in cat_items if i.get("status") == SurveillanceItemStatus.REALISE.value])
            by_category[cat] = {
                "total": len(cat_items),
                "realises": cat_realises,
                "pourcentage": round((cat_realises / len(cat_items) * 100) if cat_items else 0, 1)
            }
        
        # Par responsable
        by_responsable = {}
        for resp in SurveillanceResponsible:
            resp_items = [i for i in items if i.get("responsable") == resp.value]
            resp_realises = len([i for i in resp_items if i.get("status") == SurveillanceItemStatus.REALISE.value])
            by_responsable[resp.value] = {
                "total": len(resp_items),
                "realises": resp_realises,
                "pourcentage": round((resp_realises / len(resp_items) * 100) if resp_items else 0, 1)
            }
        
        return {
            "global": {
                "total": total,
                "realises": realises,
                "planifies": planifies,
                "a_planifier": a_planifier,
                "pourcentage_realisation": round((realises / total * 100) if total > 0 else 0, 1)
            },
            "by_category": by_category,
            "by_responsable": by_responsable
        }
    except Exception as e:
        logger.error(f"Erreur récupération statistiques: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Alertes et Notifications ====================

@router.get("/alerts")
async def get_surveillance_alerts(current_user: dict = Depends(get_current_user)):
    """Récupérer les items nécessitant une alerte (échéance proche)"""
    try:
        items = await db.surveillance_items.find().to_list(length=None)
        
        alerts = []
        today = datetime.now(timezone.utc).date()
        
        for item in items:
            if item.get("prochain_controle") and item.get("status") != SurveillanceItemStatus.REALISE.value:
                try:
                    prochain_controle = datetime.fromisoformat(item["prochain_controle"]).date()
                    days_until = (prochain_controle - today).days
                    duree_rappel = item.get("duree_rappel_echeance", 30)
                    
                    # Alerte si moins de la durée de rappel configurée
                    if days_until <= duree_rappel:
                        if "_id" in item:
                            del item["_id"]
                        item["days_until"] = days_until
                        item["urgence"] = "critique" if days_until <= 7 else "important" if days_until <= 14 else "normal"
                        alerts.append(item)
                except:
                    pass
        
        # Trier par urgence (plus proche en premier)
        alerts.sort(key=lambda x: x.get("days_until", 999))
        
        return {
            "count": len(alerts),
            "alerts": alerts
        }
    except Exception as e:
        logger.error(f"Erreur récupération alertes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/badge-stats")
async def get_badge_stats(current_user: dict = Depends(get_current_user)):
    """
    Récupérer les statistiques pour le badge de notification du header
    - Nombre de contrôles à échéance proche (selon duree_rappel_echeance de chaque item)
    - Pourcentage de réalisation global
    """
    try:
        items = await db.surveillance_items.find().to_list(length=None)
        
        total = len(items)
        if total == 0:
            return {
                "echeances_proches": 0,
                "pourcentage_realisation": 0
            }
        
        # Compter les items réalisés
        realises = len([i for i in items if i.get("status") == SurveillanceItemStatus.REALISE.value])
        pourcentage_realisation = round((realises / total * 100), 1)
        
        # Compter les échéances proches (selon la durée de rappel de chaque item)
        echeances_proches = 0
        today = datetime.now(timezone.utc).date()
        
        for item in items:
            # Ignorer les items déjà réalisés
            if item.get("status") == SurveillanceItemStatus.REALISE.value:
                continue
                
            if item.get("prochain_controle"):
                try:
                    prochain_controle = datetime.fromisoformat(item["prochain_controle"]).date()
                    days_until = (prochain_controle - today).days
                    duree_rappel = item.get("duree_rappel_echeance", 30)
                    
                    # Compter si l'échéance est proche selon la durée de rappel
                    if days_until <= duree_rappel:
                        echeances_proches += 1
                except Exception as e:
                    logger.warning(f"Erreur parsing date pour item {item.get('id')}: {str(e)}")
                    pass
        
        return {
            "echeances_proches": echeances_proches,
            "pourcentage_realisation": pourcentage_realisation
        }
    except Exception as e:
        logger.error(f"Erreur récupération badge stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rapport-stats")
async def get_rapport_stats(
    annee: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Récupérer les statistiques complètes pour la page Rapport.
    Filtrable par année. Inclut écart moyen et % dans les temps (tolérance ±8%).
    """
    try:
        import math
        
        query = {}
        if annee:
            query["annee"] = annee
        
        items = await db.surveillance_items.find(query).to_list(length=None)
        
        empty_response = {
            "global": {
                "total": 0, "realises": 0, "planifies": 0, "a_planifier": 0,
                "pourcentage_realisation": 0, "en_retard": 0, "a_temps": 0,
                "ecart_moyen": None, "dans_les_temps": 0, "dans_les_temps_total": 0,
                "pourcentage_dans_les_temps": 0
            },
            "by_category": {}, "by_batiment": {}, "by_periodicite": {},
            "by_responsable": {}, "anomalies": 0
        }
        
        total = len(items)
        if total == 0:
            return empty_response
        
        today = datetime.now(timezone.utc).date()
        
        realises = [i for i in items if i.get("status") == SurveillanceItemStatus.REALISE.value]
        planifies = [i for i in items if i.get("status") == SurveillanceItemStatus.PLANIFIE.value]
        a_planifier = [i for i in items if i.get("status") == SurveillanceItemStatus.PLANIFIER.value]
        
        en_retard = 0
        a_temps = 0
        for item in items:
            if item.get("status") != SurveillanceItemStatus.REALISE.value and item.get("prochain_controle"):
                try:
                    prochain_controle = datetime.fromisoformat(item["prochain_controle"]).date()
                    if prochain_controle < today:
                        en_retard += 1
                    else:
                        a_temps += 1
                except:
                    pass
        
        # Statistiques d'écart : écart moyen + % dans les temps (tolérance ±8%)
        ecarts = []
        dans_les_temps = 0
        dans_les_temps_total = 0
        for item in realises:
            ecart = item.get("ecart_jours")
            if ecart is not None:
                ecarts.append(ecart)
                dans_les_temps_total += 1
                tolerance = calculate_tolerance_days(item.get("periodicite", "1 an"))
                if abs(ecart) <= tolerance:
                    dans_les_temps += 1
        
        ecart_moyen = round(sum(ecarts) / len(ecarts), 1) if ecarts else None
        pourcentage_dans_les_temps = round((dans_les_temps / dans_les_temps_total * 100), 1) if dans_les_temps_total > 0 else 0
        
        # Anomalies
        anomalies = 0
        for item in items:
            commentaire = (item.get("commentaire") or "").lower()
            if any(kw in commentaire for kw in ["anomalie", "problème", "défaut", "dysfonctionnement", "intervention", "réparation"]):
                anomalies += 1
        
        # Par catégorie
        by_category = {}
        for cat in set(i.get("category") for i in items if i.get("category")):
            cat_items = [i for i in items if i.get("category") == cat]
            cat_realises = [i for i in cat_items if i.get("status") == SurveillanceItemStatus.REALISE.value]
            cat_ecarts = [i.get("ecart_jours") for i in cat_realises if i.get("ecart_jours") is not None]
            by_category[cat] = {
                "total": len(cat_items),
                "realises": len(cat_realises),
                "pourcentage": round((len(cat_realises) / len(cat_items) * 100) if cat_items else 0, 1),
                "ecart_moyen": round(sum(cat_ecarts) / len(cat_ecarts), 1) if cat_ecarts else None
            }
        
        # Par bâtiment
        by_batiment = {}
        for bat in set(i.get("batiment", "Non spécifié") for i in items):
            bat_items = [i for i in items if i.get("batiment") == bat]
            bat_realises = len([i for i in bat_items if i.get("status") == SurveillanceItemStatus.REALISE.value])
            by_batiment[bat] = {
                "total": len(bat_items),
                "realises": bat_realises,
                "pourcentage": round((bat_realises / len(bat_items) * 100) if bat_items else 0, 1)
            }
        
        # Par périodicité
        by_periodicite = {}
        for per in set(i.get("periodicite", "Non spécifié") for i in items):
            per_items = [i for i in items if i.get("periodicite") == per]
            per_realises = len([i for i in per_items if i.get("status") == SurveillanceItemStatus.REALISE.value])
            by_periodicite[per] = {
                "total": len(per_items),
                "realises": per_realises,
                "pourcentage": round((per_realises / len(per_items) * 100) if per_items else 0, 1)
            }
        
        # Par responsable
        by_responsable = {}
        for resp in SurveillanceResponsible:
            resp_items = [i for i in items if i.get("responsable") == resp.value]
            resp_realises = len([i for i in resp_items if i.get("status") == SurveillanceItemStatus.REALISE.value])
            by_responsable[resp.value] = {
                "total": len(resp_items),
                "realises": resp_realises,
                "pourcentage": round((resp_realises / len(resp_items) * 100) if resp_items else 0, 1)
            }
        
        return {
            "global": {
                "total": total,
                "realises": len(realises),
                "planifies": len(planifies),
                "a_planifier": len(a_planifier),
                "pourcentage_realisation": round((len(realises) / total * 100), 1),
                "en_retard": en_retard,
                "a_temps": a_temps,
                "ecart_moyen": ecart_moyen,
                "dans_les_temps": dans_les_temps,
                "dans_les_temps_total": dans_les_temps_total,
                "pourcentage_dans_les_temps": pourcentage_dans_les_temps
            },
            "by_category": by_category,
            "by_batiment": by_batiment,
            "by_periodicite": by_periodicite,
            "by_responsable": by_responsable,
            "anomalies": anomalies
        }
    except Exception as e:
        logger.error(f"Erreur récupération rapport stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Envoi manuel d'email de rappel ====================

@router.post("/items/{item_id}/send-reminder")
async def send_manual_reminder(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Envoyer manuellement un email de rappel pour un contrôle"""
    try:
        # Récupérer l'item
        item = await db.surveillance_items.find_one({"id": item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Contrôle non trouvé")
        
        # Vérifier qu'un responsable est défini
        responsable_id = item.get("responsable_notification_id")
        if not responsable_id:
            raise HTTPException(status_code=400, detail="Aucun responsable de notification défini pour ce contrôle")
        
        # Récupérer l'utilisateur
        from bson import ObjectId
        user = await db.users.find_one({"id": responsable_id})
        if not user:
            try:
                user = await db.users.find_one({"_id": ObjectId(responsable_id)})
            except:
                pass
        
        if not user:
            raise HTTPException(status_code=404, detail="Responsable non trouvé")
        
        if not user.get("email"):
            raise HTTPException(status_code=400, detail="Le responsable n'a pas d'adresse email")
        
        # Envoyer l'email
        user_name = f"{user.get('prenom', '')} {user.get('nom', '')}".strip() or user.get("email")
        success = await send_surveillance_reminder_email(
            user_email=user.get("email"),
            user_name=user_name,
            item=item
        )
        
        if success:
            # Optionnel: marquer l'email comme envoyé
            await db.surveillance_items.update_one(
                {"id": item_id},
                {"$set": {
                    "email_rappel_envoye": True,
                    "alerte_date": datetime.now(timezone.utc).isoformat(),
                    "rappel_manuel_par": current_user.get("id"),
                    "rappel_manuel_date": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            return {
                "success": True,
                "message": f"Email de rappel envoyé à {user.get('email')}"
            }
        else:
            raise HTTPException(status_code=500, detail="Échec de l'envoi de l'email. Vérifiez la configuration SMTP.")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur envoi manuel rappel: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Upload de pièces jointes ====================

@router.post("/items/{item_id}/upload")
async def upload_piece_jointe(
    item_id: str,
    files: list[UploadFile] = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload une ou plusieurs pièces jointes pour un item"""
    try:
        item = await db.surveillance_items.find_one({"id": item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Item non trouvé")
        
        upload_dir = Path("uploads/surveillance")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        existing_attachments = item.get("attachments", [])
        new_attachments = []
        
        for file in files:
            file_ext = Path(file.filename).suffix
            file_id = str(uuid.uuid4())
            unique_filename = f"{item_id}_{file_id}{file_ext}"
            file_path = upload_dir / unique_filename
            
            content = await file.read()
            with open(file_path, "wb") as f:
                f.write(content)
            
            attachment = {
                "id": file_id,
                "filename": file.filename,
                "url": f"/uploads/surveillance/{unique_filename}",
                "size": len(content),
                "uploaded_at": datetime.now(timezone.utc).isoformat()
            }
            new_attachments.append(attachment)
        
        all_attachments = existing_attachments + new_attachments
        
        await db.surveillance_items.update_one(
            {"id": item_id},
            {
                "$set": {
                    "attachments": all_attachments,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user.get("id")
                }
            }
        )
        
        return {
            "success": True,
            "attachments": new_attachments,
            "total_attachments": len(all_attachments)
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur upload pièce jointe: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/items/{item_id}/attachments/{attachment_id}")
async def delete_attachment(
    item_id: str,
    attachment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une pièce jointe"""
    try:
        item = await db.surveillance_items.find_one({"id": item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Item non trouvé")
        
        attachments = item.get("attachments", [])
        attachment = next((a for a in attachments if a.get("id") == attachment_id), None)
        if not attachment:
            raise HTTPException(status_code=404, detail="Pièce jointe non trouvée")
        
        # Supprimer le fichier physique
        file_path = Path(attachment["url"].lstrip("/"))
        if file_path.exists():
            file_path.unlink()
        
        # Retirer de la liste
        new_attachments = [a for a in attachments if a.get("id") != attachment_id]
        await db.surveillance_items.update_one(
            {"id": item_id},
            {
                "$set": {
                    "attachments": new_attachments,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user.get("id")
                }
            }
        )
        
        return {"success": True, "message": "Pièce jointe supprimée"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur suppression pièce jointe: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Recherche ====================

@router.post("/search")
async def search_surveillance_items(
    search_request: dict,
    current_user: dict = Depends(get_current_user)
):
    """Recherche dans les contrôles du plan de surveillance (style Manuel)"""
    try:
        query = search_request.get("query", "").lower().strip()
        if not query:
            return {"results": []}
        
        items = await db.surveillance_items.find({}, {"_id": 0}).to_list(length=None)
        
        results = []
        for item in items:
            score = 0.0
            
            # Champs à rechercher avec pondération
            fields = {
                "classe_type": 3.0,
                "category": 2.5,
                "executant": 2.0,
                "organisme_controle": 2.0,
                "batiment": 1.5,
                "description": 1.0,
                "commentaire": 1.0,
                "reference_reglementaire": 1.5,
                "numero_rapport": 2.0,
                "periodicite": 1.0,
                "resultat_controle": 1.0
            }
            
            matched_fields = []
            for field, weight in fields.items():
                value = str(item.get(field, "") or "").lower()
                if query in value:
                    score += weight
                    matched_fields.append(field)
            
            if score > 0:
                # Construire un extrait pertinent
                excerpt_parts = []
                if item.get("classe_type"):
                    excerpt_parts.append(item["classe_type"])
                if item.get("description"):
                    desc = item["description"]
                    idx = desc.lower().find(query)
                    if idx >= 0:
                        start = max(0, idx - 40)
                        excerpt_parts.append("..." + desc[start:start + 120] + "...")
                    else:
                        excerpt_parts.append(desc[:120])
                
                results.append({
                    "id": item.get("id"),
                    "classe_type": item.get("classe_type", ""),
                    "category": item.get("category", ""),
                    "batiment": item.get("batiment", ""),
                    "executant": item.get("executant", ""),
                    "periodicite": item.get("periodicite", ""),
                    "resultat_controle": item.get("resultat_controle"),
                    "derniere_visite": item.get("derniere_visite"),
                    "prochain_controle": item.get("prochain_controle"),
                    "status": item.get("status"),
                    "excerpt": " | ".join(excerpt_parts),
                    "matched_fields": matched_fields,
                    "relevance_score": score
                })
        
        results.sort(key=lambda x: x["relevance_score"], reverse=True)
        return {"results": results[:20]}
    
    except Exception as e:
        logger.error(f"Erreur recherche surveillance: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Import/Export ====================

@router.post("/import")
async def import_surveillance_data(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_admin_user)
):
    """Importer des données depuis un fichier CSV/Excel"""
    try:
        import pandas as pd
        from io import BytesIO
        
        content = await file.read()
        
        # Lire le fichier selon l'extension
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Format de fichier non supporté")
        
        # Mapper les colonnes
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                item = SurveillanceItem(
                    classe_type=str(row.get('classe_type', '')),
                    category=str(row.get('category', 'AUTRE')),
                    batiment=str(row.get('batiment', '')),
                    periodicite=str(row.get('periodicite', '')),
                    responsable=str(row.get('responsable', 'MAINT')),
                    executant=str(row.get('executant', '')),
                    description=str(row.get('description', '')) if pd.notna(row.get('description')) else None,
                    derniere_visite=str(row.get('derniere_visite', '')) if pd.notna(row.get('derniere_visite')) else None,
                    prochain_controle=str(row.get('prochain_controle', '')) if pd.notna(row.get('prochain_controle')) else None,
                    commentaire=str(row.get('commentaire', '')) if pd.notna(row.get('commentaire')) else None,
                    created_by=current_user.get("id"),
                    updated_by=current_user.get("id")
                )
                
                await db.surveillance_items.insert_one(item.model_dump())
                imported_count += 1
            except Exception as e:
                errors.append(f"Ligne {index + 2}: {str(e)}")
        
        return {
            "success": True,
            "imported_count": imported_count,
            "errors": errors[:10]  # Limiter à 10 erreurs
        }
    except Exception as e:
        logger.error(f"Erreur import données: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/template")
async def export_template(current_user: dict = Depends(get_current_user)):
    """Télécharger un template CSV pour l'import"""
    try:
        import pandas as pd
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        # Créer un DataFrame avec les colonnes attendues
        template_data = {
            "classe_type": ["Protection incendie", "Installations électriques"],
            "category": ["INCENDIE", "ELECTRIQUE"],
            "batiment": ["BATIMENT 1", "BATIMENT 2"],
            "periodicite": ["6 mois", "1 an"],
            "responsable": ["MAINT", "PROD"],
            "executant": ["DESAUTEL", "APAVE"],
            "description": ["Contrôle des liaisons, zones, batterie", "Contrôle réglementaire"],
            "derniere_visite": ["2024-01-15", "2024-02-20"],
            "prochain_controle": ["2024-07-15", "2025-02-20"],
            "commentaire": ["RAS", "À planifier"]
        }
        
        df = pd.DataFrame(template_data)
        
        # Créer un buffer
        buffer = BytesIO()
        df.to_csv(buffer, index=False, encoding='utf-8-sig')
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=template_plan_surveillance.csv"
            }
        )
    except Exception as e:
        logger.error(f"Erreur export template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Vérification automatique des échéances ====================

@router.post("/check-due-dates")
async def check_due_dates(current_user: dict = Depends(get_current_user)):
    """
    Vérifier les dates d'échéance et retourner les items nécessitant une alerte.
    
    Logique (notification uniquement, PAS de changement de statut) :
    - Pour chaque item NON réalisé avec un prochain_controle
    - Si aujourd'hui est dans la période de rappel (duree_rappel_echeance jours AVANT prochain_controle)
    - Alors signaler l'item comme nécessitant une alerte
    
    Le délai de rappel (duree_rappel_echeance) est paramétrable par item (défaut: 30 jours).
    Le décompte ne se déclenche que AVANT la date du contrôle, jamais après.
    """
    try:
        today = datetime.now(timezone.utc).date()
        alerts_needed = []
        
        # Récupérer les items non réalisés ayant une date de prochain contrôle
        items = await db.surveillance_items.find({
            "status": {"$ne": SurveillanceItemStatus.REALISE.value},
            "prochain_controle": {"$ne": None, "$exists": True}
        }).to_list(length=None)
        
        for item in items:
            try:
                prochain_controle = datetime.fromisoformat(item["prochain_controle"]).date()
                duree_rappel = item.get("duree_rappel_echeance", 30)
                
                # Calculer la date de début de la période de rappel
                date_rappel = prochain_controle - timedelta(days=duree_rappel)
                
                # Alerte uniquement en décompte : today >= date_rappel ET today <= prochain_controle
                if today >= date_rappel and today <= prochain_controle:
                    days_remaining = (prochain_controle - today).days
                    alerts_needed.append({
                        "id": item["id"],
                        "classe_type": item.get("classe_type"),
                        "category": item.get("category"),
                        "batiment": item.get("batiment"),
                        "prochain_controle": item["prochain_controle"],
                        "days_remaining": days_remaining,
                        "duree_rappel_echeance": duree_rappel,
                        "alerte_envoyee": item.get("alerte_envoyee", False),
                        "responsable_notification_id": item.get("responsable_notification_id")
                    })
            except Exception as e:
                logger.warning(f"Erreur traitement item {item.get('id')}: {str(e)}")
                continue
        
        return {
            "success": True,
            "alerts_count": len(alerts_needed),
            "alerts": alerts_needed,
            "message": f"{len(alerts_needed)} contrôle(s) à échéance proche"
        }
    except Exception as e:
        logger.error(f"Erreur vérification échéances: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# ==================== Extraction IA de documents de contrôle ====================

@router.post("/ai/extract")
async def extract_surveillance_from_document(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Analyse un document de contrôle (PDF) via IA et extrait les informations
    pour créer un ou plusieurs contrôles dans le plan de surveillance.
    L'IA recherche aussi la périodicité réglementaire.
    """
    import tempfile
    import os
    import json

    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType

        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Clé LLM non configurée")

        # Sauvegarder temporairement le fichier ET de façon permanente
        ext = os.path.splitext(file.filename)[1].lower()
        
        # Sauvegarde permanente pour rattachement aux contrôles
        upload_dir = Path("uploads/surveillance")
        upload_dir.mkdir(parents=True, exist_ok=True)
        permanent_file_id = str(uuid.uuid4())
        permanent_filename = f"ai_source_{permanent_file_id}{ext}"
        permanent_path = upload_dir / permanent_filename
        permanent_url = f"/uploads/surveillance/{permanent_filename}"
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        # Sauvegarder aussi de façon permanente
        with open(permanent_path, "wb") as f:
            f.write(content)

        mime_map = {
            ".pdf": "application/pdf",
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".webp": "image/webp"
        }
        mime_type = mime_map.get(ext, "application/pdf")

        # Étape 1 : Extraction des informations du document
        chat = LlmChat(
            api_key=api_key,
            session_id=f"surveillance_extract_{uuid.uuid4().hex[:8]}",
            system_message="""Tu es un expert en réglementation française de sécurité au travail et en contrôles réglementaires.
Analyse le document de contrôle/vérification fourni et extrais TOUS les types de contrôles distincts qu'il contient.

Un rapport d'organisme de contrôle (APAVE, SOCOTEC, DEKRA, BUREAU VERITAS) peut contenir :
- Plusieurs types de vérifications dans un même rapport (levage, électrique, incendie, etc.)
- Plusieurs équipements inspectés pour un même type de contrôle
- Des observations, anomalies ou réserves pour chaque point

RÈGLE : Crée une entrée SÉPARÉE pour chaque type de contrôle/vérification réglementaire distinct.
Si le rapport mentionne à la fois une VGP de machines ET une vérification d'installations électriques, ce sont 2 entrées.
Si le rapport couvre plusieurs équipements du MÊME type, cela reste UNE SEULE entrée.

Réponds UNIQUEMENT avec un JSON valide, sans texte autour ni backticks.

Format attendu:
{
  "document_info": {
    "numero_rapport": "string - numéro du rapport ou null",
    "organisme_controle": "string - nom de l'organisme",
    "date_intervention": "YYYY-MM-DD - date du contrôle",
    "site_controle": "string - nom/adresse du site contrôlé"
  },
  "controles": [
    {
      "classe_type": "string - type précis du contrôle",
      "category": "string parmi: ELECTRIQUE, INCENDIE, MANUTENTION, SECURITE_ENVIRONNEMENT, MMRI, EXTRACTION, AUTRE",
      "batiment": "string - bâtiment ou zone concernée ('' si non précisé, JAMAIS null)",
      "executant": "string - nom de l'organisme",
      "description": "string - description de ce qui a été contrôlé",
      "derniere_visite": "YYYY-MM-DD - date de la visite",
      "references_reglementaires": "string - références légales",
      "resultat": "CONFORME|NON_CONFORME|AVEC_RESERVES",
      "anomalies": "string ou null",
      "periodicite_detectee": "UNIQUEMENT en format simple: '1 an', '6 mois', '3 mois', '2 ans'. NE PAS inclure de références réglementaires ici. null si non trouvée",
      "equipements_concernes": "string - liste des équipements"
    }
  ]
}

IMPORTANT:
- Sépare les différents types de contrôles
- periodicite_detectee en format SIMPLE uniquement (ex: '1 an', '6 mois')
- batiment JAMAIS null, utiliser '' si non précisé
- Si anomalies: résultat = NON_CONFORME ou AVEC_RESERVES"""
        ).with_model("gemini", "gemini-2.5-flash")

        file_content = FileContentWithMimeType(
            file_path=tmp_path,
            mime_type=mime_type
        )

        response = await chat.send_message(UserMessage(
            text="Analyse ce document de contrôle réglementaire et extrais toutes les informations pour chaque type de contrôle distinct.",
            file_contents=[file_content]
        ))

        # Nettoyer le fichier temporaire
        os.unlink(tmp_path)

        # Parser la réponse JSON
        response_text = response.strip()
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        response_text = response_text.strip()
        
        extracted = json.loads(response_text)

        # Étape 2 : Pour chaque contrôle sans périodicité, rechercher via IA + connaissances réglementaires
        controles = extracted.get("controles", [])
        for ctrl in controles:
            if not ctrl.get("periodicite_detectee"):
                # Utiliser une seconde requête IA pour rechercher la périodicité réglementaire
                refs = ctrl.get("references_reglementaires", "")
                classe = ctrl.get("classe_type", "")
                category = ctrl.get("category", "")
                
                periodicity_chat = LlmChat(
                    api_key=api_key,
                    session_id=f"periodicity_search_{uuid.uuid4().hex[:8]}",
                    system_message="""Tu es un expert en réglementation française de sécurité au travail.
On te donne un type de contrôle et ses références réglementaires.
Tu dois déterminer la périodicité réglementaire obligatoire en France pour ce type de contrôle.

Réponds UNIQUEMENT avec un JSON valide, sans texte ni backticks:
{
  "periodicite": "string - la périodicité (ex: '1 an', '6 mois', '3 mois', '5 ans') ou null si introuvable",
  "source_reglementaire": "string - le texte de loi qui impose cette périodicité ou null",
  "confiance": "HAUTE|MOYENNE|BASSE - ton niveau de confiance dans cette réponse",
  "explication": "string - brève explication de la réglementation applicable"
}

Voici les principales périodicités réglementaires françaises:
- Installations électriques (vérification initiale + périodique): 1 an (Arrêté du 26/12/2011, Art. R4226-14 Code du Travail)
- Thermographie infrarouge APSAD D19: 1 an (recommandation APSAD)
- Appareils de levage (chariots élévateurs, ponts roulants): 1 an pour la VGP (Arrêté du 01/03/2004)
- Ascenseurs/monte-charges: 1 an (Arrêté du 29/12/2010)
- Portes et portails automatiques: 6 mois (Arrêté du 21/12/1993)
- Échafaudages roulants: 3 mois ou avant chaque utilisation (Arrêté du 21/12/2004)
- EPI contre les chutes: 12 mois (Arrêté du 19/03/1993)
- Extincteurs: 1 an (Arrêté du 20/05/1963, R4227-39 Code du Travail)
- SSI/détection incendie: 1 an (MS 73, PE 4)
- Installations de gaz: 1 an (Arrêté du 21/12/1993)
- Appareils à pression: selon catégorie (Arrêté du 20/11/2017)"""
                ).with_model("gemini", "gemini-2.5-flash")

                period_response = await periodicity_chat.send_message(UserMessage(
                    text=f"""Détermine la périodicité réglementaire pour ce contrôle:
- Type: {classe}
- Catégorie: {category}
- Références réglementaires trouvées dans le document: {refs}
- Équipements: {ctrl.get('equipements_concernes', 'Non précisé')}"""
                ))

                period_text = period_response.strip()
                if period_text.startswith("```"):
                    period_text = period_text.split("\n", 1)[1] if "\n" in period_text else period_text[3:]
                if period_text.endswith("```"):
                    period_text = period_text[:-3]
                period_text = period_text.strip()

                try:
                    period_data = json.loads(period_text)
                    ctrl["periodicite_detectee"] = period_data.get("periodicite")
                    ctrl["periodicite_source"] = period_data.get("source_reglementaire")
                    ctrl["periodicite_confiance"] = period_data.get("confiance", "BASSE")
                    ctrl["periodicite_explication"] = period_data.get("explication")
                except json.JSONDecodeError:
                    ctrl["periodicite_detectee"] = None
                    ctrl["periodicite_confiance"] = "BASSE"
                    ctrl["periodicite_explication"] = "Impossible de déterminer automatiquement"
            else:
                ctrl["periodicite_confiance"] = "HAUTE"
                ctrl["periodicite_source"] = ctrl.get("references_reglementaires")
                ctrl["periodicite_explication"] = "Périodicité trouvée directement dans le document"

        extracted["controles"] = controles
        return {
            "success": True, 
            "data": extracted,
            "source_file": {
                "id": permanent_file_id,
                "filename": file.filename,
                "url": permanent_url,
                "size": len(content)
            }
        }

    except json.JSONDecodeError:
        logger.error(f"Erreur parsing JSON de l'IA: {response_text[:300] if 'response_text' in dir() else 'N/A'}")
        return {"success": False, "error": "L'IA n'a pas pu extraire les informations correctement. Veuillez réessayer."}
    except ImportError:
        raise HTTPException(status_code=500, detail="Module emergentintegrations non installé")
    except Exception as e:
        logger.error(f"Erreur extraction IA surveillance: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {"success": False, "error": f"Erreur lors de l'extraction: {str(e)}"}


# ==================== Fonctions utilitaires pour create_batch_from_ai ====================

def _normalize_periodicite(raw: str) -> str:
    """Normalise une périodicité brute en format standard.
    Ex: 'Annuelle (réf. Arrêtés du 5 mars)' → '1 an'
    """
    p = raw.lower().strip()
    if any(w in p for w in ['annuel', 'annual']):
        return "2 ans" if ('bi' in p or 'biennal' in p) else "1 an"
    if 'semestriel' in p:
        return "6 mois"
    if 'trimestriel' in p:
        return "3 mois"
    if 'bimestriel' in p:
        return "2 mois"
    if 'mensuel' in p:
        return "1 mois"
    if 'hebdomadaire' in p:
        return "1 semaine"
    if 'quotidien' in p:
        return "1 jour"
    return raw


def _calculate_next_control_date(derniere_visite: str, periodicite: str) -> Optional[str]:
    """Calcule la date du prochain contrôle = dernière visite + périodicité."""
    if not derniere_visite or not periodicite:
        return None
    try:
        from dateutil.relativedelta import relativedelta
        import re
        base_date = datetime.fromisoformat(derniere_visite)
        p = periodicite.lower().strip()
        
        # Table de correspondance textuelle
        text_map = {
            'annuel': ('years', 1), 'annual': ('years', 1),
            'semestriel': ('months', 6), 'trimestriel': ('months', 3),
            'bimestriel': ('months', 2), 'mensuel': ('months', 1),
            'hebdomadaire': ('weeks', 1), 'quotidien': ('days', 1),
        }
        for keyword, (unit, num) in text_map.items():
            if keyword in p:
                if keyword in ('annuel', 'annual') and ('bi' in p or 'biennal' in p):
                    unit, num = 'years', 2
                delta = relativedelta(years=num) if unit == 'years' else \
                        relativedelta(months=num) if unit == 'months' else \
                        timedelta(weeks=num) if unit == 'weeks' else timedelta(days=num)
                return (base_date + delta).strftime("%Y-%m-%d")
        
        # Format numérique: "1 an", "6 mois", "3 semaines", "5 jours"
        patterns = [
            (r'(\d+)\s*an', 'years'), (r'(\d+)\s*mois', 'months'),
            (r'(\d+)\s*semaine', 'weeks'), (r'(\d+)\s*jour', 'days'),
        ]
        for pattern, unit in patterns:
            m = re.search(pattern, p)
            if m:
                n = int(m.group(1))
                delta = relativedelta(years=n) if unit == 'years' else \
                        relativedelta(months=n) if unit == 'months' else \
                        timedelta(weeks=n) if unit == 'weeks' else timedelta(days=n)
                return (base_date + delta).strftime("%Y-%m-%d")
        
        # Fallback: 1 an
        logger.warning(f"Périodicité non reconnue '{periodicite}', fallback 1 an")
        return (base_date + relativedelta(years=1)).strftime("%Y-%m-%d")
    except Exception as e:
        logger.warning(f"Erreur calcul prochain contrôle: {e}")
        return None


def _build_control_comment(ctrl: dict) -> Optional[str]:
    """Construit le commentaire à partir des anomalies et équipements."""
    parts = []
    if ctrl.get("anomalies"):
        parts.append(f"ANOMALIES DÉTECTÉES:\n{ctrl['anomalies']}")
    if ctrl.get("equipements_concernes"):
        parts.append(f"Équipements: {ctrl['equipements_concernes']}")
    return "\n\n".join(parts) if parts else None


_RESULTAT_MAP = {"CONFORME": "Conforme", "NON_CONFORME": "Non conforme", "AVEC_RESERVES": "Avec réserves"}

def _map_resultat(resultat: Optional[str]) -> Optional[str]:
    """Mappe un code résultat IA vers un libellé lisible."""
    return _RESULTAT_MAP.get(resultat, resultat)


def _prepare_source_attachment(source_file: Optional[dict]) -> Optional[dict]:
    """Prépare la pièce jointe source si fournie."""
    if not source_file or not source_file.get("url"):
        return None
    return {
        "id": source_file.get("id", str(uuid.uuid4())),
        "filename": source_file.get("filename", "document.pdf"),
        "url": source_file.get("url"),
        "size": source_file.get("size", 0),
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }


async def _update_matched_occurrence(
    matched: dict, ctrl: dict, document_info: dict,
    derniere_visite: str, prochain_controle: str, periodicite: str,
    resultat: str, commentaire: Optional[str],
    source_attachment: Optional[dict], current_user: dict
) -> dict:
    """Met à jour une occurrence planifiée existante qui correspond au rapport analysé.
    Calcule l'écart, met à jour l'item, régénère les occurrences futures, et logge l'audit.
    """
    # Calculer l'écart en jours
    ecart_jours = None
    try:
        date_prevue = datetime.fromisoformat(matched["prochain_controle"]).date()
        date_reelle = datetime.fromisoformat(derniere_visite).date()
        ecart_jours = (date_reelle - date_prevue).days
    except Exception:
        pass
    
    # Préparer la mise à jour
    set_data = {
        "status": SurveillanceItemStatus.REALISE.value,
        "derniere_visite": derniere_visite,
        "date_realisation": derniere_visite,
        "prochain_controle": prochain_controle,
        "ecart_jours": ecart_jours,
        "resultat_controle": resultat,
        "numero_rapport": document_info.get("numero_rapport"),
        "organisme_controle": document_info.get("organisme_controle"),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.get("id")
    }
    if commentaire:
        set_data["commentaire"] = commentaire
    if ctrl.get("references_reglementaires"):
        set_data["reference_reglementaire"] = ctrl["references_reglementaires"]
    
    update_query = {"$set": set_data}
    if source_attachment:
        update_query["$push"] = {"attachments": source_attachment}
    
    await db.surveillance_items.update_one({"id": matched["id"]}, update_query)
    
    # Récupérer l'item mis à jour
    updated_item = await db.surveillance_items.find_one({"id": matched["id"]}, {"_id": 0})
    
    # Régénérer les occurrences futures
    group_id = matched.get("groupe_controle_id")
    if derniere_visite and periodicite:
        if group_id:
            await db.surveillance_items.delete_many({
                "groupe_controle_id": group_id,
                "status": {"$in": [SurveillanceItemStatus.PLANIFIER.value, SurveillanceItemStatus.PLANIFIE.value]},
                "id": {"$ne": matched["id"]}
            })
        base = updated_item or matched
        base["groupe_controle_id"] = group_id or str(uuid.uuid4())
        recurring = generate_recurring_controls(base, derniere_visite, periodicite)
        if recurring:
            for r in recurring:
                r["created_by"] = current_user.get("id")
                r["updated_by"] = current_user.get("id")
            await db.surveillance_items.insert_many(recurring)
            logger.info(f"✅ {len(recurring)} occurrence(s) régénérée(s) pour {matched.get('classe_type')}")
    
    logger.info(f"✅ Occurrence matchée: {matched['id'][:8]} (écart: {ecart_jours}j)")
    
    # Audit
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=f"{current_user['prenom']} {current_user['nom']}",
        user_email=current_user["email"],
        action=ActionType.UPDATE,
        entity_type=EntityType.SURVEILLANCE,
        entity_id=matched["id"],
        entity_name=f"Plan surveillance (IA match): {matched.get('classe_type')}"
    )
    
    return {
        **(updated_item or {}),
        "_matched_from": matched["id"],
        "_ecart_jours": ecart_jours,
        "_match_score": matched.get("_match_score"),
        "_action": "updated_occurrence"
    }


async def _create_new_surveillance_item(
    ctrl: dict, document_info: dict,
    derniere_visite: str, prochain_controle: str, periodicite: str,
    resultat: str, commentaire: Optional[str],
    source_attachment: Optional[dict], annee: int, current_user: dict
) -> dict:
    """Crée un nouvel item de surveillance (réalisé) + occurrences récurrentes futures."""
    groupe_id = str(uuid.uuid4())
    
    item = SurveillanceItem(
        classe_type=ctrl.get("classe_type", ""),
        category=ctrl.get("category", "AUTRE"),
        batiment=ctrl.get("batiment") or "",
        periodicite=ctrl.get("periodicite", "Non déterminée"),
        responsable=SurveillanceResponsible.EXTERNE,
        executant=ctrl.get("executant", document_info.get("organisme_controle", "")),
        description=ctrl.get("description"),
        derniere_visite=derniere_visite,
        prochain_controle=prochain_controle,
        status=SurveillanceItemStatus.REALISE,
        date_realisation=derniere_visite,
        commentaire=commentaire,
        reference_reglementaire=ctrl.get("references_reglementaires"),
        numero_rapport=document_info.get("numero_rapport"),
        organisme_controle=document_info.get("organisme_controle"),
        resultat_controle=resultat,
        attachments=[source_attachment] if source_attachment else [],
        annee=annee,
        groupe_controle_id=groupe_id,
        created_by=current_user.get("id"),
        updated_by=current_user.get("id")
    )
    
    item_dict = item.model_dump()
    await db.surveillance_items.insert_one(item_dict)
    item_dict.pop("_id", None)
    
    # Générer les récurrences futures
    if derniere_visite and periodicite:
        recurring = generate_recurring_controls(item_dict, derniere_visite, periodicite)
        if recurring:
            for r in recurring:
                r["created_by"] = current_user.get("id")
                r["updated_by"] = current_user.get("id")
            await db.surveillance_items.insert_many(recurring)
            logger.info(f"✅ {len(recurring)} récurrence(s) créée(s) pour {item.classe_type}")
    
    # Audit
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=f"{current_user['prenom']} {current_user['nom']}",
        user_email=current_user["email"],
        action=ActionType.CREATE,
        entity_type=EntityType.SURVEILLANCE,
        entity_id=item.id,
        entity_name=f"Plan surveillance (IA): {item.classe_type}"
    )
    
    return item_dict


async def _create_curative_work_order(
    ctrl: dict, document_info: dict,
    surveillance_item_id: str, current_user: dict
) -> Optional[dict]:
    """Crée un bon de travail curatif pour une non-conformité détectée."""
    try:
        from bson import ObjectId as BsonObjectId
        from routes.shared import get_next_work_order_numero
        
        wo_numero = await get_next_work_order_numero()
        
        wo_dict = {
            "_id": BsonObjectId(),
            "titre": f"[Curatif] {ctrl.get('classe_type', 'Contrôle')} - Non-conformité",
            "description": (
                f"Non-conformité détectée lors du contrôle réglementaire.\n\n"
                f"Organisme: {document_info.get('organisme_controle', 'N/A')}\n"
                f"Date du contrôle: {ctrl.get('derniere_visite', 'N/A')}\n"
                f"Rapport: {document_info.get('numero_rapport', 'N/A')}\n\n"
                f"ANOMALIES:\n{ctrl.get('anomalies', '')}"
            ),
            "statut": "OUVERT", "priorite": "HAUTE", "categorie": "TRAVAUX_CURATIF",
            "equipement_id": None, "assigne_a_id": None, "emplacement_id": None,
            "dateLimite": None, "tempsEstime": None, "tempsReel": None,
            "createdBy": current_user.get("id"), "numero": wo_numero,
            "dateCreation": datetime.now(timezone.utc), "dateTermine": None,
            "attachments": [], "comments": [], "parts_used": [],
            "service": None, "surveillance_item_id": surveillance_item_id
        }
        wo_dict["id"] = str(wo_dict["_id"])
        
        await db.work_orders.insert_one(wo_dict)
        
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.CREATE,
            entity_type=EntityType.WORK_ORDER,
            entity_id=wo_dict["id"],
            entity_name=wo_dict["titre"],
            details=f"BT curatif #{wo_numero} créé auto depuis contrôle IA"
        )
        
        return {
            "id": wo_dict["id"], "numero": wo_numero,
            "titre": wo_dict["titre"], "anomalies": ctrl.get("anomalies")
        }
    except Exception as e:
        logger.error(f"Erreur création BT curatif: {e}")
        return None


async def _archive_ai_analysis(
    items_data: dict, document_info: dict, controles: list,
    created_items: list, created_work_orders: list, matched_items: list,
    current_user: dict
) -> AIAnalysisHistory:
    """Archive l'analyse IA dans l'historique."""
    result_counts = {"CONFORME": 0, "NON_CONFORME": 0, "AVEC_RESERVES": 0}
    categories_set = set()
    for ctrl in controles:
        r = ctrl.get("resultat", "")
        if r in result_counts:
            result_counts[r] += 1
        cat = ctrl.get("category")
        if cat:
            categories_set.add(cat)
    
    entry = AIAnalysisHistory(
        filename=items_data.get("filename", "document.pdf"),
        file_size=items_data.get("file_size"),
        organisme_controle=document_info.get("organisme_controle"),
        date_intervention=document_info.get("date_intervention"),
        numero_rapport=document_info.get("numero_rapport"),
        site_controle=document_info.get("site_controle"),
        controles_count=len(created_items) + len(matched_items),
        conformes_count=result_counts["CONFORME"],
        non_conformes_count=result_counts["NON_CONFORME"],
        avec_reserves_count=result_counts["AVEC_RESERVES"],
        created_item_ids=[item["id"] for item in created_items],
        created_work_order_ids=[wo["id"] for wo in created_work_orders],
        raw_extracted_data=items_data,
        categories=list(categories_set),
        analyzed_by=current_user.get("id"),
        analyzed_by_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip()
    )
    
    entry_dict = entry.model_dump()
    await db.ai_analysis_history.insert_one(entry_dict)
    if "_id" in entry_dict:
        del entry_dict["_id"]
    
    return entry



@router.post("/ai/create-batch")
async def create_batch_from_ai(
    items_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Crée plusieurs contrôles de surveillance à partir des données extraites par l'IA.
    Gère le matching intelligent (mise à jour d'occurrences existantes) et la création.
    Crée automatiquement un bon de travail curatif pour chaque non-conformité.
    """
    try:
        controles = items_data.get("controles", [])
        document_info = items_data.get("document_info", {})
        source_attachment = _prepare_source_attachment(items_data.get("source_file"))
        
        created_items = []
        created_work_orders = []
        matched_items = []
        ambiguous_items = []
        errors = []
        
        for ctrl_index, ctrl in enumerate(controles):
          try:
            derniere_visite = ctrl.get("derniere_visite") or document_info.get("date_intervention")
            periodicite = _normalize_periodicite(ctrl.get("periodicite") or "Non déterminée")
            prochain_controle = _calculate_next_control_date(derniere_visite, periodicite)
            annee = get_year_from_date_str(derniere_visite) if derniere_visite else datetime.now().year
            if not annee:
                annee = datetime.now().year
            commentaire = _build_control_comment(ctrl)
            resultat = _map_resultat(ctrl.get("resultat"))
            
            # Chercher une occurrence existante qui correspond
            matched_occurrence = await find_matching_occurrence(ctrl, document_info, db)
            
            # CAS AMBIGUÏTÉ: confiance moyenne → demander à l'utilisateur
            if matched_occurrence and matched_occurrence.get("_match_confidence") == "medium":
                ambiguous_items.append({
                    "ctrl_index": ctrl_index,
                    "ctrl": ctrl,
                    "candidate": {
                        "id": matched_occurrence["id"],
                        "classe_type": matched_occurrence.get("classe_type"),
                        "category": matched_occurrence.get("category"),
                        "batiment": matched_occurrence.get("batiment"),
                        "prochain_controle": matched_occurrence.get("prochain_controle"),
                        "periodicite": matched_occurrence.get("periodicite"),
                        "executant": matched_occurrence.get("executant"),
                        "groupe_controle_id": matched_occurrence.get("groupe_controle_id"),
                        "match_score": matched_occurrence.get("_match_score"),
                    },
                    "report_date": derniere_visite,
                    "periodicite": periodicite,
                    "prochain_controle": prochain_controle,
                })
                continue
            
            # CAS MATCH: haute confiance → mise à jour automatique
            if matched_occurrence and matched_occurrence.get("_match_confidence") == "high":
                match_result = await _update_matched_occurrence(
                    matched_occurrence, ctrl, document_info, derniere_visite,
                    prochain_controle, periodicite, resultat, commentaire,
                    source_attachment, current_user
                )
                matched_items.append(match_result)
                continue
            
            # CAS CRÉATION: pas de correspondance → nouveau contrôle
            item_dict = await _create_new_surveillance_item(
                ctrl, document_info, derniere_visite, prochain_controle,
                periodicite, resultat, commentaire, source_attachment,
                annee, current_user
            )
            created_items.append(item_dict)
            
            # BT curatif si non-conformité
            if ctrl.get("resultat") in ("NON_CONFORME", "AVEC_RESERVES") and ctrl.get("anomalies"):
                wo = await _create_curative_work_order(ctrl, document_info, item_dict["id"], current_user)
                if wo:
                    created_work_orders.append(wo)
            
            # Broadcast WebSocket
            if realtime_manager:
                await realtime_manager.emit_event("surveillance_plans", "created", item_dict, user_id=current_user["id"])
                
          except Exception as ctrl_error:
            logger.error(f"Erreur création contrôle {ctrl_index} ({ctrl.get('classe_type', '?')}): {ctrl_error}")
            errors.append(f"Contrôle {ctrl_index + 1} ({ctrl.get('classe_type', '?')[:30]}): {str(ctrl_error)}")
        
        # Archiver l'analyse IA
        history_entry = await _archive_ai_analysis(
            items_data, document_info, controles,
            created_items, created_work_orders, matched_items, current_user
        )
        
        return {
            "success": True,
            "created_count": len(created_items),
            "matched_count": len(matched_items),
            "ambiguous_count": len(ambiguous_items),
            "created_items": created_items,
            "matched_items": matched_items,
            "ambiguous_items": ambiguous_items,
            "work_orders_created": created_work_orders,
            "analysis_id": history_entry.id,
            "errors": errors if errors else None,
            "message": (
                (f"{len(matched_items)} occurrence(s) mise(s) à jour, " if matched_items else "") +
                (f"{len(ambiguous_items)} correspondance(s) à confirmer, " if ambiguous_items else "") +
                f"{len(created_items)} contrôle(s) créé(s)" + 
                (f", {len(created_work_orders)} bon(s) de travail curatif(s)" if created_work_orders else "") +
                (f" ({len(errors)} erreur(s))" if errors else "")
            )
        }
    
    except Exception as e:
        logger.error(f"Erreur création batch surveillance: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Matching manuel (icône robot) ====================
@router.post("/items/{item_id}/analyze-report")
async def analyze_report_for_occurrence(
    item_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Analyse un rapport PDF/Excel/image via IA et met à jour l'occurrence spécifiée.
    Utilisé depuis l'icône robot sur chaque occurrence À PLANIFIER.
    """
    import tempfile
    import json as json_mod
    
    # Vérifier que l'item existe et est bien une occurrence non réalisée
    item = await db.surveillance_items.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item non trouvé")
    
    if item.get("status") == SurveillanceItemStatus.REALISE.value:
        raise HTTPException(status_code=400, detail="Cet item est déjà réalisé")
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Clé LLM non configurée")
        
        ext = os.path.splitext(file.filename)[1].lower()
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        # Préparer le contenu pour l'IA
        chat = LlmChat(
            api_key=api_key,
            session_id=f"report_match_{uuid.uuid4().hex[:8]}",
            system_message=f"""Tu analyses un rapport de contrôle pour mettre à jour un contrôle planifié.

Le contrôle planifié est:
- Type: {item.get('classe_type')}
- Catégorie: {item.get('category')}
- Bâtiment: {item.get('batiment', 'Non spécifié')}
- Date prévue: {item.get('prochain_controle')}
- Périodicité: {item.get('periodicite')}

Extrais du document:
1. La date de réalisation du contrôle
2. Le résultat (CONFORME, NON_CONFORME, AVEC_RESERVES)
3. L'organisme de contrôle
4. Le numéro de rapport
5. Les anomalies éventuelles
6. Les observations

Réponds UNIQUEMENT en JSON:
{{
  "date_realisation": "YYYY-MM-DD",
  "resultat": "CONFORME|NON_CONFORME|AVEC_RESERVES",
  "organisme_controle": "string ou null",
  "numero_rapport": "string ou null",
  "anomalies": "string ou null",
  "observations": "string ou null",
  "corresponds_to_planned": true ou false
}}

corresponds_to_planned = true si le rapport correspond bien au contrôle planifié ci-dessus."""
        ).with_model("gemini", "gemini-2.5-flash")
        
        # Envoyer le fichier ou le texte
        spreadsheet_formats = {".xlsx", ".xls", ".csv"}
        
        if ext in spreadsheet_formats:
            import openpyxl
            text_content = ""
            try:
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text_content += f"=== {sheet_name} ===\n"
                    for row in ws.iter_rows(values_only=False):
                        row_texts = [str(c.value).strip() for c in row if c.value is not None]
                        if row_texts:
                            text_content += " | ".join(row_texts) + "\n"
            except Exception:
                with open(tmp_path, 'r', errors='replace') as f:
                    text_content = f.read()
            
            response = await chat.send_message(
                UserMessage(text=f"Analyse ce rapport:\n\n{text_content[:15000]}")
            )
        else:
            mime_map = {".pdf": "application/pdf", ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".webp": "image/webp"}
            response = await chat.send_message(
                UserMessage(
                    text="Analyse ce rapport de contrôle.",
                    file_contents=[FileContentWithMimeType(file_path=tmp_path, mime_type=mime_map.get(ext, "application/octet-stream"))]
                )
            )
        
        try:
            os.unlink(tmp_path)
        except Exception:
            pass
        
        # Parser la réponse
        response_text = response.strip()
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        
        extracted = json_mod.loads(response_text.strip())
        
        if not extracted.get("corresponds_to_planned", True):
            return {
                "success": False,
                "message": "L'IA estime que ce rapport ne correspond pas au contrôle planifié",
                "extracted": extracted
            }
        
        date_realisation = extracted.get("date_realisation")
        if not date_realisation:
            raise HTTPException(status_code=422, detail="L'IA n'a pas pu extraire la date de réalisation")
        
        # Calculer l'écart
        ecart_jours = None
        try:
            date_prevue = datetime.fromisoformat(item["prochain_controle"]).date()
            date_reelle = datetime.fromisoformat(date_realisation).date()
            ecart_jours = (date_reelle - date_prevue).days
        except Exception:
            pass
        
        # Calculer le prochain contrôle
        periodicite = item.get("periodicite", "1 an")
        prochain = None
        try:
            from dateutil.relativedelta import relativedelta
            months = parse_periodicite_to_months(periodicite)
            base = datetime.fromisoformat(date_realisation)
            next_date = add_months_to_date(base, months)
            prochain = next_date.strftime("%Y-%m-%d")
        except Exception:
            pass
        
        # Résultat
        resultat_map = {"CONFORME": "Conforme", "NON_CONFORME": "Non conforme", "AVEC_RESERVES": "Avec réserves"}
        resultat = resultat_map.get(extracted.get("resultat"), extracted.get("resultat"))
        
        # Upload le fichier en pièce jointe
        source_attachment = {
            "id": str(uuid.uuid4()),
            "filename": file.filename,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Mettre à jour l'occurrence
        update_fields = {
            "status": SurveillanceItemStatus.REALISE.value,
            "derniere_visite": date_realisation,
            "date_realisation": date_realisation,
            "prochain_controle": prochain or item.get("prochain_controle"),
            "ecart_jours": ecart_jours,
            "resultat_controle": resultat,
            "numero_rapport": extracted.get("numero_rapport"),
            "organisme_controle": extracted.get("organisme_controle"),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.get("id")
        }
        
        if extracted.get("anomalies"):
            update_fields["commentaire"] = f"ANOMALIES:\n{extracted['anomalies']}"
        if extracted.get("observations"):
            existing = update_fields.get("commentaire", "")
            update_fields["commentaire"] = (existing + f"\n\nOBSERVATIONS:\n{extracted['observations']}").strip()
        
        await db.surveillance_items.update_one(
            {"id": item_id},
            {"$set": update_fields, "$push": {"attachments": source_attachment}}
        )
        
        # Régénérer les occurrences futures
        group_id = item.get("groupe_controle_id")
        if group_id and date_realisation and periodicite:
            await db.surveillance_items.delete_many({
                "groupe_controle_id": group_id,
                "status": {"$in": [SurveillanceItemStatus.PLANIFIER.value, SurveillanceItemStatus.PLANIFIE.value]},
                "id": {"$ne": item_id}
            })
            
            updated_item = await db.surveillance_items.find_one({"id": item_id}, {"_id": 0})
            if updated_item:
                recurring = generate_recurring_controls(updated_item, date_realisation, periodicite)
                if recurring:
                    for r in recurring:
                        r["created_by"] = current_user.get("id")
                    await db.surveillance_items.insert_many(recurring)
        
        # Émettre l'événement temps réel
        await realtime_manager.emit_event("surveillance_plans", "updated", update_fields, exclude_user=current_user.get("id"))
        
        return {
            "success": True,
            "message": f"Occurrence mise à jour en RÉALISÉ (écart: {ecart_jours:+d}j)" if ecart_jours is not None else "Occurrence mise à jour en RÉALISÉ",
            "item_id": item_id,
            "ecart_jours": ecart_jours,
            "extracted": extracted
        }
    
    except json_mod.JSONDecodeError as e:
        raise HTTPException(status_code=422, detail=f"L'IA n'a pas retourné un JSON valide: {str(e)}")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur analyse rapport pour occurrence: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Confirmation d'une correspondance ambiguë ====================

@router.post("/ai/confirm-match")
async def confirm_ambiguous_match(
    match_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Confirme manuellement une correspondance ambiguë détectée par l'IA.
    L'utilisateur choisit de matcher le rapport à l'occurrence planifiée proposée,
    ou de créer un nouveau contrôle.
    """
    try:
        action = match_data.get("action")  # "match" ou "create_new"
        item_id = match_data.get("item_id")  # ID de l'occurrence planifiée
        ctrl = match_data.get("ctrl", {})
        document_info = match_data.get("document_info", {})
        report_date = match_data.get("report_date")
        periodicite = match_data.get("periodicite", "1 an")
        prochain_controle = match_data.get("prochain_controle")
        source_file = match_data.get("source_file")
        
        if action == "match" and item_id:
            # Matcher : mettre à jour l'occurrence existante
            item = await db.surveillance_items.find_one({"id": item_id}, {"_id": 0})
            if not item:
                raise HTTPException(status_code=404, detail="Occurrence non trouvée")
            
            # Calculer l'écart
            ecart_jours = None
            try:
                date_prevue = datetime.fromisoformat(item["prochain_controle"]).date()
                date_reelle = datetime.fromisoformat(report_date).date()
                ecart_jours = (date_reelle - date_prevue).days
            except Exception:
                pass
            
            # Résultat
            resultat_map = {"CONFORME": "Conforme", "NON_CONFORME": "Non conforme", "AVEC_RESERVES": "Avec réserves"}
            resultat = resultat_map.get(ctrl.get("resultat"), ctrl.get("resultat"))
            
            # Commentaire
            commentaire_parts = []
            if ctrl.get("anomalies"):
                commentaire_parts.append(f"ANOMALIES DÉTECTÉES:\n{ctrl['anomalies']}")
            if ctrl.get("equipements_concernes"):
                commentaire_parts.append(f"Équipements: {ctrl['equipements_concernes']}")
            commentaire = "\n\n".join(commentaire_parts) if commentaire_parts else None
            
            update_data = {
                "status": SurveillanceItemStatus.REALISE.value,
                "derniere_visite": report_date,
                "date_realisation": report_date,
                "prochain_controle": prochain_controle,
                "ecart_jours": ecart_jours,
                "resultat_controle": resultat,
                "numero_rapport": document_info.get("numero_rapport"),
                "organisme_controle": document_info.get("organisme_controle"),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user.get("id")
            }
            if commentaire:
                update_data["commentaire"] = commentaire
            
            await db.surveillance_items.update_one({"id": item_id}, {"$set": update_data})
            
            # Régénérer les occurrences futures
            group_id = item.get("groupe_controle_id")
            if group_id and report_date and periodicite:
                await db.surveillance_items.delete_many({
                    "groupe_controle_id": group_id,
                    "status": {"$in": [SurveillanceItemStatus.PLANIFIER.value, SurveillanceItemStatus.PLANIFIE.value]},
                    "id": {"$ne": item_id}
                })
                updated_item = await db.surveillance_items.find_one({"id": item_id}, {"_id": 0})
                if updated_item:
                    recurring = generate_recurring_controls(updated_item, report_date, periodicite)
                    if recurring:
                        for r in recurring:
                            r["created_by"] = current_user.get("id")
                        await db.surveillance_items.insert_many(recurring)
            
            return {
                "success": True,
                "action": "matched",
                "item_id": item_id,
                "ecart_jours": ecart_jours,
                "message": f"Occurrence mise à jour (écart: {ecart_jours:+d}j)" if ecart_jours is not None else "Occurrence mise à jour"
            }
        
        elif action == "create_new":
            # Créer un nouveau contrôle
            groupe_id = str(uuid.uuid4())
            annee = get_year_from_date_str(report_date) if report_date else datetime.now().year
            
            resultat_map = {"CONFORME": "Conforme", "NON_CONFORME": "Non conforme", "AVEC_RESERVES": "Avec réserves"}
            resultat = resultat_map.get(ctrl.get("resultat"), ctrl.get("resultat"))
            
            commentaire_parts = []
            if ctrl.get("anomalies"):
                commentaire_parts.append(f"ANOMALIES DÉTECTÉES:\n{ctrl['anomalies']}")
            if ctrl.get("equipements_concernes"):
                commentaire_parts.append(f"Équipements: {ctrl['equipements_concernes']}")
            commentaire = "\n\n".join(commentaire_parts) if commentaire_parts else None
            
            new_item = SurveillanceItem(
                classe_type=ctrl.get("classe_type", ""),
                category=ctrl.get("category", "AUTRE"),
                batiment=ctrl.get("batiment") or "",
                periodicite=periodicite,
                responsable=SurveillanceResponsible.EXTERNE,
                executant=ctrl.get("executant", document_info.get("organisme_controle", "")),
                description=ctrl.get("description"),
                derniere_visite=report_date,
                prochain_controle=prochain_controle,
                status=SurveillanceItemStatus.REALISE,
                date_realisation=report_date,
                commentaire=commentaire,
                reference_reglementaire=ctrl.get("references_reglementaires"),
                numero_rapport=document_info.get("numero_rapport"),
                organisme_controle=document_info.get("organisme_controle"),
                resultat_controle=resultat,
                annee=annee or datetime.now().year,
                groupe_controle_id=groupe_id,
                created_by=current_user.get("id"),
                updated_by=current_user.get("id")
            )
            
            item_dict = new_item.model_dump()
            await db.surveillance_items.insert_one(item_dict)
            if "_id" in item_dict:
                del item_dict["_id"]
            
            # Générer récurrences
            if report_date and periodicite:
                recurring = generate_recurring_controls(item_dict, report_date, periodicite)
                if recurring:
                    for r in recurring:
                        r["created_by"] = current_user.get("id")
                    await db.surveillance_items.insert_many(recurring)
            
            return {
                "success": True,
                "action": "created",
                "item": item_dict,
                "message": "Nouveau contrôle créé"
            }
        
        else:
            raise HTTPException(status_code=400, detail="Action invalide. Utilisez 'match' ou 'create_new'.")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur confirmation match: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# ==================== Historique des analyses IA (Phase 1 & 2) ====================

@router.get("/ai/history")
async def get_ai_analysis_history(
    limit: int = 50,
    skip: int = 0,
    organisme: Optional[str] = None,
    category: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer l'historique des analyses IA"""
    try:
        query = {}
        if organisme:
            query["organisme_controle"] = {"$regex": organisme, "$options": "i"}
        if category:
            query["categories"] = category
        
        total = await db.ai_analysis_history.count_documents(query)
        items = await db.ai_analysis_history.find(
            query, {"_id": 0, "raw_extracted_data": 0}
        ).sort("created_at", -1).skip(skip).limit(limit).to_list(length=None)
        
        return {
            "total": total,
            "items": items
        }
    except Exception as e:
        logger.error(f"Erreur récupération historique IA: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ai/history/{analysis_id}")
async def get_ai_analysis_detail(
    analysis_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer le détail d'une analyse IA"""
    try:
        item = await db.ai_analysis_history.find_one({"id": analysis_id}, {"_id": 0})
        if not item:
            raise HTTPException(status_code=404, detail="Analyse non trouvée")
        return item
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur récupération détail analyse: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Tableau de bord & Tendances IA (Phase 3) ====================

@router.get("/ai/analytics")
async def get_ai_analytics(
    current_user: dict = Depends(get_current_user)
):
    """Statistiques et tendances des analyses IA"""
    try:
        all_analyses = await db.ai_analysis_history.find(
            {}, {"_id": 0, "raw_extracted_data": 0}
        ).sort("created_at", -1).to_list(length=None)
        
        total_analyses = len(all_analyses)
        if total_analyses == 0:
            return {
                "kpis": {
                    "total_analyses": 0,
                    "total_controles": 0,
                    "taux_conformite": 0,
                    "total_non_conformites": 0,
                    "total_work_orders": 0
                },
                "evolution_mensuelle": [],
                "par_organisme": [],
                "par_categorie": [],
                "par_resultat": [],
                "tendances_degradation": []
            }
        
        # KPIs globaux
        total_controles = sum(a.get("controles_count", 0) for a in all_analyses)
        total_conformes = sum(a.get("conformes_count", 0) for a in all_analyses)
        total_non_conformes = sum(a.get("non_conformes_count", 0) for a in all_analyses)
        total_avec_reserves = sum(a.get("avec_reserves_count", 0) for a in all_analyses)
        total_wo = sum(len(a.get("created_work_order_ids", [])) for a in all_analyses)
        
        taux_conformite = round((total_conformes / total_controles * 100) if total_controles > 0 else 0, 1)
        
        # Évolution mensuelle (12 derniers mois)
        evolution = {}
        now = datetime.now(timezone.utc)
        for i in range(11, -1, -1):
            month_date = now - timedelta(days=i * 30)
            key = month_date.strftime("%Y-%m")
            evolution[key] = {"mois": key, "analyses": 0, "controles": 0, "conformes": 0, "non_conformes": 0}
        
        for a in all_analyses:
            try:
                date_str = a.get("created_at", "")
                if date_str:
                    month_key = date_str[:7]  # YYYY-MM
                    if month_key in evolution:
                        evolution[month_key]["analyses"] += 1
                        evolution[month_key]["controles"] += a.get("controles_count", 0)
                        evolution[month_key]["conformes"] += a.get("conformes_count", 0)
                        evolution[month_key]["non_conformes"] += a.get("non_conformes_count", 0)
            except Exception:
                pass
        
        evolution_list = list(evolution.values())
        # Calculer taux conformité mensuel
        for e in evolution_list:
            total_m = e["conformes"] + e["non_conformes"]
            e["taux_conformite"] = round((e["conformes"] / total_m * 100) if total_m > 0 else 0, 1)
        
        # Par organisme
        organismes = {}
        for a in all_analyses:
            org = a.get("organisme_controle") or "Non précisé"
            if org not in organismes:
                organismes[org] = {"organisme": org, "analyses": 0, "controles": 0, "non_conformites": 0}
            organismes[org]["analyses"] += 1
            organismes[org]["controles"] += a.get("controles_count", 0)
            organismes[org]["non_conformites"] += a.get("non_conformes_count", 0)
        
        # Par catégorie
        categories = {}
        for a in all_analyses:
            for cat in a.get("categories", []):
                if cat not in categories:
                    categories[cat] = {"categorie": cat, "analyses": 0, "conformes": 0, "non_conformes": 0, "avec_reserves": 0}
                categories[cat]["analyses"] += 1
            for cat in a.get("categories", []):
                categories[cat]["conformes"] += a.get("conformes_count", 0)
                categories[cat]["non_conformes"] += a.get("non_conformes_count", 0)
                categories[cat]["avec_reserves"] += a.get("avec_reserves_count", 0)
        
        # Par résultat (pour pie chart)
        par_resultat = [
            {"id": "Conforme", "label": "Conforme", "value": total_conformes, "color": "#10b981"},
            {"id": "Non conforme", "label": "Non conforme", "value": total_non_conformes, "color": "#ef4444"},
            {"id": "Avec réserves", "label": "Avec réserves", "value": total_avec_reserves, "color": "#f59e0b"}
        ]
        
        # Phase 4: Détection de tendances de dégradation
        tendances = await _detect_degradation_trends(all_analyses)
        
        return {
            "kpis": {
                "total_analyses": total_analyses,
                "total_controles": total_controles,
                "taux_conformite": taux_conformite,
                "total_non_conformites": total_non_conformes + total_avec_reserves,
                "total_work_orders": total_wo
            },
            "evolution_mensuelle": evolution_list,
            "par_organisme": list(organismes.values()),
            "par_categorie": list(categories.values()),
            "par_resultat": [r for r in par_resultat if r["value"] > 0],
            "tendances_degradation": tendances
        }
    except Exception as e:
        logger.error(f"Erreur analytics IA: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Alertes intelligentes (Phase 4) ====================

async def _detect_degradation_trends(all_analyses):
    """Détecte les tendances de dégradation à partir de l'historique"""
    trends = []
    
    # Regrouper par catégorie + organisme pour voir les évolutions
    type_history = {}
    for a in sorted(all_analyses, key=lambda x: x.get("created_at", "")):
        for cat in a.get("categories", []):
            key = cat
            if key not in type_history:
                type_history[key] = []
            type_history[key].append({
                "date": a.get("created_at", "")[:10],
                "conformes": a.get("conformes_count", 0),
                "non_conformes": a.get("non_conformes_count", 0),
                "avec_reserves": a.get("avec_reserves_count", 0),
                "organisme": a.get("organisme_controle")
            })
    
    for cat, history in type_history.items():
        if len(history) < 2:
            continue
        
        # Vérifier les 2 dernières analyses
        recent = history[-2:]
        last_nc = recent[-1]["non_conformes"] + recent[-1]["avec_reserves"]
        prev_nc = recent[-2]["non_conformes"] + recent[-2]["avec_reserves"]
        
        if last_nc > 0 and prev_nc > 0:
            trends.append({
                "type": "degradation_consecutive",
                "severity": "HAUTE",
                "categorie": cat,
                "message": f"Non-conformités consécutives sur {cat} ({recent[-2]['date']} et {recent[-1]['date']})",
                "details": f"Dernière analyse: {last_nc} NC, Précédente: {prev_nc} NC",
                "last_date": recent[-1]["date"]
            })
        elif last_nc > prev_nc and last_nc > 0:
            trends.append({
                "type": "degradation_increase",
                "severity": "MOYENNE",
                "categorie": cat,
                "message": f"Augmentation des non-conformités sur {cat}",
                "details": f"Passage de {prev_nc} à {last_nc} non-conformité(s)",
                "last_date": recent[-1]["date"]
            })
    
    return trends


@router.get("/ai/alerts")
async def get_ai_smart_alerts(
    current_user: dict = Depends(get_current_user)
):
    """Récupérer les alertes intelligentes basées sur l'historique IA"""
    try:
        all_analyses = await db.ai_analysis_history.find(
            {}, {"_id": 0, "raw_extracted_data": 0}
        ).sort("created_at", -1).to_list(length=None)
        
        alerts = []
        
        # 1. Tendances de dégradation
        tendances = await _detect_degradation_trends(all_analyses)
        for t in tendances:
            alerts.append({
                "type": "degradation",
                "severity": t["severity"],
                "title": t["message"],
                "details": t["details"],
                "categorie": t["categorie"],
                "date": t["last_date"]
            })
        
        # 2. Catégories avec taux de conformité bas
        cat_stats = {}
        for a in all_analyses:
            for cat in a.get("categories", []):
                if cat not in cat_stats:
                    cat_stats[cat] = {"conformes": 0, "total": 0}
                cat_stats[cat]["conformes"] += a.get("conformes_count", 0)
                cat_stats[cat]["total"] += a.get("controles_count", 0)
        
        for cat, stats in cat_stats.items():
            if stats["total"] >= 2:
                taux = (stats["conformes"] / stats["total"] * 100) if stats["total"] > 0 else 0
                if taux < 70:
                    alerts.append({
                        "type": "low_conformity",
                        "severity": "HAUTE" if taux < 50 else "MOYENNE",
                        "title": f"Taux de conformité bas: {cat} ({round(taux, 1)}%)",
                        "details": f"{stats['conformes']}/{stats['total']} contrôles conformes",
                        "categorie": cat,
                        "date": None
                    })
        
        # 3. Vérifier les contrôles non-conformes sans BT curatif
        recent_nc = [a for a in all_analyses if a.get("non_conformes_count", 0) > 0 and not a.get("created_work_order_ids")]
        for a in recent_nc[:5]:
            alerts.append({
                "type": "missing_wo",
                "severity": "HAUTE",
                "title": "Non-conformité sans bon de travail curatif",
                "details": f"Analyse du {a.get('created_at', '')[:10]} - {a.get('organisme_controle', 'N/A')}",
                "categorie": ", ".join(a.get("categories", [])),
                "date": a.get("created_at", "")[:10]
            })
        
        # Trier par sévérité
        severity_order = {"HAUTE": 0, "MOYENNE": 1, "BASSE": 2}
        alerts.sort(key=lambda x: severity_order.get(x["severity"], 3))
        
        return {
            "count": len(alerts),
            "alerts": alerts
        }
    except Exception as e:
        logger.error(f"Erreur alertes intelligentes: {e}")
        raise HTTPException(status_code=500, detail=str(e))

async def send_surveillance_reminder_email(user_email: str, user_name: str, item: dict):
    """
    Envoie un email de rappel d'échéance pour un contrôle de surveillance.
    """
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    import os
    
    try:
        # Configuration SMTP
        smtp_server = os.environ.get("SMTP_HOST", "smtp.gmail.com")
        smtp_port = int(os.environ.get("SMTP_PORT", 587))
        smtp_user = os.environ.get("SMTP_USER", "")
        smtp_password = os.environ.get("SMTP_PASSWORD", "")
        
        if not smtp_user or not smtp_password:
            logger.warning("Configuration SMTP manquante pour l'envoi d'email de rappel surveillance")
            return False
        
        # Extraire les informations du contrôle
        classe_type = item.get("classe_type", "Contrôle")
        batiment = item.get("batiment", "Non spécifié")
        prochain_controle = item.get("prochain_controle", "Non spécifié")
        
        # Formater la date
        if prochain_controle and prochain_controle != "Non spécifié":
            try:
                date_obj = datetime.fromisoformat(prochain_controle)
                prochain_controle_formatted = date_obj.strftime("%d/%m/%Y")
            except:
                prochain_controle_formatted = prochain_controle
        else:
            prochain_controle_formatted = "Non spécifié"
        
        # Créer le message
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = user_email
        msg['Subject'] = f"[FSAO] Rappel - Contrôle à venir : {classe_type}"
        
        # Corps du message HTML
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="background-color: #3b82f6; color: white; padding: 20px; text-align: center; border-radius: 8px 8px 0 0;">
                    <h1 style="margin: 0; font-size: 24px;">📋 Rappel de Contrôle</h1>
                </div>
                
                <div style="background-color: #f8fafc; padding: 30px; border: 1px solid #e2e8f0; border-radius: 0 0 8px 8px;">
                    <p>Bonjour {user_name},</p>
                    
                    <p>Ce message vous est envoyé pour vous informer qu'un contrôle du plan de surveillance arrive à échéance prochainement.</p>
                    
                    <div style="background-color: white; border-left: 4px solid #3b82f6; padding: 15px; margin: 20px 0; border-radius: 4px;">
                        <h3 style="margin: 0 0 10px 0; color: #1e40af;">📌 Détails du contrôle</h3>
                        <table style="width: 100%;">
                            <tr>
                                <td style="padding: 5px 0; font-weight: bold; width: 150px;">Nom du contrôle :</td>
                                <td style="padding: 5px 0;">{classe_type}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px 0; font-weight: bold;">Équipement/Bâtiment :</td>
                                <td style="padding: 5px 0;">{batiment}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px 0; font-weight: bold;">Date d'échéance :</td>
                                <td style="padding: 5px 0; color: #dc2626; font-weight: bold;">{prochain_controle_formatted}</td>
                            </tr>
                        </table>
                    </div>
                    
                    <p>Merci de prendre les dispositions nécessaires pour planifier ce contrôle dans les délais.</p>
                    
                    <p style="color: #64748b; font-size: 12px; margin-top: 30px;">
                        Ce message est généré automatiquement par le système FSAO.<br>
                        Merci de ne pas répondre à cet email.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Envoyer l'email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)
        server.quit()
        
        logger.info(f"✅ Email de rappel surveillance envoyé à {user_email} pour le contrôle {classe_type}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Erreur envoi email rappel surveillance: {str(e)}")
        return False


async def check_surveillance_reminders():
    """
    Vérifie les contrôles de surveillance et envoie des emails de rappel.
    Cette fonction est appelée quotidiennement par le scheduler.
    
    Logique:
    - Pour chaque item avec un responsable_notification_id et un prochain_controle
    - Si la date actuelle = prochain_controle - duree_rappel_echeance
    - Et que l'email n'a pas déjà été envoyé (email_rappel_envoye = False)
    - Alors envoyer l'email et marquer email_rappel_envoye = True
    """
    try:
        logger.info("🔔 Vérification des rappels de surveillance...")
        today = datetime.now(timezone.utc).date()
        emails_sent = 0
        
        # Récupérer tous les items avec un responsable de notification
        items = await db.surveillance_items.find({
            "responsable_notification_id": {"$ne": None, "$exists": True},
            "email_rappel_envoye": {"$ne": True},
            "prochain_controle": {"$ne": None, "$exists": True}
        }).to_list(length=None)
        
        for item in items:
            try:
                # Vérifier si l'item a une date de prochain contrôle
                prochain_controle_str = item.get("prochain_controle")
                if not prochain_controle_str:
                    continue
                
                prochain_controle = datetime.fromisoformat(prochain_controle_str).date()
                duree_rappel = item.get("duree_rappel_echeance", 30)
                
                # Calculer la date de rappel
                date_rappel = prochain_controle - timedelta(days=duree_rappel)
                
                # Si aujourd'hui est le jour du rappel
                if today == date_rappel:
                    # Récupérer les informations de l'utilisateur
                    user_id = item.get("responsable_notification_id")
                    user = await db.users.find_one({"id": user_id})
                    
                    if not user:
                        # Essayer avec _id
                        from bson import ObjectId
                        try:
                            user = await db.users.find_one({"_id": ObjectId(user_id)})
                        except:
                            pass
                    
                    if user and user.get("email"):
                        user_name = f"{user.get('prenom', '')} {user.get('nom', '')}".strip() or user.get("email")
                        
                        # Envoyer l'email
                        success = await send_surveillance_reminder_email(
                            user_email=user.get("email"),
                            user_name=user_name,
                            item=item
                        )
                        
                        if success:
                            # Marquer l'email comme envoyé
                            await db.surveillance_items.update_one(
                                {"id": item["id"]},
                                {"$set": {
                                    "email_rappel_envoye": True,
                                    "alerte_date": datetime.now(timezone.utc).isoformat()
                                }}
                            )
                            emails_sent += 1
                    else:
                        logger.warning(f"Utilisateur {user_id} non trouvé ou sans email pour le contrôle {item.get('id')}")
                
            except Exception as e:
                logger.warning(f"Erreur traitement rappel item {item.get('id')}: {str(e)}")
                continue
        
        logger.info(f"🔔 {emails_sent} email(s) de rappel de surveillance envoyé(s)")
        return emails_sent
        
    except Exception as e:
        logger.error(f"❌ Erreur vérification rappels surveillance: {str(e)}")
        return 0
