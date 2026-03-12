from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Response, Request
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.exceptions import RequestValidationError
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.openapi.docs import get_swagger_ui_html, get_redoc_html
from fastapi.openapi.utils import get_openapi
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import aiofiles
import uuid
import mimetypes
import pandas as pd
import io
import secrets
import string
import asyncio
from pathlib import Path
from datetime import datetime, timedelta, timezone
from bson import ObjectId
import pytz
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from openapi_config import API_DESCRIPTION, OPENAPI_TAGS, STANDARD_ERRORS, CRUD_ERRORS, AUTH_ERRORS

# Import our models and dependencies
from models import *
# Rename pour éviter conflits avec realtime_events
EntityType_Audit = EntityType
from auth import get_password_hash, verify_password, create_access_token, decode_access_token
import dependencies
from dependencies import get_current_user, get_current_admin_user, check_permission, require_permission
import email_service
from audit_service import AuditService
from category_mapping import get_category_from_article_dm6

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'gmao_iris')]

# Initialize dependencies with database
dependencies.set_database(db)

# Initialize audit service
audit_service = AuditService(db)

# Create the main app (docs desactivees par defaut, servies manuellement avec auth)
app = FastAPI(
    title="FSAO Atlas API",
    description=API_DESCRIPTION,
    version="2.2.0",
    openapi_tags=OPENAPI_TAGS,
    docs_url=None,
    redoc_url=None,
    openapi_url="/api/openapi.json",
    contact={
        "name": "FSAO Atlas Support",
        "email": "support@gmao-atlas.fr"
    },
    license_info={
        "name": "Proprietary",
    }
)

# --- Protection docs Swagger par HTTP Basic Auth ---
docs_security = HTTPBasic()

def verify_docs_credentials(credentials: HTTPBasicCredentials = Depends(docs_security)):
    """Verifie les identifiants pour acceder a la documentation API"""
    correct_user = secrets.compare_digest(credentials.username, os.environ.get("DOCS_USER", "admin"))
    correct_pass = secrets.compare_digest(credentials.password, os.environ.get("DOCS_PASS", "atlas2024"))
    if not (correct_user and correct_pass):
        raise HTTPException(
            status_code=401,
            detail="Identifiants incorrects",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials

@app.get("/api/docs", include_in_schema=False)
async def custom_swagger_ui(credentials: HTTPBasicCredentials = Depends(verify_docs_credentials)):
    return get_swagger_ui_html(
        openapi_url="/api/openapi.json",
        title="FSAO Atlas - Documentation API",
        swagger_favicon_url="https://fastapi.tiangolo.com/img/favicon.png"
    )

@app.get("/api/redoc", include_in_schema=False)
async def custom_redoc(credentials: HTTPBasicCredentials = Depends(verify_docs_credentials)):
    return get_redoc_html(
        openapi_url="/api/openapi.json",
        title="FSAO Atlas - Documentation API (ReDoc)",
        redoc_favicon_url="https://fastapi.tiangolo.com/img/favicon.png"
    )

# Gestionnaire d'erreur pour les erreurs de validation Pydantic
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    """Capturer et logger les erreurs de validation Pydantic"""
    logger.error("=" * 80)
    logger.error("❌ ERREUR DE VALIDATION PYDANTIC 422")
    logger.error(f"📍 URL: {request.method} {request.url.path}")
    logger.error(f"📦 Body reçu: {await request.body()}")
    logger.error(f"🔍 Erreurs de validation: {exc.errors()}")
    logger.error("=" * 80)
    
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors(), "body": str(await request.body())}
    )

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Initialiser le scheduler pour les tâches automatiques
scheduler = AsyncIOScheduler()


# ==================== CRON JOB: GESTION DES MAINTENANCES PLANIFIÉES ====================

async def manage_planned_maintenance_status():
    """
    Cron job quotidien pour gérer les transitions de statut des maintenances planifiées.
    - Démarre les maintenances qui commencent aujourd'hui (statut → EN_MAINTENANCE)
    - Déclenche les emails de fin pour les maintenances qui se terminent aujourd'hui
    """
    try:
        from demande_arret_emails import send_end_maintenance_email
        
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        logger.info(f"🔄 [CRON] Gestion des maintenances planifiées pour le {today}...")
        
        started_count = 0
        ending_count = 0
        
        # 1. Trouver les maintenances qui DÉMARRENT aujourd'hui
        maintenances_starting = await db.planning_equipement.find({
            "date_debut": today,
            "maintenance_started": {"$ne": True}  # Pas encore démarrée
        }).to_list(length=None)
        
        for entry in maintenances_starting:
            eq_id = entry.get("equipement_id")
            
            # Mettre à jour le statut de l'équipement
            now = datetime.now(timezone.utc)
            rounded_hour = now.replace(minute=0, second=0, microsecond=0)
            
            await db.equipments.update_one(
                {"_id": ObjectId(eq_id)},
                {"$set": {
                    "statut": "EN_MAINTENANCE",
                    "statut_changed_at": rounded_hour,
                    "updated_at": now.isoformat()
                }}
            )
            
            # Enregistrer dans l'historique
            history_entry = {
                "equipment_id": eq_id,
                "statut": "EN_MAINTENANCE",
                "changed_at": rounded_hour,
                "changed_by": "cron_maintenance",
                "changed_by_name": "Maintenance planifiée (automatique)",
                "demande_arret_id": entry.get("demande_arret_id"),
                "is_start_of_maintenance": True
            }
            await db.equipment_status_history.update_one(
                {"equipment_id": eq_id, "changed_at": rounded_hour},
                {"$set": history_entry},
                upsert=True
            )
            
            # Marquer comme démarrée
            await db.planning_equipement.update_one(
                {"_id": entry["_id"]},
                {"$set": {"maintenance_started": True}}
            )
            
            started_count += 1
            logger.info(f"✅ [CRON] Maintenance démarrée pour équipement {eq_id}")
        
        # 2. Trouver les maintenances qui SE TERMINENT aujourd'hui
        maintenances_ending = await db.planning_equipement.find({
            "date_fin": today,
            "end_maintenance_email_sent": {"$ne": True}  # Email pas encore envoyé
        }).to_list(length=None)
        
        for entry in maintenances_ending:
            demande_id = entry.get("demande_arret_id")
            if not demande_id:
                continue
                
            # Récupérer la demande associée
            demande = await db.demandes_arret.find_one({"id": demande_id})
            if not demande:
                continue
            
            # Ne pas envoyer si déjà terminée ou email déjà envoyé
            if demande.get("statut") == "TERMINEE" or demande.get("end_maintenance_email_sent"):
                continue
            
            # Envoyer l'email de fin de maintenance
            equipement_noms = demande.get("equipement_noms", [])
            success = await send_end_maintenance_email(demande, equipement_noms)
            
            if success:
                # Marquer l'email comme envoyé
                await db.demandes_arret.update_one(
                    {"id": demande_id},
                    {"$set": {
                        "end_maintenance_email_sent": True,
                        "end_maintenance_email_sent_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                await db.planning_equipement.update_one(
                    {"_id": entry["_id"]},
                    {"$set": {"end_maintenance_email_sent": True}}
                )
                ending_count += 1
                logger.info(f"✅ [CRON] Email de fin de maintenance envoyé pour demande {demande_id}")
        
        logger.info(f"✅ [CRON] Terminé: {started_count} maintenance(s) démarrée(s), {ending_count} email(s) de fin envoyé(s)")
        
    except Exception as e:
        logger.error(f"❌ [CRON] Erreur gestion maintenances planifiées: {str(e)}")


# Fonction pour vérifier et créer automatiquement les bons de travail pour les maintenances échues
async def auto_check_preventive_maintenance():
    """Fonction exécutée automatiquement chaque jour pour vérifier les maintenances échues"""
    try:
        logger.info("🔄 Vérification automatique des maintenances préventives échues...")
        
        today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Trouver toutes les maintenances actives dont la date est aujourd'hui ou passée
        pm_list = await db.preventive_maintenances.find({
            "statut": "ACTIF",
            "prochaineMaintenance": {"$lte": today + timedelta(days=1)}
        }).to_list(length=None)
        
        created_count = 0
        updated_count = 0
        errors = []
        
        for pm in pm_list:
            try:
                # Récupérer l'équipement
                equipement = await db.equipments.find_one({"_id": ObjectId(pm["equipement_id"])})
                
                # Créer le bon de travail
                wo_id = str(uuid.uuid4())
                work_order = {
                    "_id": ObjectId(),
                    "id": wo_id,
                    "numero": f"PM-{datetime.utcnow().strftime('%Y%m%d')}-{secrets.token_hex(3).upper()}",
                    "titre": f"Maintenance préventive: {pm['titre']}",
                    "description": f"Maintenance automatique générée depuis la planification préventive '{pm['titre']}'",
                    "type": "PREVENTIF",
                    "priorite": "NORMALE",
                    "statut": "OUVERT",
                    "equipement_id": pm["equipement_id"],
                    "emplacement_id": equipement.get("emplacement_id") if equipement else None,
                    "assigne_a_id": pm.get("assigne_a_id"),
                    "tempsEstime": pm.get("duree"),
                    "dateLimite": datetime.utcnow() + timedelta(days=7),
                    "dateCreation": datetime.utcnow(),
                    "createdBy": "system-auto",
                    "comments": [],
                    "attachments": [],
                    "historique": []
                }
                
                await db.work_orders.insert_one(work_order)
                created_count += 1
                logger.info(f"✅ Bon de travail créé: {work_order['numero']} pour PM '{pm['titre']}'")
                
                # Calculer la prochaine date de maintenance
                next_date = calculate_next_maintenance_date(pm["prochaineMaintenance"], pm["frequence"])
                
                # Mettre à jour la maintenance préventive
                await db.preventive_maintenances.update_one(
                    {"_id": pm["_id"]},
                    {
                        "$set": {
                            "prochaineMaintenance": next_date,
                            "derniereMaintenance": datetime.utcnow()
                        }
                    }
                )
                updated_count += 1
                logger.info(f"✅ Prochaine maintenance mise à jour: {next_date.strftime('%Y-%m-%d')} (fréquence: {pm['frequence']})")
                
            except Exception as e:
                error_msg = f"Erreur pour PM '{pm.get('titre', 'Unknown')}': {str(e)}"
                errors.append(error_msg)
                logger.error(f"❌ {error_msg}")
        
        logger.info(f"✅ Vérification terminée: {created_count} bons créés, {updated_count} maintenances mises à jour, {len(errors)} erreurs")
        
    except Exception as e:
        logger.error(f"❌ Erreur lors de la vérification automatique des maintenances: {str(e)}")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Helper functions
def serialize_doc(doc, _is_root=True):
    """Convert MongoDB document to JSON serializable format"""
    if doc is None:
        return None
    
    # Convertir le _id principal
    if "_id" in doc:
        doc["id"] = str(doc["_id"])
        del doc["_id"]
    
    # Supprimer les champs sensibles si présents
    sensitive_fields = ["password", "hashed_password", "reset_token", "reset_token_created"]
    for field in sensitive_fields:
        if field in doc:
            del doc[field]
    
    # Convertir récursivement tous les ObjectId et types non sérialisables
    for key, value in list(doc.items()):
        if isinstance(value, ObjectId):
            doc[key] = str(value)
        elif isinstance(value, (int, float)) and key in ["telephone", "phone", "numero"]:
            # Convertir les numéros de téléphone et numéros en strings
            doc[key] = str(value)
        elif isinstance(value, list):
            doc[key] = [
                str(item) if isinstance(item, ObjectId) 
                else serialize_doc(item, _is_root=False) if isinstance(item, dict) 
                else str(item) if isinstance(item, (int, float)) and key in ["telephone", "phone", "numero"]
                else item 
                for item in value
            ]
        elif isinstance(value, dict):
            doc[key] = serialize_doc(value, _is_root=False)
    
    # Ajouter dateCreation et attachments uniquement au niveau racine du document
    if _is_root:
        if "dateCreation" not in doc:
            doc["dateCreation"] = datetime.utcnow()
        if "attachments" not in doc:
            doc["attachments"] = []
    
    return doc

async def get_user_by_id(user_id: str):
    """Get user details by ID"""
    try:
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if user:
            return {
                "id": str(user["_id"]),
                "nom": user.get("nom"),
                "prenom": user.get("prenom"),
                "email": user.get("email"),
                "role": user.get("role")
            }
    except:
        return None

async def get_location_by_id(location_id: str):
    """Get location details by ID"""
    try:
        location = await db.locations.find_one({"_id": ObjectId(location_id)})
        if location:
            return {
                "id": str(location["_id"]),
                "nom": location.get("nom")
            }
    except:
        return None

async def get_equipment_by_id(equipment_id: str):
    """Get equipment details by ID"""
    try:
        equipment = await db.equipments.find_one({"_id": ObjectId(equipment_id)})
        if equipment:
            return {
                "id": str(equipment["_id"]),
                "nom": equipment.get("nom")
            }
    except:
        return None

# ==================== AUTH ROUTES ====================
@api_router.post("/auth/register", response_model=User, tags=["Authentification"],
    summary="Inscrire un nouvel utilisateur",
    description="Cree un compte utilisateur avec les permissions par defaut selon le role choisi.",
    responses={**AUTH_ERRORS, 400: {"description": "Email deja utilise"}}
)
async def register(user_create: UserCreate):
    """Créer un nouveau compte utilisateur"""
    # Check if user already exists
    existing_user = await db.users.find_one({"email": user_create.email})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Un utilisateur avec cet email existe déjà"
        )
    
    # Hash password
    hashed_password = get_password_hash(user_create.password)
    
    # Définir les permissions par défaut selon le rôle (utilisation centralisée)
    permissions = get_default_permissions_by_role(user_create.role).model_dump()
    
    # Create user
    user_dict = user_create.model_dump()
    del user_dict["password"]
    user_dict["hashed_password"] = hashed_password
    user_dict["statut"] = "actif"
    user_dict["dateCreation"] = datetime.utcnow()
    user_dict["derniereConnexion"] = None
    user_dict["permissions"] = permissions
    user_dict["_id"] = ObjectId()
    
    await db.users.insert_one(user_dict)
    
    return User(**serialize_doc(user_dict))

@api_router.get("/version", response_model=VersionResponse, tags=["Systeme"],
    summary="Version de l'application",
    description="Retourne la version actuelle, le nom de version et la date de publication.",
    responses={200: {"description": "Version retournee avec succes", "content": {"application/json": {"example": {"version": "2.2.0", "versionName": "Documentation Enrichie", "releaseDate": "2025-01-18"}}}}}
)
async def get_version():
    """Obtenir la version actuelle de l'application (endpoint public).
    Lit la version depuis updates/version.json (source unique de vérité)."""
    try:
        version_file = os.path.join(os.path.dirname(__file__), "updates", "version.json")
        if not os.path.exists(version_file):
            version_file = os.path.join(os.path.dirname(__file__), "..", "updates", "version.json")
        if os.path.exists(version_file):
            import json as json_mod
            with open(version_file, 'r') as f:
                data = json_mod.load(f)
            return {
                "version": data.get("version", "1.0.0"),
                "versionName": data.get("versionName", ""),
                "releaseDate": data.get("releaseDate", "")
            }
    except Exception:
        pass
    return {
        "version": "1.0.0",
        "versionName": "",
        "releaseDate": ""
    }


@api_router.get("/bell-counts", tags=["Systeme"],
    summary="Compteurs pour l'icône cloche du header",
    description="Retourne les compteurs d'OT en attente, améliorations en attente et maintenances préventives échues."
)
async def get_bell_counts(current_user: dict = Depends(get_current_user)):
    """Compteurs pour les badges de la cloche du header."""
    now = datetime.utcnow()

    # 1. Ordres de travail en attente (statut EN_ATTENTE uniquement)
    wo_count = await db.work_orders.count_documents({
        "statut": "EN_ATTENTE"
    })

    # 2. Améliorations en attente (statut EN_ATTENTE uniquement)
    imp_count = await db.improvements.count_documents({
        "statut": "EN_ATTENTE"
    })

    # 3. Maintenances préventives planifiées mais pas encore réalisées (date dépassée)
    pm_count = await db.preventive_maintenances.count_documents({
        "statut": "ACTIF",
        "prochaineMaintenance": {"$lte": now}
    })

    return {
        "work_orders": wo_count,
        "improvements": imp_count,
        "preventive": pm_count
    }


@api_router.get("/dashboard/widget-data", tags=["Dashboard"])
async def get_dashboard_widget_data(current_user: dict = Depends(get_current_user)):
    """Retourne les donnees pour les widgets du dashboard principal."""
    from datetime import timedelta
    now = datetime.utcnow()
    thirty_days_ago = now - timedelta(days=30)
    seven_days_ago = now - timedelta(days=7)

    try:
        # Stock bas
        inventory = await db.inventory.find({}, {"_id": 0, "quantite": 1, "quantiteMin": 1, "seuil_alerte": 1, "stock_monitoring_enabled": 1}).to_list(5000)
        low_stock = 0
        out_of_stock = 0
        for item in inventory:
            if not item.get("stock_monitoring_enabled", True):
                continue
            qty = item.get("quantite", 0) or 0
            qty_min = item.get("quantiteMin", item.get("seuil_alerte", 0)) or 0
            if qty <= 0:
                out_of_stock += 1
            elif qty_min > 0 and qty <= qty_min:
                low_stock += 1

        # Incidents recents (presqu'accidents 30 derniers jours)
        thirty_days_str = thirty_days_ago.strftime("%Y-%m-%d")
        recent_incidents = await db.presqu_accident_items.count_documents({
            "date_incident": {"$gte": thirty_days_str}
        })
        # Total presqu'accidents
        total_incidents = await db.presqu_accident_items.count_documents({})

        # Maintenances a venir (7 prochains jours)
        seven_days_later = now + timedelta(days=7)
        upcoming_maintenance = await db.preventive_maintenances.count_documents({
            "statut": "ACTIF",
            "prochaineMaintenance": {"$gte": now, "$lte": seven_days_later}
        })
        # M.Prev en retard
        overdue_mprev = await db.preventive_maintenances.count_documents({
            "statut": "ACTIF",
            "prochaineMaintenance": {"$lte": now}
        })

        # Changements de statut recents (7 derniers jours)
        recent_changes = await db.equipment_status_history.count_documents({
            "changed_at": {"$gte": seven_days_ago.isoformat()}
        }) if await db.list_collection_names() else 0
        try:
            recent_changes = await db.equipment_status_history.count_documents({
                "changed_at": {"$gte": seven_days_ago.isoformat()}
            })
        except Exception:
            recent_changes = 0

        return {
            "low_stock": low_stock,
            "out_of_stock": out_of_stock,
            "recent_incidents_30d": recent_incidents,
            "total_incidents": total_incidents,
            "upcoming_maintenance_7d": upcoming_maintenance,
            "overdue_mprev": overdue_mprev,
            "recent_status_changes_7d": recent_changes
        }
    except Exception as e:
        logger.error(f"Erreur dashboard widget-data: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@api_router.post("/auth/login", response_model=Token, tags=["Authentification"],
    summary="Connexion utilisateur",
    description="Authentifie un utilisateur et retourne un token JWT valide 7 jours. Le token doit etre inclus dans le header `Authorization: Bearer <token>` pour les requetes protegees.",
    responses={401: {"description": "Identifiants invalides", "content": {"application/json": {"example": {"detail": "Identifiants invalides"}}}}}
)
async def login(login_request: LoginRequest):
    """Se connecter et obtenir un token JWT"""
    # Debug logging
    logger.info(f"🔍 LOGIN ATTEMPT - Email: {login_request.email}")
    
    # Find user
    user = await db.users.find_one({"email": login_request.email})
    logger.info(f"🔍 User found in DB: {user is not None}")
    
    if not user:
        logger.warning(f"❌ User not found for email: {login_request.email}")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email ou mot de passe incorrect"
        )
    
    # Verify password
    logger.info(f"🔍 Attempting password verification...")
    logger.info(f"   Password length: {len(login_request.password)}")
    
    # Support both 'password' and 'hashed_password' field names
    password_hash = user.get("hashed_password") or user.get("password")
    if not password_hash:
        logger.warning(f"❌ No password hash found for email: {login_request.email}")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email ou mot de passe incorrect"
        )
    
    logger.info(f"   Hash prefix: {password_hash[:20]}...")
    password_valid = verify_password(login_request.password, password_hash)
    logger.info(f"🔍 Password valid: {password_valid} (type: {type(password_valid)})")
    
    if not password_valid:
        logger.warning(f"❌ Invalid password for email: {login_request.email}")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Email ou mot de passe incorrect"
        )
    
    # Update last login
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {"derniereConnexion": datetime.utcnow()}}
    )
    
    # Log dans l'audit
    await audit_service.log_action(
        user_id=user.get("id", str(user["_id"])),
        user_name=f"{user['prenom']} {user['nom']}",
        user_email=user["email"],
        action=ActionType.LOGIN,
        entity_type=EntityType_Audit.USER,
        entity_id=user.get("id", str(user["_id"])),
        entity_name=f"{user['prenom']} {user['nom']}"
    )
    
    # Create access token (valide 1 heure)
    access_token = create_access_token(
        data={"sub": str(user["_id"])},
        expires_delta=timedelta(hours=1)
    )
    
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=User(**serialize_doc(user))
    )

@api_router.get("/auth/me", response_model=User, tags=["Authentification"],
    summary="Profil utilisateur connecte",
    description="Retourne les informations completes de l'utilisateur actuellement authentifie.",
    responses={**STANDARD_ERRORS}
)
async def get_me(current_user: dict = Depends(get_current_user)):
    """Obtenir l'utilisateur connecté"""
    return User(**current_user)


@api_router.post("/auth/forgot-password", response_model=MessageResponse, tags=["Authentification"],
    summary="Mot de passe oublie",
    description="Envoie un email avec un lien de reinitialisation du mot de passe. Le token est valable 1 heure.",
    responses={200: {"description": "Email envoye (meme si l'adresse n'existe pas, pour des raisons de securite)"}}
)
async def forgot_password(request: ForgotPasswordRequest):
    """Demander une réinitialisation de mot de passe"""
    # Vérifier si l'utilisateur existe
    user = await db.users.find_one({"email": request.email})
    
    if user:
        # Créer un token de réinitialisation (valide 1 heure)
        reset_token = create_access_token(
            data={"sub": str(user["_id"]), "type": "reset"},
            expires_delta=timedelta(hours=1)
        )
        
        # Construire l'URL de réinitialisation
        APP_URL = os.environ.get('APP_URL', 'http://localhost:3000')
        reset_url = f"{APP_URL}/reset-password?token={reset_token}"
        
        # Envoyer l'email de réinitialisation
        try:
            email_sent = email_service.send_password_reset_email(
                to_email=request.email,
                prenom=user.get('prenom', 'Utilisateur'),
                reset_url=reset_url
            )
            
            if email_sent:
                logger.info(f"Email de réinitialisation envoyé à {request.email}")
            else:
                logger.error(f"Échec de l'envoi de l'email de réinitialisation à {request.email}")
        except Exception as email_error:
            logger.error(f"Erreur lors de l'envoi de l'email de réinitialisation : {str(email_error)}")
        
        # Sauvegarder le token dans la base (pour invalider après usage)
        await db.users.update_one(
            {"_id": user["_id"]},
            {"$set": {"reset_token": reset_token, "reset_token_created": datetime.utcnow()}}
        )
    
    # Toujours retourner succès pour ne pas révéler si l'email existe
    return {"message": "Si cet email existe, un lien de réinitialisation a été envoyé"}

@api_router.post("/auth/reset-password", response_model=MessageResponse, tags=["Authentification"],
    summary="Reinitialiser le mot de passe",
    description="Reinitialise le mot de passe avec un token recu par email. Le token expire apres 1 heure.",
    responses={400: {"description": "Token invalide ou expire"}}
)
async def reset_password(request: ResetPasswordRequest):
    """Réinitialiser le mot de passe avec un token"""
    try:
        # Vérifier le token
        payload = jwt.decode(request.token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        token_type = payload.get("type")
        
        if token_type != "reset":
            raise HTTPException(status_code=400, detail="Token invalide")
        
        # Trouver l'utilisateur
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
        
        # Vérifier que le token correspond (si sauvegardé)
        if user.get("reset_token") != request.token:
            raise HTTPException(status_code=400, detail="Token invalide ou déjà utilisé")
        
        # Hacher le nouveau mot de passe
        hashed_password = get_password_hash(request.new_password)
        
        # Mettre à jour le mot de passe et supprimer le token
        await db.users.update_one(
            {"_id": ObjectId(user_id)},
            {
                "$set": {"hashed_password": hashed_password},
                "$unset": {"reset_token": "", "reset_token_created": ""}
            }
        )
        
        return {"message": "Mot de passe réinitialisé avec succès"}
        
    except JWTError:
        raise HTTPException(status_code=400, detail="Token invalide ou expiré")


# ==================== INVITATION & REGISTRATION ROUTES ====================

def generate_temp_password(length: int = 12) -> str:
    """Génère un mot de passe temporaire aléatoire"""
    characters = string.ascii_letters + string.digits + "!@#$%"
    return ''.join(secrets.choice(characters) for _ in range(length))

@api_router.post("/users/invite-member", response_model=InviteMemberResponse, tags=["Utilisateurs"],
    summary="Inviter un membre",
    description="Cree un compte utilisateur et envoie un email d'invitation avec les identifiants. Necessite le role ADMIN.",
    responses={**STANDARD_ERRORS, 400: {"description": "Email deja utilise"}}
)
async def invite_member(request: InviteMemberRequest, current_user: dict = Depends(get_current_admin_user)):
    """
    Envoyer une invitation par email (Admin uniquement)
    L'utilisateur recevra un lien pour compléter son inscription
    """
    # Vérifier si l'email existe déjà
    existing_user = await db.users.find_one({"email": request.email})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Un utilisateur avec cet email existe déjà"
        )
    
    # Créer un token d'invitation (valide 7 jours)
    invitation_data = {
        "sub": request.email,
        "type": "invitation",
        "role": request.role,
        "invited_by": current_user.get("_id")
    }
    invitation_token = create_access_token(
        data=invitation_data,
        expires_delta=timedelta(days=7)
    )
    
    # Envoyer l'email d'invitation
    email_sent = email_service.send_invitation_email(
        to_email=request.email,
        token=invitation_token,
        role=request.role
    )
    
    if not email_sent:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors de l'envoi de l'email d'invitation"
        )
    
    # Log l'invitation
    logger.info(f"Invitation envoyée à {request.email} par {current_user.get('email')}")
    
    return {
        "message": f"Invitation envoyée à {request.email}",
        "email": request.email,
        "role": request.role
    }

@api_router.post("/users/create-member", response_model=User, tags=["Utilisateurs"])
async def create_member(request: CreateMemberRequest, current_user: dict = Depends(get_current_admin_user)):
    """
    Créer un membre directement avec mot de passe temporaire (Admin uniquement)
    L'utilisateur recevra un email avec ses identifiants
    """
    # Vérifier si l'email existe déjà
    existing_user = await db.users.find_one({"email": request.email})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Un utilisateur avec cet email existe déjà"
        )
    
    # Hasher le mot de passe fourni
    hashed_password = get_password_hash(request.password)
    
    # Obtenir les permissions par défaut selon le rôle
    default_permissions = get_default_permissions_by_role(request.role)
    permissions = default_permissions.model_dump()
    
    # Si des permissions personnalisées sont fournies, les utiliser
    if hasattr(request, 'permissions') and request.permissions:
        permissions = request.permissions
    
    # Créer l'utilisateur
    user_dict = {
        "id": str(uuid.uuid4()),
        "nom": request.nom,
        "prenom": request.prenom,
        "email": request.email,
        "telephone": request.telephone or "",
        "role": request.role,
        "service": request.service,
        "regime": request.regime if request.regime else "Journée",  # Régime de travail
        "hashed_password": hashed_password,
        "statut": "actif",
        "dateCreation": datetime.utcnow(),
        "derniereConnexion": datetime.utcnow(),
        "permissions": permissions,
        "firstLogin": True  # Doit changer son mot de passe à la première connexion
    }
    
    await db.users.insert_one(user_dict)
    
    # Émettre l'événement WebSocket pour la création
    try:
        from realtime_events import EntityType as RealtimeEntityType, EventType as RealtimeEventType
        await realtime_manager.emit_event(
            RealtimeEntityType.USERS.value,
            RealtimeEventType.CREATED.value,
            serialize_doc(user_dict),
            current_user.get("id")
        )
    except Exception as e:
        logger.error(f"Erreur émission événement WebSocket users create: {e}")
    
    # Envoyer l'email avec les identifiants
    email_sent = email_service.send_account_created_email(
        to_email=request.email,
        temp_password=request.password,
        prenom=request.prenom
    )
    
    if not email_sent:
        logger.warning(f"Email non envoyé à {request.email}, mais compte créé")
    
    logger.info(f"Membre créé: {request.email} par {current_user.get('email')}")
    
    return User(**serialize_doc(user_dict))

@api_router.get("/auth/validate-invitation/{token}", response_model=ValidateInvitationResponse, tags=["Authentification"],
    summary="Valider un token d'invitation",
    description="Verifie la validite d'un token d'invitation. Utilise lors du processus d'acceptation d'invitation.",
    responses={400: {"description": "Token invalide ou expire"}}
)
async def validate_invitation(token: str):
    """
    Valider un token d'invitation et retourner les informations
    """
    try:
        payload = decode_access_token(token)
        if not payload or payload.get("type") != "invitation":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Token d'invitation invalide"
            )
        
        # Vérifier que l'utilisateur n'existe pas déjà
        email = payload.get("sub")
        existing_user = await db.users.find_one({"email": email})
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cet utilisateur existe déjà"
            )
        
        return {
            "valid": True,
            "email": email,
            "role": payload.get("role")
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Token d'invitation invalide ou expiré"
        )

@api_router.post("/auth/complete-registration", response_model=User, tags=["Authentification"])
async def complete_registration(request: CompleteRegistrationRequest):
    """
    Compléter l'inscription après avoir reçu une invitation
    """
    try:
        # Valider le token
        payload = decode_access_token(request.token)
        if not payload or payload.get("type") != "invitation":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Token d'invitation invalide"
            )
        
        email = payload.get("sub")
        role = payload.get("role")
        
        # Vérifier que l'utilisateur n'existe pas déjà
        existing_user = await db.users.find_one({"email": email})
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cet utilisateur existe déjà"
            )
        
        # Hasher le mot de passe
        hashed_password = get_password_hash(request.password)
        
        # Obtenir les permissions par défaut selon le rôle
        default_permissions = get_default_permissions_by_role(role)
        permissions = default_permissions.model_dump()
        
        # Créer l'utilisateur
        user_dict = {
            "id": str(uuid.uuid4()),
            "nom": request.nom,
            "prenom": request.prenom,
            "email": email,
            "telephone": request.telephone or "",
            "role": role,
            "service": None,
            "hashed_password": hashed_password,
            "statut": "actif",
            "dateCreation": datetime.utcnow(),
            "derniereConnexion": datetime.utcnow(),
            "permissions": permissions,
            "firstLogin": False  # A déjà défini son mot de passe
        }
        
        await db.users.insert_one(user_dict)
        
        logger.info(f"Inscription complétée pour {email}")
        
        return User(**serialize_doc(user_dict))
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de la completion de l'inscription: {e}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Erreur lors de l'inscription"
        )

@api_router.post("/auth/change-password-first-login", response_model=MessageResponse, tags=["Authentification"],
    summary="Changer le mot de passe (premiere connexion)",
    description="Permet a un utilisateur invite de definir son mot de passe definitif lors de sa premiere connexion.",
    responses={**AUTH_ERRORS}
)
async def change_password_first_login(request: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    """
    Changer le mot de passe lors de la première connexion
    """
    user_id = current_user.get("id")  # Changé de "_id" à "id"
    
    # Vérifier l'ancien mot de passe
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    if not verify_password(request.old_password, user["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Mot de passe actuel incorrect"
        )
    
    # Hasher le nouveau mot de passe
    new_hashed_password = get_password_hash(request.new_password)
    
    # Mettre à jour le mot de passe et marquer firstLogin comme False
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {
            "$set": {
                "hashed_password": new_hashed_password,
                "firstLogin": False
            }
        }
    )
    
    logger.info(f"Mot de passe changé pour {user.get('email')}")
    
    return {"message": "Mot de passe changé avec succès"}


@api_router.get("/auth/me", response_model=User, tags=["Authentification"])
async def get_current_user_profile(current_user: dict = Depends(get_current_user)):
    """
    Récupérer le profil de l'utilisateur connecté
    """
    user_id = current_user.get("id")
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return User(**serialize_doc(user))


@api_router.put("/auth/me", tags=["Authentification"],
    summary="Mettre a jour le profil",
    description="Met a jour les informations du profil de l'utilisateur connecte (nom, prenom, email, telephone, photo).",
    responses={**STANDARD_ERRORS}
)
async def update_current_user_profile(user_update: UserProfileUpdate, current_user: dict = Depends(get_current_user)):
    """
    Mettre à jour le profil de l'utilisateur connecté
    """
    user_id = current_user.get("id")
    
    # Préparer les données à mettre à jour (exclure None)
    update_data = {k: v for k, v in user_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnée à mettre à jour")
    
    # Mettre à jour l'utilisateur
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": update_data}
    )
    
    # Récupérer l'utilisateur mis à jour
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    logger.info(f"Profil mis à jour pour {user.get('email')}")
    
    return {"message": "Profil mis à jour avec succès", "user": serialize_doc(user)}


@api_router.post("/auth/change-password", response_model=MessageResponse, tags=["Authentification"],
    summary="Changer le mot de passe",
    description="Permet a l'utilisateur connecte de changer son mot de passe en fournissant l'ancien et le nouveau.",
    responses={**AUTH_ERRORS, 400: {"description": "Ancien mot de passe incorrect"}}
)
async def change_password(request: ChangePasswordRequest, current_user: dict = Depends(get_current_user)):
    """
    Changer le mot de passe de l'utilisateur connecté
    """
    user_id = current_user.get("id")
    
    # Vérifier l'ancien mot de passe
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    if not verify_password(request.old_password, user["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Mot de passe actuel incorrect"
        )
    
    # Hasher le nouveau mot de passe
    new_hashed_password = get_password_hash(request.new_password)
    
    # Mettre à jour le mot de passe
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"hashed_password": new_hashed_password}}
    )
    
    logger.info(f"Mot de passe changé pour {user.get('email')}")
    
    return {"message": "Mot de passe changé avec succès"}


# ==================== SERVICE MANAGER ROUTES ====================

@api_router.get("/service-manager/status", tags=["Service Manager"],
    summary="Statut du service",
    description="Retourne le statut operationnel du service : equipements en maintenance, ordres de travail en cours, alertes actives.",
    responses={**STANDARD_ERRORS}
)
async def get_service_manager_status(current_user: dict = Depends(get_current_user)):
    """Vérifie si l'utilisateur est un responsable de service et retourne ses services"""
    from service_filter import is_service_manager, get_user_managed_services, get_user_service_filter
    
    is_manager = await is_service_manager(current_user)
    managed_services = await get_user_managed_services(current_user)
    service_filter = await get_user_service_filter(current_user)
    
    return {
        "is_service_manager": is_manager,
        "managed_services": managed_services,
        "service_filter": service_filter,
        "user_service": current_user.get("service"),
        "user_role": current_user.get("role")
    }


@api_router.get("/service-manager/team", tags=["Service Manager"],
    summary="Equipe du service",
    description="Retourne la liste des membres de l'equipe du service avec leur activite recente.",
    responses={**STANDARD_ERRORS}
)
async def get_service_team(current_user: dict = Depends(get_current_user)):
    """Récupère les membres de l'équipe du responsable de service"""
    from service_filter import is_service_manager, get_service_team_members
    
    is_manager = await is_service_manager(current_user)
    if not is_manager:
        raise HTTPException(status_code=403, detail="Vous n'êtes pas responsable de service")
    
    team = await get_service_team_members(current_user)
    
    # Nettoyer les données sensibles
    for member in team:
        member.pop("password", None)
        member.pop("hashed_password", None)
    
    return {
        "team_count": len(team),
        "team_members": team
    }


@api_router.get("/service-manager/stats", tags=["Service Manager"],
    summary="Statistiques du service",
    description="Retourne les KPIs du service : nombre d'OT, temps moyen de resolution, taux de completion.",
    responses={**STANDARD_ERRORS}
)
async def get_service_manager_stats(current_user: dict = Depends(get_current_user)):
    """Statistiques du service pour le responsable"""
    from service_filter import is_service_manager, get_user_service_filter
    
    is_manager = await is_service_manager(current_user)
    if not is_manager:
        raise HTTPException(status_code=403, detail="Vous n'êtes pas responsable de service")
    
    service_filter = await get_user_service_filter(current_user)
    
    # Construire la requête de filtrage
    query = {}
    if service_filter:
        query["service"] = service_filter
    
    # Statistiques des ordres de travail
    ot_total = await db.work_orders.count_documents(query)
    ot_en_cours = await db.work_orders.count_documents({**query, "status": "EN_COURS"})
    ot_en_attente = await db.work_orders.count_documents({**query, "status": "EN_ATTENTE"})
    ot_termines = await db.work_orders.count_documents({**query, "status": {"$in": ["TERMINE", "CLOTURE"]}})
    
    # Statistiques des équipements
    eq_total = await db.equipments.count_documents(query)
    eq_panne = await db.equipments.count_documents({**query, "status": "EN_PANNE"})
    
    # Statistiques des demandes d'intervention
    di_en_attente = await db.intervention_requests.count_documents({**query, "status": "EN_ATTENTE"})
    
    # Membres de l'équipe
    team_count = await db.users.count_documents(query) if service_filter else 0
    
    return {
        "service": service_filter or "Tous",
        "work_orders": {
            "total": ot_total,
            "en_cours": ot_en_cours,
            "en_attente": ot_en_attente,
            "termines": ot_termines,
            "taux_completion": round((ot_termines / ot_total * 100) if ot_total > 0 else 0, 1)
        },
        "equipments": {
            "total": eq_total,
            "en_panne": eq_panne,
            "taux_disponibilite": round(((eq_total - eq_panne) / eq_total * 100) if eq_total > 0 else 100, 1)
        },
        "demandes_intervention": {
            "en_attente": di_en_attente
        },
        "team": {
            "count": team_count
        }
    }


# ==================== WORK ORDERS ROUTES ====================
@api_router.get("/work-orders", response_model=List[WorkOrder], tags=["Ordres de Travail"],
    summary="Lister les ordres de travail",
    description="Retourne la liste des ordres de travail avec filtres optionnels par date. Supporte le filtrage par date de creation ou d'echeance.",
    responses={**STANDARD_ERRORS}
)
async def get_work_orders(
    date_debut: str = None,
    date_fin: str = None,
    date_type: str = "creation",
    current_user: dict = Depends(require_permission("workOrders", "view"))
):
    """Liste tous les ordres de travail avec filtrage par date"""
    
    query = {}
    
    # Filtrage par date
    if date_debut and date_fin:
        date_field = "dateCreation" if date_type == "creation" else "dateLimite"
        query[date_field] = {
            "$gte": datetime.fromisoformat(date_debut),
            "$lte": datetime.fromisoformat(date_fin)
        }
    
    work_orders = await db.work_orders.find(query).to_list(1000)
    
    # Mapping statut/priorite valides
    VALID_STATUTS = {"OUVERT", "EN_COURS", "EN_ATTENTE", "TERMINE"}
    STATUT_MAP = {"en_attente": "EN_ATTENTE", "en_cours": "EN_COURS", "ouvert": "OUVERT", "termine": "TERMINE"}
    VALID_PRIORITES = {"URGENTE", "HAUTE", "MOYENNE", "NORMALE", "BASSE", "AUCUNE"}
    
    # Populate references
    for wo in work_orders:
        # Serialiser le document pour convertir tous les types non JSON
        wo = serialize_doc(wo)
        
        # Normaliser statut et priorite en majuscules
        raw_statut = wo.get("statut", "")
        if raw_statut and raw_statut not in VALID_STATUTS:
            wo["statut"] = STATUT_MAP.get(raw_statut.lower(), raw_statut.upper() if raw_statut.upper() in VALID_STATUTS else "EN_ATTENTE")
        raw_prio = wo.get("priorite", "")
        if raw_prio and raw_prio.upper() in VALID_PRIORITES and raw_prio not in VALID_PRIORITES:
            wo["priorite"] = raw_prio.upper()
        
        # S'assurer que attachments existe et convertir les ObjectId
        if "attachments" not in wo:
            wo["attachments"] = []
        else:
            # Convertir tous les ObjectId dans attachments
            cleaned_attachments = []
            for att in wo["attachments"]:
                # Ignorer si att n'est pas un dict
                if not isinstance(att, dict):
                    continue
                if "_id" in att and isinstance(att["_id"], ObjectId):
                    att["_id"] = str(att["_id"])
                for key, value in att.items():
                    if isinstance(value, ObjectId):
                        att[key] = str(value)
                cleaned_attachments.append(att)
            wo["attachments"] = cleaned_attachments
        
        if wo.get("assigne_a_id"):
            wo["assigneA"] = await get_user_by_id(wo["assigne_a_id"])
        if wo.get("emplacement_id"):
            wo["emplacement"] = await get_location_by_id(wo["emplacement_id"])
        if wo.get("equipement_id"):
            wo["equipement"] = await get_equipment_by_id(wo["equipement_id"])
        
        # Ajouter le nom du créateur
        if wo.get("createdBy"):
            try:
                # Essayer de chercher par ObjectId
                creator = await db.users.find_one({"_id": ObjectId(wo["createdBy"])})
                if creator:
                    wo["createdByName"] = f"{creator.get('prenom', '')} {creator.get('nom', '')}".strip()
                else:
                    # Sinon essayer par le champ id (UUID)
                    creator = await db.users.find_one({"id": wo["createdBy"]})
                    if creator:
                        wo["createdByName"] = f"{creator.get('prenom', '')} {creator.get('nom', '')}".strip()
            except Exception as e:
                # Si ça échoue, laisser vide
                logger.error(f"Erreur lors de la recherche du créateur {wo.get('createdBy')}: {e}")
                pass
    
    # Ajouter un numero par défaut si manquant (pour compatibilité avec anciens ordres)
    for wo in work_orders:
        if "numero" not in wo or not wo["numero"]:
            wo["numero"] = "N/A"
    
    return [WorkOrder(**wo) for wo in work_orders]

@api_router.get("/work-orders/{wo_id}",
    summary="Detail d'un ordre de travail", response_model=WorkOrder, tags=["Ordres de Travail"])
async def get_work_order(wo_id: str, current_user: dict = Depends(require_permission("workOrders", "view"))):
    """Détails d'un ordre de travail"""
    try:
        # Chercher par le champ id d'abord
        wo = await db.work_orders.find_one({"id": wo_id})
        
        # Si pas trouvé, essayer par _id (ObjectId)
        if not wo:
            try:
                wo = await db.work_orders.find_one({"_id": ObjectId(wo_id)})
            except:
                pass
        
        if not wo:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        wo = serialize_doc(wo)
        
        # Normaliser statut/priorite
        VALID_STATUTS = {"OUVERT", "EN_COURS", "EN_ATTENTE", "TERMINE"}
        STATUT_MAP = {"en_attente": "EN_ATTENTE", "en_cours": "EN_COURS", "ouvert": "OUVERT", "termine": "TERMINE"}
        raw_statut = wo.get("statut", "")
        if raw_statut and raw_statut not in VALID_STATUTS:
            wo["statut"] = STATUT_MAP.get(raw_statut.lower(), raw_statut.upper() if raw_statut.upper() in VALID_STATUTS else "EN_ATTENTE")
        
        if wo.get("assigne_a_id"):
            wo["assigneA"] = await get_user_by_id(wo["assigne_a_id"])
        if wo.get("emplacement_id"):
            wo["emplacement"] = await get_location_by_id(wo["emplacement_id"])
        if wo.get("equipement_id"):
            wo["equipement"] = await get_equipment_by_id(wo["equipement_id"])
        
        # Ajouter le nom du créateur
        if wo.get("createdBy"):
            try:
                creator = await db.users.find_one({"_id": ObjectId(wo["createdBy"])})
                if creator:
                    wo["createdByName"] = f"{creator.get('prenom', '')} {creator.get('nom', '')}".strip()
                else:
                    creator = await db.users.find_one({"id": wo["createdBy"]})
                    if creator:
                        wo["createdByName"] = f"{creator.get('prenom', '')} {creator.get('nom', '')}".strip()
            except Exception:
                pass
        
        # Ajouter un numero par défaut si manquant
        if "numero" not in wo or not wo["numero"]:
            wo["numero"] = "N/A"
        
        return WorkOrder(**wo)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/work-orders", response_model=WorkOrder,
    summary="Creer un ordre de travail", tags=["Ordres de Travail"])
async def create_work_order(wo_create: WorkOrderCreate, current_user: dict = Depends(require_permission("workOrders", "edit"))):
    """Créer un nouvel ordre de travail"""
    # Generate numero
    count = await db.work_orders.count_documents({})
    numero = str(5800 + count + 1)
    
    wo_dict = wo_create.model_dump()
    wo_dict["numero"] = numero
    wo_dict["dateCreation"] = datetime.utcnow()
    wo_dict["tempsReel"] = None
    wo_dict["dateTermine"] = None
    wo_dict["attachments"] = []
    wo_dict["comments"] = []  # Initialiser les commentaires
    wo_dict["parts_used"] = []  # Initialiser les pièces utilisées
    wo_dict["createdBy"] = current_user.get("id")  # Ajouter le créateur
    wo_dict["_id"] = ObjectId()
    wo_dict["id"] = str(wo_dict["_id"])  # Stocker aussi le champ id pour les recherches
    
    await db.work_orders.insert_one(wo_dict)
    
    # Log dans l'audit
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=f"{current_user['prenom']} {current_user['nom']}",
        user_email=current_user["email"],
        action=ActionType.CREATE,
        entity_type=EntityType_Audit.WORK_ORDER,
        entity_id=wo_dict.get("id"),
        entity_name=wo_dict["titre"],
        details=f"Ordre de travail #{numero} créé"
    )
    
    wo = serialize_doc(wo_dict)
    if wo.get("assigne_a_id"):
        wo["assigneA"] = await get_user_by_id(wo["assigne_a_id"])
    if wo.get("emplacement_id"):
        wo["emplacement"] = await get_location_by_id(wo["emplacement_id"])
    if wo.get("equipement_id"):
        wo["equipement"] = await get_equipment_by_id(wo["equipement_id"])
    
    # Notification push si un utilisateur est assigne
    logger.info(f"[PUSH TRIGGER CREATE] assigne_a_id={wo_create.assigne_a_id}, current_user_id={current_user.get('id')}")
    if wo_create.assigne_a_id and wo_create.assigne_a_id != current_user.get("id"):
        from notifications import notify_work_order_assigned
        logger.info(f"[PUSH TRIGGER CREATE] Envoi notification a {wo_create.assigne_a_id}")
        asyncio.create_task(
            notify_work_order_assigned(
                db=db,
                work_order_id=wo.get("id", ""),
                work_order_title=wo_create.titre,
                work_order_numero=numero,
                assigned_user_id=wo_create.assigne_a_id
            )
        )
        # Web Push PWA
        asyncio.create_task(
            notify_work_order_assigned_web(db, wo, wo_create.assigne_a_id, current_user.get("id"))
        )
    elif wo_create.assigne_a_id == current_user.get("id"):
        logger.info(f"[PUSH TRIGGER CREATE] Auto-assignation, pas de notification")
    
    # Émettre événement temps réel
    from realtime_manager import realtime_manager
    from realtime_events import EntityType as RealtimeEntityType, EventType as RealtimeEventType
    await realtime_manager.emit_event(
        RealtimeEntityType.WORK_ORDERS.value,
        RealtimeEventType.CREATED.value,
        wo,
        user_id=current_user.get("id")
    )
    
    return WorkOrder(**wo)

@api_router.put("/work-orders/{wo_id}",
    summary="Modifier un ordre de travail", response_model=WorkOrder, tags=["Ordres de Travail"])
async def update_work_order(wo_id: str, wo_update: WorkOrderUpdate, current_user: dict = Depends(require_permission("workOrders", "edit"))):
    """Modifier un ordre de travail"""
    from dependencies import can_edit_work_order_status
    
    try:
        # Récupérer l'ordre de travail existant - essayer par string ID d'abord
        existing_wo = await db.work_orders.find_one({"id": wo_id})
        
        # Si pas trouvé, essayer par ObjectId
        if not existing_wo:
            try:
                existing_wo = await db.work_orders.find_one({"_id": ObjectId(wo_id)})
            except:
                pass
        
        if not existing_wo:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        existing_wo["id"] = str(existing_wo["_id"])
        
        # Vérifier les permissions
        user_role = current_user.get("role")
        user_id = current_user.get("id")
        created_by = existing_wo.get("createdBy")
        assigne_a_id = existing_wo.get("assigne_a_id")
        
        # Admin : peut tout modifier
        if user_role == "ADMIN":
            can_full_edit = True
        # Technicien : peut modifier ce qu'il a créé
        elif user_role == "TECHNICIEN":
            can_full_edit = (created_by == user_id)
        # Visualiseur assigné : peut seulement modifier le statut
        elif user_role == "VISUALISEUR":
            can_full_edit = False
            # Vérifier si le visualiseur est assigné
            if assigne_a_id != user_id:
                raise HTTPException(
                    status_code=403,
                    detail="Vous ne pouvez pas modifier cet ordre de travail"
                )
            # Le visualiseur ne peut modifier que le statut
            if wo_update.model_dump(exclude_unset=True).keys() != {'statut'}:
                raise HTTPException(
                    status_code=403,
                    detail="Les visualiseurs ne peuvent modifier que le statut"
                )
        else:
            raise HTTPException(status_code=403, detail="Permission refusée")
        
        # Si pas de permission complète et qu'on essaie de modifier autre chose que le statut
        if not can_full_edit:
            update_dict = wo_update.model_dump(exclude_unset=True)
            if len(update_dict) > 1 or (len(update_dict) == 1 and 'statut' not in update_dict):
                raise HTTPException(
                    status_code=403,
                    detail="Vous ne pouvez modifier que le statut de cet ordre de travail"
                )
        
        # Appliquer les modifications
        # Note: On utilise exclude_unset=True pour ne modifier que les champs explicitement envoyés
        # Mais on doit gérer le cas où assigne_a_id est explicitement mis à null (pour retirer l'assignation)
        update_data = {}
        sent_data = wo_update.model_dump(exclude_unset=True)
        for k, v in wo_update.model_dump().items():
            # Inclure le champ si:
            # 1. Il a une valeur non-None, OU
            # 2. Il a été explicitement envoyé (même si None) pour les champs qui peuvent être "vidés"
            if v is not None:
                update_data[k] = v
            elif k in sent_data and k in ['assigne_a_id', 'equipement_id', 'emplacement_id', 'dateLimite']:
                # Ces champs peuvent être explicitement mis à null pour les "vider"
                update_data[k] = None
        
        # Si on vide assigne_a_id, il faut aussi vider assigneA
        if 'assigne_a_id' in update_data and update_data['assigne_a_id'] is None:
            update_data['assigneA'] = None
        
        # Si on vide equipement_id, il faut aussi vider equipement
        if 'equipement_id' in update_data and update_data['equipement_id'] is None:
            update_data['equipement'] = None
        
        # Si on vide emplacement_id, il faut aussi vider emplacement
        if 'emplacement_id' in update_data and update_data['emplacement_id'] is None:
            update_data['emplacement'] = None
        
        if wo_update.statut == WorkOrderStatus.TERMINE and "dateTermine" not in update_data:
            update_data["dateTermine"] = datetime.utcnow()
        
        # Utiliser _id (toujours présent) pour le filtre MongoDB
        wo_filter = {"_id": existing_wo["_id"]}
        
        await db.work_orders.update_one(
            wo_filter,
            {"$set": update_data}
        )
        
        # Log dans l'audit
        changes_desc = ", ".join([f"{k}: {v}" for k, v in update_data.items()])
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType_Audit.WORK_ORDER,
            entity_id=existing_wo.get("id"),
            entity_name=existing_wo["titre"],
            details=f"Modifications: {changes_desc}",
            changes=update_data
        )
        
        wo = await db.work_orders.find_one(wo_filter)
        wo = serialize_doc(wo)
        
        # Normaliser statut
        VALID_STATUTS = {"OUVERT", "EN_COURS", "EN_ATTENTE", "TERMINE"}
        STATUT_MAP = {"en_attente": "EN_ATTENTE", "en_cours": "EN_COURS", "ouvert": "OUVERT", "termine": "TERMINE"}
        raw_statut = wo.get("statut", "")
        if raw_statut and raw_statut not in VALID_STATUTS:
            wo["statut"] = STATUT_MAP.get(raw_statut.lower(), raw_statut.upper() if raw_statut.upper() in VALID_STATUTS else "EN_ATTENTE")
        if "numero" not in wo or not wo["numero"]:
            wo["numero"] = "N/A"
        
        if wo.get("assigne_a_id"):
            wo["assigneA"] = await get_user_by_id(wo["assigne_a_id"])
        if wo.get("emplacement_id"):
            wo["emplacement"] = await get_location_by_id(wo["emplacement_id"])
        if wo.get("equipement_id"):
            wo["equipement"] = await get_equipment_by_id(wo["equipement_id"])
        
        # Notifications push
        from notifications import notify_work_order_assigned, notify_work_order_status_changed
        
        # Notification si changement d'assignation
        logger.info(f"[PUSH TRIGGER UPDATE] update_data keys={list(update_data.keys())}, assigne_a_id in update={update_data.get('assigne_a_id')}")
        if "assigne_a_id" in update_data and update_data.get("assigne_a_id"):
            new_assigne = update_data["assigne_a_id"]
            old_assigne = existing_wo.get("assigne_a_id")
            logger.info(f"[PUSH TRIGGER UPDATE] Assignation: old={old_assigne} -> new={new_assigne}, current={current_user.get('id')}")
            if new_assigne != old_assigne:
                logger.info(f"[PUSH TRIGGER UPDATE] Envoi notification assignation a {new_assigne}")
                asyncio.create_task(
                    notify_work_order_assigned(
                        db=db,
                        work_order_id=wo.get("id", ""),
                        work_order_title=existing_wo.get("titre", ""),
                        work_order_numero=existing_wo.get("numero", ""),
                        assigned_user_id=new_assigne
                    )
                )
                # Web Push PWA
                asyncio.create_task(
                    notify_work_order_assigned_web(db, wo, new_assigne, current_user.get("id"))
                )
        
        # Notification si changement de statut
        if "statut" in update_data and existing_wo.get("statut") != update_data["statut"]:
            notify_ids = []
            if existing_wo.get("createdBy"):
                notify_ids.append(str(existing_wo["createdBy"]))
            if existing_wo.get("assigne_a_id"):
                notify_ids.append(str(existing_wo["assigne_a_id"]))
            notify_ids = list(set(notify_ids) - {str(current_user.get("id"))})
            logger.info(f"[PUSH TRIGGER UPDATE] Statut change: {existing_wo.get('statut')} -> {update_data['statut']}, notify_ids={notify_ids}")
            if notify_ids:
                asyncio.create_task(
                    notify_work_order_status_changed(
                        db=db,
                        work_order_id=wo.get("id", ""),
                        work_order_title=existing_wo.get("titre", ""),
                        work_order_numero=existing_wo.get("numero", ""),
                        old_status=existing_wo.get("statut", ""),
                        new_status=update_data["statut"],
                        notify_user_ids=notify_ids
                    )
                )
                # Web Push PWA
                asyncio.create_task(
                    notify_work_order_status_changed_web(db, wo, existing_wo.get("statut", ""), update_data["statut"], current_user.get("id"))
                )
        
        # Émettre événement temps réel
        from realtime_manager import realtime_manager
        from realtime_events import EntityType as RealtimeEntityType, EventType as RealtimeEventType
        
        # Si changement de statut, émettre événement spécifique
        if "statut" in update_data and existing_wo.get("statut") != update_data["statut"]:
            await realtime_manager.emit_event(
                RealtimeEntityType.WORK_ORDERS.value,
                RealtimeEventType.STATUS_CHANGED.value,
                {
                    "id": wo["id"],
                    "old_status": existing_wo.get("statut"),
                    "new_status": update_data["statut"],
                    "work_order": wo
                },
                user_id=current_user.get("id")
            )
        
        # Émettre événement de mise à jour générale
        await realtime_manager.emit_event(
            RealtimeEntityType.WORK_ORDERS.value,
            RealtimeEventType.UPDATED.value,
            wo,
            user_id=current_user.get("id")
        )
        
        return WorkOrder(**wo)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.post("/work-orders/{wo_id}/add-time",
    summary="Ajouter du temps passe", response_model=WorkOrder, tags=["Ordres de Travail"])
async def add_time_to_work_order(wo_id: str, time_data: AddTimeSpent, current_user: dict = Depends(require_permission("workOrders", "edit"))):
    """Ajouter du temps passé à un ordre de travail"""
    try:
        # Récupérer l'ordre de travail existant
        existing_wo = await db.work_orders.find_one({"_id": ObjectId(wo_id)})
        if not existing_wo:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        # Convertir le temps en heures décimales
        time_to_add = time_data.hours + (time_data.minutes / 60.0)
        
        # Récupérer le temps réel actuel (0 si None)
        current_time = existing_wo.get("tempsReel", 0) or 0
        
        # Calculer le nouveau temps réel
        new_time = current_time + time_to_add
        
        # Créer une entrée d'historique de temps avec l'utilisateur qui l'a saisi
        time_entry = {
            "id": str(uuid.uuid4()),
            "user_id": current_user["id"],
            "user_name": f"{current_user['prenom']} {current_user['nom']}",
            "hours": time_to_add,
            "timestamp": datetime.now(timezone.utc)
        }
        
        # Mettre à jour l'ordre de travail avec le temps total ET l'entrée d'historique
        await db.work_orders.update_one(
            {"_id": ObjectId(wo_id)},
            {
                "$set": {"tempsReel": new_time},
                "$push": {"time_entries": time_entry}
            }
        )
        
        # Log dans l'audit
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType_Audit.WORK_ORDER,
            entity_id=str(existing_wo["_id"]),
            entity_name=existing_wo["titre"],
            details=f"Ajout de temps passé: {time_data.hours}h{time_data.minutes:02d}min",
            changes={"tempsReel_old": current_time, "tempsReel_new": new_time, "time_added": time_to_add}
        )
        
        # Récupérer l'ordre de travail mis à jour
        wo = await db.work_orders.find_one({"_id": ObjectId(wo_id)})
        wo = serialize_doc(wo)
        
        if wo.get("assigne_a_id"):
            wo["assigneA"] = await get_user_by_id(wo["assigne_a_id"])
        if wo.get("emplacement_id"):
            wo["emplacement"] = await get_location_by_id(wo["emplacement_id"])
        if wo.get("equipement_id"):
            wo["equipement"] = await get_equipment_by_id(wo["equipement_id"])
        
        return WorkOrder(**wo)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de l'ajout de temps : {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/work-orders/{wo_id}", response_model=MessageResponse,
    summary="Supprimer un ordre de travail", tags=["Ordres de Travail"])
async def delete_work_order(wo_id: str, current_user: dict = Depends(require_permission("workOrders", "delete"))):
    """Supprimer un ordre de travail"""
    try:
        # Récupérer l'ordre de travail avant suppression pour le log
        wo = await db.work_orders.find_one({"_id": ObjectId(wo_id)})
        if not wo:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        result = await db.work_orders.delete_one({"_id": ObjectId(wo_id)})
        
        # Log dans l'audit
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.DELETE,
            entity_type=EntityType_Audit.WORK_ORDER,
            entity_id=wo.get("id"),
            entity_name=wo.get("titre", ""),
            details=f"Ordre de travail #{wo.get('numero')} supprimé"
        )
        
        # Émettre événement temps réel
        from realtime_manager import realtime_manager
        from realtime_events import EntityType as RealtimeEntityType, EventType as RealtimeEventType
        await realtime_manager.emit_event(
            RealtimeEntityType.WORK_ORDERS.value,
            RealtimeEventType.DELETED.value,
            {"id": str(wo["_id"])},
            user_id=current_user.get("id")
        )
        
        return {"message": "Ordre de travail supprimé"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ==================== WORK ORDER ATTACHMENTS ====================
UPLOAD_DIR = Path("/app/backend/uploads/work-orders")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MAX_FILE_SIZE = 25 * 1024 * 1024  # 25MB

@api_router.post("/work-orders/{wo_id}/attachments",
    summary="Uploader une piece jointe", response_model=AttachmentResponse, tags=["Ordres de Travail"])
async def upload_attachment(
    wo_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission("workOrders", "edit"))
):
    """Uploader une pièce jointe (max 25MB)"""
    try:
        # Vérifier que l'ordre de travail existe
        wo = await db.work_orders.find_one({"_id": ObjectId(wo_id)})
        if not wo:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        # Vérifier la taille du fichier
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 25MB)")
        
        # Générer un nom de fichier unique
        file_ext = Path(file.filename).suffix
        unique_filename = f"{uuid.uuid4()}{file_ext}"
        file_path = UPLOAD_DIR / unique_filename
        
        # Sauvegarder le fichier
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(content)
        
        # Créer l'entrée attachment
        attachment = {
            "_id": ObjectId(),
            "filename": unique_filename,
            "original_filename": file.filename,
            "size": len(content),
            "mime_type": file.content_type or mimetypes.guess_type(file.filename)[0] or "application/octet-stream",
            "uploaded_at": datetime.utcnow()
        }
        
        # Ajouter à la base de données
        await db.work_orders.update_one(
            {"_id": ObjectId(wo_id)},
            {"$push": {"attachments": attachment}}
        )
        
        attachment_response = {
            "id": str(attachment["_id"]),
            "filename": attachment["filename"],
            "original_filename": attachment["original_filename"],
            "size": attachment["size"],
            "mime_type": attachment["mime_type"],
            "uploaded_at": attachment["uploaded_at"],
            "url": f"/api/work-orders/{wo_id}/attachments/{str(attachment['_id'])}"
        }
        
        return attachment_response
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/work-orders/{wo_id}/attachments",
    summary="Lister les pieces jointes", response_model=List[AttachmentResponse], tags=["Ordres de Travail"])
async def get_attachments(wo_id: str, current_user: dict = Depends(require_permission("workOrders", "view"))):
    """Lister les pièces jointes d'un ordre de travail"""
    try:
        # Chercher par UUID (id) ou par ObjectId (_id)
        wo = await db.work_orders.find_one({"id": wo_id})
        if not wo:
            try:
                wo = await db.work_orders.find_one({"_id": ObjectId(wo_id)})
            except:
                pass
        
        if not wo:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        attachments = wo.get("attachments", [])
        result = []
        for att in attachments:
            # Gérer les deux formats d'attachment (ancien avec _id et nouveau avec id)
            att_id = str(att.get("_id", att.get("id", "")))
            result.append({
                "id": att_id,
                "filename": att.get("filename", ""),
                "original_filename": att.get("original_filename", att.get("filename", "")),
                "size": att.get("size", 0),
                "mime_type": att.get("mime_type", att.get("type", "application/octet-stream")),
                "uploaded_at": att.get("uploaded_at", att.get("uploadedAt", "")),
                "url": f"/api/work-orders/{wo_id}/attachments/{att_id}"
            })
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/work-orders/{wo_id}/attachments/{attachment_id}",
    summary="Telecharger une piece jointe", tags=["Ordres de Travail"])
async def download_attachment(
    wo_id: str,
    attachment_id: str,
    preview: bool = False,
    current_user: dict = Depends(require_permission("workOrders", "view"))
):
    """Télécharger ou prévisualiser une pièce jointe"""
    try:
        # Chercher par UUID (id) ou par ObjectId (_id)
        wo = await db.work_orders.find_one({"id": wo_id})
        if not wo:
            try:
                wo = await db.work_orders.find_one({"_id": ObjectId(wo_id)})
            except:
                pass
        
        if not wo:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        # Trouver l'attachment (par _id ou id)
        attachment = None
        for att in wo.get("attachments", []):
            att_id = str(att.get("_id", att.get("id", "")))
            if att_id == attachment_id:
                attachment = att
                break
        
        if not attachment:
            raise HTTPException(status_code=404, detail="Pièce jointe non trouvée")
        
        # Gérer les deux formats de chemin
        file_path = attachment.get("path")
        if not file_path:
            file_path = str(UPLOAD_DIR / attachment.get("filename", ""))
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Fichier non trouvé sur le serveur")
        
        disposition = "inline" if preview else "attachment"
        return FileResponse(
            path=file_path,
            filename=attachment.get("original_filename", attachment.get("filename", "file")),
            media_type=attachment.get("mime_type", attachment.get("type", "application/octet-stream")),
            content_disposition_type=disposition
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/work-orders/{wo_id}/attachments/{attachment_id}",
    summary="Supprimer une piece jointe", response_model=MessageResponse, tags=["Ordres de Travail"])
async def delete_attachment(
    wo_id: str,
    attachment_id: str,
    current_user: dict = Depends(require_permission("workOrders", "edit"))
):
    """Supprimer une pièce jointe"""
    try:
        # Chercher par UUID (id) ou par ObjectId (_id)
        wo = await db.work_orders.find_one({"id": wo_id})
        if not wo:
            try:
                wo = await db.work_orders.find_one({"_id": ObjectId(wo_id)})
            except:
                pass
        
        if not wo:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        # Trouver l'attachment (par _id ou id)
        attachment = None
        for att in wo.get("attachments", []):
            att_id = str(att.get("_id", att.get("id", "")))
            if att_id == attachment_id:
                attachment = att
                break
        
        if not attachment:
            raise HTTPException(status_code=404, detail="Pièce jointe non trouvée")
        
        # Supprimer le fichier physique
        file_path = attachment.get("path")
        if not file_path:
            file_path = str(UPLOAD_DIR / attachment.get("filename", ""))
        
        if os.path.exists(file_path):
            os.remove(file_path)
        
        # Retirer de la base de données (gérer les deux formats)
        # Utiliser le même filtre que celui utilisé pour trouver l'OT
        wo_filter = {"id": wo.get("id")} if wo.get("id") else {"_id": wo["_id"]}
        
        if "_id" in attachment:
            await db.work_orders.update_one(
                wo_filter,
                {"$pull": {"attachments": {"_id": attachment["_id"]}}}
            )
        else:
            await db.work_orders.update_one(
                wo_filter,
                {"$pull": {"attachments": {"id": attachment_id}}}
            )
        
        return {"message": "Pièce jointe supprimée"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ==================== EQUIPMENTS ROUTES ====================
@api_router.get("/equipments",
    summary="Lister les equipements", tags=["Equipements"])
async def get_equipments(current_user: dict = Depends(get_current_user)):
    """Liste tous les équipements avec filtrage par service
    
    Note : Accessible à tous les utilisateurs authentifiés pour permettre
    la sélection d'équipements dans les ordres de travail (Prélevée Sur), 
    même sans permission 'assets'.
    """
    from service_filter import apply_service_filter
    
    query = {}
    # Appliquer le filtre par service pour les responsables de service
    query = await apply_service_filter(query, current_user, "service")
    
    equipments = await db.equipments.find(query).sort("display_order", 1).to_list(1000)
    
    result = []
    for eq in equipments:
        eq = serialize_doc(eq)
        
        # Convertir parent_id en string
        if eq.get("parent_id") and not isinstance(eq["parent_id"], str):
            eq["parent_id"] = str(eq["parent_id"])
        
        # Convertir emplacement_id en string
        if eq.get("emplacement_id") and not isinstance(eq["emplacement_id"], str):
            eq["emplacement_id"] = str(eq["emplacement_id"])
        
        if eq.get("emplacement_id"):
            eq["emplacement"] = await get_location_by_id(eq["emplacement_id"])
        
        if eq.get("parent_id"):
            eq["parent"] = await get_equipment_by_id(eq["parent_id"])
        
        # Vérifier si l'équipement a des enfants
        try:
            children_count = await db.equipments.count_documents({
                "$or": [
                    {"parent_id": eq["id"]},
                    {"parent_id": ObjectId(eq["id"])}
                ]
            })
            eq["hasChildren"] = children_count > 0
        except Exception:
            eq["hasChildren"] = False
        
        result.append(eq)
    
    return result


@api_router.put("/equipments/reorder",
    summary="Reordonner les equipements", tags=["Equipements"])
async def reorder_equipments(
    items: List[dict],
    current_user: dict = Depends(require_permission("assets", "edit"))
):
    """Mettre à jour l'ordre d'affichage des équipements (admin uniquement)"""
    if current_user.get("role") != "ADMIN":
        raise HTTPException(status_code=403, detail="Seuls les administrateurs peuvent réorganiser les équipements")
    
    for item in items:
        eq_id = item.get("id")
        order = item.get("display_order", 0)
        if eq_id:
            await db.equipments.update_one(
                {"_id": ObjectId(eq_id)},
                {"$set": {"display_order": order}}
            )
    
    return {"message": "Ordre mis à jour", "count": len(items)}


@api_router.post("/equipments",
    summary="Creer un equipement", response_model=Equipment, tags=["Equipements"])
async def create_equipment(eq_create: EquipmentCreate, current_user: dict = Depends(require_permission("assets", "edit"))):
    """Créer un nouvel équipement"""
    eq_dict = eq_create.model_dump()
    
    # Si un parent est spécifié et qu'aucun emplacement n'est fourni, hériter de l'emplacement du parent
    if eq_dict.get("parent_id"):
        parent = await db.equipments.find_one({"_id": ObjectId(eq_dict["parent_id"])})
        if parent:
            # Hériter de l'emplacement du parent
            if not eq_dict.get("emplacement_id"):
                eq_dict["emplacement_id"] = parent.get("emplacement_id")
        else:
            raise HTTPException(status_code=404, detail="Équipement parent non trouvé")
    
    # Vérifier qu'on a un emplacement_id valide après héritage
    if not eq_dict.get("emplacement_id"):
        raise HTTPException(status_code=400, detail="Un emplacement est requis (directement ou hérité du parent)")
    
    eq_dict["dateCreation"] = datetime.utcnow()
    eq_dict["derniereMaintenance"] = None
    eq_dict["createdBy"] = current_user.get("id")  # Ajouter le créateur
    eq_dict["_id"] = ObjectId()
    
    await db.equipments.insert_one(eq_dict)
    
    eq = serialize_doc(eq_dict)
    if eq.get("emplacement_id"):
        eq["emplacement"] = await get_location_by_id(eq["emplacement_id"])
    
    if eq.get("parent_id"):
        eq["parent"] = await get_equipment_by_id(eq["parent_id"])
    
    eq["hasChildren"] = False
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "equipments",
        "created",
        eq,
        user_id=current_user.get("id")
    )
    
    return Equipment(**eq)

@api_router.get("/equipments/status-history",
    summary="Historique des statuts", tags=["Equipements"])
async def get_equipment_status_history(
    equipment_ids: Optional[str] = None,
    date_debut: Optional[str] = None,
    date_fin: Optional[str] = None,
    current_user: dict = Depends(require_permission("assets", "view"))
):
    """Récupérer l'historique des statuts des équipements pour le Planning M.Prev"""
    try:
        query = {}
        
        # Filtrer par équipements si spécifié
        if equipment_ids:
            ids_list = equipment_ids.split(",")
            query["equipment_id"] = {"$in": ids_list}
        
        # Filtrer par date si spécifié
        if date_debut or date_fin:
            query["changed_at"] = {}
            if date_debut:
                query["changed_at"]["$gte"] = datetime.fromisoformat(date_debut.replace('Z', '+00:00'))
            if date_fin:
                query["changed_at"]["$lte"] = datetime.fromisoformat(date_fin.replace('Z', '+00:00'))
        
        # Récupérer l'historique trié par date
        history = await db.equipment_status_history.find(query).sort("changed_at", 1).to_list(10000)
        
        # Sérialiser les documents
        result = []
        for entry in history:
            entry["id"] = str(entry["_id"])
            del entry["_id"]
            # Convertir datetime en ISO string
            if isinstance(entry.get("changed_at"), datetime):
                entry["changed_at"] = entry["changed_at"].isoformat()
            result.append(entry)
        
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/equipments/{eq_id}",
    summary="Detail d'un equipement", tags=["Equipements"])
async def get_equipment_detail(eq_id: str, current_user: dict = Depends(require_permission("assets", "view"))):
    """Récupérer les détails d'un équipement"""
    try:
        eq = await db.equipments.find_one({"_id": ObjectId(eq_id)})
        if not eq:
            raise HTTPException(status_code=404, detail="Équipement non trouvé")
        
        eq = serialize_doc(eq)
        
        if eq.get("emplacement_id"):
            eq["emplacement"] = await get_location_by_id(eq["emplacement_id"])
        
        if eq.get("parent_id"):
            eq["parent_id"] = str(eq["parent_id"])
            eq["parent"] = await get_equipment_by_id(eq["parent_id"])
        
        if eq.get("emplacement_id"):
            eq["emplacement_id"] = str(eq["emplacement_id"])
        
        # Vérifier si l'équipement a des enfants
        children_count = await db.equipments.count_documents({
            "$or": [
                {"parent_id": eq["id"]},
                {"parent_id": ObjectId(eq["id"])}
            ]
        })
        eq["hasChildren"] = children_count > 0
        
        try:
            return Equipment(**eq)
        except Exception:
            return eq
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/equipments/{eq_id}/children",
    summary="Sous-equipements", tags=["Equipements"])
async def get_equipment_children(eq_id: str, current_user: dict = Depends(require_permission("assets", "view"))):
    """Récupérer tous les sous-équipements d'un équipement"""
    try:
        # Vérifier que le parent existe
        parent = await db.equipments.find_one({"_id": ObjectId(eq_id)})
        if not parent:
            raise HTTPException(status_code=404, detail="Équipement parent non trouvé")
        
        # Récupérer tous les enfants (parent_id peut être string ou ObjectId)
        children = await db.equipments.find({
            "$or": [
                {"parent_id": eq_id},
                {"parent_id": ObjectId(eq_id)}
            ]
        }).to_list(1000)
        
        result = []
        for child in children:
            child = serialize_doc(child)
            
            if child.get("emplacement_id"):
                child["emplacement_id"] = str(child["emplacement_id"])
                child["emplacement"] = await get_location_by_id(child["emplacement_id"])
            
            if child.get("parent_id"):
                child["parent_id"] = str(child["parent_id"])
                child["parent"] = await get_equipment_by_id(child["parent_id"])
            
            # Vérifier si cet enfant a lui-même des enfants
            grandchildren_count = await db.equipments.count_documents({
                "$or": [
                    {"parent_id": child["id"]},
                    {"parent_id": ObjectId(child["id"])}
                ]
            })
            child["hasChildren"] = grandchildren_count > 0
            
            try:
                result.append(Equipment(**child))
            except Exception:
                result.append(child)
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/equipments/{eq_id}/hierarchy",
    summary="Hierarchie complete", tags=["Equipements"])
async def get_equipment_hierarchy(eq_id: str, current_user: dict = Depends(require_permission("assets", "view"))):
    """Récupérer toute la hiérarchie d'un équipement (récursif)"""
    try:
        async def build_hierarchy(equipment_id: str):
            eq = await db.equipments.find_one({"_id": ObjectId(equipment_id)})
            if not eq:
                return None
            
            eq = serialize_doc(eq)
            
            if eq.get("emplacement_id"):
                eq["emplacement"] = await get_location_by_id(eq["emplacement_id"])
            
            # Récupérer les enfants
            children = await db.equipments.find({"parent_id": eq["id"]}).to_list(1000)
            eq["children"] = []
            
            for child in children:
                child_hierarchy = await build_hierarchy(str(child["_id"]))
                if child_hierarchy:
                    eq["children"].append(child_hierarchy)
            
            eq["hasChildren"] = len(eq["children"]) > 0
            
            return eq
        
        hierarchy = await build_hierarchy(eq_id)
        if not hierarchy:
            raise HTTPException(status_code=404, detail="Équipement non trouvé")
        
        return hierarchy
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.put("/equipments/{eq_id}",
    summary="Modifier un equipement", response_model=Equipment, tags=["Equipements"])
async def update_equipment(eq_id: str, eq_update: EquipmentUpdate, current_user: dict = Depends(require_permission("assets", "edit"))):
    """Modifier un équipement"""
    from dependencies import can_edit_resource
    
    try:
        # Récupérer l'équipement existant
        existing_eq = await db.equipments.find_one({"_id": ObjectId(eq_id)})
        if not existing_eq:
            raise HTTPException(status_code=404, detail="Équipement non trouvé")
        
        existing_eq["id"] = str(existing_eq["_id"])
        
        # Vérifier les permissions (sauf admin, seulement le créateur peut modifier)
        if not can_edit_resource(current_user, existing_eq):
            raise HTTPException(
                status_code=403,
                detail="Vous ne pouvez modifier que les équipements que vous avez créés"
            )
        
        update_data = {k: v for k, v in eq_update.model_dump().items() if v is not None}
        
        # Si le statut change, enregistrer dans l'historique et le journal
        old_statut = existing_eq.get("statut")
        if "statut" in update_data and old_statut != update_data["statut"]:
            now = datetime.now(timezone.utc)
            # Arrondir à l'heure inférieure (supprimer minutes, secondes, microsecondes)
            rounded_hour = now.replace(minute=0, second=0, microsecond=0)
            update_data["statut_changed_at"] = rounded_hour
            
            # Enregistrer dans l'historique des statuts
            # Si une entrée existe déjà pour cet équipement à la même heure, l'écraser
            history_entry = {
                "equipment_id": eq_id,
                "statut": update_data["statut"],
                "changed_at": rounded_hour,
                "changed_by": current_user.get("id"),
                "changed_by_name": f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip()
            }
            await db.equipment_status_history.update_one(
                {"equipment_id": eq_id, "changed_at": rounded_hour},
                {"$set": history_entry},
                upsert=True
            )
            
            # Enregistrer dans le journal d'audit
            await audit_service.log_action(
                user_id=current_user.get("id"),
                user_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip(),
                user_email=current_user.get("email", ""),
                action=ActionType.UPDATE,
                entity_type=EntityType.EQUIPMENT,
                entity_id=eq_id,
                entity_name=existing_eq.get("nom"),
                details=f"Changement de statut: {old_statut} → {update_data['statut']}",
                changes={"statut": {"old": old_statut, "new": update_data["statut"]}}
            )
        
        await db.equipments.update_one(
            {"_id": ObjectId(eq_id)},
            {"$set": update_data}
        )
        
        eq = await db.equipments.find_one({"_id": ObjectId(eq_id)})
        eq = serialize_doc(eq)
        
        if eq.get("emplacement_id"):
            eq["emplacement"] = await get_location_by_id(eq["emplacement_id"])
        
        if eq.get("parent_id"):
            eq["parent"] = await get_equipment_by_id(eq["parent_id"])
        
        children_count = await db.equipments.count_documents({"parent_id": eq["id"]})
        eq["hasChildren"] = children_count > 0
        
        # Broadcast WebSocket pour la synchronisation temps réel (sans exclure l'utilisateur pour les autres vues)
        await realtime_manager.emit_event(
            "equipments",
            "updated",
            eq
        )
        
        return Equipment(**eq)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

async def check_and_update_parent_status(equipment_id: str):
    """Vérifier et mettre à jour le statut du parent en fonction des enfants"""
    # Récupérer l'équipement
    equipment = await db.equipments.find_one({"_id": ObjectId(equipment_id)})
    if not equipment:
        return
    
    # Si cet équipement a un parent, vérifier le statut du parent
    if equipment.get("parent_id"):
        await update_parent_alert_status(equipment["parent_id"])

async def update_parent_alert_status(parent_id: str):
    """Mettre à jour le statut du parent en fonction des statuts des enfants"""
    # Récupérer tous les enfants
    children = await db.equipments.find({"parent_id": parent_id}).to_list(1000)
    
    if not children:
        return
    
    # Vérifier si au moins un enfant est EN_MAINTENANCE ou HORS_SERVICE
    has_problematic_child = any(
        child.get("statut") in ["EN_MAINTENANCE", "HORS_SERVICE"] 
        for child in children
    )
    
    parent = await db.equipments.find_one({"_id": ObjectId(parent_id)})
    if not parent:
        return
    
    if has_problematic_child:
        # Mettre le parent en ALERTE_S_EQUIP (alerte automatique)
        await db.equipments.update_one(
            {"_id": ObjectId(parent_id)},
            {"$set": {"statut": "ALERTE_S_EQUIP"}}
        )
    else:
        # Si tous les enfants sont OPERATIONNEL et le parent est en ALERTE, remettre à OPERATIONNEL
        if parent.get("statut") == "ALERTE_S_EQUIP":
            all_operational = all(
                child.get("statut") == "OPERATIONNEL" 
                for child in children
            )
            if all_operational:
                await db.equipments.update_one(
                    {"_id": ObjectId(parent_id)},
                    {"$set": {"statut": "OPERATIONNEL"}}
                )

@api_router.patch("/equipments/{eq_id}/status",
    summary="Changer le statut", tags=["Equipements"])
async def update_equipment_status(
    eq_id: str, 
    statut: EquipmentStatus, 
    force: bool = False,  # Paramètre pour forcer le changement malgré maintenance en cours
    current_user: dict = Depends(require_permission("assets", "edit"))
):
    """Mettre à jour rapidement le statut d'un équipement"""
    try:
        equipment = await db.equipments.find_one({"_id": ObjectId(eq_id)})
        if not equipment:
            raise HTTPException(status_code=404, detail="Équipement non trouvé")
        
        # Vérifier si l'équipement a une maintenance préventive en cours
        # IMPORTANT: Exclure les maintenances déjà terminées de manière anticipée
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        active_maintenance = await db.planning_equipement.find_one({
            "equipement_id": eq_id,
            "date_debut": {"$lte": today},
            "date_fin": {"$gte": today},
            "fin_anticipee": {"$ne": True}  # Exclure les maintenances déjà terminées
        })
        
        # Si maintenance en cours et pas de force, retourner une demande de confirmation
        if active_maintenance and not force:
            # Récupérer les infos de la demande d'arrêt associée
            demande_arret = await db.demandes_arret.find_one({"id": active_maintenance.get("demande_arret_id")})
            return {
                "requires_confirmation": True,
                "message": "Cet équipement est actuellement en maintenance préventive planifiée",
                "maintenance_info": {
                    "id": active_maintenance.get("id"),
                    "date_debut": active_maintenance.get("date_debut"),
                    "date_fin": active_maintenance.get("date_fin"),
                    "demande_id": active_maintenance.get("demande_arret_id"),
                    "motif": demande_arret.get("motif") if demande_arret else None
                },
                "current_status": equipment.get("statut"),
                "new_status": statut
            }
        
        # Si maintenance en cours et force=true, terminer la maintenance anticipée
        if active_maintenance and force:
            # Mettre fin à TOUTES les maintenances préventives actives pour cet équipement
            # (il peut y avoir plusieurs entrées de planning pour la même demande ou des demandes différentes)
            update_result = await db.planning_equipement.update_many(
                {
                    "equipement_id": eq_id,
                    "date_debut": {"$lte": today},
                    "date_fin": {"$gte": today},
                    "fin_anticipee": {"$ne": True}  # Ne pas re-mettre à jour celles déjà terminées
                },
                {"$set": {
                    "date_fin": today,
                    "fin_anticipee": True,
                    "fin_anticipee_par": current_user.get("id"),
                    "fin_anticipee_par_nom": f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip(),
                    "fin_anticipee_le": datetime.now(timezone.utc).isoformat()
                }}
            )
            logger.info(f"Fin anticipée: {update_result.modified_count} entrée(s) de planning mises à jour pour équipement {eq_id}")
            
            # Mettre à jour TOUTES les demandes d'arrêt associées
            demande_arret_id = active_maintenance.get("demande_arret_id")
            if demande_arret_id:
                await db.demandes_arret.update_one(
                    {"id": demande_arret_id},
                    {"$set": {
                        "fin_anticipee": True,
                        "date_fin_effective": today,
                        "fin_anticipee_par_id": current_user.get("id"),
                        "fin_anticipee_par_nom": f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip(),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
            
            # Émettre un événement WebSocket pour notifier les autres clients
            try:
                await broadcast_update("equipments", {
                    "type": "equipment_status_changed",
                    "equipment_id": eq_id,
                    "new_status": statut.value if hasattr(statut, 'value') else statut,
                    "maintenance_ended": True
                })
            except Exception as ws_error:
                logger.warning(f"Erreur WebSocket broadcast: {ws_error}")
        
        # Note: La validation des sous-équipements a été retirée pour permettre
        # aux utilisateurs de changer librement le statut des équipements parents
        
        # Si le statut change, enregistrer dans l'historique et le journal
        old_statut = equipment.get("statut")
        if old_statut != statut:
            now = datetime.now(timezone.utc)
            # Arrondir à l'heure inférieure
            rounded_hour = now.replace(minute=0, second=0, microsecond=0)
            
            # Notification push si equipement passe hors service
            new_statut_value_check = statut.value if hasattr(statut, 'value') else statut
            if new_statut_value_check == "HORS_SERVICE" and old_statut != "HORS_SERVICE":
                from notifications import notify_equipment_alert
                asyncio.create_task(
                    notify_equipment_alert(
                        db=db,
                        equipment_id=eq_id,
                        equipment_name=equipment.get("nom", ""),
                        alert_type="PANNE",
                        alert_message="L'equipement est hors service"
                    )
                )
                # Web Push PWA
                asyncio.create_task(
                    notify_equipment_alert_web(db, equipment, "PANNE")
                )
            
            # Enregistrer dans l'historique (upsert pour écraser si même heure)
            history_entry = {
                "equipment_id": eq_id,
                "statut": statut,
                "changed_at": rounded_hour,
                "changed_by": current_user.get("id"),
                "changed_by_name": f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip()
            }
            await db.equipment_status_history.update_one(
                {"equipment_id": eq_id, "changed_at": rounded_hour},
                {"$set": history_entry},
                upsert=True
            )
            
            # Enregistrer dans le journal d'audit
            new_statut_value = statut.value if hasattr(statut, 'value') else statut
            await audit_service.log_action(
                user_id=current_user.get("id"),
                user_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip(),
                user_email=current_user.get("email", ""),
                action=ActionType.UPDATE,
                entity_type=EntityType.EQUIPMENT,
                entity_id=eq_id,
                entity_name=equipment.get("nom"),
                details=f"Changement de statut: {old_statut} → {new_statut_value}",
                changes={"statut": {"old": old_statut, "new": new_statut_value}}
            )
            
            # Mettre à jour le statut ET la date de changement
            result = await db.equipments.update_one(
                {"_id": ObjectId(eq_id)},
                {"$set": {"statut": statut, "statut_changed_at": rounded_hour}}
            )
        else:
            # Mettre à jour seulement le statut (pas de changement réel)
            result = await db.equipments.update_one(
                {"_id": ObjectId(eq_id)},
                {"$set": {"statut": statut}}
            )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Équipement non trouvé")
        
        # Mettre à jour le statut du parent si nécessaire
        await check_and_update_parent_status(eq_id)
        
        # Broadcast WebSocket pour la synchronisation temps réel
        # Ne pas exclure l'utilisateur courant pour que les autres vues (Planning) soient aussi mises à jour
        updated_eq = await db.equipments.find_one({"_id": ObjectId(eq_id)})
        if updated_eq:
            updated_eq = serialize_doc(updated_eq)
            await realtime_manager.emit_event(
                "equipments",
                "status_changed",
                updated_eq
            )
        
        return {"message": "Statut mis à jour", "statut": statut}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/equipments/{eq_id}", response_model=MessageResponse,
    summary="Supprimer un equipement", tags=["Equipements"])
async def delete_equipment(eq_id: str, current_user: dict = Depends(require_permission("assets", "delete"))):
    """Supprimer un équipement"""
    try:
        # Récupérer l'équipement avant suppression pour le broadcast
        equipment = await db.equipments.find_one({"_id": ObjectId(eq_id)})
        eq_name = equipment.get("nom", "Inconnu") if equipment else "Inconnu"
        
        result = await db.equipments.delete_one({"_id": ObjectId(eq_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Équipement non trouvé")
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "equipments",
            "deleted",
            {"id": eq_id, "nom": eq_name},
            user_id=current_user.get("id")
        )
        
        return {"message": "Équipement supprimé"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ==================== AVAILABILITY ROUTES ====================
@api_router.get("/availabilities",
    summary="Lister les disponibilites", tags=["Disponibilites"])
async def get_availabilities(
    start_date: str = None,
    end_date: str = None,
    user_id: str = None,
    current_user: dict = Depends(require_permission("planning", "view"))
):
    """Récupérer les disponibilités du personnel"""
    query = {}
    
    if user_id:
        query["user_id"] = user_id
    
    if start_date and end_date:
        query["date"] = {
            "$gte": datetime.fromisoformat(start_date),
            "$lte": datetime.fromisoformat(end_date)
        }
    
    availabilities = await db.availabilities.find(query).to_list(1000)
    
    for avail in availabilities:
        avail["id"] = str(avail["_id"])
        del avail["_id"]
        if avail.get("user_id"):
            avail["user"] = await get_user_by_id(avail["user_id"])
    
    return availabilities

@api_router.post("/availabilities",
    summary="Creer une disponibilite", tags=["Disponibilites"])
async def create_availability(
    availability: UserAvailabilityCreate,
    current_user: dict = Depends(get_current_admin_user)
):
    """Créer une disponibilité (admin uniquement)"""
    avail_dict = availability.model_dump()
    avail_dict["_id"] = ObjectId()
    
    await db.availabilities.insert_one(avail_dict)
    
    avail = serialize_doc(avail_dict)
    if avail.get("user_id"):
        avail["user"] = await get_user_by_id(avail["user_id"])
    
    # Émettre l'événement WebSocket
    try:
        from realtime_events import EntityType as RealtimeEntityType, EventType as RealtimeEventType
        await realtime_manager.emit_event(
            RealtimeEntityType.AVAILABILITIES.value,
            RealtimeEventType.CREATED.value,
            avail,
            current_user.get("id")
        )
    except Exception as e:
        logger.error(f"Erreur émission événement WebSocket availabilities: {e}")
    
    return avail

@api_router.put("/availabilities/{avail_id}",
    summary="Modifier une disponibilite", tags=["Disponibilites"])
async def update_availability(
    avail_id: str,
    availability_update: UserAvailabilityUpdate,
    current_user: dict = Depends(get_current_admin_user)
):
    """Mettre à jour une disponibilité (admin uniquement)"""
    try:
        # Utiliser model_dump() pour obtenir toutes les valeurs, y compris None
        # On accepte explicitement les valeurs null pour pouvoir remettre à blanc
        update_data = {}
        raw_dict = availability_update.model_dump()
        
        # Pour chaque champ de disponibilité, vérifier s'il a été envoyé
        for field in ['disponible', 'disponible_matin', 'disponible_aprem', 'disponible_nuit', 'motif']:
            if field in raw_dict:
                update_data[field] = raw_dict[field]
        
        if update_data:
            await db.availabilities.update_one(
                {"_id": ObjectId(avail_id)},
                {"$set": update_data}
            )
        
        avail = await db.availabilities.find_one({"_id": ObjectId(avail_id)})
        avail = serialize_doc(avail)
        
        if avail.get("user_id"):
            avail["user"] = await get_user_by_id(avail["user_id"])
        
        # Émettre l'événement WebSocket
        try:
            from realtime_events import EntityType as RealtimeEntityType, EventType as RealtimeEventType
            await realtime_manager.emit_event(
                RealtimeEntityType.AVAILABILITIES.value,
                RealtimeEventType.UPDATED.value,
                avail,
                current_user.get("id")
            )
        except Exception as e:
            logger.error(f"Erreur émission événement WebSocket availabilities: {e}")
        
        return avail
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/availabilities/{avail_id}", response_model=MessageResponse,
    summary="Supprimer une disponibilite", tags=["Disponibilites"])
async def delete_availability(
    avail_id: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """Supprimer une disponibilité (admin uniquement)"""
    try:
        result = await db.availabilities.delete_one({"_id": ObjectId(avail_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Disponibilité non trouvée")
        
        # Émettre l'événement WebSocket
        try:
            from realtime_events import EntityType as RealtimeEntityType, EventType as RealtimeEventType
            await realtime_manager.emit_event(
                RealtimeEntityType.AVAILABILITIES.value,
                RealtimeEventType.DELETED.value,
                {"id": avail_id},
                current_user.get("id")
            )
        except Exception as e:
            logger.error(f"Erreur émission événement WebSocket availabilities: {e}")
        
        return {"message": "Disponibilité supprimée"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ==================== LOCATIONS ROUTES ====================
@api_router.get("/locations",
    summary="Lister les emplacements", response_model=List[Location], tags=["Emplacements"])
async def get_locations(current_user: dict = Depends(require_permission("locations", "view"))):
    """Liste toutes les zones avec hiérarchie"""
    locations = await db.locations.find().to_list(1000)
    
    # Enrichir avec les informations de hiérarchie
    result = []
    for loc in locations:
        loc_data = serialize_doc(loc)
        
        # Calculer le niveau dans la hiérarchie
        level = 0
        parent_id = loc.get('parent_id')
        while parent_id and level < 3:
            parent = await db.locations.find_one({"_id": ObjectId(parent_id)})
            if parent:
                level += 1
                parent_id = parent.get('parent_id')
            else:
                break
        loc_data['level'] = level
        
        # Vérifier si cette zone a des enfants
        has_children = await db.locations.count_documents({"parent_id": loc_data['id']}) > 0
        loc_data['hasChildren'] = has_children
        
        # Ajouter les infos du parent si présent
        if loc.get('parent_id'):
            parent = await db.locations.find_one({"_id": ObjectId(loc.get('parent_id'))})
            if parent:
                loc_data['parent'] = {
                    "id": str(parent["_id"]),
                    "nom": parent.get("nom")
                }
        
        result.append(Location(**loc_data))
    
    return result

@api_router.get("/locations/{loc_id}/children", response_model=List[Location], tags=["Emplacements"])
async def get_location_children(loc_id: str, current_user: dict = Depends(require_permission("locations", "view"))):
    """Récupérer les sous-zones d'une zone"""
    children = await db.locations.find({"parent_id": loc_id}).to_list(100)
    result = []
    for child in children:
        child_data = serialize_doc(child)
        child_data['level'] = 1  # Simplifié pour l'instant
        child_data['hasChildren'] = await db.locations.count_documents({"parent_id": child_data['id']}) > 0
        result.append(Location(**child_data))
    return result

@api_router.post("/locations",
    summary="Creer un emplacement", response_model=Location, tags=["Emplacements"])
async def create_location(loc_create: LocationCreate, current_user: dict = Depends(require_permission("locations", "edit"))):
    """Créer une nouvelle zone"""
    loc_dict = loc_create.model_dump()
    loc_dict["dateCreation"] = datetime.utcnow()
    loc_dict["_id"] = ObjectId()
    
    # Vérifier le niveau de hiérarchie si parent_id est fourni
    if loc_dict.get('parent_id'):
        parent_id = loc_dict['parent_id']
        level = 0
        
        # Remonter la hiérarchie pour calculer le niveau
        while parent_id and level < 3:
            parent = await db.locations.find_one({"_id": ObjectId(parent_id)})
            if parent:
                level += 1
                parent_id = parent.get('parent_id')
            else:
                break
        
        # Limiter à 3 niveaux (0, 1, 2)
        if level >= 3:
            raise HTTPException(
                status_code=400, 
                detail="Limite de hiérarchie atteinte. Maximum 3 niveaux de sous-zones."
            )
    
    await db.locations.insert_one(loc_dict)
    
    loc_data = serialize_doc(loc_dict)
    loc_data['level'] = 0
    loc_data['hasChildren'] = False
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "zones",
        "created",
        loc_data,
        user_id=current_user["id"]
    )
    
    return Location(**loc_data)

@api_router.put("/locations/{loc_id}",
    summary="Modifier un emplacement", response_model=Location, tags=["Emplacements"])
async def update_location(loc_id: str, loc_update: LocationUpdate, current_user: dict = Depends(require_permission("locations", "edit"))):
    """Modifier une zone"""
    try:
        update_data = {k: v for k, v in loc_update.model_dump().items() if v is not None}
        
        # Si on change le parent_id, vérifier la hiérarchie
        if 'parent_id' in update_data and update_data['parent_id']:
            parent_id = update_data['parent_id']
            level = 0
            
            while parent_id and level < 3:
                parent = await db.locations.find_one({"_id": ObjectId(parent_id)})
                if parent:
                    level += 1
                    parent_id = parent.get('parent_id')
                else:
                    break
            
            if level >= 3:
                raise HTTPException(
                    status_code=400,
                    detail="Limite de hiérarchie atteinte. Maximum 3 niveaux de sous-zones."
                )
        
        await db.locations.update_one(
            {"_id": ObjectId(loc_id)},
            {"$set": update_data}
        )
        
        loc = await db.locations.find_one({"_id": ObjectId(loc_id)})
        loc_data = serialize_doc(loc)
        
        # Calculer le niveau
        level = 0
        parent_id = loc.get('parent_id')
        while parent_id and level < 3:
            parent = await db.locations.find_one({"_id": ObjectId(parent_id)})
            if parent:
                level += 1
                parent_id = parent.get('parent_id')
            else:
                break
        loc_data['level'] = level
        loc_data['hasChildren'] = await db.locations.count_documents({"parent_id": loc_id}) > 0
        
        if loc.get('parent_id'):
            parent = await db.locations.find_one({"_id": ObjectId(loc.get('parent_id'))})
            if parent:
                loc_data['parent'] = {
                    "id": str(parent["_id"]),
                    "nom": parent.get("nom")
                }
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "zones",
            "updated",
            loc_data,
            user_id=current_user["id"]
        )
        
        return Location(**loc_data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/locations/{loc_id}", response_model=MessageResponse,
    summary="Supprimer un emplacement", tags=["Emplacements"])
async def delete_location(loc_id: str, current_user: dict = Depends(require_permission("locations", "delete"))):
    """Supprimer une zone et ses sous-zones"""
    try:
        # Vérifier s'il y a des sous-zones
        children_count = await db.locations.count_documents({"parent_id": loc_id})
        if children_count > 0:
            raise HTTPException(
                status_code=400,
                detail=f"Impossible de supprimer cette zone car elle contient {children_count} sous-zone(s). Supprimez d'abord les sous-zones."
            )
        
        # Vérifier s'il y a des équipements liés
        equipment_count = await db.equipments.count_documents({"emplacement_id": loc_id})
        if equipment_count > 0:
            raise HTTPException(
                status_code=400,
                detail=f"Impossible de supprimer cette zone car elle contient {equipment_count} équipement(s)."
            )
        
        result = await db.locations.delete_one({"_id": ObjectId(loc_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Zone non trouvée")
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "zones",
            "deleted",
            {"id": loc_id},
            user_id=current_user["id"]
        )
        
        return {"message": "Zone supprimée"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ==================== INVENTORY ROUTES ====================
@api_router.get("/inventory",
    summary="Lister l'inventaire", tags=["Inventaire"])
async def get_inventory(current_user: dict = Depends(get_current_user)):
    """Liste tous les articles de l'inventaire
    
    Note : Accessible à tous les utilisateurs authentifiés pour permettre
    la sélection de pièces dans les ordres de travail, même sans permission 'inventory'.
    """
    inventory = await db.inventory.find({}).to_list(1000)
    # Sérialiser chaque document pour convertir _id en id
    return [serialize_doc(item) for item in inventory]

@api_router.post("/inventory",
    summary="Ajouter un article", response_model=Inventory, tags=["Inventaire"])
async def create_inventory_item(inv_create: InventoryCreate, current_user: dict = Depends(require_permission("inventory", "edit"))):
    """Créer un nouvel article dans l'inventaire"""
    inv_dict = inv_create.model_dump()
    inv_dict["dateCreation"] = datetime.utcnow()
    inv_dict["derniereModification"] = datetime.utcnow()
    inv_dict["_id"] = ObjectId()
    
    await db.inventory.insert_one(inv_dict)
    
    inv_data = serialize_doc(inv_dict)
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "inventory",
        "created",
        inv_data,
        user_id=current_user.get("id")
    )
    
    return Inventory(**inv_data)

@api_router.put("/inventory/{inv_id}",
    summary="Modifier un article", response_model=Inventory, tags=["Inventaire"])
async def update_inventory_item(inv_id: str, inv_update: InventoryUpdate, current_user: dict = Depends(require_permission("inventory", "edit"))):
    """Modifier un article de l'inventaire"""
    try:
        # Récupérer l'article actuel pour vérifier la quantité avant modification
        current_item = await db.inventory.find_one({"_id": ObjectId(inv_id)})
        if not current_item:
            raise HTTPException(status_code=404, detail="Article non trouvé")
        
        old_quantity = current_item.get("quantite", 0)
        
        update_data = {k: v for k, v in inv_update.model_dump().items() if v is not None}
        update_data["derniereModification"] = datetime.utcnow()
        
        await db.inventory.update_one(
            {"_id": ObjectId(inv_id)},
            {"$set": update_data}
        )
        
        inv = await db.inventory.find_one({"_id": ObjectId(inv_id)})
        inv_data = serialize_doc(inv)
        
        # Vérifier si la quantité est passée à 0
        new_quantity = inv_data.get("quantite", 0)
        if new_quantity == 0 and old_quantity > 0:
            # Créer automatiquement une demande d'achat
            await create_auto_purchase_request(inv_data, current_user)
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "inventory",
            "updated",
            inv_data,
            user_id=current_user.get("id")
        )
        
        return Inventory(**inv_data)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


async def create_auto_purchase_request(inventory_item: dict, current_user: dict):
    """Créer automatiquement une demande d'achat pour un article en rupture de stock"""
    try:
        from models import PurchaseRequest, PurchaseRequestType, PurchaseRequestUrgency, PurchaseRequestStatus, PurchaseRequestHistoryEntry
        
        # Générer un numéro unique pour la demande
        year = datetime.utcnow().year
        count = await db.purchase_requests.count_documents({
            "numero": {"$regex": f"^DA-{year}-"}
        })
        numero = f"DA-{year}-{str(count + 1).zfill(5)}"
        
        # Créer la demande d'achat avec le modèle Pydantic (génère automatiquement l'UUID)
        purchase_request = PurchaseRequest(
            numero=numero,
            type=PurchaseRequestType.CONSOMMABLE,
            designation=inventory_item.get("nom", "Article inconnu"),
            description=f"Demande automatique - Rupture de stock détectée pour l'article '{inventory_item.get('nom')}'",
            quantite=inventory_item.get("quantiteMin", 10),  # Commander au moins le seuil minimum
            unite="Unité",
            reference=inventory_item.get("reference", ""),
            fournisseur_suggere=inventory_item.get("fournisseur", ""),
            urgence=PurchaseRequestUrgency.URGENT,
            justification=f"Rupture de stock automatiquement détectée. L'article '{inventory_item.get('nom')}' (Réf: {inventory_item.get('reference', 'N/A')}) a atteint une quantité de 0. Emplacement: {inventory_item.get('emplacement', 'N/A')}",
            destinataire_id=None,
            destinataire_nom="Service Maintenance",
            inventory_item_id=inventory_item.get("id"),
            attached_files=[],
            demandeur_id="SYSTEM",
            demandeur_nom="Système automatique",
            demandeur_email="system@gmao.local",
            status=PurchaseRequestStatus.SOUMISE,
            responsable_n1_id=None,
            responsable_n1_nom=None,
            history=[
                PurchaseRequestHistoryEntry(
                    user_id="SYSTEM",
                    user_name="Système automatique",
                    action="Création automatique - Rupture de stock",
                    new_status=PurchaseRequestStatus.SOUMISE.value,
                    comment=f"Article '{inventory_item.get('nom')}' en rupture de stock"
                )
            ]
        )
        
        # Sauvegarder dans la DB
        await db.purchase_requests.insert_one(purchase_request.model_dump())
        
        # Broadcast WebSocket pour notifier
        await realtime_manager.emit_event(
            "purchase_requests",
            "created",
            purchase_request.model_dump(),
            user_id=current_user.get("id")
        )
        
        logger.info(f"Demande d'achat automatique créée: {numero} pour article {inventory_item.get('nom')}")
        
    except Exception as e:
        logger.error(f"Erreur création demande d'achat automatique: {e}")
        import traceback
        logger.error(traceback.format_exc())


@api_router.delete("/inventory/{inv_id}", response_model=MessageResponse,
    summary="Supprimer un article", tags=["Inventaire"])
async def delete_inventory_item(inv_id: str, current_user: dict = Depends(require_permission("inventory", "delete"))):
    """Supprimer un article de l'inventaire"""
    try:
        # Récupérer l'article avant suppression pour le broadcast
        item = await db.inventory.find_one({"_id": ObjectId(inv_id)})
        item_name = item.get("nom", "Inconnu") if item else "Inconnu"
        
        result = await db.inventory.delete_one({"_id": ObjectId(inv_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Article non trouvé")
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "inventory",
            "deleted",
            {"id": inv_id, "nom": item_name},
            user_id=current_user.get("id")
        )
        
        return {"message": "Article supprimé"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.patch("/inventory/{inv_id}/toggle-monitoring",
    summary="Activer/desactiver la surveillance", response_model=ToggleMonitoringResponse, tags=["Inventaire"])
async def toggle_inventory_monitoring(inv_id: str, current_user: dict = Depends(require_permission("inventory", "edit"))):
    """Active/Désactive la surveillance du stock d'un article"""
    try:
        # Récupérer l'article actuel
        item = await db.inventory.find_one({"_id": ObjectId(inv_id)})
        if not item:
            raise HTTPException(status_code=404, detail="Article non trouvé")
        
        # Inverser le statut de surveillance (par défaut True si n'existe pas)
        current_monitoring = item.get("stock_monitoring_enabled", True)
        new_monitoring = not current_monitoring
        
        # Mettre à jour
        await db.inventory.update_one(
            {"_id": ObjectId(inv_id)},
            {
                "$set": {
                    "stock_monitoring_enabled": new_monitoring,
                    "derniere_modification": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        logger.info(f"📊 Surveillance stock {'activée' if new_monitoring else 'désactivée'} pour {item.get('nom', 'Article')}")
        
        return {
            "message": f"Surveillance {'activée' if new_monitoring else 'désactivée'}",
            "stock_monitoring_enabled": new_monitoring
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur toggle monitoring: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/inventory/stats",
    summary="Statistiques inventaire", response_model=InventoryStatsResponse, tags=["Inventaire"])
async def get_inventory_stats(current_user: dict = Depends(require_permission("inventory", "view"))):
    """Récupère les statistiques de l'inventaire (rupture et niveau bas)"""
    try:
        inventory = await db.inventory.find().to_list(1000)
        
        rupture = 0
        niveau_bas = 0
        
        for item in inventory:
            # Ignorer les articles dont la surveillance est désactivée
            if not item.get("stock_monitoring_enabled", True):
                continue
            
            quantite = item.get("quantite", 0)
            quantite_min = item.get("quantiteMin", item.get("seuil_alerte", 0))
            
            if quantite <= 0:
                rupture += 1
            elif quantite <= quantite_min:
                niveau_bas += 1
        
        return {
            "rupture": rupture,
            "niveau_bas": niveau_bas
        }
    except Exception as e:
        logging.error(f"Erreur lors du calcul des stats inventaire: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ===== Services d'inventaire (onglets par service) =====

@api_router.get("/inventory/services", tags=["Inventaire - Services"])
async def get_inventory_services(current_user: dict = Depends(get_current_user)):
    """Liste tous les services d'inventaire (onglets), synchronises avec les services des roles."""
    # Synchroniser avec la liste des services (meme source que Dashboard Service)
    try:
        from roles_routes import SERVICES as roles_services
        
        existing_names = set()
        existing_services = await db.inventory_services.find({}, {"_id": 0}).to_list(200)
        for svc in existing_services:
            existing_names.add(svc.get("name", "").upper())
        
        # Auto-creer les services manquants
        for role_svc in roles_services:
            if role_svc.upper() not in existing_names:
                new_svc = {
                    "id": str(uuid.uuid4()),
                    "name": role_svc,
                    "created_by": "system",
                    "created_by_name": "Synchronisation automatique",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.inventory_services.insert_one(new_svc)
        
        # Assurer "Non classe" existe
        if "NON CLASSÉ" not in existing_names and "NON CLASSE" not in existing_names:
            nc_doc = {
                "id": str(uuid.uuid4()),
                "name": "Non classé",
                "created_by": "system",
                "created_by_name": "Système",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.inventory_services.insert_one(nc_doc)
    except Exception as e:
        logger.warning(f"Erreur sync services inventaire: {e}")
    
    services = await db.inventory_services.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    return services


@api_router.post("/inventory/services", tags=["Inventaire - Services"])
async def create_inventory_service(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Créer un nouvel onglet de service d'inventaire (Admin ou responsable de service)."""
    user_role = current_user.get("role", "")
    is_admin = user_role in ["ADMIN", "admin", "Administrateur"]
    
    # Vérifier si responsable de service
    is_manager = False
    try:
        from service_filter import is_service_manager
        is_manager = await is_service_manager(current_user)
    except Exception:
        pass
    
    if not is_admin and not is_manager:
        raise HTTPException(403, "Seuls les administrateurs et responsables de service peuvent créer des onglets")
    
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(400, "Le nom du service est requis")
    
    # Vérifier unicité
    existing = await db.inventory_services.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(400, f"Le service '{name}' existe déjà")
    
    service_doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "created_by": current_user.get("id"),
        "created_by_name": f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.inventory_services.insert_one(service_doc)
    del service_doc["_id"]
    return service_doc


@api_router.delete("/inventory/services/{service_id}", tags=["Inventaire - Services"])
async def delete_inventory_service(
    service_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer un onglet de service d'inventaire. Les articles seront déplacés vers 'Non classé'."""
    user_role = current_user.get("role", "")
    is_admin = user_role in ["ADMIN", "admin", "Administrateur"]
    
    is_manager = False
    try:
        from service_filter import is_service_manager
        is_manager = await is_service_manager(current_user)
    except Exception:
        pass
    
    if not is_admin and not is_manager:
        raise HTTPException(403, "Seuls les administrateurs et responsables de service peuvent supprimer des onglets")
    
    service = await db.inventory_services.find_one({"id": service_id}, {"_id": 0})
    if not service:
        raise HTTPException(404, "Service introuvable")
    
    if service.get("name") == "Non classé":
        raise HTTPException(400, "Impossible de supprimer le service 'Non classé'")
    
    # Trouver le service "Non classé"
    non_classe = await db.inventory_services.find_one({"name": "Non classé"}, {"_id": 0})
    nc_id = non_classe["id"] if non_classe else None
    
    # Déplacer les articles vers "Non classé"
    if nc_id:
        await db.inventory.update_many(
            {"service_id": service_id},
            {"$set": {"service_id": nc_id}}
        )
        # Retirer ce service des articles partagés
        await db.inventory.update_many(
            {"shared_service_ids": service_id},
            {"$pull": {"shared_service_ids": service_id}}
        )
    
    await db.inventory_services.delete_one({"id": service_id})
    
    article_count = await db.inventory.count_documents({"service_id": nc_id}) if nc_id else 0
    return {"success": True, "message": f"Service supprimé. Articles déplacés vers 'Non classé'.", "moved_count": article_count}


@api_router.post("/inventory/{inv_id}/share", tags=["Inventaire - Partage"])
async def share_inventory_item(
    inv_id: str,
    data: dict,
    current_user: dict = Depends(require_permission("inventory", "edit"))
):
    """Importer/partager un article dans un autre service (lien partagé, même stock)."""
    target_service_id = data.get("target_service_id")
    if not target_service_id:
        raise HTTPException(400, "target_service_id requis")
    
    # Vérifier que l'article existe
    item = await db.inventory.find_one({"_id": ObjectId(inv_id)})
    if not item:
        raise HTTPException(404, "Article introuvable")
    
    # Vérifier que le service cible existe
    target_service = await db.inventory_services.find_one({"id": target_service_id}, {"_id": 0})
    if not target_service:
        raise HTTPException(404, "Service cible introuvable")
    
    # Vérifier que ce n'est pas déjà partagé
    shared = item.get("shared_service_ids", [])
    if target_service_id in shared or item.get("service_id") == target_service_id:
        raise HTTPException(400, "Cet article est déjà dans ce service")
    
    await db.inventory.update_one(
        {"_id": ObjectId(inv_id)},
        {"$addToSet": {"shared_service_ids": target_service_id}}
    )
    
    updated = await db.inventory.find_one({"_id": ObjectId(inv_id)})
    return serialize_doc(updated)


@api_router.delete("/inventory/{inv_id}/unshare/{service_id}", tags=["Inventaire - Partage"])
async def unshare_inventory_item(
    inv_id: str,
    service_id: str,
    current_user: dict = Depends(require_permission("inventory", "edit"))
):
    """Retirer le partage d'un article d'un service."""
    item = await db.inventory.find_one({"_id": ObjectId(inv_id)})
    if not item:
        raise HTTPException(404, "Article introuvable")
    
    await db.inventory.update_one(
        {"_id": ObjectId(inv_id)},
        {"$pull": {"shared_service_ids": service_id}}
    )
    
    return {"success": True, "message": "Partage retiré"}


@api_router.get("/inventory/by-service/{service_id}", tags=["Inventaire - Services"])
async def get_inventory_by_service(
    service_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Liste les articles d'un service (propriétaires + partagés)."""
    items = await db.inventory.find(
        {"$or": [
            {"service_id": service_id},
            {"shared_service_ids": service_id}
        ]}
    ).to_list(1000)
    return [serialize_doc(item) for item in items]



# ==================== PREVENTIVE MAINTENANCE ROUTES ====================
@api_router.get("/preventive-maintenance",
    summary="Lister les maintenances preventives", response_model=List[PreventiveMaintenance], tags=["Maintenance Preventive"])
async def get_preventive_maintenance(current_user: dict = Depends(require_permission("preventiveMaintenance", "view"))):
    """Liste toutes les maintenances préventives avec filtrage par service"""
    from service_filter import apply_service_filter
    
    query = {}
    # Appliquer le filtre par service
    query = await apply_service_filter(query, current_user, "service")
    
    pm_list = await db.preventive_maintenances.find(query).to_list(1000)
    
    for pm in pm_list:
        pm["id"] = str(pm["_id"])
        del pm["_id"]
        
        if pm.get("equipement_id"):
            pm["equipement"] = await get_equipment_by_id(pm["equipement_id"])
        if pm.get("assigne_a_id"):
            pm["assigneA"] = await get_user_by_id(pm["assigne_a_id"])
    
    return [PreventiveMaintenance(**pm) for pm in pm_list]

@api_router.post("/preventive-maintenance",
    summary="Creer une maintenance preventive", response_model=PreventiveMaintenance, tags=["Maintenance Preventive"])
async def create_preventive_maintenance(pm_create: PreventiveMaintenanceCreate, current_user: dict = Depends(require_permission("preventiveMaintenance", "edit"))):
    """Créer une nouvelle maintenance préventive"""
    pm_dict = pm_create.model_dump()
    pm_dict["dateCreation"] = datetime.utcnow()
    pm_dict["derniereMaintenance"] = None
    pm_dict["_id"] = ObjectId()
    
    await db.preventive_maintenances.insert_one(pm_dict)
    
    pm = serialize_doc(pm_dict)
    if pm.get("equipement_id"):
        pm["equipement"] = await get_equipment_by_id(pm["equipement_id"])
    if pm.get("assigne_a_id"):
        pm["assigneA"] = await get_user_by_id(pm["assigne_a_id"])
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "preventive_maintenance",
        "created",
        pm,
        user_id=current_user.get("id")
    )
    
    return PreventiveMaintenance(**pm)

@api_router.put("/preventive-maintenance/{pm_id}",
    summary="Modifier une maintenance preventive", response_model=PreventiveMaintenance, tags=["Maintenance Preventive"])
async def update_preventive_maintenance(pm_id: str, pm_update: PreventiveMaintenanceUpdate, current_user: dict = Depends(require_permission("preventiveMaintenance", "edit"))):
    """Modifier une maintenance préventive"""
    try:
        update_data = {k: v for k, v in pm_update.model_dump().items() if v is not None}
        
        await db.preventive_maintenances.update_one(
            {"_id": ObjectId(pm_id)},
            {"$set": update_data}
        )
        
        pm = await db.preventive_maintenances.find_one({"_id": ObjectId(pm_id)})
        pm = serialize_doc(pm)
        
        if pm.get("equipement_id"):
            pm["equipement"] = await get_equipment_by_id(pm["equipement_id"])
        if pm.get("assigne_a_id"):
            pm["assigneA"] = await get_user_by_id(pm["assigne_a_id"])
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "preventive_maintenance",
            "updated",
            pm,
            user_id=current_user.get("id")
        )
        
        return PreventiveMaintenance(**pm)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/preventive-maintenance/{pm_id}", response_model=MessageResponse,
    summary="Supprimer une maintenance preventive", tags=["Maintenance Preventive"])
async def delete_preventive_maintenance(pm_id: str, current_user: dict = Depends(require_permission("preventiveMaintenance", "delete"))):
    """Supprimer une maintenance préventive"""
    try:
        # Récupérer la maintenance avant suppression pour le broadcast
        pm = await db.preventive_maintenances.find_one({"_id": ObjectId(pm_id)})
        pm_nom = pm.get("nom", "Inconnu") if pm else "Inconnu"
        
        result = await db.preventive_maintenances.delete_one({"_id": ObjectId(pm_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Maintenance préventive non trouvée")
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "preventive_maintenance",
            "deleted",
            {"id": pm_id, "nom": pm_nom},
            user_id=current_user.get("id")
        )
        
        return {"message": "Maintenance préventive supprimée"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ==================== PREVENTIVE MAINTENANCE - ATTACHMENTS ====================

@api_router.post("/preventive-maintenance/{pm_id}/attachments", tags=["Maintenance Preventive"])
async def upload_pm_attachment(
    pm_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission("preventiveMaintenance", "edit"))
):
    """Upload une pièce jointe pour une maintenance préventive"""
    import os
    import uuid as uuid_mod
    
    try:
        # Vérifier que la maintenance existe
        pm = await db.preventive_maintenances.find_one({"_id": ObjectId(pm_id)})
        if not pm:
            raise HTTPException(status_code=404, detail="Maintenance préventive non trouvée")
        
        # Créer le répertoire uploads/preventive-maintenance si nécessaire
        upload_dir = "/app/backend/uploads/preventive-maintenance"
        os.makedirs(upload_dir, exist_ok=True)
        
        # Générer un nom de fichier unique
        file_ext = os.path.splitext(file.filename)[1] if file.filename else ""
        attachment_id = str(uuid_mod.uuid4())
        unique_filename = f"{attachment_id}{file_ext}"
        file_path = os.path.join(upload_dir, unique_filename)
        
        # Lire et sauvegarder le fichier
        content = await file.read()
        file_size = len(content)
        
        with open(file_path, "wb") as f:
            f.write(content)
        
        # Créer l'objet attachment
        new_attachment = {
            "id": attachment_id,
            "filename": unique_filename,
            "original_filename": file.filename,
            "path": file_path,
            "mime_type": file.content_type or "application/octet-stream",
            "size": file_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "uploaded_by": current_user.get("id")
        }
        
        # Ajouter au tableau attachments
        await db.preventive_maintenances.update_one(
            {"_id": ObjectId(pm_id)},
            {"$push": {"attachments": new_attachment}}
        )
        
        logger.info(f"Pièce jointe ajoutée à la maintenance préventive {pm_id}: {file.filename}")
        
        return {
            "success": True,
            "attachment": new_attachment
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur upload pièce jointe maintenance préventive: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/preventive-maintenance/{pm_id}/attachments", tags=["Maintenance Preventive"])
async def get_pm_attachments(
    pm_id: str,
    current_user: dict = Depends(require_permission("preventiveMaintenance", "view"))
):
    """Récupérer les pièces jointes d'une maintenance préventive"""
    try:
        pm = await db.preventive_maintenances.find_one({"_id": ObjectId(pm_id)})
        if not pm:
            raise HTTPException(status_code=404, detail="Maintenance préventive non trouvée")
        
        return pm.get("attachments", [])
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur récupération pièces jointes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/preventive-maintenance/{pm_id}/attachments/{attachment_id}", tags=["Maintenance Preventive"])
async def download_pm_attachment(
    pm_id: str,
    attachment_id: str,
    preview: bool = False,
    current_user: dict = Depends(require_permission("preventiveMaintenance", "view"))
):
    """Télécharger ou prévisualiser une pièce jointe d'une maintenance préventive"""
    import os
    
    try:
        pm = await db.preventive_maintenances.find_one({"_id": ObjectId(pm_id)})
        if not pm:
            raise HTTPException(status_code=404, detail="Maintenance préventive non trouvée")
        
        # Trouver la pièce jointe
        attachment = None
        for att in pm.get("attachments", []):
            if att.get("id") == attachment_id:
                attachment = att
                break
        
        if not attachment:
            raise HTTPException(status_code=404, detail="Pièce jointe non trouvée")
        
        file_path = attachment.get("path")
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Fichier non trouvé sur le serveur")
        
        disposition = "inline" if preview else "attachment"
        return FileResponse(
            path=file_path,
            filename=attachment.get("original_filename", attachment.get("filename")),
            media_type=attachment.get("mime_type", "application/octet-stream"),
            content_disposition_type=disposition
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur téléchargement pièce jointe: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/preventive-maintenance/{pm_id}/attachments/{attachment_id}", response_model=SuccessResponse, tags=["Maintenance Preventive"])
async def delete_pm_attachment(
    pm_id: str,
    attachment_id: str,
    current_user: dict = Depends(require_permission("preventiveMaintenance", "edit"))
):
    """Supprimer une pièce jointe d'une maintenance préventive"""
    import os
    
    try:
        pm = await db.preventive_maintenances.find_one({"_id": ObjectId(pm_id)})
        if not pm:
            raise HTTPException(status_code=404, detail="Maintenance préventive non trouvée")
        
        # Trouver la pièce jointe
        attachment = None
        for att in pm.get("attachments", []):
            if att.get("id") == attachment_id:
                attachment = att
                break
        
        if not attachment:
            raise HTTPException(status_code=404, detail="Pièce jointe non trouvée")
        
        # Supprimer le fichier physique
        file_path = attachment.get("path")
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
        
        # Retirer du tableau attachments
        await db.preventive_maintenances.update_one(
            {"_id": ObjectId(pm_id)},
            {"$pull": {"attachments": {"id": attachment_id}}}
        )
        
        logger.info(f"Pièce jointe supprimée de la maintenance préventive {pm_id}: {attachment_id}")
        
        return {"success": True, "message": "Pièce jointe supprimée"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur suppression pièce jointe: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


def calculate_next_maintenance_date(current_date: datetime, frequency: str) -> datetime:
    """Calcule la prochaine date de maintenance selon la fréquence"""
    if frequency == "QUOTIDIENNE":
        return current_date + timedelta(days=1)
    elif frequency == "HEBDOMADAIRE":
        return current_date + timedelta(weeks=1)
    elif frequency == "MENSUELLE":
        # Ajouter un mois
        month = current_date.month
        year = current_date.year
        if month == 12:
            month = 1
            year += 1
        else:
            month += 1
        return current_date.replace(year=year, month=month)
    elif frequency == "ANNUELLE":
        return current_date.replace(year=current_date.year + 1)
    else:
        # Par défaut, mensuelle
        return current_date + timedelta(days=30)

@api_router.post("/preventive-maintenance/check-and-execute", response_model=SuccessResponse, tags=["Maintenance Preventive"])
async def check_and_execute_due_maintenances(current_user: dict = Depends(get_current_admin_user)):
    """Vérifie et exécute MANUELLEMENT les maintenances échues (admin uniquement)"""
    try:
        logger.info(f"🔄 Vérification MANUELLE déclenchée par {current_user.get('email', 'Unknown')}")
        await auto_check_preventive_maintenance()
        return {"success": True, "message": "Vérification manuelle effectuée - Consultez les logs pour les détails"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/preventive-maintenance/check-and-execute-OLD", tags=["Maintenance Preventive"])
async def check_and_execute_due_maintenances_old(current_user: dict = Depends(get_current_admin_user)):
    """Version détaillée pour debug (admin uniquement)"""
    try:
        today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Trouver toutes les maintenances actives dont la date est aujourd'hui ou passée
        pm_list = await db.preventive_maintenances.find({
            "statut": "ACTIF",
            "prochaineMaintenance": {"$lte": today + timedelta(days=1)}
        }).to_list(length=None)
        
        created_count = 0
        updated_count = 0
        errors = []
        
        for pm in pm_list:
            try:
                # Récupérer l'équipement
                equipement = await db.equipments.find_one({"_id": ObjectId(pm["equipement_id"])})
                
                # Créer le bon de travail
                wo_id = str(uuid.uuid4())
                work_order = {
                    "_id": ObjectId(),
                    "id": wo_id,
                    "numero": f"PM-{datetime.utcnow().strftime('%Y%m%d')}-{secrets.token_hex(3).upper()}",
                    "titre": f"Maintenance préventive: {pm['titre']}",
                    "description": f"Maintenance automatique générée depuis la planification préventive",
                    "type": "PREVENTIF",
                    "priorite": "NORMALE",
                    "statut": "OUVERT",
                    "equipement_id": pm["equipement_id"],
                    "emplacement_id": equipement.get("emplacement_id") if equipement else None,
                    "assigne_a_id": pm.get("assigne_a_id"),
                    "tempsEstime": pm.get("duree"),
                    "dateLimite": datetime.utcnow() + timedelta(days=7),
                    "dateCreation": datetime.utcnow(),
                    "createdBy": "system",
                    "comments": [],
                    "attachments": [],
                    "historique": []
                }
                
                await db.work_orders.insert_one(work_order)
                created_count += 1
                
                # Calculer la prochaine date de maintenance
                next_date = calculate_next_maintenance_date(pm["prochaineMaintenance"], pm["frequence"])
                
                # Mettre à jour la maintenance préventive
                await db.preventive_maintenances.update_one(
                    {"_id": pm["_id"]},
                    {
                        "$set": {
                            "prochaineMaintenance": next_date,
                            "derniereMaintenance": datetime.utcnow()
                        }
                    }
                )
                updated_count += 1
                
            except Exception as e:
                errors.append(f"Erreur pour PM {pm.get('titre', 'Unknown')}: {str(e)}")
        
        return {
            "success": True,
            "workOrdersCreated": created_count,
            "maintenancesUpdated": updated_count,
            "errors": errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== CHECKLIST ROUTES ====================

@api_router.get("/checklists/templates",
    summary="Lister les modeles de checklist", response_model=List[ChecklistTemplate], tags=["Checklists"])
async def get_checklist_templates(current_user: dict = Depends(require_permission("preventiveMaintenance", "view"))):
    """Liste tous les modèles de checklists"""
    templates = await db.checklist_templates.find().to_list(1000)
    return [ChecklistTemplate(**serialize_doc(t)) for t in templates]

@api_router.get("/checklists/templates/{template_id}",
    summary="Detail d'un modele", response_model=ChecklistTemplate, tags=["Checklists"])
async def get_checklist_template(template_id: str, current_user: dict = Depends(require_permission("preventiveMaintenance", "view"))):
    """Récupère un modèle de checklist par ID"""
    try:
        template = await db.checklist_templates.find_one({"_id": ObjectId(template_id)})
        if not template:
            # Essayer avec l'ID string
            template = await db.checklist_templates.find_one({"id": template_id})
        if not template:
            raise HTTPException(status_code=404, detail="Modèle de checklist non trouvé")
        return ChecklistTemplate(**serialize_doc(template))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/checklists/templates",
    summary="Creer un modele", response_model=ChecklistTemplate, tags=["Checklists"])
async def create_checklist_template(template_create: ChecklistTemplateCreate, current_user: dict = Depends(require_permission("preventiveMaintenance", "edit"))):
    """Créer un nouveau modèle de checklist"""
    try:
        template_dict = template_create.model_dump()
        template_dict["id"] = str(uuid.uuid4())
        template_dict["created_by_id"] = current_user.get("id")
        template_dict["created_by_name"] = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip()
        template_dict["created_at"] = datetime.now(timezone.utc).isoformat()
        template_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.checklist_templates.insert_one(template_dict)
        return ChecklistTemplate(**template_dict)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.put("/checklists/templates/{template_id}",
    summary="Modifier un modele", response_model=ChecklistTemplate, tags=["Checklists"])
async def update_checklist_template(template_id: str, template_update: ChecklistTemplateUpdate, current_user: dict = Depends(require_permission("preventiveMaintenance", "edit"))):
    """Mettre à jour un modèle de checklist"""
    try:
        update_data = {k: v for k, v in template_update.model_dump().items() if v is not None}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        # Essayer avec ObjectId d'abord
        result = await db.checklist_templates.update_one(
            {"_id": ObjectId(template_id)},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            # Essayer avec l'ID string
            result = await db.checklist_templates.update_one(
                {"id": template_id},
                {"$set": update_data}
            )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Modèle de checklist non trouvé")
        
        template = await db.checklist_templates.find_one({"_id": ObjectId(template_id)})
        if not template:
            template = await db.checklist_templates.find_one({"id": template_id})
        return ChecklistTemplate(**serialize_doc(template))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/checklists/templates/{template_id}", response_model=MessageResponse,
    summary="Supprimer un modele", tags=["Checklists"])
async def delete_checklist_template(template_id: str, current_user: dict = Depends(require_permission("preventiveMaintenance", "delete"))):
    """Supprimer un modèle de checklist"""
    try:
        result = await db.checklist_templates.delete_one({"_id": ObjectId(template_id)})
        if result.deleted_count == 0:
            result = await db.checklist_templates.delete_one({"id": template_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Modèle de checklist non trouvé")
        return {"message": "Modèle de checklist supprimé"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# === Exécutions de checklists ===

@api_router.get("/checklists/executions", response_model=List[ChecklistExecution], tags=["Checklists"])
async def get_checklist_executions(
    work_order_id: Optional[str] = None,
    preventive_maintenance_id: Optional[str] = None,
    current_user: dict = Depends(require_permission("preventiveMaintenance", "view"))
):
    """Liste les exécutions de checklists, avec filtres optionnels"""
    query = {}
    if work_order_id:
        query["work_order_id"] = work_order_id
    if preventive_maintenance_id:
        query["preventive_maintenance_id"] = preventive_maintenance_id
    
    executions = await db.checklist_executions.find(query).sort("started_at", -1).to_list(1000)
    return [ChecklistExecution(**serialize_doc(e)) for e in executions]

@api_router.get("/checklists/executions/{execution_id}", response_model=ChecklistExecution, tags=["Checklists"])
async def get_checklist_execution(execution_id: str, current_user: dict = Depends(require_permission("preventiveMaintenance", "view"))):
    """Récupère une exécution de checklist par ID"""
    try:
        execution = await db.checklist_executions.find_one({"_id": ObjectId(execution_id)})
        if not execution:
            execution = await db.checklist_executions.find_one({"id": execution_id})
        if not execution:
            raise HTTPException(status_code=404, detail="Exécution de checklist non trouvée")
        return ChecklistExecution(**serialize_doc(execution))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/checklists/executions", response_model=ChecklistExecution, tags=["Checklists"])
async def create_checklist_execution(execution_create: ChecklistExecutionCreate, current_user: dict = Depends(require_permission("preventiveMaintenance", "edit"))):
    """Démarrer une nouvelle exécution de checklist"""
    try:
        # Récupérer le template
        template = await db.checklist_templates.find_one({"id": execution_create.checklist_template_id})
        if not template:
            template = await db.checklist_templates.find_one({"_id": ObjectId(execution_create.checklist_template_id)})
        if not template:
            raise HTTPException(status_code=404, detail="Modèle de checklist non trouvé")
        
        # Récupérer l'équipement si spécifié
        equipment_name = None
        if execution_create.equipment_id:
            equipment = await db.equipments.find_one({"_id": ObjectId(execution_create.equipment_id)})
            if equipment:
                equipment_name = equipment.get("nom", "")
        
        execution_dict = {
            "id": str(uuid.uuid4()),
            "checklist_template_id": execution_create.checklist_template_id,
            "checklist_name": template.get("name", ""),
            "work_order_id": execution_create.work_order_id,
            "preventive_maintenance_id": execution_create.preventive_maintenance_id,
            "equipment_id": execution_create.equipment_id,
            "equipment_name": equipment_name,
            "responses": [],
            "total_items": len(template.get("items", [])),
            "completed_items": 0,
            "compliant_items": 0,
            "non_compliant_items": 0,
            "general_comment": None,
            "general_photos": [],
            "status": "in_progress",
            "executed_by_id": current_user.get("id"),
            "executed_by_name": f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip(),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "completed_at": None
        }
        
        await db.checklist_executions.insert_one(execution_dict)
        return ChecklistExecution(**execution_dict)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.put("/checklists/executions/{execution_id}", response_model=ChecklistExecution, tags=["Checklists"])
async def update_checklist_execution(execution_id: str, execution_update: ChecklistExecutionUpdate, current_user: dict = Depends(require_permission("preventiveMaintenance", "edit"))):
    """Mettre à jour une exécution de checklist (ajouter des réponses)"""
    try:
        update_data = {k: v for k, v in execution_update.model_dump().items() if v is not None}
        
        # Si on passe des réponses, convertir en dict
        if "responses" in update_data:
            update_data["responses"] = [r.model_dump() if hasattr(r, 'model_dump') else r for r in update_data["responses"]]
            
            # Calculer les statistiques
            responses = update_data["responses"]
            completed = sum(1 for r in responses if r.get("value_yes_no") is not None or r.get("value_numeric") is not None or r.get("value_text"))
            compliant = sum(1 for r in responses if r.get("is_compliant", True))
            non_compliant = sum(1 for r in responses if not r.get("is_compliant", True))
            
            update_data["completed_items"] = completed
            update_data["compliant_items"] = compliant
            update_data["non_compliant_items"] = non_compliant
        
        # Si le statut passe à "completed", ajouter la date de fin
        if update_data.get("status") == "completed":
            update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
        
        # Essayer d'abord avec l'ID custom
        result = await db.checklist_executions.update_one(
            {"id": execution_id},
            {"$set": update_data}
        )
        
        # Si pas trouvé, essayer avec ObjectId (pour compatibilité)
        if result.matched_count == 0:
            try:
                result = await db.checklist_executions.update_one(
                    {"_id": ObjectId(execution_id)},
                    {"$set": update_data}
                )
            except:
                pass
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Exécution de checklist non trouvée")
        
        # Récupérer l'exécution mise à jour
        execution = await db.checklist_executions.find_one({"id": execution_id})
        if not execution:
            try:
                execution = await db.checklist_executions.find_one({"_id": ObjectId(execution_id)})
            except:
                pass
        if not execution:
            raise HTTPException(status_code=404, detail="Exécution de checklist non trouvée")
        return ChecklistExecution(**serialize_doc(execution))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/checklists/history",
    summary="Historique des checklists executees", tags=["Checklists"])
async def get_checklist_history(
    equipment_id: Optional[str] = None,
    template_id: Optional[str] = None,
    limit: int = 50,
    current_user: dict = Depends(require_permission("preventiveMaintenance", "view"))
):
    """Récupère l'historique des exécutions de checklists"""
    query = {"status": "completed"}
    if equipment_id:
        query["equipment_id"] = equipment_id
    if template_id:
        query["checklist_template_id"] = template_id
    
    executions = await db.checklist_executions.find(query).sort("completed_at", -1).to_list(limit)
    return [serialize_doc(e) for e in executions]

# ==================== USERS ROUTES ====================
@api_router.get("/users",
    summary="Lister les utilisateurs", tags=["Utilisateurs"])
async def get_users(current_user: dict = Depends(get_current_user)):
    """Liste tous les utilisateurs"""
    # Vérifier les permissions - Admin a toujours accès, sinon vérifier permission people.view
    if current_user.get("role") != "ADMIN":
        permissions = current_user.get("permissions", {})
        people_perms = permissions.get("people", {})
        if not people_perms.get("view", False):
            raise HTTPException(
                status_code=403,
                detail="Vous n'avez pas la permission de voir les utilisateurs"
            )
    
    users = await db.users.find().to_list(1000)
    result = []
    for user in users:
        doc = serialize_doc(user)
        # Fix permissions: doit être un dict, pas une liste
        if isinstance(doc.get("permissions"), list):
            doc["permissions"] = {}
        # Fix mqtt fields: convertir en string
        for field in ["mqtt_action_ok", "mqtt_action_reception"]:
            if field in doc and not isinstance(doc[field], str):
                doc[field] = str(doc[field])
        # Assurer les champs obligatoires
        if "nom" not in doc:
            doc["nom"] = doc.get("name", "Inconnu")
        if "prenom" not in doc:
            doc["prenom"] = ""
        result.append(doc)
    return result

@api_router.put("/users/{user_id}",
    summary="Modifier un utilisateur", response_model=User, tags=["Utilisateurs"])
async def update_user(user_id: str, user_update: UserUpdate, current_user: dict = Depends(get_current_admin_user)):
    """Modifier un utilisateur (admin uniquement)"""
    try:
        update_data = {k: v for k, v in user_update.model_dump().items() if v is not None}
        
        # Si le rôle change, mettre à jour automatiquement les permissions par défaut
        if "role" in update_data:
            new_role = update_data["role"]
            default_permissions = get_default_permissions_by_role(new_role).model_dump()
            update_data["permissions"] = default_permissions
        
        await db.users.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": update_data}
        )
        
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        user_response = User(**serialize_doc(user))
        
        # Émettre l'événement WebSocket à TOUS les utilisateurs (y compris celui qui fait la modification)
        # Important pour la synchronisation du Planning quand on modifie un service
        try:
            from realtime_events import EntityType as RealtimeEntityType, EventType as RealtimeEventType
            await realtime_manager.emit_event(
                RealtimeEntityType.USERS.value,
                RealtimeEventType.UPDATED.value,
                user_response.model_dump(),
                None  # Ne pas exclure l'utilisateur actuel pour assurer la synchro du Planning
            )
        except Exception as e:
            logger.error(f"Erreur émission événement WebSocket users: {e}")
        
        return user_response
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/users/{user_id}/header-visibility",
    summary="Obtenir la visibilité des icônes header d'un utilisateur", tags=["Utilisateurs"])
async def get_user_header_visibility(user_id: str, current_user: dict = Depends(get_current_user)):
    """Obtenir les paramètres de visibilité des icônes header pour un utilisateur"""
    user = await db.users.find_one({"_id": ObjectId(user_id)}, {"header_icons_visibility": 1})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return user.get("header_icons_visibility", {})

@api_router.put("/users/{user_id}/header-visibility",
    summary="Modifier la visibilité des icônes header d'un utilisateur", tags=["Utilisateurs"])
async def update_user_header_visibility(
    user_id: str,
    visibility: dict,
    current_user: dict = Depends(get_current_admin_user)
):
    """Modifier les paramètres de visibilité des icônes header (admin uniquement)"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"header_icons_visibility": visibility}}
    )
    
    return {"message": "Visibilité des icônes mise à jour", "visibility": visibility}


@api_router.delete("/users/{user_id}", response_model=MessageResponse,
    summary="Supprimer un utilisateur", tags=["Utilisateurs"])
async def delete_user(user_id: str, current_user: dict = Depends(get_current_admin_user)):
    """Supprimer un utilisateur (admin uniquement)"""
    try:
        # Empêcher de se supprimer soi-même
        if str(user_id) == str(current_user.get('id')):
            raise HTTPException(status_code=400, detail="Vous ne pouvez pas vous supprimer vous-même")
        
        result = await db.users.delete_one({"_id": ObjectId(user_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
        
        # Émettre l'événement WebSocket
        try:
            from realtime_events import EntityType as RealtimeEntityType, EventType as RealtimeEventType
            await realtime_manager.emit_event(
                RealtimeEntityType.USERS.value,
                RealtimeEventType.DELETED.value,
                {"id": user_id},
                current_user.get("id")
            )
        except Exception as e:
            logger.error(f"Erreur émission événement WebSocket users: {e}")
        
        return {"message": "Utilisateur supprimé"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/users/invite", response_model=User, tags=["Utilisateurs"])
async def invite_user(user_invite: UserInvite, current_user: dict = Depends(get_current_admin_user)):
    """Inviter un nouveau membre (admin uniquement)"""
    # Vérifier si l'utilisateur existe déjà
    existing_user = await db.users.find_one({"email": user_invite.email})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Un utilisateur avec cet email existe déjà"
        )
    
    # Générer un mot de passe temporaire
    import secrets
    import string
    temp_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
    hashed_password = get_password_hash(temp_password)
    
    # Définir les permissions par défaut selon le rôle
    if user_invite.permissions is None:
        # Utiliser la fonction centralisée pour obtenir les permissions par défaut
        permissions = get_default_permissions_by_role(user_invite.role).model_dump()
    else:
        permissions = user_invite.permissions.model_dump()
    
    # Créer l'utilisateur
    user_dict = {
        "nom": user_invite.nom,
        "prenom": user_invite.prenom,
        "email": user_invite.email,
        "telephone": user_invite.telephone,
        "role": user_invite.role,
        "hashed_password": hashed_password,
        "statut": "actif",
        "dateCreation": datetime.utcnow(),
        "derniereConnexion": None,
        "permissions": permissions,
        "_id": ObjectId()
    }
    
    await db.users.insert_one(user_dict)
    
    # TODO: Envoyer un email avec le mot de passe temporaire
    # Pour l'instant, on log juste le mot de passe (À REMPLACER EN PRODUCTION)
    logger.info(f"Utilisateur {user_invite.email} créé avec mot de passe temporaire: {temp_password}")
    
    return User(**serialize_doc(user_dict))

@api_router.get("/users/{user_id}/permissions", response_model=UserPermissions, tags=["Utilisateurs"])
async def get_user_permissions(user_id: str, current_user: dict = Depends(get_current_admin_user)):
    """Obtenir les permissions d'un utilisateur"""
    try:
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
        
        permissions = user.get("permissions", {})
        return UserPermissions(**permissions)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api_router.get("/users/default-permissions/{role}",
    summary="Permissions par defaut d'un role", tags=["Utilisateurs"])
async def get_default_permissions_for_role(
    role: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """Obtenir les permissions par défaut pour un rôle spécifique (admin uniquement)"""
    try:
        default_permissions = get_default_permissions_by_role(role)
        return {"role": role, "permissions": default_permissions.model_dump()}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors de la récupération des permissions: {str(e)}")


@api_router.get("/users/service-manager/{service}",
    summary="Responsable d'un service", tags=["Utilisateurs"])
async def get_service_manager_for_user(
    service: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer le responsable de service pour un service donné"""
    try:
        # Chercher dans service_responsables
        manager_entry = await db.service_responsables.find_one({"service": service})
        
        if not manager_entry:
            raise HTTPException(status_code=404, detail="Aucun responsable assigné pour ce service")
        
        user_id = manager_entry["user_id"]
        
        # Récupérer les infos du responsable (chercher par id OU par _id)
        manager = await db.users.find_one(
            {"id": user_id, "statut": "actif"},
            {"_id": 0, "id": 1, "nom": 1, "prenom": 1, "email": 1, "role": 1}
        )
        
        # Si non trouvé, essayer avec _id (ObjectId)
        if not manager:
            try:
                manager = await db.users.find_one(
                    {"_id": ObjectId(user_id), "statut": "actif"},
                    {"_id": 0, "id": 1, "nom": 1, "prenom": 1, "email": 1, "role": 1}
                )
            except:
                pass
        
        if not manager:
            raise HTTPException(status_code=404, detail="Responsable non trouvé ou inactif")
        
        return manager
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur récupération responsable service: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.put("/users/{user_id}/permissions", response_model=User, tags=["Utilisateurs"])
async def update_user_permissions(
    user_id: str, 
    permissions_update: UserPermissionsUpdate, 
    current_user: dict = Depends(get_current_admin_user)
):
    """Mettre à jour les permissions d'un utilisateur (admin uniquement)"""
    try:
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
        
        permissions_dict = permissions_update.permissions.model_dump()
        
        await db.users.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": {"permissions": permissions_dict}}
        )
        
        updated_user = await db.users.find_one({"_id": ObjectId(user_id)})
        return User(**serialize_doc(updated_user))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/users/init-time-tracking-permissions", tags=["Utilisateurs"])
async def init_time_tracking_permissions(
    current_user: dict = Depends(get_current_admin_user)
):
    """Initialiser les permissions timeTracking pour tous les utilisateurs selon leur rôle"""
    try:
        # Mettre à jour les ADMIN avec toutes les permissions timeTracking
        admin_result = await db.users.update_many(
            {"role": "ADMIN"},
            {"$set": {"permissions.timeTracking": {"view": True, "edit": True, "delete": True}}}
        )
        
        # Mettre à jour les TECHNICIEN et DIRECTEUR avec view et edit
        tech_result = await db.users.update_many(
            {"role": {"$in": ["TECHNICIEN", "DIRECTEUR"]}},
            {"$set": {"permissions.timeTracking": {"view": True, "edit": True, "delete": False}}}
        )
        
        # Mettre à jour les autres rôles avec view seulement
        other_result = await db.users.update_many(
            {"role": {"$nin": ["ADMIN", "TECHNICIEN", "DIRECTEUR", "AFFICHAGE"]}},
            {"$set": {"permissions.timeTracking": {"view": True, "edit": False, "delete": False}}}
        )
        
        return {
            "message": "Permissions timeTracking initialisées",
            "updated": {
                "admin": admin_result.modified_count,
                "technicien_directeur": tech_result.modified_count,
                "others": other_result.modified_count
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/users/init-cameras-permissions", tags=["Utilisateurs"])
async def init_cameras_permissions(
    current_user: dict = Depends(get_current_admin_user)
):
    """Initialiser les permissions caméras pour tous les utilisateurs selon leur rôle"""
    try:
        # ADMIN : toutes les permissions
        admin_result = await db.users.update_many(
            {"role": "ADMIN"},
            {"$set": {"permissions.cameras": {"view": True, "edit": True, "delete": True}}}
        )
        
        # Responsables de service (DIRECTEUR, responsable) : view seulement
        responsable_result = await db.users.update_many(
            {"$or": [
                {"role": "DIRECTEUR"},
                {"is_service_manager": True}
            ]},
            {"$set": {"permissions.cameras": {"view": True, "edit": False, "delete": False}}}
        )
        
        return {
            "message": "Permissions caméras initialisées",
            "updated": {
                "admin": admin_result.modified_count,
                "responsables": responsable_result.modified_count
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/users/migrate-all-permissions", tags=["Utilisateurs"])
async def migrate_all_user_permissions(
    current_user: dict = Depends(get_current_admin_user)
):
    """Migrer les permissions de TOUS les utilisateurs selon leur rôle actuel.
    Réinitialise les permissions par défaut pour chaque utilisateur selon son rôle."""
    try:
        all_users = await db.users.find({}).to_list(length=None)
        updated_count = 0
        for u in all_users:
            user_role = u.get("role", "VISUALISEUR")
            default_perms = get_default_permissions_by_role(user_role).model_dump()
            await db.users.update_one(
                {"_id": u["_id"]},
                {"$set": {"permissions": default_perms}}
            )
            updated_count += 1
        return {
            "success": True,
            "message": f"Permissions mises à jour pour {updated_count} utilisateur(s)",
            "updated_count": updated_count
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/users/{user_id}/set-password-permanent",
    summary="Definir un mot de passe permanent", response_model=SuccessResponse, tags=["Utilisateurs"])
async def set_password_permanent(
    user_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Marquer le mot de passe temporaire comme permanent (désactiver le changement obligatoire au premier login)
    L'utilisateur peut uniquement modifier son propre statut, sauf si c'est un admin
    """
    try:
        # Vérifier que l'utilisateur existe
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
        
        # Vérifier que l'utilisateur modifie son propre compte OU qu'il est admin
        current_user_id = current_user.get("id")
        is_admin = current_user.get("role") == "ADMIN"
        
        if str(user_id) != str(current_user_id) and not is_admin:
            raise HTTPException(
                status_code=403, 
                detail="Vous ne pouvez modifier que votre propre statut"
            )
        
        # Mettre à jour le champ firstLogin à False
        await db.users.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": {"firstLogin": False}}
        )
        
        # Enregistrer l'action dans le journal d'audit
        await audit_service.log_action(
            user_id=current_user_id,
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType_Audit.USER,
            entity_id=user_id,
            entity_name=f"{user.get('prenom', '')} {user.get('nom', '')}".strip(),
            details=f"Mot de passe temporaire conservé comme permanent",
            changes={"firstLogin": False}
        )
        
        return {
            "success": True,
            "message": "Mot de passe conservé avec succès"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")



@api_router.post("/users/{user_id}/reset-password-admin",
    summary="Reinitialiser le mot de passe (admin)", response_model=ResetPasswordAdminResponse, tags=["Utilisateurs"])
async def reset_password_admin(
    user_id: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Réinitialiser le mot de passe d'un utilisateur (Admin uniquement)
    Génère un nouveau mot de passe temporaire et force le changement au prochain login
    """
    try:
        # Vérifier que l'utilisateur existe
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
        
        # Générer un nouveau mot de passe temporaire
        temp_password = generate_temp_password()
        
        # Hasher le mot de passe
        hashed_password = get_password_hash(temp_password)
        
        # Mettre à jour le mot de passe et forcer le changement au prochain login
        await db.users.update_one(
            {"_id": ObjectId(user_id)},
            {
                "$set": {
                    "hashed_password": hashed_password,
                    "firstLogin": True
                }
            }
        )
        
        # Enregistrer l'action dans le journal d'audit
        await audit_service.log_action(
            user_id=current_user.get("id"),
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType_Audit.USER,
            entity_id=user_id,
            entity_name=f"{user.get('prenom', '')} {user.get('nom', '')}".strip(),
            details=f"Réinitialisation du mot de passe par l'administrateur",
            changes={"firstLogin": True, "password_reset": "admin_action"}
        )
        
        # Envoyer un email à l'utilisateur avec le nouveau mot de passe
        try:
            email_sent = email_service.send_account_created_email(
                to_email=user['email'],
                prenom=user.get('prenom', ''),
                temp_password=temp_password
            )
            
            if email_sent:
                logger.info(f"Email de réinitialisation envoyé à {user['email']}")
        except Exception as email_error:
            logger.error(f"Erreur lors de l'envoi de l'email de réinitialisation : {str(email_error)}")
        
        return {
            "success": True,
            "message": "Mot de passe réinitialisé avec succès",
            "tempPassword": temp_password,
            "emailSent": email_sent if 'email_sent' in locals() else False
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de la réinitialisation du mot de passe : {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")


# ==================== SETTINGS ROUTES ====================
@api_router.get("/settings",
    summary="Configuration systeme", response_model=SystemSettings, tags=["Parametres"])
async def get_system_settings(current_user: dict = Depends(get_current_admin_user)):
    """Récupérer les paramètres système"""
    try:
        settings = await db.system_settings.find_one({"_id": "default"})
        if not settings:
            # Paramètres par défaut
            default_settings = {
                "_id": "default",
                "inactivity_timeout_minutes": 15
            }
            await db.system_settings.insert_one(default_settings)
            return SystemSettings(**default_settings)
        
        return SystemSettings(**settings)
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des paramètres : {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")

@api_router.put("/settings",
    summary="Modifier la configuration", response_model=SystemSettings, tags=["Parametres"])
async def update_system_settings(
    settings_update: SystemSettingsUpdate,
    current_user: dict = Depends(get_current_admin_user)
):
    """Mettre à jour les paramètres système (Admin uniquement)"""
    try:
        # Vérifier que la valeur est dans une plage acceptable (entre 1 et 120 minutes)
        if settings_update.inactivity_timeout_minutes is not None:
            if settings_update.inactivity_timeout_minutes < 1 or settings_update.inactivity_timeout_minutes > 120:
                raise HTTPException(
                    status_code=400, 
                    detail="Le temps d'inactivité doit être entre 1 et 120 minutes"
                )
        
        # Mettre à jour ou créer les paramètres
        update_data = {k: v for k, v in settings_update.model_dump().items() if v is not None}
        
        settings = await db.system_settings.find_one({"_id": "default"})
        if not settings:
            # Créer les paramètres par défaut
            default_settings = {
                "_id": "default",
                "inactivity_timeout_minutes": settings_update.inactivity_timeout_minutes or 15
            }
            await db.system_settings.insert_one(default_settings)
            settings = default_settings
        else:
            # Mettre à jour
            await db.system_settings.update_one(
                {"_id": "default"},
                {"$set": update_data}
            )
            settings.update(update_data)
        
        # Journaliser l'action
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType.SETTINGS,
            entity_id="default",
            entity_name="System Settings",
            details="Modification des paramètres système",
            changes=update_data
        )
        
        return SystemSettings(**settings)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour des paramètres : {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")



# ==================== USER PREFERENCES ROUTES ====================
@api_router.get("/user-preferences", response_model=UserPreferences)
async def get_user_preferences(current_user: dict = Depends(get_current_user)):
    """Récupérer les préférences de l'utilisateur connecté"""
    try:
        user_id = current_user.get("id")
        preferences = await db.user_preferences.find_one({"user_id": user_id})
        
        if not preferences:
            # Créer des préférences par défaut
            default_prefs = {
                "user_id": user_id,
                "theme_mode": "light",
                "primary_color": "#2563eb",
                "secondary_color": "#64748b",
                "sidebar_bg_color": "#1f2937",
                "sidebar_position": "left",
                "sidebar_behavior": "minimizable",
                "sidebar_width": 256,
                "sidebar_icon_color": "#ffffff",
                "display_density": "normal",
                "font_size": "normal",
                "menu_categories": [],
                "menu_items": [],
                "header_icon_order": [],
                "default_home_page": "/dashboard",
                "date_format": "DD/MM/YYYY",
                "time_format": "24h",
                "currency": "€",
                "language": "fr",
                "dashboard_widgets": [],
                "dashboard_layout": {},
                "notifications_enabled": True,
                "email_notifications": True,
                "push_notifications": True,
                "sound_enabled": True,
                "stock_alert_threshold": 5,
                "customization_view_mode": "tabs",
                "preset_theme": None,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            preferences_obj = UserPreferences(**default_prefs)
            prefs_dict = preferences_obj.model_dump()
            await db.user_preferences.insert_one(prefs_dict)
            return preferences_obj
        
        return UserPreferences(**serialize_doc(preferences))
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des préférences : {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")

@api_router.put("/user-preferences", response_model=UserPreferences)
async def update_user_preferences(
    preferences_update: UserPreferencesUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour les préférences de l'utilisateur connecté"""
    try:
        user_id = current_user.get("id")
        logger.info(f"[PREFS] Mise à jour pour user_id: {user_id}")
        
        # Préparer les données de mise à jour
        update_data = {k: v for k, v in preferences_update.model_dump().items() if v is not None}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        logger.info(f"[PREFS] Données à mettre à jour: {list(update_data.keys())}")
        
        # Vérifier si les préférences existent
        existing = await db.user_preferences.find_one({"user_id": user_id})
        
        if not existing:
            # Créer les préférences si elles n'existent pas
            default_prefs = {
                "user_id": user_id,
                "theme_mode": "light",
                "primary_color": "#2563eb",
                "secondary_color": "#64748b",
                "sidebar_bg_color": "#1f2937",
                "sidebar_position": "left",
                "sidebar_behavior": "minimizable",
                "sidebar_width": 256,
                "sidebar_icon_color": "#ffffff",
                "display_density": "normal",
                "font_size": "normal",
                "menu_categories": [],
                "menu_items": [],
                "header_icon_order": [],
                "default_home_page": "/dashboard",
                "date_format": "DD/MM/YYYY",
                "time_format": "24h",
                "currency": "€",
                "language": "fr",
                "dashboard_widgets": [],
                "dashboard_layout": {},
                "notifications_enabled": True,
                "email_notifications": True,
                "push_notifications": True,
                "sound_enabled": True,
                "stock_alert_threshold": 5,
                "customization_view_mode": "tabs",
                "preset_theme": None,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            default_prefs.update(update_data)
            preferences_obj = UserPreferences(**default_prefs)
            prefs_dict = preferences_obj.model_dump()
            await db.user_preferences.insert_one(prefs_dict)
            
            # Journaliser l'action
            await audit_service.log_action(
                user_id=user_id,
                user_name=current_user.get("name", ""),
                user_email=current_user.get("email", ""),
                action=ActionType.UPDATE,
                entity_type=EntityType.SETTINGS,
                entity_id=user_id,
                details=f"Préférences utilisateur créées"
            )
            
            return preferences_obj
        else:
            # Mettre à jour les préférences existantes
            await db.user_preferences.update_one(
                {"user_id": user_id},
                {"$set": update_data}
            )
            
            # Récupérer les préférences mises à jour
            updated_prefs = await db.user_preferences.find_one({"user_id": user_id})
            
            # Journaliser l'action
            await audit_service.log_action(
                user_id=user_id,
                user_name=current_user.get("name", ""),
                user_email=current_user.get("email", ""),
                action=ActionType.UPDATE,
                entity_type=EntityType.SETTINGS,
                entity_id=user_id,
                details=f"Préférences utilisateur mises à jour"
            )
            
            return UserPreferences(**serialize_doc(updated_prefs))
    except Exception as e:
        import traceback
        logger.error(f"Erreur lors de la mise à jour des préférences : {str(e)}")
        logger.error(f"Traceback complet: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")

@api_router.post("/user-preferences/reset")
async def reset_user_preferences(current_user: dict = Depends(get_current_user)):
    """Réinitialiser les préférences aux valeurs par défaut"""
    try:
        user_id = current_user.get("id")
        
        # Supprimer les préférences existantes
        await db.user_preferences.delete_one({"user_id": user_id})
        
        # Créer des préférences par défaut
        default_prefs = {
            "user_id": user_id,
            "theme_mode": "light",
            "primary_color": "#2563eb",
            "secondary_color": "#64748b",
            "sidebar_bg_color": "#1f2937",
            "sidebar_position": "left",
            "sidebar_behavior": "minimizable",
            "sidebar_width": 256,
            "sidebar_icon_color": "#ffffff",
            "display_density": "normal",
            "font_size": "normal",
            "menu_categories": [],
            "menu_items": [],
            "header_icon_order": [],
            "default_home_page": "/dashboard",
            "date_format": "DD/MM/YYYY",
            "time_format": "24h",
            "currency": "€",
            "language": "fr",
            "dashboard_widgets": [],
            "dashboard_layout": {},
            "notifications_enabled": True,
            "email_notifications": True,
            "push_notifications": True,
            "sound_enabled": True,
            "stock_alert_threshold": 5,
            "customization_view_mode": "tabs",
            "preset_theme": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        preferences_obj = UserPreferences(**default_prefs)
        prefs_dict = preferences_obj.model_dump()
        await db.user_preferences.insert_one(prefs_dict)
        
        # Journaliser l'action
        await audit_service.log_action(
            user_id=user_id,
            user_name=current_user.get("name", ""),
            user_email=current_user.get("email", ""),
            action=ActionType.UPDATE,
            entity_type=EntityType.SETTINGS,
            entity_id=user_id,
            details=f"Préférences utilisateur réinitialisées"
        )
        
        return {"message": "Préférences réinitialisées avec succès", "preferences": preferences_obj}
    except Exception as e:
        logger.error(f"Erreur lors de la réinitialisation des préférences : {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")


@api_router.post("/user-preferences/migrate-menus")
async def migrate_menu_preferences(current_user: dict = Depends(get_current_user)):
    """Mettre à jour automatiquement les préférences pour ajouter les menus manquants"""
    try:
        user_id = current_user.get("id")
        
        # Liste complète des menus par défaut
        complete_menu_items = [
            { "id": "dashboard", "label": "Tableau de bord", "path": "/dashboard", "icon": "LayoutDashboard", "module": "dashboard", "visible": True, "favorite": False, "order": 0 },
            { "id": "service-dashboard", "label": "Dashboard Service", "path": "/service-dashboard", "icon": "Presentation", "module": "serviceDashboard", "visible": True, "favorite": False, "order": 0.5 },
            { "id": "chat-live", "label": "Chat Live", "path": "/chat-live", "icon": "Mail", "module": "chatLive", "visible": True, "favorite": False, "order": 0.8 },
            { "id": "intervention-requests", "label": "Demandes d'inter.", "path": "/intervention-requests", "icon": "MessageSquare", "module": "interventionRequests", "visible": True, "favorite": False, "order": 1 },
            { "id": "work-orders", "label": "Ordres de travail", "path": "/work-orders", "icon": "ClipboardList", "module": "workOrders", "visible": True, "favorite": False, "order": 2 },
            { "id": "improvement-requests", "label": "Demandes d'amél.", "path": "/improvement-requests", "icon": "Lightbulb", "module": "improvementRequests", "visible": True, "favorite": False, "order": 3 },
            { "id": "improvements", "label": "Améliorations", "path": "/improvements", "icon": "Sparkles", "module": "improvements", "visible": True, "favorite": False, "order": 4 },
            { "id": "preventive-maintenance", "label": "Maintenance prev.", "path": "/preventive-maintenance", "icon": "Calendar", "module": "preventiveMaintenance", "visible": True, "favorite": False, "order": 5 },
            { "id": "planning-mprev", "label": "Planning M.Prev.", "path": "/planning-mprev", "icon": "Calendar", "module": "planningMprev", "visible": True, "favorite": False, "order": 6 },
            { "id": "assets", "label": "Équipements", "path": "/assets", "icon": "Wrench", "module": "assets", "visible": True, "favorite": False, "order": 7 },
            { "id": "inventory", "label": "Inventaire", "path": "/inventory", "icon": "Package", "module": "inventory", "visible": True, "favorite": False, "order": 8 },
            { "id": "purchase-requests", "label": "Demandes d'Achat", "path": "/purchase-requests", "icon": "ShoppingCart", "module": "purchaseRequests", "visible": True, "favorite": False, "order": 8.5 },
            { "id": "locations", "label": "Zones", "path": "/locations", "icon": "MapPin", "module": "locations", "visible": True, "favorite": False, "order": 9 },
            { "id": "meters", "label": "Compteurs", "path": "/meters", "icon": "Gauge", "module": "meters", "visible": True, "favorite": False, "order": 10 },
            { "id": "surveillance-plan", "label": "Plan de Surveillance", "path": "/surveillance-plan", "icon": "Eye", "module": "surveillance", "visible": True, "favorite": False, "order": 11 },
            { "id": "surveillance-rapport", "label": "Rapport Surveillance", "path": "/surveillance-rapport", "icon": "FileText", "module": "surveillanceRapport", "visible": True, "favorite": False, "order": 12 },
            { "id": "weekly-reports", "label": "Rapports Hebdo.", "path": "/weekly-reports", "icon": "FileText", "module": "weeklyReports", "visible": True, "favorite": False, "order": 12.5 },
            { "id": "presqu-accident", "label": "Presqu'accident", "path": "/presqu-accident", "icon": "AlertTriangle", "module": "presquaccident", "visible": True, "favorite": False, "order": 13 },
            { "id": "presqu-accident-rapport", "label": "Rapport P.accident", "path": "/presqu-accident-rapport", "icon": "FileText", "module": "presquaccidentRapport", "visible": True, "favorite": False, "order": 14 },
            { "id": "documentations", "label": "Documentations", "path": "/documentations", "icon": "FolderOpen", "module": "documentations", "visible": True, "favorite": False, "order": 15 },
            { "id": "reports", "label": "Rapports", "path": "/reports", "icon": "BarChart3", "module": "reports", "visible": True, "favorite": False, "order": 16 },
            { "id": "team-management", "label": "Gestion d'équipe", "path": "/team-management", "icon": "UserCog", "module": "timeTracking", "visible": True, "favorite": False, "order": 16.5 },
            { "id": "cameras", "label": "Caméras", "path": "/cameras", "icon": "Camera", "module": "cameras", "visible": True, "favorite": False, "order": 16.6 },
            { "id": "mes", "label": "M.E.S.", "path": "/mes", "icon": "Zap", "module": "mes", "visible": True, "favorite": False, "order": 16.7 },
            { "id": "mes-reports", "label": "Rapports M.E.S.", "path": "/mes-reports", "icon": "FileBarChart", "module": "mesReports", "visible": True, "favorite": False, "order": 16.8 },
            { "id": "analytics-checklists", "label": "Analytics Checklists", "path": "/analytics/checklists", "icon": "BarChart3", "module": "analyticsChecklists", "visible": True, "favorite": False, "order": 16.9 },
            { "id": "people", "label": "Utilisateurs", "path": "/people", "icon": "Users", "module": "people", "visible": True, "favorite": False, "order": 17 },
            { "id": "planning", "label": "Planning", "path": "/planning", "icon": "Calendar", "module": "planning", "visible": True, "favorite": False, "order": 18 },
            { "id": "vendors", "label": "Fournisseurs", "path": "/vendors", "icon": "ShoppingCart", "module": "vendors", "visible": True, "favorite": False, "order": 19 },
            { "id": "contrats", "label": "Contrats", "path": "/contrats", "icon": "FileSignature", "module": "contrats", "visible": True, "favorite": False, "order": 19.5 },
            { "id": "purchase-history", "label": "Historique Achat", "path": "/purchase-history", "icon": "ShoppingBag", "module": "purchaseHistory", "visible": True, "favorite": False, "order": 20 },
            { "id": "import-export", "label": "Import / Export", "path": "/import-export", "icon": "Database", "module": "importExport", "visible": True, "favorite": False, "order": 21 },
            { "id": "sensors", "label": "Capteurs MQTT", "path": "/sensors", "icon": "Activity", "module": "sensors", "visible": True, "favorite": False, "order": 22 },
            { "id": "iot-dashboard", "label": "Dashboard IoT", "path": "/iot-dashboard", "icon": "BarChart3", "module": "iotDashboard", "visible": True, "favorite": False, "order": 23 },
            { "id": "mqtt-logs", "label": "Logs MQTT", "path": "/mqtt-logs", "icon": "Terminal", "module": "mqttLogs", "visible": True, "favorite": False, "order": 24 },
            { "id": "whiteboard", "label": "Tableau d'affichage", "path": "/whiteboard", "icon": "Presentation", "module": "whiteboard", "visible": True, "favorite": False, "order": 25 }
        ]
        
        # Récupérer les préférences actuelles
        preferences = await db.user_preferences.find_one({"user_id": user_id})
        
        if not preferences:
            # Si aucune préférence, utiliser la liste complète
            update_data = {
                "menu_items": complete_menu_items,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            await db.user_preferences.insert_one({
                "user_id": user_id,
                **update_data
            })
            return {"message": "Préférences créées avec tous les menus", "added_count": len(complete_menu_items)}
        
        # Récupérer les menus actuels
        current_menus = preferences.get("menu_items", [])
        current_menu_ids = {menu["id"] for menu in current_menus}
        
        # Trouver les menus manquants
        missing_menus = [menu for menu in complete_menu_items if menu["id"] not in current_menu_ids]
        
        if not missing_menus:
            return {"message": "Aucun menu manquant", "added_count": 0}
        
        # Ajouter les menus manquants en préservant l'ordre
        updated_menus = current_menus + missing_menus
        
        # Réordonner tous les menus
        for idx, menu in enumerate(updated_menus):
            menu["order"] = idx
        
        # Mettre à jour dans la base de données
        await db.user_preferences.update_one(
            {"user_id": user_id},
            {
                "$set": {
                    "menu_items": updated_menus,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        # Journaliser l'action
        await audit_service.log_action(
            user_id=user_id,
            user_name=current_user.get("name", ""),
            user_email=current_user.get("email", ""),
            action=ActionType.UPDATE,
            entity_type=EntityType.SETTINGS,
            entity_id=user_id,
            details=f"Migration des menus - {len(missing_menus)} menu(s) ajouté(s)"
        )
        
        return {
            "message": f"{len(missing_menus)} menu(s) ajouté(s) avec succès",
            "added_count": len(missing_menus),
            "added_menus": [menu["label"] for menu in missing_menus]
        }
        
    except Exception as e:
        logger.error(f"Erreur lors de la migration des menus : {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur serveur: {str(e)}")

# ==================== SMTP CONFIGURATION ROUTES ====================
@api_router.get("/smtp/config",
    summary="Configuration SMTP", response_model=SMTPConfig, tags=["Parametres"])
async def get_smtp_config(current_user: dict = Depends(get_current_admin_user)):
    """Récupérer la configuration SMTP actuelle (Admin uniquement)"""
    try:
        # Lire depuis les variables d'environnement
        config = SMTPConfig(
            smtp_host=os.environ.get('SMTP_HOST', 'smtp.gmail.com'),
            smtp_port=int(os.environ.get('SMTP_PORT', '587')),
            smtp_user=os.environ.get('SMTP_USER', ''),
            smtp_password='****' if os.environ.get('SMTP_PASSWORD') else '',  # Masquer le mot de passe
            smtp_from_email=os.environ.get('SMTP_FROM_EMAIL', ''),
            smtp_from_name=os.environ.get('SMTP_FROM_NAME', 'FSAO Iris'),
            smtp_use_tls=os.environ.get('SMTP_USE_TLS', 'true').lower() == 'true',
            frontend_url=os.environ.get('FRONTEND_URL', ''),
            backend_url=os.environ.get('BACKEND_URL', '')
        )
        return config
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la config SMTP: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/smtp/config",
    summary="Modifier la config SMTP", response_model=SuccessResponse, tags=["Parametres"])
async def update_smtp_config(
    smtp_update: SMTPConfigUpdate,
    current_user: dict = Depends(get_current_admin_user)
):
    """Mettre à jour la configuration SMTP (Admin uniquement)"""
    try:
        env_path = ROOT_DIR / '.env'
        
        # Lire le fichier .env actuel
        env_vars = {}
        if env_path.exists():
            with open(env_path, 'r') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, value = line.split('=', 1)
                        env_vars[key] = value
        
        # Mettre à jour les variables
        if smtp_update.smtp_host is not None:
            env_vars['SMTP_HOST'] = smtp_update.smtp_host
        if smtp_update.smtp_port is not None:
            env_vars['SMTP_PORT'] = str(smtp_update.smtp_port)
        if smtp_update.smtp_user is not None:
            env_vars['SMTP_USER'] = smtp_update.smtp_user
        if smtp_update.smtp_password is not None and smtp_update.smtp_password != '****':
            env_vars['SMTP_PASSWORD'] = smtp_update.smtp_password
        if smtp_update.smtp_from_email is not None:
            env_vars['SMTP_FROM_EMAIL'] = smtp_update.smtp_from_email
        if smtp_update.smtp_from_name is not None:
            env_vars['SMTP_FROM_NAME'] = smtp_update.smtp_from_name
        if smtp_update.smtp_use_tls is not None:
            env_vars['SMTP_USE_TLS'] = 'true' if smtp_update.smtp_use_tls else 'false'
        if smtp_update.frontend_url is not None:
            env_vars['FRONTEND_URL'] = smtp_update.frontend_url
        if smtp_update.backend_url is not None:
            env_vars['BACKEND_URL'] = smtp_update.backend_url
        
        # Écrire le fichier .env mis à jour
        with open(env_path, 'w') as f:
            for key, value in env_vars.items():
                f.write(f"{key}={value}\n")
        
        # Mettre à jour les variables d'environnement en mémoire
        for key, value in env_vars.items():
            os.environ[key] = value
        
        # Réinitialiser le service email avec la nouvelle configuration
        email_service.init_email_service()
        
        # Journaliser l'action
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType.SETTINGS,
            entity_id="smtp",
            entity_name="Configuration SMTP",
            details="Modification de la configuration SMTP"
        )
        
        return {"success": True, "message": "Configuration SMTP mise à jour avec succès"}
    except Exception as e:
        logger.error(f"Erreur lors de la mise à jour de la config SMTP: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/smtp/test",
    summary="Tester la config SMTP", response_model=SuccessResponse, tags=["Parametres"])
async def test_smtp_config(
    test_request: SMTPTestRequest,
    current_user: dict = Depends(get_current_admin_user)
):
    """Tester la configuration SMTP en envoyant un email de test (Admin uniquement)"""
    try:
        # Envoyer un email de test
        success = email_service.send_test_email(test_request.test_email)
        
        if success:
            # Journaliser l'action
            await audit_service.log_action(
                user_id=current_user["id"],
                user_name=f"{current_user['prenom']} {current_user['nom']}",
                user_email=current_user["email"],
                action=ActionType.UPDATE,
                entity_type=EntityType.SETTINGS,
                entity_id="smtp_test",
                entity_name="Test SMTP",
                details=f"Test d'envoi d'email vers {test_request.test_email}"
            )
            
            return {"success": True, "message": f"Email de test envoyé avec succès à {test_request.test_email}"}
        else:
            return {"success": False, "message": "Échec de l'envoi de l'email de test"}
    except Exception as e:
        logger.error(f"Erreur lors du test SMTP: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# ==================== SUPPORT HELP ROUTES ====================

# Stockage en mémoire des demandes d'aide par utilisateur (anti-spam)
help_request_tracker = {}


class SimpleSupportRequest(BaseModel):
    """Modèle pour une demande d'aide simple depuis la page Paramètres"""
    subject: Optional[str] = "Demande d'assistance"
    message: str


@api_router.post("/support/request",
    summary="Soumettre une demande de support", response_model=SuccessResponse, tags=["Support"])
async def submit_support_request(
    request: SimpleSupportRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Envoyer une demande d'aide simple aux administrateurs (depuis la page Paramètres)
    """
    try:
        user_id = current_user.get("id")
        user_name = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip()
        user_email = current_user.get('email', '')
        user_service = current_user.get('service', 'Non défini')
        user_role = current_user.get('role', 'N/A')
        
        # Anti-spam : Vérifier le nombre de demandes dans la dernière heure
        now = datetime.now(timezone.utc)
        one_hour_ago = now - timedelta(hours=1)
        
        if user_id in help_request_tracker:
            help_request_tracker[user_id] = [
                req_time for req_time in help_request_tracker[user_id] 
                if req_time > one_hour_ago
            ]
            if len(help_request_tracker[user_id]) >= 10:
                raise HTTPException(
                    status_code=429, 
                    detail="Limite de demandes atteinte. Veuillez réessayer dans 1 heure."
                )
        else:
            help_request_tracker[user_id] = []
        
        help_request_tracker[user_id].append(now)
        
        # Récupérer les emails des administrateurs
        admins = await db.users.find({"role": "ADMIN", "statut": "actif"}).to_list(100)
        admin_emails = [admin['email'] for admin in admins if admin.get('email')]
        
        if not admin_emails:
            raise HTTPException(
                status_code=500,
                detail="Aucun administrateur disponible pour recevoir la demande"
            )
        
        # Préparer les valeurs
        subject_display = request.subject or "Demande d'assistance"
        user_display = user_name or user_email
        date_display = now.strftime('%d/%m/%Y à %H:%M')
        
        # Créer l'email
        email_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <!-- En-tête -->
                <div style="background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%); padding: 20px; border-radius: 10px 10px 0 0; text-align: center;">
                    <h1 style="color: white; margin: 0; font-size: 22px;">Demande d'assistance</h1>
                </div>
                
                <!-- Corps -->
                <div style="background: white; padding: 25px; border: 1px solid #e0e0e0; border-top: none;">
                    <p style="margin: 0 0 20px 0;">Un utilisateur a envoyé une demande d'assistance via le Centre d'aide.</p>
                    
                    <!-- Informations utilisateur -->
                    <div style="background: #f8fafc; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                        <h3 style="margin: 0 0 10px 0; color: #1e40af; font-size: 14px;">Informations de l'utilisateur</h3>
                        <table style="width: 100%; font-size: 14px;">
                            <tr>
                                <td style="padding: 5px 0; color: #64748b; width: 100px;">Nom</td>
                                <td style="padding: 5px 0; font-weight: 500;">{user_display}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px 0; color: #64748b;">Email</td>
                                <td style="padding: 5px 0;">{user_email}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px 0; color: #64748b;">Service</td>
                                <td style="padding: 5px 0;">{user_service}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px 0; color: #64748b;">Rôle</td>
                                <td style="padding: 5px 0;">{user_role}</td>
                            </tr>
                        </table>
                    </div>
                    
                    <!-- Sujet -->
                    <div style="margin-bottom: 15px;">
                        <h3 style="margin: 0 0 5px 0; color: #1e40af; font-size: 14px;">Sujet</h3>
                        <p style="margin: 0; padding: 10px; background: #eff6ff; border-radius: 5px; font-weight: 500;">
                            {subject_display}
                        </p>
                    </div>
                    
                    <!-- Message -->
                    <div style="margin-bottom: 20px;">
                        <h3 style="margin: 0 0 5px 0; color: #1e40af; font-size: 14px;">Message</h3>
                        <div style="padding: 15px; background: #fefce8; border-left: 4px solid #eab308; border-radius: 0 5px 5px 0;">
                            <p style="margin: 0; white-space: pre-wrap;">{request.message}</p>
                        </div>
                    </div>
                    
                    <!-- Action -->
                    <div style="text-align: center; padding: 15px; background: #f0fdf4; border-radius: 8px;">
                        <p style="margin: 0 0 10px 0; color: #166534;">
                            Veuillez répondre directement à cet utilisateur par email.
                        </p>
                        <a href="mailto:{user_email}?subject=Re: {subject_display}" 
                           style="display: inline-block; padding: 10px 25px; background-color: #22c55e; 
                                  color: white; text-decoration: none; border-radius: 5px; font-weight: 500;">
                            Répondre à {user_display}
                        </a>
                    </div>
                </div>
                
                <!-- Pied de page -->
                <div style="background: #f5f5f5; padding: 15px; border-radius: 0 0 10px 10px; border: 1px solid #e0e0e0; border-top: none;">
                    <p style="color: #aaa; font-size: 10px; margin: 0; text-align: center;">
                        Demande envoyée le {date_display} depuis FSAO Iris - Centre d'aide
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Envoyer l'email à tous les admins
        subject = f"[FSAO Support] {request.subject} - {user_name or user_email}"
        
        for admin_email in admin_emails:
            try:
                email_service.send_email(
                    to_email=admin_email,
                    subject=subject,
                    html_content=email_html
                )
            except Exception as e:
                logger.warning(f"Erreur envoi email support à {admin_email}: {e}")
        
        # Sauvegarder la demande en base de données
        support_request_data = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "user_name": user_name,
            "user_email": user_email,
            "user_service": user_service,
            "subject": request.subject,
            "message": request.message,
            "status": "pending",
            "created_at": now.isoformat(),
            "notified_admins": admin_emails
        }
        
        await db.support_requests.insert_one(support_request_data)
        
        logger.info(f"📬 Demande de support reçue de {user_email}: {request.subject}")
        
        return {
            "success": True,
            "message": "Votre demande a été envoyée aux administrateurs"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur envoi demande support: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/support/request-help", response_model=HelpRequestResponse, tags=["Support"])
async def request_help(
    help_request: HelpRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Envoyer une demande d'aide aux administrateurs
    Limitation : 15 demandes par heure par utilisateur
    """
    try:
        user_id = current_user.get("id")
        user_name = f"{current_user['prenom']} {current_user['nom']}"
        user_email = current_user['email']
        
        # Anti-spam : Vérifier le nombre de demandes dans la dernière heure
        now = datetime.now(timezone.utc)
        one_hour_ago = now - timedelta(hours=1)
        
        if user_id in help_request_tracker:
            # Nettoyer les anciennes requêtes
            help_request_tracker[user_id] = [
                req_time for req_time in help_request_tracker[user_id] 
                if req_time > one_hour_ago
            ]
            
            # Vérifier la limite
            if len(help_request_tracker[user_id]) >= 15:
                raise HTTPException(
                    status_code=429, 
                    detail="Limite de demandes d'aide atteinte. Veuillez réessayer dans 1 heure."
                )
        else:
            help_request_tracker[user_id] = []
        
        # Enregistrer cette demande
        help_request_tracker[user_id].append(now)
        
        # Générer un ID unique pour cette demande
        request_id = str(uuid.uuid4())
        
        # Récupérer tous les administrateurs
        admins = await db.users.find({"role": "ADMIN"}).to_list(100)
        admin_emails = [admin['email'] for admin in admins if admin.get('email')]
        
        if not admin_emails:
            raise HTTPException(
                status_code=500,
                detail="Aucun administrateur trouvé pour recevoir la demande"
            )
        
        # Préparer les données du screenshot (décoder base64 si nécessaire)
        screenshot_data = help_request.screenshot
        if screenshot_data.startswith('data:image'):
            # Extraire seulement les données base64
            screenshot_data = screenshot_data.split(',')[1]
        
        # Créer le contenu HTML de l'email
        email_html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 800px; margin: 0 auto; padding: 20px; }}
                .header {{ background-color: #dc2626; color: white; padding: 20px; border-radius: 8px 8px 0 0; }}
                .content {{ background-color: #f9fafb; padding: 20px; border: 1px solid #e5e7eb; }}
                .info-section {{ background-color: white; padding: 15px; margin: 10px 0; border-radius: 6px; border-left: 4px solid #2563eb; }}
                .label {{ font-weight: bold; color: #1f2937; }}
                .value {{ color: #4b5563; margin-left: 10px; }}
                .message-box {{ background-color: #fef3c7; padding: 15px; border-left: 4px solid #f59e0b; margin: 15px 0; }}
                .logs-box {{ background-color: #fee2e2; padding: 15px; border-left: 4px solid #dc2626; margin: 15px 0; font-family: monospace; font-size: 12px; }}
                .footer {{ text-align: center; color: #6b7280; font-size: 12px; margin-top: 20px; }}
                img {{ max-width: 100%; height: auto; border: 2px solid #e5e7eb; border-radius: 8px; margin-top: 15px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🆘 Demande d'Aide - FSAO Iris</h1>
                    <p style="margin: 5px 0;">ID: {request_id}</p>
                </div>
                
                <div class="content">
                    <div class="info-section">
                        <p><span class="label">👤 Utilisateur:</span><span class="value">{user_name} ({user_email})</span></p>
                        <p><span class="label">📄 Page:</span><span class="value">{help_request.page_url}</span></p>
                        <p><span class="label">🌐 Navigateur:</span><span class="value">{help_request.browser_info}</span></p>
                        <p><span class="label">🕐 Date/Heure:</span><span class="value">{now.strftime('%d/%m/%Y %H:%M:%S')} UTC</span></p>
                    </div>
                    
                    {f'''
                    <div class="message-box">
                        <h3 style="margin-top: 0;">💬 Message de l'utilisateur:</h3>
                        <p>{help_request.user_message}</p>
                    </div>
                    ''' if help_request.user_message else ''}
                    
                    {f'''
                    <div class="logs-box">
                        <h3 style="margin-top: 0; color: #dc2626;">⚠️ Logs Console (Erreurs):</h3>
                        {"<br>".join(help_request.console_logs[:10])}
                    </div>
                    ''' if help_request.console_logs else ''}
                    
                    <h3>📸 Capture d'écran:</h3>
                    <p style="color: #6b7280;">Voir la pièce jointe : screenshot.png</p>
                </div>
                
                <div class="footer">
                    <p>Cette demande d'aide a été générée automatiquement par FSAO Iris</p>
                    <p>Pour répondre à l'utilisateur, envoyez un email à: {user_email}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Envoyer l'email à tous les administrateurs avec la capture d'écran en pièce jointe
        try:
            subject = f"🆘 Demande d'Aide - {user_name} - {help_request.page_url}"
            
            # Décoder le screenshot base64 en bytes
            import base64
            screenshot_bytes = base64.b64decode(screenshot_data)
            screenshot_filename = f'screenshot_{request_id[:8]}.png'
            
            for admin_email in admin_emails:
                email_service.send_email_with_attachment(
                    to_email=admin_email,
                    subject=subject,
                    html_content=email_html,
                    attachment_data=screenshot_bytes,
                    attachment_filename=screenshot_filename
                )
            
            # Journaliser l'action
            await audit_service.log_action(
                user_id=user_id,
                user_name=user_name,
                user_email=user_email,
                action=ActionType.CREATE,
                entity_type=EntityType.SETTINGS,  # Utiliser SETTINGS comme type générique
                entity_id=request_id,
                entity_name="Demande d'aide",
                details=f"Demande d'aide envoyée depuis {help_request.page_url} à {len(admin_emails)} administrateur(s)"
            )
            
            logger.info(f"✅ Demande d'aide {request_id} envoyée à {len(admin_emails)} administrateur(s)")
            
            return HelpRequestResponse(
                success=True,
                message=f"Demande d'aide envoyée avec succès à {len(admin_emails)} administrateur(s)",
                request_id=request_id
            )
            
        except Exception as email_error:
            logger.error(f"❌ Erreur lors de l'envoi de l'email d'aide: {str(email_error)}")
            raise HTTPException(
                status_code=500,
                detail=f"Erreur lors de l'envoi de l'email: {str(email_error)}"
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur lors du traitement de la demande d'aide: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== VENDORS ROUTES ====================
@api_router.get("/vendors",
    summary="Lister les fournisseurs", response_model=List[Vendor], tags=["Fournisseurs"])
async def get_vendors(current_user: dict = Depends(require_permission("vendors", "view"))):
    """Liste tous les fournisseurs"""
    vendors = await db.vendors.find().to_list(1000)
    return [Vendor(**serialize_doc(vendor)) for vendor in vendors]

@api_router.post("/vendors",
    summary="Creer un fournisseur", response_model=Vendor, tags=["Fournisseurs"])
async def create_vendor(vendor_create: VendorCreate, current_user: dict = Depends(require_permission("vendors", "edit"))):
    """Créer un nouveau fournisseur"""
    vendor_dict = vendor_create.model_dump()
    vendor_dict["dateCreation"] = datetime.utcnow()
    vendor_dict["_id"] = ObjectId()
    
    await db.vendors.insert_one(vendor_dict)
    
    vendor_data = serialize_doc(vendor_dict)
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "suppliers",
        "created",
        vendor_data,
        user_id=current_user.get("id")
    )
    
    return Vendor(**vendor_data)

@api_router.put("/vendors/{vendor_id}",
    summary="Modifier un fournisseur", response_model=Vendor, tags=["Fournisseurs"])
async def update_vendor(vendor_id: str, vendor_update: VendorUpdate, current_user: dict = Depends(require_permission("vendors", "edit"))):
    """Modifier un fournisseur"""
    try:
        update_data = {k: v for k, v in vendor_update.model_dump().items() if v is not None}
        
        await db.vendors.update_one(
            {"_id": ObjectId(vendor_id)},
            {"$set": update_data}
        )
        
        vendor = await db.vendors.find_one({"_id": ObjectId(vendor_id)})
        vendor_data = serialize_doc(vendor)
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "suppliers",
            "updated",
            vendor_data,
            user_id=current_user.get("id")
        )
        
        return Vendor(**vendor_data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/vendors/{vendor_id}", response_model=MessageResponse,
    summary="Supprimer un fournisseur", tags=["Fournisseurs"])
async def delete_vendor(vendor_id: str, current_user: dict = Depends(require_permission("vendors", "delete"))):
    """Supprimer un fournisseur"""
    try:
        # Récupérer le fournisseur avant suppression pour le broadcast
        vendor = await db.vendors.find_one({"_id": ObjectId(vendor_id)})
        vendor_name = vendor.get("nom", "Inconnu") if vendor else "Inconnu"
        
        result = await db.vendors.delete_one({"_id": ObjectId(vendor_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Fournisseur non trouvé")
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "suppliers",
            "deleted",
            {"id": vendor_id, "nom": vendor_name},
            user_id=current_user.get("id")
        )
        
        return {"message": "Fournisseur supprimé"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ==================== VENDOR AI EXTRACT ====================
@api_router.post("/vendors/ai/extract",
    summary="Extraire les informations fournisseur d'un document via IA", tags=["Fournisseurs"])
async def extract_vendor_from_document(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission("vendors", "edit"))
):
    """
    Analyse un document (Excel, PDF, image) via IA et extrait les informations
    pour créer une fiche fournisseur.
    Supporte: PDF, images, Excel (converti en texte avant envoi à l'IA)
    """
    import tempfile
    import json as json_mod

    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType

        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Clé LLM non configurée")

        ext = os.path.splitext(file.filename)[1].lower()

        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        # Formats supportés nativement par Gemini (fichier binaire)
        native_formats = {".pdf", ".png", ".jpg", ".jpeg", ".webp"}
        # Formats à convertir en texte
        spreadsheet_formats = {".xlsx", ".xls", ".csv"}
        
        system_prompt = """Tu es un assistant spécialisé dans l'extraction d'informations fournisseurs à partir de documents administratifs et commerciaux.

Analyse le document fourni et extrais TOUTES les informations relatives au fournisseur.
Le document peut être un formulaire de création fournisseur, un devis, une facture, un bon de commande, un contrat, ou tout autre document commercial.

Réponds UNIQUEMENT avec un JSON valide, sans texte autour ni backticks.

Format attendu:
{
  "nom": "Nom de la société/entreprise",
  "contact": "Nom du contact principal (Prénom Nom) ou null",
  "contact_fonction": "Fonction/poste du contact ou null",
  "email": "Email du contact ou de l'entreprise ou null",
  "telephone": "Numéro de téléphone ou null",
  "adresse": "Adresse complète (rue) ou null",
  "code_postal": "Code postal ou null",
  "ville": "Ville ou null",
  "pays": "Code pays (FR, DE, LU, etc.) ou null",
  "specialite": "Domaine d'activité/spécialité déduit du document",
  "tva_intra": "N° TVA intracommunautaire ou null",
  "siret": "N° SIRET/SIREN ou numéro d'enregistrement ou null",
  "conditions_paiement": "valeur parmi: 30J_NET, 30J_FDM, 45J_FDM, 60J_FDM, 90J_FDM ou null",
  "devise": "EUR, USD, GBP, etc. ou null",
  "categorie": "valeur parmi: MAINTENANCE, FOURNITURES, SERVICES, EQUIPEMENTS, SOUS_TRAITANCE, ENERGIE, INFORMATIQUE, LOGISTIQUE, NETTOYAGE, SECURITE, AUTRE ou null",
  "sous_traitant": false,
  "site_web": "URL du site web ou null",
  "notes": "Informations complémentaires utiles extraites du document ou null",
  "confidence": 0.8
}

RÈGLES:
- Si une information n'est pas trouvée, mets null
- Pour le nom de société, cherche: raison sociale, nom commercial, dénomination
- Pour le contact, cherche: interlocuteur, responsable, signataire
- Déduis la spécialité et la catégorie à partir du contenu du document
- Le champ conditions_paiement doit correspondre EXACTEMENT à une des valeurs listées
- Le champ categorie doit correspondre EXACTEMENT à une des valeurs listées
- Extrais le maximum d'informations possibles"""

        chat = LlmChat(
            api_key=api_key,
            session_id=f"vendor_extract_{uuid.uuid4().hex[:8]}",
            system_message=system_prompt
        ).with_model("gemini", "gemini-2.5-flash")

        if ext in spreadsheet_formats:
            # Convertir Excel/CSV en texte pour l'envoyer à Gemini
            text_content = ""
            try:
                import openpyxl
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text_content += f"=== Feuille: {sheet_name} ===\n"
                    for row in ws.iter_rows(values_only=False):
                        row_texts = []
                        for cell in row:
                            if cell.value is not None:
                                row_texts.append(f"{str(cell.value).strip()}")
                        if any(t for t in row_texts):
                            text_content += " | ".join(row_texts) + "\n"
                    text_content += "\n"
            except Exception:
                # Fallback: lire comme CSV
                try:
                    with open(tmp_path, 'r', encoding='utf-8', errors='replace') as f:
                        text_content = f.read()
                except Exception:
                    text_content = "Impossible de lire le fichier"

            response = await chat.send_message(
                UserMessage(
                    text=f"Voici le contenu extrait d'un document fournisseur ({file.filename}). Analyse-le et extrais les informations du fournisseur. Réponds uniquement en JSON.\n\n---\n{text_content[:15000]}"
                )
            )
        else:
            # Formats natifs (PDF, images) — envoi direct du fichier
            mime_map = {
                ".pdf": "application/pdf",
                ".png": "image/png",
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".webp": "image/webp",
            }
            mime_type = mime_map.get(ext, "application/octet-stream")
            
            response = await chat.send_message(
                UserMessage(
                    text="Analyse ce document et extrais les informations du fournisseur. Réponds uniquement en JSON.",
                    file_contents=[FileContentWithMimeType(file_path=tmp_path, mime_type=mime_type)]
                )
            )

        # Nettoyer le fichier temporaire
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

        # Parser la réponse JSON (response est un string directement)
        response_text = response.strip()
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        response_text = response_text.strip()

        extracted_data = json_mod.loads(response_text)

        return {
            "success": True,
            "extracted_data": extracted_data,
            "source_filename": file.filename
        }

    except json_mod.JSONDecodeError as e:
        logger.error(f"Erreur parsing JSON IA: {str(e)}")
        raise HTTPException(status_code=422, detail=f"L'IA n'a pas retourné un JSON valide: {str(e)}")
    except ImportError:
        raise HTTPException(status_code=500, detail="Module IA non disponible (emergentintegrations)")
    except Exception as e:
        logger.error(f"Erreur extraction IA fournisseur: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'analyse IA: {str(e)}")



# ==================== PURCHASE HISTORY ROUTES ====================
@api_router.get("/purchase-history/template", tags=["Historique Achats"])
async def download_purchase_history_template(
    format: str = "csv",
    current_user: dict = Depends(require_permission("purchaseHistory", "view"))
):
    """Telecharger le template CSV pour l'import d'historique d'achat"""
    import io
    import csv
    
    headers = [
        "fournisseur", "numeroCommande", "numeroReception", "dateCreation",
        "article", "description", "groupeStatistique", "quantite",
        "montantLigneHT", "quantiteRetournee", "site", "creationUser"
    ]
    
    example_row = [
        "Fournisseur ABC", "CMD-2026-001", "REC-2026-001", "2026-01-15",
        "Roulement SKF 6205", "Roulement a billes", "Pieces mecaniques", "10",
        "150.00", "0", "Site principal", "admin"
    ]
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(headers)
    writer.writerow(example_row)
    
    content = output.getvalue()
    output.close()
    
    from starlette.responses import Response
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_historique_achat.csv"}
    )

@api_router.get("/purchase-history/grouped", tags=["Historique Achats"])
async def get_purchase_history_grouped(current_user: dict = Depends(require_permission("purchaseHistory", "view"))):
    """Liste tous les achats groupés par N° Commande"""
    purchases = await db.purchase_history.find().sort("dateCreation", -1).to_list(5000)
    
    # Grouper par numeroCommande
    grouped = {}
    for p in purchases:
        num_cmd = p.get('numeroCommande')
        if not num_cmd:
            continue
            
        if num_cmd not in grouped:
            # Utiliser Fournisseur2 (colonne M) si disponible, sinon fournisseur
            fournisseur_display = p.get('Fournisseur2') or p.get('fournisseur', 'Inconnu')
            
            grouped[num_cmd] = {
                'numeroCommande': num_cmd,
                'fournisseur': fournisseur_display,
                'numeroReception': p.get('numeroReception'),  # Premier N° reception de la commande
                'dateCreation': p.get('dateCreation'),
                'site': p.get('site'),
                'items': [],
                'montantTotal': 0.0,
                'itemCount': 0
            }
        
        # Ajouter l'item au groupe
        item_data = {
            'article': p.get('article'),
            'description': p.get('description'),
            'quantite': p.get('quantite', 0.0),
            'montantLigneHT': p.get('montantLigneHT', 0.0),
            'numeroReception': p.get('numeroReception'),
            'groupeStatistique': p.get('groupeStatistique')
        }
        
        grouped[num_cmd]['items'].append(item_data)
        grouped[num_cmd]['montantTotal'] += item_data['montantLigneHT']
        grouped[num_cmd]['itemCount'] += 1
    
    # Convertir en liste
    result = list(grouped.values())
    return result


@api_router.delete("/purchase-history/all", tags=["Historique Achats"])
async def delete_all_purchase_history(current_user: dict = Depends(get_current_admin_user)):
    """Supprimer tout l'historique d'achat (admin uniquement)"""
    result = await db.purchase_history.delete_many({})
    return {
        "message": f"{result.deleted_count} achats supprimés",
        "deleted_count": result.deleted_count
    }


@api_router.get("/purchase-history", response_model=List[PurchaseHistory], tags=["Historique Achats"])
async def get_purchase_history(current_user: dict = Depends(require_permission("purchaseHistory", "view"))):
    """Liste tous les achats"""
    purchases = await db.purchase_history.find().sort("dateCreation", -1).to_list(5000)
    
    # Filtrer pour ne garder que les champs du modèle PurchaseHistory
    allowed_fields = {
        '_id', 'id', 'fournisseur', 'numeroCommande', 'numeroReception', 
        'dateCreation', 'article', 'description', 'groupeStatistique',
        'quantite', 'montantLigneHT', 'quantiteRetournee', 'site', 
        'creationUser', 'dateEnregistrement'
    }
    
    result = []
    for p in purchases:
        # Ne garder que les champs autorisés
        filtered_doc = {k: v for k, v in p.items() if k in allowed_fields}
        
        # S'assurer que les champs obligatoires existent avec des valeurs par défaut
        if 'montantLigneHT' not in filtered_doc or filtered_doc['montantLigneHT'] is None:
            filtered_doc['montantLigneHT'] = 0.0
        if 'quantite' not in filtered_doc or filtered_doc['quantite'] is None:
            filtered_doc['quantite'] = 0.0
        if 'quantiteRetournee' not in filtered_doc:
            filtered_doc['quantiteRetournee'] = 0.0
        
        try:
            result.append(PurchaseHistory(**serialize_doc(filtered_doc)))
        except Exception as e:
            logger.error(f"Erreur serialization purchase {filtered_doc.get('numeroCommande')}: {e}")
            continue
    
    return result

@api_router.get("/purchase-history/stats", tags=["Historique Achats"])
async def get_purchase_stats(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(require_permission("purchaseHistory", "view"))
):
    """Statistiques complètes des achats"""
    
    # Filtres de date
    match_filter = {}
    if start_date:
        match_filter["dateCreation"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "dateCreation" in match_filter:
            match_filter["dateCreation"]["$lte"] = datetime.fromisoformat(end_date)
        else:
            match_filter["dateCreation"] = {"$lte": datetime.fromisoformat(end_date)}
    
    # Total des achats
    all_purchases = await db.purchase_history.find(match_filter).to_list(10000)
    
    if not all_purchases:
        return {
            "totalAchats": 0,
            "montantTotal": 0,
            "commandesTotales": 0,
            "parFournisseur": [],
            "parMois": [],
            "parSite": [],
            "parGroupeStatistique": [],
            "articlesTop": [],
            "par_utilisateur": [],
            "par_mois": [],
            "par_mois_categories": []
        }
    
    total_achats = len(all_purchases)
    montant_total = sum(p.get("montantLigneHT", 0) for p in all_purchases)
    
    # Compter les commandes uniques (pas les lignes)
    commandes_uniques = set()
    for p in all_purchases:
        num_cmd = p.get("numeroCommande")
        if num_cmd:
            commandes_uniques.add(num_cmd)
    
    commandes_totales = len(commandes_uniques)
    
    # NOUVELLES STATS - Par utilisateur (créateur colonne L)
    user_stats = {}
    for purchase in all_purchases:
        user = purchase.get('creationUser', 'Inconnu')
        num_commande = purchase.get('numeroCommande')
        montant = purchase.get('montantLigneHT', 0)
        
        if user not in user_stats:
            user_stats[user] = {
                'utilisateur': user,
                'commandes': set(),
                'montant_total': 0,
                'nb_lignes': 0
            }
        
        if num_commande:
            user_stats[user]['commandes'].add(num_commande)
        user_stats[user]['montant_total'] += montant
        user_stats[user]['nb_lignes'] += 1
    
    # Convertir en liste
    users_list = []
    for user, data in user_stats.items():
        nb_commandes = len(data['commandes'])
        montant = data['montant_total']
        pourcentage = (montant / montant_total * 100) if montant_total > 0 else 0
        
        users_list.append({
            'utilisateur': user,
            'nb_commandes': nb_commandes,
            'nb_lignes': data['nb_lignes'],
            'montant_total': round(montant, 2),
            'pourcentage': round(pourcentage, 2)
        })
    
    users_list.sort(key=lambda x: x['montant_total'], reverse=True)
    
    # NOUVELLES STATS - Par mois avec catégorisation PAR (ARTICLE + DM6)
    monthly_stats = {}
    monthly_article_dm6_stats = {}  # Structure: {mois: {(article, dm6): {category, montant, etc.}}}
    
    for purchase in all_purchases:
        date = purchase.get('dateCreation')
        if date:
            if isinstance(date, str):
                date = datetime.fromisoformat(date.replace('Z', '+00:00'))
            month_key = date.strftime('%Y-%m')
            num_commande = purchase.get('numeroCommande')
            montant = purchase.get('montantLigneHT', 0)
            article = purchase.get('article', '')
            dm6 = purchase.get('DM6', 'Non défini')
            
            # Stats globales par mois
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {
                    'mois': month_key,
                    'commandes': set(),
                    'montant': 0,
                    'nb_lignes': 0
                }
            
            if num_commande:
                monthly_stats[month_key]['commandes'].add(num_commande)
            monthly_stats[month_key]['montant'] += montant
            monthly_stats[month_key]['nb_lignes'] += 1
            
            # Stats par (ARTICLE, DM6) - CHAQUE COMBINAISON est unique!
            try:
                category = get_category_from_article_dm6(article, dm6)
            except Exception as e:
                category = "Non catégorisé"
            
            if month_key not in monthly_article_dm6_stats:
                monthly_article_dm6_stats[month_key] = {}
            
            # Clé unique = (ARTICLE, DM6) - PAS juste DM6!
            unique_key = (article, dm6)
            
            if unique_key not in monthly_article_dm6_stats[month_key]:
                monthly_article_dm6_stats[month_key][unique_key] = {
                    'article': article,
                    'dm6': dm6,
                    'categorie': category,
                    'montant': 0,
                    'nb_lignes': 0,
                    'commandes': set()
                }
            
            monthly_article_dm6_stats[month_key][unique_key]['montant'] += montant
            monthly_article_dm6_stats[month_key][unique_key]['nb_lignes'] += 1
            if num_commande:
                monthly_article_dm6_stats[month_key][unique_key]['commandes'].add(num_commande)
    
    # Créer la liste mensuelle globale
    monthly_list = []
    for month, data in monthly_stats.items():
        monthly_list.append({
            'mois': month,
            'nb_commandes': len(data['commandes']),
            'nb_lignes': data['nb_lignes'],
            'montant_total': round(data['montant'], 2)
        })
    monthly_list.sort(key=lambda x: x['mois'])
    
    # Créer la liste mensuelle par (ARTICLE, DM6)
    monthly_category_list = []
    for month in sorted(monthly_article_dm6_stats.keys()):
        month_data = {
            'mois': month,
            'categories': []  # Garder 'categories' pour compatibilité frontend
        }
        
        for (article, dm6), data in monthly_article_dm6_stats[month].items():
            month_data['categories'].append({
                'article': article,
                'dm6': dm6,
                'nom': data['categorie'],
                'montant': round(data['montant'], 2),
                'nb_lignes': data['nb_lignes'],
                'nb_commandes': len(data['commandes'])
            })
        
        # Trier par montant décroissant
        month_data['categories'].sort(key=lambda x: x['montant'], reverse=True)
        monthly_category_list.append(month_data)
    
    # Par fournisseur (ancienne stat - gardée)
    fournisseurs = {}
    for p in all_purchases:
        fournisseur = p.get("Fournisseur2") or p.get("fournisseur", "Inconnu")
        if fournisseur not in fournisseurs:
            fournisseurs[fournisseur] = {"montant": 0, "quantite": 0, "count": 0}
        fournisseurs[fournisseur]["montant"] += p.get("montantLigneHT", 0)
        fournisseurs[fournisseur]["quantite"] += p.get("quantite", 0)
        fournisseurs[fournisseur]["count"] += 1
    
    par_fournisseur = [
        {
            "fournisseur": k,
            "montant": v["montant"],
            "quantite": v["quantite"],
            "count": v["count"],
            "pourcentage": round((v["montant"] / montant_total * 100) if montant_total > 0 else 0, 2)
        }
        for k, v in sorted(fournisseurs.items(), key=lambda x: x[1]["montant"], reverse=True)
    ]
    
    # Par mois (ancien format - gardé pour compatibilité)
    mois_dict = {}
    for p in all_purchases:
        date_creation = p.get("dateCreation")
        if date_creation:
            if isinstance(date_creation, str):
                date_creation = datetime.fromisoformat(date_creation.replace('Z', '+00:00'))
            mois_annee = date_creation.strftime("%Y-%m")
            if mois_annee not in mois_dict:
                mois_dict[mois_annee] = {"montant": 0, "quantite": 0, "count": 0}
            mois_dict[mois_annee]["montant"] += p.get("montantLigneHT", 0)
            mois_dict[mois_annee]["quantite"] += p.get("quantite", 0)
            mois_dict[mois_annee]["count"] += 1
    
    par_mois = [
        {"mois": k, "montant": v["montant"], "quantite": v["quantite"], "count": v["count"]}
        for k, v in sorted(mois_dict.items())
    ]
    
    # Par site
    sites = {}
    for p in all_purchases:
        site = p.get("site", "Non défini")
        if site not in sites:
            sites[site] = {"montant": 0, "quantite": 0, "count": 0}
        sites[site]["montant"] += p.get("montantLigneHT", 0)
        sites[site]["quantite"] += p.get("quantite", 0)
        sites[site]["count"] += 1
    
    par_site = [
        {"site": k, "montant": v["montant"], "quantite": v["quantite"], "count": v["count"]}
        for k, v in sorted(sites.items(), key=lambda x: x[1]["montant"], reverse=True)
    ]
    
    # Par groupe statistique
    groupes = {}
    for p in all_purchases:
        groupe = p.get("groupeStatistique", "Non défini")
        if groupe not in groupes:
            groupes[groupe] = {"montant": 0, "quantite": 0, "count": 0}
        groupes[groupe]["montant"] += p.get("montantLigneHT", 0)
        groupes[groupe]["quantite"] += p.get("quantite", 0)
        groupes[groupe]["count"] += 1
    
    par_groupe = [
        {"groupe": k, "montant": v["montant"], "quantite": v["quantite"], "count": v["count"]}
        for k, v in sorted(groupes.items(), key=lambda x: x[1]["montant"], reverse=True)
    ]
    
    # Articles top
    articles = {}
    for p in all_purchases:
        article = p.get("article", "Inconnu")
        if article not in articles:
            articles[article] = {"montant": 0, "quantite": 0, "count": 0, "description": p.get("description", "")}
        articles[article]["montant"] += p.get("montantLigneHT", 0)
        articles[article]["quantite"] += p.get("quantite", 0)
        articles[article]["count"] += 1
    
    articles_top = [
        {"article": k, **v}
        for k, v in sorted(articles.items(), key=lambda x: x[1]["montant"], reverse=True)[:20]
    ]
    
    return {
        "totalAchats": total_achats,
        "montantTotal": round(montant_total, 2),
        "commandesTotales": commandes_totales,
        "parFournisseur": par_fournisseur,
        "parMois": par_mois,
        "parSite": par_site,
        "parGroupeStatistique": par_groupe,
        "articlesTop": articles_top,
        "par_utilisateur": users_list,  # NOUVELLES STATS
        "par_mois": monthly_list,  # NOUVELLES STATS (format différent)
        "par_mois_categories": monthly_category_list  # NOUVELLES STATS - Catégorisation mensuelle
    }

@api_router.post("/purchase-history", response_model=PurchaseHistory, tags=["Historique Achats"])
async def create_purchase(purchase: PurchaseHistoryCreate, current_user: dict = Depends(require_permission("purchaseHistory", "edit"))):
    """Créer un nouvel achat"""
    purchase_dict = purchase.model_dump()
    
    # Convertir datetime en ISO string si nécessaire
    if isinstance(purchase_dict.get("dateCreation"), datetime):
        purchase_dict["dateCreation"] = purchase_dict["dateCreation"].isoformat()
    
    purchase_dict["dateEnregistrement"] = datetime.utcnow()
    purchase_dict["_id"] = ObjectId()
    
    # Ajouter l'utilisateur créateur si non fourni
    if not purchase_dict.get("creationUser"):
        purchase_dict["creationUser"] = current_user.get("email")
    
    await db.purchase_history.insert_one(purchase_dict)
    
    return PurchaseHistory(**serialize_doc(purchase_dict))

@api_router.put("/purchase-history/{purchase_id}", response_model=PurchaseHistory, tags=["Historique Achats"])
async def update_purchase(purchase_id: str, purchase_update: PurchaseHistoryUpdate, current_user: dict = Depends(require_permission("purchaseHistory", "edit"))):
    """Modifier un achat"""
    try:
        update_data = {k: v for k, v in purchase_update.model_dump().items() if v is not None}
        
        # Convertir datetime en ISO string si nécessaire
        if "dateCreation" in update_data and isinstance(update_data["dateCreation"], datetime):
            update_data["dateCreation"] = update_data["dateCreation"].isoformat()
        
        await db.purchase_history.update_one(
            {"_id": ObjectId(purchase_id)},
            {"$set": update_data}
        )
        
        purchase = await db.purchase_history.find_one({"_id": ObjectId(purchase_id)})
        return PurchaseHistory(**serialize_doc(purchase))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.delete("/purchase-history/{purchase_id}", tags=["Historique Achats"])
async def delete_purchase(purchase_id: str, current_user: dict = Depends(require_permission("purchaseHistory", "delete"))):
    """Supprimer un achat"""
    try:
        result = await db.purchase_history.delete_one({"_id": ObjectId(purchase_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Achat non trouvé")
        return {"message": "Achat supprimé"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ==================== REPORTS/ANALYTICS ROUTES ====================
@api_router.get("/reports/analytics")
async def get_analytics(current_user: dict = Depends(require_permission("reports", "view"))):
    """Obtenir les données analytiques générales"""
    # Work orders stats
    total_wo = await db.work_orders.count_documents({})
    wo_by_status = {}
    for status in ["OUVERT", "EN_COURS", "EN_ATTENTE", "TERMINE"]:
        count = await db.work_orders.count_documents({"statut": status})
        wo_by_status[status] = count
    
    wo_by_priority = {}
    for priority in ["HAUTE", "MOYENNE", "BASSE", "AUCUNE"]:
        count = await db.work_orders.count_documents({"priorite": priority})
        wo_by_priority[priority] = count
    
    # Equipment stats
    eq_by_status = {}
    for status in ["OPERATIONNEL", "EN_MAINTENANCE", "HORS_SERVICE"]:
        count = await db.equipments.count_documents({"statut": status})
        eq_by_status[status] = count
    
    # Simple mock data for costs and time response
    analytics = {
        "workOrdersParStatut": wo_by_status,
        "workOrdersParPriorite": wo_by_priority,
        "equipementsParStatut": eq_by_status,
        "coutsMaintenance": {
            "janvier": 4500,
            "decembre": 3200,
            "novembre": 2800,
            "octobre": 3500,
            "septembre": 2900,
            "aout": 3100
        },
        "tempsReponse": {
            "moyen": 2.5,
            "median": 2,
            "min": 1,
            "max": 6
        },
        "tauxRealisation": 87,
        "nombreMaintenancesPrev": await db.preventive_maintenances.count_documents({"statut": "ACTIF"}),
        "nombreMaintenancesCorrectives": await db.work_orders.count_documents({"priorite": {"$ne": "AUCUNE"}})
    }
    
    return analytics


@api_router.get("/reports/time-by-category")
async def get_time_by_category(start_month: str, current_user: dict = Depends(require_permission("reports", "view"))):
    """
    Obtenir le temps passé par catégorie sur 12 mois glissants
    start_month format: YYYY-MM (ex: 2025-09)
    """
    try:
        from datetime import datetime
        from dateutil.relativedelta import relativedelta
        
        # Parser le mois de départ
        start_date = datetime.strptime(start_month + "-01", "%Y-%m-%d")
        
        # Créer 12 mois de données
        months_data = []
        for i in range(12):
            current_month = start_date + relativedelta(months=i)
            month_start = current_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            month_end = (month_start + relativedelta(months=1)) - relativedelta(seconds=1)
            
            # Requête pour récupérer tous les ordres de travail dans ce mois
            pipeline = [
                {
                    "$match": {
                        "dateCreation": {
                            "$gte": month_start,
                            "$lte": month_end
                        },
                        "categorie": {"$ne": None}  # Exclure les ordres sans catégorie
                    }
                },
                {
                    "$group": {
                        "_id": "$categorie",
                        "totalTime": {"$sum": {"$ifNull": ["$tempsReel", 0]}}
                    }
                }
            ]
            
            results = await db.work_orders.aggregate(pipeline).to_list(length=None)
            
            # Organiser par catégorie
            time_by_category = {
                "CHANGEMENT_FORMAT": 0,
                "TRAVAUX_PREVENTIFS": 0,
                "TRAVAUX_CURATIF": 0,
                "TRAVAUX_DIVERS": 0,
                "FORMATION": 0,
                "REGLAGE": 0
            }
            
            # Debug logging
            logger.info(f"Mois {current_month.strftime('%Y-%m')} - Résultats MongoDB: {results}")
            
            for result in results:
                category = result.get("_id")
                if category and category in time_by_category:
                    time_by_category[category] = round(result["totalTime"], 2)
                    logger.info(f"  Catégorie {category}: {result['totalTime']}h")
                else:
                    logger.warning(f"  Catégorie inconnue ou None: {category}")
            
            months_data.append({
                "month": current_month.strftime("%Y-%m"),
                "monthLabel": current_month.strftime("%B %Y"),
                "categories": time_by_category
            })
        
        return {
            "startMonth": start_month,
            "months": months_data
        }
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des stats par catégorie : {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/reports/user-time-tracking")
async def get_user_time_tracking(
    user_ids: str = None,  # Comma-separated list of user IDs
    period: str = "weekly",  # daily, weekly, monthly, yearly, custom
    start_date: str = None,  # Format: YYYY-MM-DD
    end_date: str = None,  # Format: YYYY-MM-DD
    categories: str = None,  # Comma-separated list of categories, None = all
    current_user: dict = Depends(require_permission("reports", "view"))
):
    """
    Obtenir le temps passé par utilisateur par catégorie
    - user_ids: Liste des IDs utilisateurs séparés par des virgules (si vide, utilisateur courant)
    - period: daily, weekly, monthly, yearly, custom
    - start_date, end_date: Pour la période personnalisée
    - categories: Liste des catégories à inclure (si vide, toutes)
    """
    try:
        from datetime import datetime, timedelta
        from dateutil.relativedelta import relativedelta
        
        # Vérifier les permissions pour voir d'autres utilisateurs
        can_view_others = False
        user_role = current_user.get("role", "")
        user_permissions = current_user.get("permissions", {})
        
        # Admin ou responsable peuvent voir tous les utilisateurs
        if user_role == "ADMIN":
            can_view_others = True
        elif isinstance(user_permissions, dict):
            time_tracking_perm = user_permissions.get("timeTracking", {})
            if isinstance(time_tracking_perm, dict) and time_tracking_perm.get("view", False):
                can_view_others = True
        
        # Parser les user_ids
        if user_ids:
            requested_user_ids = [uid.strip() for uid in user_ids.split(",") if uid.strip()]
        else:
            requested_user_ids = [current_user["id"]]
        
        # Si l'utilisateur ne peut pas voir les autres, forcer son propre ID
        if not can_view_others:
            requested_user_ids = [current_user["id"]]
        
        # Déterminer les dates selon la période
        now = datetime.now()
        today = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
        if period == "daily":
            date_start = today
            date_end = today + timedelta(days=1) - timedelta(seconds=1)
            time_labels = [f"{h}h" for h in range(24)]
            group_by = "hour"
        elif period == "weekly":
            # Début de la semaine (lundi)
            days_since_monday = today.weekday()
            date_start = today - timedelta(days=days_since_monday)
            date_end = date_start + timedelta(days=7) - timedelta(seconds=1)
            time_labels = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
            group_by = "dayOfWeek"
        elif period == "monthly":
            date_start = today.replace(day=1)
            date_end = (date_start + relativedelta(months=1)) - timedelta(seconds=1)
            # Générer les labels pour chaque jour du mois
            days_in_month = (date_end.replace(day=1) + relativedelta(months=1) - date_end.replace(day=1)).days
            if hasattr(date_end, 'day'):
                days_in_month = date_end.day
            time_labels = [str(d) for d in range(1, days_in_month + 1)]
            group_by = "dayOfMonth"
        elif period == "yearly":
            date_start = today.replace(month=1, day=1)
            date_end = today.replace(month=12, day=31, hour=23, minute=59, second=59)
            time_labels = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Août", "Sep", "Oct", "Nov", "Déc"]
            group_by = "month"
        elif period == "custom" and start_date and end_date:
            date_start = datetime.strptime(start_date, "%Y-%m-%d")
            date_end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            # Calculer le nombre de jours
            delta = (date_end - date_start).days + 1
            if delta <= 7:
                time_labels = [(date_start + timedelta(days=i)).strftime("%d/%m") for i in range(delta)]
                group_by = "dayOfMonth"
            elif delta <= 31:
                time_labels = [(date_start + timedelta(days=i)).strftime("%d") for i in range(delta)]
                group_by = "dayOfMonth"
            else:
                # Grouper par mois
                time_labels = []
                current = date_start
                while current <= date_end:
                    time_labels.append(current.strftime("%b %Y"))
                    current = current + relativedelta(months=1)
                group_by = "month"
        else:
            # Par défaut: hebdomadaire
            days_since_monday = today.weekday()
            date_start = today - timedelta(days=days_since_monday)
            date_end = date_start + timedelta(days=7) - timedelta(seconds=1)
            time_labels = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
            group_by = "dayOfWeek"
        
        # Parser les catégories
        all_categories = ["CHANGEMENT_FORMAT", "TRAVAUX_PREVENTIFS", "TRAVAUX_CURATIF", "TRAVAUX_DIVERS", "FORMATION", "REGLAGE", "AMELIORATIONS"]
        if categories:
            selected_categories = [cat.strip() for cat in categories.split(",") if cat.strip() in all_categories]
        else:
            selected_categories = all_categories
        
        # Récupérer les informations des utilisateurs demandés
        users_info = {}
        for uid in requested_user_ids:
            try:
                user_doc = await db.users.find_one({"_id": ObjectId(uid)}, {"_id": 1, "nom": 1, "prenom": 1, "email": 1})
                if user_doc:
                    users_info[uid] = {
                        "id": uid,
                        "name": f"{user_doc.get('prenom', '')} {user_doc.get('nom', '')}".strip() or user_doc.get('email', 'Inconnu')
                    }
            except:
                pass
        
        # Si aucun utilisateur trouvé, utiliser l'utilisateur courant
        if not users_info:
            users_info[current_user["id"]] = {
                "id": current_user["id"],
                "name": current_user.get("name", current_user.get("email", "Moi"))
            }
        
        # Construire les données pour chaque utilisateur
        results = {}
        
        for user_id in users_info.keys():
            user_data = {cat: [0] * len(time_labels) for cat in selected_categories}
            
            # Requête pour les ordres de travail - basée sur time_entries (qui a saisi le temps)
            wo_categories = [cat for cat in selected_categories if cat != "AMELIORATIONS"]
            if wo_categories:
                # Chercher les OT qui ont des time_entries pour cet utilisateur dans la période
                wo_match = {
                    "categorie": {"$in": wo_categories},
                    "time_entries": {
                        "$elemMatch": {
                            "user_id": user_id,
                            "timestamp": {"$gte": date_start, "$lte": date_end}
                        }
                    }
                }
                
                wo_cursor = db.work_orders.find(wo_match, {"categorie": 1, "time_entries": 1})
                async for wo in wo_cursor:
                    category = wo.get("categorie")
                    time_entries = wo.get("time_entries", [])
                    if category and time_entries:
                        for entry in time_entries:
                            # Ne compter que les entrées de cet utilisateur dans la période
                            if entry.get("user_id") == user_id:
                                entry_timestamp = entry.get("timestamp")
                                if entry_timestamp and date_start <= entry_timestamp <= date_end:
                                    idx = get_time_index(entry_timestamp, date_start, group_by, len(time_labels))
                                    if 0 <= idx < len(time_labels) and category in user_data:
                                        user_data[category][idx] += entry.get("hours", 0)
            
            # Requête pour les améliorations - basée sur time_entries
            if "AMELIORATIONS" in selected_categories:
                imp_match = {
                    "time_entries": {
                        "$elemMatch": {
                            "user_id": user_id,
                            "timestamp": {"$gte": date_start, "$lte": date_end}
                        }
                    }
                }
                
                imp_cursor = db.improvements.find(imp_match, {"time_entries": 1})
                async for imp in imp_cursor:
                    time_entries = imp.get("time_entries", [])
                    for entry in time_entries:
                        if entry.get("user_id") == user_id:
                            entry_timestamp = entry.get("timestamp")
                            if entry_timestamp and date_start <= entry_timestamp <= date_end:
                                idx = get_time_index(entry_timestamp, date_start, group_by, len(time_labels))
                                if 0 <= idx < len(time_labels):
                                    user_data["AMELIORATIONS"][idx] += entry.get("hours", 0)
            
            # Arrondir les valeurs
            for cat in user_data:
                user_data[cat] = [round(v, 2) for v in user_data[cat]]
            
            results[user_id] = {
                "user": users_info[user_id],
                "data": user_data
            }
        
        # Récupérer la liste de tous les utilisateurs (pour le filtre)
        all_users = []
        if can_view_others:
            users_cursor = db.users.find({}, {"_id": 1, "nom": 1, "prenom": 1, "email": 1})
            async for user in users_cursor:
                all_users.append({
                    "id": str(user["_id"]),
                    "name": f"{user.get('prenom', '')} {user.get('nom', '')}".strip() or user.get('email', 'Inconnu')
                })
        
        return {
            "period": period,
            "startDate": date_start.strftime("%Y-%m-%d"),
            "endDate": date_end.strftime("%Y-%m-%d"),
            "timeLabels": time_labels,
            "categories": selected_categories,
            "users": results,
            "allUsers": all_users if can_view_others else [],
            "canViewOthers": can_view_others
        }
    except Exception as e:
        logger.error(f"Erreur lors de la récupération du pointage horaire : {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


def get_time_index(date, start_date, group_by, max_len):
    """Helper pour calculer l'index dans le tableau de temps"""
    from datetime import timedelta
    
    if group_by == "hour":
        return date.hour
    elif group_by == "dayOfWeek":
        return date.weekday()
    elif group_by == "dayOfMonth":
        return (date - start_date).days
    elif group_by == "month":
        months_diff = (date.year - start_date.year) * 12 + (date.month - start_date.month)
        return min(months_diff, max_len - 1)
    return 0


# ==================== IMPORT/EXPORT ROUTES ====================
# NOTE: Ces routes ont été modularisées dans import_export_routes.py
# Voir l'inclusion du router dans la section des includes (api_router.include_router)


# ==================== UPDATE ROUTES ====================
from update_manager import UpdateManager

update_manager = UpdateManager(db)

@api_router.get("/updates/current")
async def get_current_version(current_user: dict = Depends(get_current_admin_user)):
    """Récupère la version actuelle (admin uniquement)"""
    version = await update_manager.get_current_version()
    return {
        "version": version,
        "date": datetime.now().isoformat()
    }

@api_router.get("/updates/check")
async def check_updates(current_user: dict = Depends(get_current_admin_user)):
    """Vérifie si une mise à jour est disponible (admin uniquement)"""
    current = await update_manager.get_current_version()
    latest = await update_manager.check_github_version()
    
    return {
        "current_version": current,
        "latest_version": latest,
        "update_available": latest is not None and latest.get("available", False)
    }

@api_router.get("/updates/changelog")
async def get_changelog(
    from_version: Optional[str] = None,
    current_user: dict = Depends(get_current_admin_user)
):
    """Récupère le changelog (admin uniquement)"""
    changelog = await update_manager.get_changelog(from_version)
    return {"changelog": changelog}

@api_router.get("/updates/history")
async def get_update_history(current_user: dict = Depends(get_current_admin_user)):
    """Récupère l'historique des mises à jour depuis la BDD (admin uniquement)"""
    try:
        # Récupérer depuis la nouvelle collection system_update_history
        history = await db.system_update_history.find(
            {},
            {"_id": 0}
        ).sort("started_at", -1).limit(50).to_list(50)
        
        return {"history": history}
    except Exception as e:
        logger.error(f"❌ Erreur récupération historique: {str(e)}")
        # Fallback vers l'ancienne méthode si erreur
        history = await update_manager.get_update_history()
        return {"history": history}


@api_router.post("/updates/backup")
async def create_backup(current_user: dict = Depends(get_current_admin_user)):
    """Crée un backup de la base de données (admin uniquement)"""
    result = await update_manager.create_backup()
    
    if not result["success"]:
        raise HTTPException(
            status_code=500,
            detail=result.get("error", "Erreur lors de la création du backup")
        )
    
    return result

@api_router.post("/updates/rollback")
async def rollback_update(
    backup_path: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """Restaure une version précédente (admin uniquement)"""
    result = await update_manager.rollback_to_version(backup_path)
    
    if not result["success"]:
        raise HTTPException(
            status_code=500,
            detail=result.get("message", "Erreur lors du rollback")
        )
    
    return result

@api_router.get("/updates/git-history")
async def get_git_history(
    limit: int = 20,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Récupère l'historique des commits Git (versions précédentes) (admin uniquement)
    Permet de voir et restaurer des versions antérieures du code
    """
    try:
        commits = await update_manager.get_git_history(limit)
        return {
            "success": True,
            "commits": commits,
            "total": len(commits)
        }
    except Exception as e:
        logger.error(f"❌ Erreur récupération historique Git: {str(e)}")
        return {
            "success": False,
            "commits": [],
            "error": str(e)
        }

@api_router.post("/updates/git-rollback")
async def rollback_to_git_commit(
    commit_hash: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Effectue un rollback Git vers un commit spécifique (admin uniquement)
    ⚠️ ATTENTION: Cette action modifie le code source de l'application
    """
    try:
        result = await update_manager.rollback_to_commit(commit_hash)
        
        if not result["success"]:
            raise HTTPException(
                status_code=500,
                detail=result.get("message", "Erreur lors du rollback Git")
            )
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur rollback Git: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors du rollback: {str(e)}"
        )

# ==================== AUDIT LOG ROUTES (JOURNAL) ====================
@api_router.get("/audit-logs",
    summary="Journal d'audit", tags=["Audit"])
async def get_audit_logs(
    skip: int = 0,
    limit: int = 100,
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Récupère les logs d'audit (admin uniquement)
    Supporte les filtres: user_id, action, entity_type, start_date, end_date
    """
    try:
        # Convertir les strings en enums si fournis
        action_enum = None
        entity_type_enum = None
        try:
            action_enum = ActionType(action) if action else None
        except ValueError:
            pass
        try:
            entity_type_enum = EntityType(entity_type) if entity_type else None
        except ValueError:
            pass
        
        # Convertir les dates si fournies
        start_dt = None
        end_dt = None
        try:
            start_dt = datetime.fromisoformat(start_date) if start_date else None
            end_dt = datetime.fromisoformat(end_date) if end_date else None
        except (ValueError, TypeError):
            pass
        
        logs, total = await audit_service.get_logs(
            skip=skip,
            limit=limit,
            user_id=user_id,
            action=action_enum,
            entity_type=entity_type_enum,
            start_date=start_dt,
            end_date=end_dt
        )
        
        return {
            "logs": logs,
            "total": total,
            "skip": skip,
            "limit": limit
        }
        
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des logs d'audit: {e}")
        # Retourner une réponse vide au lieu d'une erreur 500
        return {
            "logs": [],
            "total": 0,
            "skip": skip,
            "limit": limit
        }

@api_router.get("/audit-logs/entity/{entity_type}/{entity_id}", tags=["Audit"])
async def get_entity_audit_history(
    entity_type: str,
    entity_id: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Récupère l'historique complet d'une entité spécifique (admin uniquement)
    """
    try:
        entity_type_enum = EntityType(entity_type)
        logs = await audit_service.get_entity_history(entity_type_enum, entity_id)
        
        return {
            "entity_type": entity_type,
            "entity_id": entity_id,
            "history": logs
        }
        
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Type d'entité invalide: {entity_type}"
        )
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de l'historique: {e}")
        raise HTTPException(
            status_code=500,
            detail="Erreur lors de la récupération de l'historique"
        )

@api_router.get("/audit-logs/export",
    summary="Exporter le journal d'audit", tags=["Audit"])
async def export_audit_logs(
    format: str = "csv",
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Exporte les logs d'audit en CSV ou Excel (admin uniquement)
    """
    try:
        # Récupérer tous les logs avec filtres
        action_enum = ActionType(action) if action else None
        entity_type_enum = EntityType(entity_type) if entity_type else None
        start_dt = datetime.fromisoformat(start_date) if start_date else None
        end_dt = datetime.fromisoformat(end_date) if end_date else None
        
        logs, _ = await audit_service.get_logs(
            skip=0,
            limit=10000,  # Limite haute pour export
            user_id=user_id,
            action=action_enum,
            entity_type=entity_type_enum,
            start_date=start_dt,
            end_date=end_dt
        )
        
        # Préparer les données pour l'export
        paris_tz = pytz.timezone('Europe/Paris')
        export_data = []
        for log in logs:
            # Convertir UTC vers Europe/Paris
            timestamp_utc = log["timestamp"]
            if timestamp_utc.tzinfo is None:
                timestamp_utc = pytz.utc.localize(timestamp_utc)
            timestamp_paris = timestamp_utc.astimezone(paris_tz)
            
            export_data.append({
                "Date/Heure": timestamp_paris.strftime("%d/%m/%Y %H:%M:%S"),
                "Utilisateur": log["user_name"],
                "Email": log["user_email"],
                "Action": log["action"],
                "Type": log["entity_type"],
                "Entité": log.get("entity_name", ""),
                "Détails": log.get("details", "")
            })
        
        df = pd.DataFrame(export_data)
        
        # Créer le fichier selon le format demandé
        if format.lower() == "csv":
            output = io.StringIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                }
            )
        else:  # Excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Audit Logs')
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                }
            )
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export des logs: {e}")
        raise HTTPException(
            status_code=500,
            detail="Erreur lors de l'export des logs"
        )

# ==================== WORK ORDER COMMENTS ROUTES ====================
@api_router.post("/work-orders/{work_order_id}/comments", tags=["Ordres de Travail"])
async def add_work_order_comment(
    work_order_id: str,
    comment: CommentWithPartsCreate,
    current_user: dict = Depends(require_permission("workOrders", "edit"))
):
    """Ajoute un commentaire et des pièces utilisées à un ordre de travail"""
    try:
        # Vérifier que l'ordre de travail existe (chercher par _id ObjectId)
        work_order = await db.work_orders.find_one({"_id": ObjectId(work_order_id)})
        if not work_order:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        # Créer le commentaire
        new_comment = {
            "id": str(uuid.uuid4()),
            "user_id": current_user.get("id", str(work_order["_id"])),
            "user_name": f"{current_user['prenom']} {current_user['nom']}",
            "text": comment.text,
            "timestamp": datetime.now(timezone.utc)
        }
        
        # Traiter les pièces utilisées
        parts_used_list = []
        logger.info(f"Traitement de {len(comment.parts_used)} pièces utilisées")
        for part in comment.parts_used:
            logger.info(f"Pièce: inventory_item_id={part.inventory_item_id}, name={part.inventory_item_name}, custom={part.custom_part_name}, quantity={part.quantity}")
            part_data = {
                "id": str(uuid.uuid4()),
                "inventory_item_id": part.inventory_item_id,
                "inventory_item_name": part.inventory_item_name,
                "custom_part_name": part.custom_part_name,
                "quantity": part.quantity,
                "user_name": f"{current_user['prenom']} {current_user['nom']}",
                "timestamp": datetime.now(timezone.utc)
            }
            
            # N'ajouter les champs "Prélevé Sur" que s'ils sont fournis
            if hasattr(part, 'source_equipment_id') and part.source_equipment_id:
                part_data["source_equipment_id"] = part.source_equipment_id
                part_data["source_equipment_name"] = part.source_equipment_name
            if hasattr(part, 'custom_source') and part.custom_source:
                part_data["custom_source"] = part.custom_source
            
            parts_used_list.append(part_data)
            
            # Si c'est une pièce d'inventaire, déduire du stock
            if part.inventory_item_id:
                inventory_item = await db.inventory.find_one({"_id": ObjectId(part.inventory_item_id)})
                if inventory_item:
                    new_quantity = inventory_item["quantite"] - part.quantity
                    await db.inventory.update_one(
                        {"_id": ObjectId(part.inventory_item_id)},
                        {"$set": {"quantite": new_quantity}}
                    )
                    logger.info(f"Stock mis à jour: {part.inventory_item_name} - {part.quantity} unité(s) déduite(s)")
        
        # Mettre à jour l'ordre de travail
        if parts_used_list:
            # Push commentaire ET pièces utilisées en une seule opération
            await db.work_orders.update_one(
                {"_id": ObjectId(work_order_id)},
                {
                    "$push": {
                        "comments": new_comment,
                        "parts_used": {"$each": parts_used_list}
                    }
                }
            )
        else:
            # Push seulement le commentaire
            await db.work_orders.update_one(
                {"_id": ObjectId(work_order_id)},
                {"$push": {"comments": new_comment}}
            )
        
        # Log dans l'audit
        details_text = f"Commentaire ajouté: {comment.text[:50]}..."
        if parts_used_list:
            details_text += f" | {len(parts_used_list)} pièce(s) utilisée(s)"
        
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType_Audit.WORK_ORDER,
            entity_id=work_order_id,
            entity_name=work_order.get("titre", ""),
            details=details_text
        )
        
        return {"comment": new_comment, "parts_used": parts_used_list}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de l'ajout du commentaire: {e}")
        raise HTTPException(
            status_code=500,
            detail="Erreur lors de l'ajout du commentaire"
        )

@api_router.post("/work-orders/{work_order_id}/parts-used", tags=["Ordres de Travail"])
async def add_work_order_parts(
    work_order_id: str,
    parts: List[PartUsedCreate],
    current_user: dict = Depends(require_permission("workOrders", "edit"))
):
    """Ajoute des pièces utilisées à un ordre de travail SANS créer de commentaire"""
    try:
        # Vérifier que l'ordre de travail existe
        work_order = await db.work_orders.find_one({"_id": ObjectId(work_order_id)})
        if not work_order:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        # Traiter les pièces utilisées
        parts_used_list = []
        logger.info(f"Ajout direct de {len(parts)} pièces utilisées")
        for part in parts:
            logger.info(f"Pièce: inventory_item_id={part.inventory_item_id}, name={part.inventory_item_name}, custom={part.custom_part_name}, quantity={part.quantity}")
            part_data = {
                "id": str(uuid.uuid4()),
                "inventory_item_id": part.inventory_item_id,
                "inventory_item_name": part.inventory_item_name,
                "custom_part_name": part.custom_part_name,
                "quantity": part.quantity,
                "user_name": f"{current_user['prenom']} {current_user['nom']}",
                "timestamp": datetime.now(timezone.utc)
            }
            
            # N'ajouter les champs "Prélevé Sur" que s'ils sont fournis
            if hasattr(part, 'source_equipment_id') and part.source_equipment_id:
                part_data["source_equipment_id"] = part.source_equipment_id
                part_data["source_equipment_name"] = part.source_equipment_name
            if hasattr(part, 'custom_source') and part.custom_source:
                part_data["custom_source"] = part.custom_source
            
            parts_used_list.append(part_data)
            
            # Si c'est une pièce d'inventaire, déduire du stock
            if part.inventory_item_id:
                inventory_item = await db.inventory.find_one({"_id": ObjectId(part.inventory_item_id)})
                if inventory_item:
                    new_quantity = inventory_item["quantite"] - part.quantity
                    await db.inventory.update_one(
                        {"_id": ObjectId(part.inventory_item_id)},
                        {"$set": {"quantite": new_quantity}}
                    )
                    logger.info(f"Stock mis à jour: {part.inventory_item_name} - {part.quantity} unité(s) déduite(s)")
        
        # Ajouter les pièces à l'ordre de travail (SANS commentaire)
        await db.work_orders.update_one(
            {"_id": ObjectId(work_order_id)},
            {"$push": {"parts_used": {"$each": parts_used_list}}}
        )
        
        # Log dans l'audit
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType_Audit.WORK_ORDER,
            entity_id=work_order_id,
            entity_name=work_order.get("titre", ""),
            details=f"{len(parts_used_list)} pièce(s) utilisée(s) ajoutée(s)"
        )
        
        return {"parts_used": parts_used_list, "count": len(parts_used_list)}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de l'ajout des pièces: {e}")
        raise HTTPException(
            status_code=500,
            detail="Erreur lors de l'ajout des pièces"
        )

@api_router.get("/work-orders/{work_order_id}/comments", tags=["Ordres de Travail"])
async def get_work_order_comments(
    work_order_id: str,
    current_user: dict = Depends(require_permission("workOrders", "view"))
):
    """Récupère tous les commentaires d'un ordre de travail"""
    try:
        work_order = await db.work_orders.find_one({"_id": ObjectId(work_order_id)})
        if not work_order:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
        
        comments = work_order.get("comments", [])
        return {"comments": comments}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des commentaires: {e}")
        raise HTTPException(
            status_code=500,
            detail="Erreur lors de la récupération des commentaires"
        )



# ==================== METERS (COMPTEURS) ENDPOINTS ====================

@api_router.post("/meters",
    summary="Creer un compteur", response_model=Meter, status_code=201, tags=["Compteurs"])
async def create_meter(meter: MeterCreate, current_user: dict = Depends(require_permission("meters", "edit"))):
    """Créer un nouveau compteur"""
    try:
        meter_id = str(uuid.uuid4())
        meter_data = meter.model_dump()
        meter_data["id"] = meter_id
        meter_data["date_creation"] = datetime.utcnow()
        meter_data["actif"] = True
        
        # Récupérer les informations de l'emplacement si fourni
        if meter_data.get("emplacement_id"):
            location = await db.locations.find_one({"id": meter_data["emplacement_id"]})
            if location:
                meter_data["emplacement"] = {"id": location["id"], "nom": location["nom"]}
        
        await db.meters.insert_one(meter_data)
        
        # Rafraîchir les abonnements MQTT si activé
        if meter_data.get("mqtt_enabled"):
            await mqtt_meter_collector.refresh_subscriptions()
        
        # Audit log
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
            user_email=current_user["email"],
            action=ActionType.CREATE,
            entity_type=EntityType_Audit.WORK_ORDER,  # Utilisons WORK_ORDER comme proxy
            entity_id=meter_id,
            entity_name=meter.nom
        )
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "counters",
            "created",
            meter_data,
            user_id=current_user["id"]
        )
        
        return Meter(**meter_data)
    except Exception as e:
        logger.error(f"Erreur création compteur: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/meters",
    summary="Lister les compteurs", response_model=List[Meter], tags=["Compteurs"])
async def get_all_meters(current_user: dict = Depends(require_permission("meters", "view"))):
    """Récupérer tous les compteurs"""
    try:
        meters = []
        async for meter in db.meters.find({"actif": True}).sort("date_creation", -1):
            meters.append(Meter(**meter))
        return meters
    except Exception as e:
        logger.error(f"Erreur récupération compteurs: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/meters/{meter_id}",
    summary="Detail d'un compteur", response_model=Meter, tags=["Compteurs"])
async def get_meter(meter_id: str, current_user: dict = Depends(require_permission("meters", "view"))):
    """Récupérer un compteur spécifique"""
    meter = await db.meters.find_one({"id": meter_id})
    if not meter:
        raise HTTPException(status_code=404, detail="Compteur non trouvé")
    return Meter(**meter)

@api_router.put("/meters/{meter_id}",
    summary="Modifier un compteur", response_model=Meter, tags=["Compteurs"])
async def update_meter(
    meter_id: str,
    meter_update: MeterUpdate,
    current_user: dict = Depends(require_permission("meters", "edit"))
):
    """Mettre à jour un compteur"""
    meter = await db.meters.find_one({"id": meter_id})
    if not meter:
        raise HTTPException(status_code=404, detail="Compteur non trouvé")
    
    update_data = {k: v for k, v in meter_update.model_dump().items() if v is not None}
    
    # Mettre à jour l'emplacement si nécessaire
    if "emplacement_id" in update_data:
        if update_data["emplacement_id"]:
            location = await db.locations.find_one({"id": update_data["emplacement_id"]})
            if location:
                update_data["emplacement"] = {"id": location["id"], "nom": location["nom"]}
        else:
            update_data["emplacement"] = None
    
    await db.meters.update_one({"id": meter_id}, {"$set": update_data})
    
    # Récupérer le compteur mis à jour
    updated_meter = await db.meters.find_one({"id": meter_id})
    
    # Rafraîchir les abonnements MQTT si MQTT activé/modifié
    if "mqtt_enabled" in update_data or "mqtt_topic" in update_data:
        await mqtt_meter_collector.refresh_subscriptions()
    
    # Audit log
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
        user_email=current_user["email"],
        action=ActionType.UPDATE,
        entity_type=EntityType_Audit.WORK_ORDER,
        entity_id=meter_id,
        entity_name=updated_meter["nom"]
    )
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "counters",
        "updated",
        dict(updated_meter),
        user_id=current_user["id"]
    )
    
    return Meter(**updated_meter)

@api_router.delete("/meters/{meter_id}", response_model=MessageResponse,
    summary="Supprimer un compteur", tags=["Compteurs"])
async def delete_meter(meter_id: str, current_user: dict = Depends(require_permission("meters", "delete"))):
    """Supprimer un compteur (soft delete)"""
    meter = await db.meters.find_one({"id": meter_id})
    if not meter:
        raise HTTPException(status_code=404, detail="Compteur non trouvé")
    
    # Soft delete
    await db.meters.update_one({"id": meter_id}, {"$set": {"actif": False}})
    
    # Audit log
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
        user_email=current_user["email"],
        action=ActionType.DELETE,
        entity_type=EntityType_Audit.WORK_ORDER,
        entity_id=meter_id,
        entity_name=meter["nom"]
    )
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "counters",
        "deleted",
        {"id": meter_id, "nom": meter["nom"]},
        user_id=current_user["id"]
    )
    
    return {"message": "Compteur supprimé"}

# ==================== METER READINGS (RELEVÉS) ENDPOINTS ====================

@api_router.post("/meters/{meter_id}/readings", response_model=MeterReading, status_code=201, tags=["Compteurs"])
async def create_reading(
    meter_id: str,
    reading: MeterReadingCreate,
    current_user: dict = Depends(require_permission("meters", "edit"))
):
    """Créer un nouveau relevé pour un compteur"""
    try:
        # Vérifier que le compteur existe
        meter = await db.meters.find_one({"id": meter_id})
        if not meter:
            raise HTTPException(status_code=404, detail="Compteur non trouvé")
        
        # Récupérer le dernier relevé pour calculer la consommation
        last_reading = await db.meter_readings.find_one(
            {"meter_id": meter_id},
            sort=[("date_releve", -1)]
        )
        
        reading_id = str(uuid.uuid4())
        reading_data = reading.model_dump()
        reading_data["id"] = reading_id
        reading_data["meter_id"] = meter_id
        reading_data["created_by"] = current_user["id"]
        reading_data["created_by_name"] = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}"
        reading_data["meter_nom"] = meter["nom"]
        reading_data["date_creation"] = datetime.utcnow()
        
        # Calculer la consommation
        if last_reading:
            consommation = reading_data["valeur"] - last_reading["valeur"]
            reading_data["consommation"] = max(0, consommation)  # Éviter les valeurs négatives
            
            # Calculer le coût si prix unitaire disponible
            prix = reading_data.get("prix_unitaire") or meter.get("prix_unitaire")
            if prix and reading_data["consommation"]:
                reading_data["cout"] = reading_data["consommation"] * prix
        else:
            reading_data["consommation"] = 0
            reading_data["cout"] = 0
        
        # Si pas de prix spécifié, utiliser celui du compteur
        if not reading_data.get("prix_unitaire"):
            reading_data["prix_unitaire"] = meter.get("prix_unitaire")
        if not reading_data.get("abonnement_mensuel"):
            reading_data["abonnement_mensuel"] = meter.get("abonnement_mensuel")
        
        await db.meter_readings.insert_one(reading_data)
        
        return MeterReading(**reading_data)
    except Exception as e:
        logger.error(f"Erreur création relevé: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/meters/{meter_id}/readings", response_model=List[MeterReading], tags=["Compteurs"])
async def get_meter_readings(
    meter_id: str,
    current_user: dict = Depends(require_permission("meters", "view")),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Récupérer tous les relevés d'un compteur"""
    try:
        query = {"meter_id": meter_id}
        
        # Filtrer par date si fourni
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            if end_date:
                date_filter["$lte"] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query["date_releve"] = date_filter
        
        readings = []
        async for reading in db.meter_readings.find(query).sort("date_releve", -1):
            readings.append(MeterReading(**reading))
        return readings
    except Exception as e:
        logger.error(f"Erreur récupération relevés: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/meters/{meter_id}/statistics", tags=["Compteurs"])
async def get_meter_statistics(
    meter_id: str,
    current_user: dict = Depends(require_permission("meters", "view")),
    period: str = "month"  # week, month, quarter, year
):
    """Obtenir les statistiques d'un compteur"""
    try:
        meter = await db.meters.find_one({"id": meter_id})
        if not meter:
            raise HTTPException(status_code=404, detail="Compteur non trouvé")
        
        # Calculer la période
        now = datetime.utcnow()
        if period == "week":
            start_date = now - timedelta(days=7)
        elif period == "month":
            start_date = now - timedelta(days=30)
        elif period == "quarter":
            start_date = now - timedelta(days=90)
        elif period == "year":
            start_date = now - timedelta(days=365)
        else:
            start_date = now - timedelta(days=30)
        
        # Récupérer les relevés de la période
        readings = []
        async for reading in db.meter_readings.find({
            "meter_id": meter_id,
            "date_releve": {"$gte": start_date}
        }).sort("date_releve", 1):
            readings.append(reading)
        
        if not readings:
            return {
                "meter_id": meter_id,
                "meter_nom": meter["nom"],
                "period": period,
                "total_consommation": 0,
                "total_cout": 0,
                "moyenne_journaliere": 0,
                "dernier_releve": None,
                "evolution": []
            }
        
        # Calculer les statistiques
        total_consommation = sum(r.get("consommation", 0) for r in readings if r.get("consommation"))
        total_cout = sum(r.get("cout", 0) for r in readings if r.get("cout"))
        
        # Calculer la moyenne journalière
        if len(readings) > 1:
            first_date = readings[0]["date_releve"]
            last_date = readings[-1]["date_releve"]
            days = (last_date - first_date).days or 1
            moyenne_journaliere = total_consommation / days
        else:
            moyenne_journaliere = 0
        
        # Préparer l'évolution
        evolution = [
            {
                "date": r["date_releve"].isoformat(),
                "valeur": r["valeur"],
                "consommation": r.get("consommation", 0),
                "cout": r.get("cout", 0)
            }
            for r in readings
        ]
        
        # Serialize the last reading to avoid ObjectId issues
        dernier_releve = None
        if readings:
            last_reading = readings[-1].copy()
            # Remove any ObjectId fields that might cause serialization issues
            if "_id" in last_reading:
                del last_reading["_id"]
            dernier_releve = last_reading
        
        return {
            "meter_id": meter_id,
            "meter_nom": meter["nom"],
            "period": period,
            "total_consommation": round(total_consommation, 2),
            "total_cout": round(total_cout, 2),
            "moyenne_journaliere": round(moyenne_journaliere, 2),
            "dernier_releve": dernier_releve,
            "evolution": evolution,
            "nombre_releves": len(readings)
        }
    except Exception as e:
        logger.error(f"Erreur calcul statistiques: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/readings/{reading_id}", response_model=MessageResponse, tags=["Compteurs"])
async def delete_reading(reading_id: str, current_user: dict = Depends(require_permission("meters", "delete"))):
    """Supprimer un relevé"""
    reading = await db.meter_readings.find_one({"id": reading_id})
    if not reading:
        raise HTTPException(status_code=404, detail="Relevé non trouvé")
    
    await db.meter_readings.delete_one({"id": reading_id})
    return {"message": "Relevé supprimé"}



# ==================== INTERVENTION REQUESTS (DEMANDES D'INTERVENTION) ENDPOINTS ====================

@api_router.post("/intervention-requests", response_model=InterventionRequest, status_code=201, tags=["Demandes Intervention"])
async def create_intervention_request(
    request: InterventionRequestCreate,
    current_user: dict = Depends(require_permission("interventionRequests", "edit"))
):
    """Créer une nouvelle demande d'intervention"""
    try:
        request_id = str(uuid.uuid4())
        request_data = request.model_dump()
        request_data["id"] = request_id
        request_data["date_creation"] = datetime.utcnow()
        request_data["created_by"] = current_user["id"]
        request_data["created_by_name"] = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}"
        request_data["work_order_id"] = None
        request_data["work_order_date_limite"] = None
        request_data["converted_at"] = None
        request_data["converted_by"] = None
        request_data["attachments"] = []
        request_data["refused"] = False
        request_data["refused_reason"] = None
        request_data["refused_at"] = None
        request_data["refused_by"] = None
        request_data["refused_by_name"] = None
        
        # Récupérer les informations de l'équipement si fourni
        if request_data.get("equipement_id"):
            eq_info = await get_equipment_by_id(request_data["equipement_id"])
            if eq_info:
                request_data["equipement"] = eq_info
        
        # Récupérer les informations de l'emplacement si fourni
        if request_data.get("emplacement_id"):
            loc_info = await get_location_by_id(request_data["emplacement_id"])
            if loc_info:
                request_data["emplacement"] = loc_info
        
        await db.intervention_requests.insert_one(request_data)
        
        # Broadcast WebSocket pour la synchronisation temps réel
        # Ne pas exclure l'utilisateur pour permettre la sync multi-appareils
        broadcast_data = {k: v for k, v in request_data.items() if k != '_id'}
        broadcast_data = _clean_ir_attachments(broadcast_data)
        await realtime_manager.emit_event(
            "intervention_requests",
            "created",
            broadcast_data
        )
        
        # Audit log
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
            user_email=current_user["email"],
            action=ActionType.CREATE,
            entity_type=EntityType_Audit.WORK_ORDER,
            entity_id=request_id,
            entity_name=request.titre,
            details=f"Création demande d'intervention"
        )
        
        return InterventionRequest(**request_data)
    except Exception as e:
        logger.error(f"Erreur création demande d'intervention: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

def _clean_ir_attachments(req):
    """Nettoyer les ObjectId dans les attachments d'une demande d'intervention"""
    if "attachments" in req and req["attachments"]:
        cleaned = []
        for att in req["attachments"]:
            clean_att = {k: v for k, v in att.items() if k != "_id"}
            if "_id" in att:
                clean_att["id"] = str(att["_id"])
                clean_att["url"] = f"/api/intervention-requests/{req.get('id', '')}/attachments/{str(att['_id'])}"
            elif "id" in att and att["id"]:
                # Old public format: keep the id, build url
                clean_att["url"] = f"/api/intervention-requests/{req.get('id', '')}/attachments/{att['id']}"
            cleaned.append(clean_att)
        req["attachments"] = cleaned
    else:
        req["attachments"] = []
    return req

@api_router.get("/intervention-requests", response_model=List[InterventionRequest], tags=["Demandes Intervention"])
async def get_all_intervention_requests(current_user: dict = Depends(require_permission("interventionRequests", "view"))):
    """Récupérer toutes les demandes d'intervention"""
    try:
        query = {}
        
        requests = []
        async for req in db.intervention_requests.find(query).sort("date_creation", -1):
            req = _clean_ir_attachments(req)
            requests.append(InterventionRequest(**req))
        return requests
    except Exception as e:
        logger.error(f"Erreur récupération demandes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/intervention-requests/stats/kpi", tags=["Demandes Intervention"])
async def get_intervention_requests_kpi(current_user: dict = Depends(require_permission("interventionRequests", "view"))):
    """KPI des demandes d'intervention : temps de reponse, taux de conversion, etc."""
    try:
        all_irs = await db.intervention_requests.find({}, {"_id": 0}).to_list(5000)
        total = len(all_irs)
        if total == 0:
            return {
                "total": 0,
                "en_attente": 0,
                "converties": 0,
                "refusees": 0,
                "taux_conversion": 0,
                "taux_refus": 0,
                "temps_moyen_reponse_heures": None,
                "temps_moyen_reponse_label": "-",
                "publiques": 0,
                "authentifiees": 0,
            }

        en_attente = sum(1 for ir in all_irs if not ir.get("work_order_id") and not ir.get("refused"))
        converties = sum(1 for ir in all_irs if ir.get("work_order_id"))
        refusees = sum(1 for ir in all_irs if ir.get("refused"))
        publiques = sum(1 for ir in all_irs if ir.get("created_by") == "PUBLIC")
        authentifiees = total - publiques

        # Temps moyen de reponse (entre date_creation et converted_at ou refused_at)
        response_times = []
        for ir in all_irs:
            creation = ir.get("date_creation")
            if not creation:
                continue
            if isinstance(creation, str):
                try:
                    creation = datetime.fromisoformat(creation.replace('Z', '+00:00'))
                except Exception:
                    continue

            response_time = None
            if ir.get("converted_at"):
                rt = ir["converted_at"]
                if isinstance(rt, str):
                    try:
                        rt = datetime.fromisoformat(rt.replace('Z', '+00:00'))
                    except Exception:
                        continue
                response_time = rt
            elif ir.get("refused_at"):
                rt = ir["refused_at"]
                if isinstance(rt, str):
                    try:
                        rt = datetime.fromisoformat(rt.replace('Z', '+00:00'))
                    except Exception:
                        continue
                response_time = rt

            if response_time:
                # Make both offset-aware or offset-naive
                if creation.tzinfo is None and response_time.tzinfo is not None:
                    creation = creation.replace(tzinfo=timezone.utc)
                elif creation.tzinfo is not None and response_time.tzinfo is None:
                    response_time = response_time.replace(tzinfo=timezone.utc)
                delta = (response_time - creation).total_seconds() / 3600
                if delta >= 0:
                    response_times.append(delta)

        temps_moyen = round(sum(response_times) / len(response_times), 1) if response_times else None
        if temps_moyen is not None:
            if temps_moyen < 1:
                label = f"{int(temps_moyen * 60)} min"
            elif temps_moyen < 24:
                label = f"{temps_moyen:.1f} h"
            else:
                label = f"{temps_moyen / 24:.1f} j"
        else:
            label = "-"

        taux_conversion = round((converties / total) * 100, 1) if total > 0 else 0
        taux_refus = round((refusees / total) * 100, 1) if total > 0 else 0

        return {
            "total": total,
            "en_attente": en_attente,
            "converties": converties,
            "refusees": refusees,
            "taux_conversion": taux_conversion,
            "taux_refus": taux_refus,
            "temps_moyen_reponse_heures": temps_moyen,
            "temps_moyen_reponse_label": label,
            "publiques": publiques,
            "authentifiees": authentifiees,
        }
    except Exception as e:
        logger.error(f"Erreur KPI DI: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/intervention-requests/{request_id}", response_model=InterventionRequest, tags=["Demandes Intervention"])
async def get_intervention_request(request_id: str, current_user: dict = Depends(require_permission("interventionRequests", "view"))):
    """Récupérer une demande d'intervention spécifique"""
    req = await db.intervention_requests.find_one({"id": request_id})
    if not req:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    req = _clean_ir_attachments(req)
    return InterventionRequest(**req)

@api_router.put("/intervention-requests/{request_id}", response_model=InterventionRequest, tags=["Demandes Intervention"])
async def update_intervention_request(
    request_id: str,
    request_update: InterventionRequestUpdate,
    current_user: dict = Depends(require_permission("interventionRequests", "edit"))
):
    """Mettre à jour une demande d'intervention"""
    req = await db.intervention_requests.find_one({"id": request_id})
    if not req:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    update_data = {k: v for k, v in request_update.model_dump().items() if v is not None}
    
    # Mettre à jour l'équipement si nécessaire
    if "equipement_id" in update_data:
        if update_data["equipement_id"]:
            eq_info = await get_equipment_by_id(update_data["equipement_id"])
            if eq_info:
                update_data["equipement"] = eq_info
        else:
            update_data["equipement"] = None
    
    # Mettre à jour l'emplacement si nécessaire
    if "emplacement_id" in update_data:
        if update_data["emplacement_id"]:
            loc_info = await get_location_by_id(update_data["emplacement_id"])
            if loc_info:
                update_data["emplacement"] = loc_info
        else:
            update_data["emplacement"] = None
    
    await db.intervention_requests.update_one({"id": request_id}, {"$set": update_data})
    
    # Récupérer la demande mise à jour
    updated_req = await db.intervention_requests.find_one({"id": request_id})
    
    # Broadcast WebSocket pour la synchronisation temps réel
    broadcast_data = {k: v for k, v in dict(updated_req).items() if k != '_id'}
    broadcast_data = _clean_ir_attachments(broadcast_data)
    await realtime_manager.emit_event(
        "intervention_requests",
        "updated",
        broadcast_data
    )
    
    # Audit log
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
        user_email=current_user["email"],
        action=ActionType.UPDATE,
        entity_type=EntityType_Audit.WORK_ORDER,
        entity_id=request_id,
        entity_name=updated_req['titre'],
        details=f"Modification demande d'intervention"
    )
    
    updated_req = _clean_ir_attachments(updated_req)
    return InterventionRequest(**updated_req)

@api_router.delete("/intervention-requests/{request_id}", response_model=MessageResponse, tags=["Demandes Intervention"])
async def delete_intervention_request(request_id: str, current_user: dict = Depends(require_permission("interventionRequests", "delete"))):
    """Supprimer une demande d'intervention"""
    req = await db.intervention_requests.find_one({"id": request_id})
    if not req:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    req_title = req.get('titre', 'Sans titre')
    
    await db.intervention_requests.delete_one({"id": request_id})
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "intervention_requests",
        "deleted",
        {"id": request_id, "titre": req_title}
    )
    
    # Audit log
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
        user_email=current_user["email"],
        action=ActionType.DELETE,
        entity_type=EntityType_Audit.WORK_ORDER,
        entity_id=request_id,
        entity_name=req_title,
        details=f"Suppression demande d'intervention"
    )
    
    return {"message": "Demande supprimée"}

# ==================== INTERVENTION REQUEST ATTACHMENTS ====================
IR_UPLOAD_DIR = Path("/app/backend/uploads/intervention-requests")
IR_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

@api_router.post("/intervention-requests/{request_id}/attachments",
    summary="Uploader une piece jointe a une demande d'intervention", response_model=AttachmentResponse, tags=["Demandes Intervention"])
async def upload_ir_attachment(
    request_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission("interventionRequests", "edit"))
):
    """Uploader une piece jointe a une demande d'intervention (max 25MB)"""
    try:
        req = await db.intervention_requests.find_one({"id": request_id})
        if not req:
            raise HTTPException(status_code=404, detail="Demande d'intervention non trouvee")
        
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 25MB)")
        
        file_ext = Path(file.filename).suffix
        unique_filename = f"{uuid.uuid4()}{file_ext}"
        file_path = IR_UPLOAD_DIR / unique_filename
        
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(content)
        
        attachment = {
            "_id": ObjectId(),
            "filename": unique_filename,
            "original_filename": file.filename,
            "size": len(content),
            "mime_type": file.content_type or mimetypes.guess_type(file.filename)[0] or "application/octet-stream",
            "uploaded_at": datetime.utcnow()
        }
        
        await db.intervention_requests.update_one(
            {"id": request_id},
            {"$push": {"attachments": attachment}}
        )
        
        attachment_response = {
            "id": str(attachment["_id"]),
            "filename": attachment["filename"],
            "original_filename": attachment["original_filename"],
            "size": attachment["size"],
            "mime_type": attachment["mime_type"],
            "uploaded_at": attachment["uploaded_at"],
            "url": f"/api/intervention-requests/{request_id}/attachments/{str(attachment['_id'])}"
        }
        
        return attachment_response
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur upload piece jointe DI: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/intervention-requests/{request_id}/attachments/{attachment_id}",
    summary="Telecharger une piece jointe", tags=["Demandes Intervention"])
async def download_ir_attachment(
    request_id: str,
    attachment_id: str,
    current_user: dict = Depends(require_permission("interventionRequests", "view"))
):
    """Telecharger une piece jointe d'une demande d'intervention"""
    try:
        req = await db.intervention_requests.find_one({"id": request_id})
        if not req:
            raise HTTPException(status_code=404, detail="Demande non trouvee")
        
        attachment = None
        for att in req.get("attachments", []):
            # Support both formats: _id (standard) and id (old public)
            att_id_str = str(att.get("_id", "")) if att.get("_id") else att.get("id", "")
            if att_id_str == attachment_id:
                attachment = att
                break
        
        if not attachment:
            raise HTTPException(status_code=404, detail="Piece jointe non trouvee")
        
        file_path = IR_UPLOAD_DIR / attachment["filename"]
        
        # Fallback: check old public upload path
        if not file_path.exists():
            old_path = Path(f"/app/backend/uploads/intervention_requests/{request_id}") / attachment["filename"]
            if old_path.exists():
                file_path = old_path
        
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Fichier non trouve sur le disque")
        
        from starlette.responses import FileResponse
        return FileResponse(
            path=str(file_path),
            filename=attachment["original_filename"],
            media_type=attachment.get("mime_type", "application/octet-stream")
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur download piece jointe DI: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/intervention-requests/{request_id}/attachments/{attachment_id}",
    summary="Supprimer une piece jointe", response_model=MessageResponse, tags=["Demandes Intervention"])
async def delete_ir_attachment(
    request_id: str,
    attachment_id: str,
    current_user: dict = Depends(require_permission("interventionRequests", "edit"))
):
    """Supprimer une piece jointe d'une demande d'intervention"""
    try:
        req = await db.intervention_requests.find_one({"id": request_id})
        if not req:
            raise HTTPException(status_code=404, detail="Demande non trouvee")
        
        attachment = None
        for att in req.get("attachments", []):
            if str(att.get("_id")) == attachment_id:
                attachment = att
                break
        
        if not attachment:
            raise HTTPException(status_code=404, detail="Piece jointe non trouvee")
        
        # Delete file from disk
        file_path = IR_UPLOAD_DIR / attachment["filename"]
        if file_path.exists():
            file_path.unlink()
        
        # Remove from DB
        await db.intervention_requests.update_one(
            {"id": request_id},
            {"$pull": {"attachments": {"_id": attachment["_id"]}}}
        )
        
        return {"message": "Piece jointe supprimee"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur suppression piece jointe DI: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== INTERVENTION REQUEST REFUSAL ====================

class RefuseInterventionRequest(BaseModel):
    motif: str

@api_router.post("/intervention-requests/{request_id}/refuse", response_model=dict, tags=["Demandes Intervention"])
async def refuse_intervention_request(
    request_id: str,
    refuse_data: RefuseInterventionRequest,
    current_user: dict = Depends(require_permission("interventionRequests", "edit"))
):
    """Refuser une demande d'intervention avec un motif"""
    try:
        req = await db.intervention_requests.find_one({"id": request_id})
        if not req:
            raise HTTPException(status_code=404, detail="Demande non trouvee")
        
        refused_by_name = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}"
        
        update_data = {
            "refused": True,
            "refused_reason": refuse_data.motif,
            "refused_at": datetime.utcnow(),
            "refused_by": current_user["id"],
            "refused_by_name": refused_by_name
        }
        
        await db.intervention_requests.update_one(
            {"id": request_id},
            {"$set": update_data}
        )
        
        # Envoyer email au demandeur
        try:
            creator_id = req.get("created_by")
            if creator_id:
                creator = await db.users.find_one({"id": creator_id})
                if creator and creator.get("email"):
                    subject = f"Demande d'intervention refusee - {req.get('titre', '')}"
                    html_content = f"""
                    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                        <div style="background-color: #dc2626; color: white; padding: 20px; border-radius: 8px 8px 0 0;">
                            <h2 style="margin: 0;">Demande d'intervention refusee</h2>
                        </div>
                        <div style="padding: 20px; background-color: #f9fafb; border: 1px solid #e5e7eb; border-radius: 0 0 8px 8px;">
                            <p>Bonjour {creator.get('prenom', '')} {creator.get('nom', '')},</p>
                            <p>Votre demande d'intervention a ete refusee.</p>
                            <div style="background: white; padding: 15px; border-radius: 8px; border: 1px solid #e5e7eb; margin: 15px 0;">
                                <p><strong>Titre :</strong> {req.get('titre', '')}</p>
                                <p><strong>Description :</strong> {req.get('description', '')}</p>
                                <p style="color: #dc2626;"><strong>Motif du refus :</strong> {refuse_data.motif}</p>
                                <p><strong>Refuse par :</strong> {refused_by_name}</p>
                            </div>
                            <p>Cordialement,<br>FSAO Iris - GMAO</p>
                        </div>
                    </div>
                    """
                    email_service.send_email(creator["email"], subject, html_content)
        except Exception as email_err:
            logger.warning(f"Erreur envoi email refus DI: {str(email_err)}")
        
        # Audit log
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType_Audit.WORK_ORDER,
            entity_id=request_id,
            entity_name=req.get('titre', ''),
            details=f"Refus demande d'intervention - Motif: {refuse_data.motif}"
        )
        
        # Broadcast WebSocket
        updated_req = await db.intervention_requests.find_one({"id": request_id})
        broadcast_data = {k: v for k, v in dict(updated_req).items() if k != '_id'}
        broadcast_data = _clean_ir_attachments(broadcast_data)
        await realtime_manager.emit_event(
            "intervention_requests",
            "updated",
            broadcast_data
        )
        
        return {"message": "Demande d'intervention refusee", "request_id": request_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur refus demande d'intervention: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/intervention-requests/{request_id}/convert-to-work-order", response_model=dict, tags=["Demandes Intervention"])
async def convert_to_work_order(
    request_id: str,
    assignee_id: Optional[str] = None,
    date_limite: Optional[str] = None,
    current_user: dict = Depends(require_permission("interventionRequests", "edit"))
):
    """Convertir une demande d'intervention en ordre de travail (Admin/Technicien uniquement)"""
    # Vérifier que l'utilisateur est admin ou technicien
    if current_user.get("role") not in ["ADMIN", "TECHNICIEN"]:
        raise HTTPException(status_code=403, detail="Accès refusé : Seuls les administrateurs et techniciens peuvent convertir des demandes")
    
    try:
        # Récupérer la demande
        req = await db.intervention_requests.find_one({"id": request_id})
        if not req:
            raise HTTPException(status_code=404, detail="Demande non trouvée")
        
        # Vérifier si déjà convertie
        if req.get("work_order_id"):
            raise HTTPException(status_code=400, detail="Cette demande a déjà été convertie en ordre de travail")
        
        # Créer l'ordre de travail
        work_order_id = str(uuid.uuid4())
        
        # Générer le numéro d'ordre (comme pour les créations normales)
        count = await db.work_orders.count_documents({})
        numero = str(5800 + count + 1)
        
        # Utiliser la date limite fournie ou celle de la demande
        date_limite_ordre = None
        if date_limite:
            date_limite_ordre = datetime.fromisoformat(date_limite.replace('Z', '+00:00'))
        elif req.get("date_limite_desiree"):
            date_limite_ordre = req.get("date_limite_desiree")
        
        work_order_data = {
            "id": work_order_id,
            "numero": numero,
            "titre": req["titre"],
            "description": req["description"],
            "statut": "OUVERT",
            "priorite": req["priorite"],
            "equipement_id": req.get("equipement_id"),
            "equipement": req.get("equipement"),
            "emplacement_id": req.get("emplacement_id"),
            "emplacement": req.get("emplacement"),
            "assigne_a_id": assignee_id,
            "assigneA": None,
            "dateLimite": date_limite_ordre,
            "tempsEstime": None,
            "dateCreation": datetime.utcnow(),
            "createdBy": req["created_by"],
            "createdByName": req.get("created_by_name"),
            "tempsReel": None,
            "dateTermine": None,
            "attachments": [],
            "comments": []
        }
        
        # Transférer les pièces jointes de la DI vers l'OT
        ir_attachments = req.get("attachments", [])
        if ir_attachments:
            import shutil
            WO_UPLOAD_DIR = Path("/app/backend/uploads/work-orders")
            WO_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            for att in ir_attachments:
                try:
                    src_path = IR_UPLOAD_DIR / att["filename"]
                    # Fallback: check old public upload path
                    if not src_path.exists():
                        old_path = Path(f"/app/backend/uploads/intervention_requests/{request_id}") / att["filename"]
                        if old_path.exists():
                            src_path = old_path
                    if src_path.exists():
                        new_filename = f"{uuid.uuid4()}{Path(att['filename']).suffix}"
                        dst_path = WO_UPLOAD_DIR / new_filename
                        shutil.copy2(str(src_path), str(dst_path))
                        wo_attachment = {
                            "_id": ObjectId(),
                            "filename": new_filename,
                            "original_filename": att.get("original_filename", att["filename"]),
                            "size": att.get("size", 0),
                            "mime_type": att.get("mime_type", "application/octet-stream"),
                            "uploaded_at": datetime.utcnow()
                        }
                        work_order_data["attachments"].append(wo_attachment)
                    else:
                        logger.warning(f"Piece jointe DI introuvable: {att['filename']}")
                except Exception as copy_err:
                    logger.warning(f"Erreur copie piece jointe DI->OT: {str(copy_err)}")
        
        # Récupérer les informations de l'assigné si fourni
        if assignee_id:
            assignee = await db.users.find_one({"id": assignee_id})
            if assignee:
                work_order_data["assigneA"] = {
                    "id": assignee["id"],
                    "nom": assignee["nom"],
                    "prenom": assignee["prenom"]
                }
        
        await db.work_orders.insert_one(work_order_data)
        
        # Mettre à jour la demande avec les informations de l'ordre créé
        await db.intervention_requests.update_one(
            {"id": request_id},
            {"$set": {
                "work_order_id": work_order_id,
                "work_order_numero": numero,
                "work_order_date_limite": date_limite_ordre,
                "converted_at": datetime.utcnow(),
                "converted_by": current_user["id"]
            }}
        )
        
        # Émettre un événement pour rafraîchir les notifications
        # Note: Dans une vraie application, on utiliserait des WebSockets
        
        return {
            "message": "Demande convertie en ordre de travail avec succès",
            "work_order_id": work_order_id,
            "request_id": request_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur conversion demande: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Configuration CORS - Autoriser toutes les origines pour accès depuis IP publique
logger.info(f"🔒 CORS configuré pour autoriser TOUTES les origines (accès IP publique)")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],  # Autoriser toutes les origines pour accès depuis IP publique
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["*"],
)

# Middleware anti-cache pour toutes les réponses API
from starlette.middleware.base import BaseHTTPMiddleware

class NoCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        if request.url.path.startswith("/api/"):
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
        return response

app.add_middleware(NoCacheMiddleware)



# ==================== IMPROVEMENT REQUESTS (DEMANDES D'AMÉLIORATION) ENDPOINTS ====================

@api_router.post("/improvement-requests", response_model=ImprovementRequest, status_code=201, tags=["Demandes Amelioration"])
async def create_improvement_request(
    request: ImprovementRequestCreate,
    current_user: dict = Depends(require_permission("improvementRequests", "edit"))
):
    """Créer une nouvelle demande d'amélioration"""
    try:
        request_id = str(uuid.uuid4())
        request_data = request.model_dump()
        request_data["id"] = request_id
        request_data["date_creation"] = datetime.utcnow()
        request_data["created_by"] = current_user["id"]
        request_data["created_by_name"] = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}"
        request_data["service"] = current_user.get("service")  # Service du demandeur
        request_data["status"] = "SOUMISE"  # Statut initial
        request_data["history"] = [{
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "user_id": current_user["id"],
            "user_name": request_data["created_by_name"],
            "action": "Création de la demande",
            "old_status": None,
            "new_status": "SOUMISE"
        }]
        request_data["improvement_id"] = None
        request_data["improvement_numero"] = None
        request_data["improvement_date_limite"] = None
        request_data["converted_at"] = None
        request_data["converted_by"] = None
        request_data["_id"] = ObjectId()
        
        if request_data.get("equipement_id"):
            equipment = await db.equipments.find_one({"id": request_data["equipement_id"]})
            if equipment:
                request_data["equipement"] = {"id": equipment["id"], "nom": equipment["nom"]}
        
        if request_data.get("emplacement_id"):
            location = await db.locations.find_one({"id": request_data["emplacement_id"]})
            if location:
                request_data["emplacement"] = {"id": location["id"], "nom": location["nom"]}
        
        await db.improvement_requests.insert_one(request_data)
        
        # Serialize for response but preserve the original UUID id
        original_id = request_data["id"]
        request_data = serialize_doc(request_data)
        request_data["id"] = original_id
        
        # Envoyer email au responsable de service pour validation
        try:
            from improvement_request_email_service import send_improvement_request_email_to_manager
            
            user_service = current_user.get("service")
            recipients = []
            
            if user_service:
                # Chercher le(s) responsable(s) de ce service
                service_managers = await db.service_responsables.find(
                    {"service": user_service}
                ).to_list(100)
                
                for manager_entry in service_managers:
                    manager = await db.users.find_one(
                        {"id": manager_entry["user_id"], "statut": "actif"},
                        {"_id": 0, "id": 1, "email": 1, "nom": 1, "prenom": 1}
                    )
                    if manager and manager.get("email"):
                        recipients.append(manager)
            
            # Si pas de responsable, envoyer aux admins
            if not recipients:
                admins = await db.users.find(
                    {"role": "ADMIN", "statut": "actif"},
                    {"_id": 0, "id": 1, "email": 1, "nom": 1, "prenom": 1}
                ).to_list(10)
                recipients = [a for a in admins if a.get("email")]
            
            # Récupérer les pièces jointes (si uploadées)
            attachments = request_data.get("attachments", [])
            
            # Envoyer l'email avec boutons d'action
            for recipient in recipients:
                await send_improvement_request_email_to_manager(
                    request_data=request_data,
                    recipient=recipient,
                    attachments=attachments
                )
                
        except Exception as email_error:
            logger.warning(f"Erreur envoi email notification création demande: {email_error}")
            # On ne bloque pas la création si l'email échoue
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "improvement_requests",
            "created",
            request_data,
            user_id=current_user["id"]
        )
        
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
            user_email=current_user["email"],
            action=ActionType.CREATE,
            entity_type=EntityType_Audit.IMPROVEMENT_REQUEST,
            entity_id=request_id,
            entity_name=request.titre,
            details=f"Création demande d'amélioration"
        )
        
        return ImprovementRequest(**request_data)
    except Exception as e:
        logger.error(f"Erreur création demande d'amélioration: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/improvement-requests", tags=["Demandes Amelioration"])
async def get_all_improvement_requests(current_user: dict = Depends(require_permission("improvementRequests", "view"))):
    """Récupérer toutes les demandes d'amélioration"""
    try:
        query = {}
        
        requests = []
        async for req in db.improvement_requests.find(query).sort("date_creation", -1):
            req_dict = serialize_doc(req)
            
            # Enrichir avec les informations du créateur
            if req.get("created_by_id"):
                creator = await db.users.find_one({"id": req["created_by_id"]})
                if creator:
                    req_dict["created_by_prenom"] = creator.get("prenom", "")
                    req_dict["created_by_nom"] = creator.get("nom", "")
            
            # Enrichir avec les informations de l'ordre de travail associé
            if req.get("work_order_id"):
                work_order = await db.work_orders.find_one({"id": req["work_order_id"]})
                if work_order:
                    req_dict["work_order_temps_reel"] = work_order.get("tempsReel", 0)
            
            requests.append(req_dict)
        return requests
    except Exception as e:
        logger.error(f"Erreur récupération demandes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/improvement-requests/pending-validation", response_model=List[dict], tags=["Demandes Amelioration"])
async def get_pending_improvement_requests(
    current_user: dict = Depends(require_permission("improvementRequests", "view"))
):
    """Récupérer les demandes d'amélioration en attente de validation pour le responsable"""
    from service_filter import is_service_manager, get_user_managed_services
    
    try:
        is_admin = current_user.get("role") == "ADMIN"
        is_manager = await is_service_manager(current_user)
        
        if not is_admin and not is_manager:
            return []  # Pas de permissions pour voir les demandes en attente
        
        query = {"$or": [{"status": "SOUMISE"}, {"status": {"$exists": False}}, {"status": None}]}
        
        # Si responsable (non admin), filtrer par service
        if not is_admin and is_manager:
            managed_services = await get_user_managed_services(current_user)
            if managed_services:
                # Récupérer les IDs des utilisateurs de ces services
                service_users = await db.users.find(
                    {"service": {"$in": managed_services}},
                    {"id": 1, "_id": 0}
                ).to_list(1000)
                user_ids = [u["id"] for u in service_users]
                
                query = {
                    "$and": [
                        {"$or": [{"status": "SOUMISE"}, {"status": {"$exists": False}}, {"status": None}]},
                        {"$or": [
                            {"created_by": {"$in": user_ids}},
                            {"service": {"$in": managed_services}}
                        ]}
                    ]
                }
        
        requests = await db.improvement_requests.find(query, {"_id": 0}).sort("date_creation", -1).to_list(1000)
        
        # Enrichir avec les infos créateur
        for req in requests:
            if req.get("created_by"):
                creator = await db.users.find_one({"id": req["created_by"]}, {"_id": 0, "service": 1})
                if creator:
                    req["service"] = creator.get("service")
        
        return requests
        
    except Exception as e:
        logger.error(f"Erreur récupération demandes en attente: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/improvement-requests/{request_id}", response_model=ImprovementRequest, tags=["Demandes Amelioration"])
async def get_improvement_request(request_id: str, current_user: dict = Depends(require_permission("improvementRequests", "view"))):
    """Récupérer une demande d'amélioration spécifique"""
    req = await db.improvement_requests.find_one({"id": request_id})
    if not req:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    req = serialize_doc(req)
    return ImprovementRequest(**req)

@api_router.put("/improvement-requests/{request_id}", response_model=ImprovementRequest, tags=["Demandes Amelioration"])
async def update_improvement_request(
    request_id: str,
    request_update: ImprovementRequestUpdate,
    current_user: dict = Depends(require_permission("improvementRequests", "edit"))
):
    """Mettre à jour une demande d'amélioration"""
    req = await db.improvement_requests.find_one({"id": request_id})
    if not req:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    update_data = {k: v for k, v in request_update.model_dump().items() if v is not None}
    
    if "equipement_id" in update_data:
        if update_data["equipement_id"]:
            equipment = await db.equipments.find_one({"id": update_data["equipement_id"]})
            if equipment:
                update_data["equipement"] = {"id": equipment["id"], "nom": equipment["nom"]}
        else:
            update_data["equipement"] = None
    
    if "emplacement_id" in update_data:
        if update_data["emplacement_id"]:
            location = await db.locations.find_one({"id": update_data["emplacement_id"]})
            if location:
                update_data["emplacement"] = {"id": location["id"], "nom": location["nom"]}
        else:
            update_data["emplacement"] = None
    
    await db.improvement_requests.update_one({"id": request_id}, {"$set": update_data})
    updated_req = await db.improvement_requests.find_one({"id": request_id})
    updated_req = serialize_doc(updated_req)
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "improvement_requests",
        "updated",
        updated_req,
        user_id=current_user["id"]
    )
    
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
        user_email=current_user["email"],
        action=ActionType.UPDATE,
        entity_type=EntityType_Audit.IMPROVEMENT_REQUEST,
        entity_id=request_id,
        entity_name=updated_req['titre'],
        details=f"Modification demande d'amélioration"
    )
    
    return ImprovementRequest(**updated_req)

@api_router.delete("/improvement-requests/{request_id}", response_model=MessageResponse, tags=["Demandes Amelioration"])
async def delete_improvement_request(request_id: str, current_user: dict = Depends(require_permission("improvementRequests", "delete"))):
    """Supprimer une demande d'amélioration"""
    # Essayer d'abord avec le champ 'id' (pour les nouveaux documents avec UUID)
    req = await db.improvement_requests.find_one({"id": request_id})
    
    # Si non trouvé, essayer avec _id (pour les anciens documents)
    if not req:
        try:
            req = await db.improvement_requests.find_one({"_id": ObjectId(request_id)})
        except:
            pass
    
    if not req:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    req_title = req.get('titre', 'Sans titre')
    
    # Supprimer avec le même critère utilisé pour trouver
    if req.get("id") == request_id:
        await db.improvement_requests.delete_one({"id": request_id})
    else:
        await db.improvement_requests.delete_one({"_id": req["_id"]})
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "improvement_requests",
        "deleted",
        {"id": request_id, "titre": req_title},
        user_id=current_user["id"]
    )
    
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=current_user.get("nom", "") + " " + current_user.get("prenom", ""),
        user_email=current_user["email"],
        action=ActionType.DELETE,
        entity_type=EntityType_Audit.IMPROVEMENT_REQUEST,
        entity_id=request_id,
        entity_name=req_title,
        details=f"Suppression demande d'amélioration"
    )
    
    return {"message": "Demande supprimée"}


@api_router.put("/improvement-requests/{request_id}/status", response_model=dict, tags=["Demandes Amelioration"])
async def update_improvement_request_status(
    request_id: str,
    status_update: ImprovementRequestStatusUpdate,
    current_user: dict = Depends(require_permission("improvementRequests", "edit"))
):
    """Valider ou rejeter une demande d'amélioration (Responsable de service ou Admin)"""
    from service_filter import is_service_manager, get_user_managed_services
    
    try:
        # Récupérer la demande
        req = await db.improvement_requests.find_one({"id": request_id})
        if not req:
            try:
                req = await db.improvement_requests.find_one({"_id": ObjectId(request_id)})
            except:
                pass
        
        if not req:
            raise HTTPException(status_code=404, detail="Demande non trouvée")
        
        # Vérifier que le statut actuel permet la validation
        current_status = req.get("status", "SOUMISE")
        if current_status not in ["SOUMISE", None]:
            raise HTTPException(
                status_code=400, 
                detail=f"Cette demande ne peut pas être modifiée (statut actuel: {current_status})"
            )
        
        # Vérifier les permissions
        is_admin = current_user.get("role") == "ADMIN"
        is_manager = await is_service_manager(current_user)
        
        if not is_admin and not is_manager:
            raise HTTPException(
                status_code=403, 
                detail="Seuls les administrateurs et responsables de service peuvent valider/rejeter"
            )
        
        # Si responsable de service, vérifier que la demande est dans son service
        if not is_admin and is_manager:
            managed_services = await get_user_managed_services(current_user)
            request_service = req.get("service")
            
            # Récupérer le service du demandeur si non présent sur la demande
            if not request_service:
                creator = await db.users.find_one({"id": req.get("created_by")}, {"service": 1})
                request_service = creator.get("service") if creator else None
            
            if request_service and request_service not in managed_services:
                raise HTTPException(
                    status_code=403, 
                    detail="Vous ne pouvez valider que les demandes de votre service"
                )
        
        # Valider le nouveau statut
        new_status = status_update.status.upper()
        if new_status not in ["VALIDEE", "REJETEE"]:
            raise HTTPException(status_code=400, detail="Statut invalide. Utilisez VALIDEE ou REJETEE")
        
        # Préparer la mise à jour
        user_name = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}"
        now = datetime.now(timezone.utc)
        
        history_entry = {
            "timestamp": now.isoformat(),
            "user_id": current_user["id"],
            "user_name": user_name,
            "action": "Validation" if new_status == "VALIDEE" else "Rejet",
            "old_status": current_status,
            "new_status": new_status,
            "comment": status_update.comment
        }
        
        update_data = {
            "status": new_status,
            "validated_at": now.isoformat(),
            "validated_by": current_user["id"],
            "validated_by_name": user_name
        }
        
        if new_status == "REJETEE" and status_update.comment:
            update_data["rejection_reason"] = status_update.comment
        
        # Mettre à jour dans la DB
        query_filter = {"id": request_id} if req.get("id") == request_id else {"_id": ObjectId(request_id)}
        await db.improvement_requests.update_one(
            query_filter,
            {
                "$set": update_data,
                "$push": {"history": history_entry}
            }
        )
        
        # Notifier le demandeur par email
        try:
            creator = await db.users.find_one({"id": req.get("created_by")}, {"_id": 0, "email": 1, "nom": 1, "prenom": 1})
            if creator and creator.get("email"):
                status_label = "validée" if new_status == "VALIDEE" else "rejetée"
                subject = f"Demande d'amélioration {status_label}: {req['titre']}"
                
                body = f"""
                <h2>Votre demande d'amélioration a été {status_label}</h2>
                <p><strong>Titre :</strong> {req['titre']}</p>
                <p><strong>Par :</strong> {user_name}</p>
                """
                if status_update.comment:
                    body += f"<p><strong>Commentaire :</strong> {status_update.comment}</p>"
                
                if new_status == "VALIDEE":
                    body += "<p>Votre demande sera prochainement convertie en projet d'amélioration.</p>"
                else:
                    body += "<p>Vous pouvez soumettre une nouvelle demande avec les modifications suggérées.</p>"
                
                email_service.send_email(
                    to_email=creator["email"],
                    subject=subject,
                    html_content=body
                )
        except Exception as e:
            logger.warning(f"Erreur envoi email notification: {e}")
        
        # Broadcast WebSocket
        updated_req = await db.improvement_requests.find_one(query_filter)
        if updated_req:
            updated_req = serialize_doc(updated_req)
            await realtime_manager.emit_event(
                "improvement_requests",
                "status_changed",
                updated_req,
                user_id=current_user["id"]
            )
        
        logger.info(f"✅ Demande d'amélioration {req['titre']} {new_status.lower()} par {user_name}")
        
        return {
            "message": f"Demande {new_status.lower()} avec succès",
            "status": new_status
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur validation demande d'amélioration: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ Endpoints pour action via email (sans authentification) ============

@api_router.get("/improvement-requests/email-action/validate/{token}", tags=["Demandes Amelioration"])
async def validate_improvement_request_email_token(token: str):
    """Valide un token d'approbation et retourne les informations de la demande"""
    from improvement_request_email_service import validate_approval_token
    
    try:
        token_data = await validate_approval_token(token)
        
        if not token_data:
            raise HTTPException(status_code=400, detail="Token invalide ou expiré")
        
        if token_data.get("request_type") != "improvement_request":
            raise HTTPException(status_code=400, detail="Token non valide pour ce type de demande")
        
        # Récupérer la demande
        request_id = token_data.get("request_id")
        req = await db.improvement_requests.find_one({"id": request_id}, {"_id": 0})
        
        if not req:
            req = await db.improvement_requests.find_one({"_id": ObjectId(request_id)})
            if req:
                req = serialize_doc(req)
        
        if not req:
            raise HTTPException(status_code=404, detail="Demande non trouvée")
        
        # Vérifier que la demande n'a pas déjà été traitée
        if req.get("status") not in ["SOUMISE", None]:
            raise HTTPException(status_code=400, detail="Cette demande a déjà été traitée")
        
        return {
            "token_data": {
                "action": token_data.get("action"),
                "expiration": token_data.get("expiration")
            },
            "request": req
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur validation token: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/improvement-requests/email-action/{token}", tags=["Demandes Amelioration"])
async def process_improvement_request_email_action(token: str, action_data: dict = None):
    """Traite une action d'approbation/rejet via le lien email"""
    from improvement_request_email_service import validate_approval_token, mark_token_used, send_validation_notification_email
    
    try:
        token_data = await validate_approval_token(token)
        
        if not token_data:
            raise HTTPException(status_code=400, detail="Token invalide ou expiré")
        
        if token_data.get("request_type") != "improvement_request":
            raise HTTPException(status_code=400, detail="Token non valide pour ce type de demande")
        
        request_id = token_data.get("request_id")
        user_id = token_data.get("user_id")
        
        # Récupérer la demande
        req = await db.improvement_requests.find_one({"id": request_id})
        query_filter = {"id": request_id}
        
        if not req:
            try:
                req = await db.improvement_requests.find_one({"_id": ObjectId(request_id)})
                query_filter = {"_id": ObjectId(request_id)}
            except:
                pass
        
        if not req:
            raise HTTPException(status_code=404, detail="Demande non trouvée")
        
        # Vérifier que la demande n'a pas déjà été traitée
        current_status = req.get("status", "SOUMISE")
        if current_status not in ["SOUMISE", None]:
            raise HTTPException(status_code=400, detail="Cette demande a déjà été traitée")
        
        # Récupérer l'utilisateur qui a validé
        validator = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not validator:
            raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
        
        validator_name = f"{validator.get('prenom', '')} {validator.get('nom', '')}"
        
        # Déterminer l'action (depuis le body ou depuis le token)
        action = "reject"
        if action_data and action_data.get("action"):
            action = action_data.get("action")
        elif token_data.get("action") == "approve":
            action = "approve"
        
        new_status = "VALIDEE" if action == "approve" else "REJETEE"
        comment = action_data.get("comment") if action_data else None
        
        now = datetime.now(timezone.utc)
        
        # Créer l'entrée d'historique
        history_entry = {
            "timestamp": now.isoformat(),
            "user_id": user_id,
            "user_name": validator_name,
            "action": "Validation via email" if new_status == "VALIDEE" else "Rejet via email",
            "old_status": current_status,
            "new_status": new_status,
            "comment": comment
        }
        
        # Mettre à jour la demande
        update_data = {
            "status": new_status,
            "validated_at": now.isoformat(),
            "validated_by": user_id,
            "validated_by_name": validator_name
        }
        
        if new_status == "REJETEE" and comment:
            update_data["rejection_reason"] = comment
        
        await db.improvement_requests.update_one(
            query_filter,
            {
                "$set": update_data,
                "$push": {"history": history_entry}
            }
        )
        
        # Marquer le token comme utilisé
        await mark_token_used(token)
        
        # Envoyer notification au demandeur
        try:
            creator = await db.users.find_one({"id": req.get("created_by")}, {"_id": 0})
            if creator:
                await send_validation_notification_email(req, creator, new_status, validator_name, comment)
        except Exception as email_error:
            logger.warning(f"Erreur envoi email notification: {email_error}")
        
        # Broadcast WebSocket
        updated_req = await db.improvement_requests.find_one(query_filter)
        if updated_req:
            updated_req = serialize_doc(updated_req)
            await realtime_manager.emit_event(
                "improvement_requests",
                "status_changed",
                updated_req
            )
        
        logger.info(f"✅ Demande d'amélioration {req['titre']} {new_status.lower()} via email par {validator_name}")
        
        return {
            "message": f"Demande {new_status.lower()} avec succès",
            "status": new_status
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur action email: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/improvement-requests/{request_id}/convert-to-improvement", response_model=dict, tags=["Demandes Amelioration"])
async def convert_to_improvement(
    request_id: str,
    assignee_id: Optional[str] = None,
    date_limite: Optional[str] = None,
    current_user: dict = Depends(require_permission("improvementRequests", "edit"))
):
    """Convertir une demande d'amélioration en amélioration (Admin/Technicien uniquement)"""
    if current_user.get("role") not in ["ADMIN", "TECHNICIEN"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Essayer d'abord avec le champ 'id' (nouveaux documents UUID)
        req = await db.improvement_requests.find_one({"id": request_id})
        query_field = "id"
        
        # Si non trouvé, essayer avec _id (anciens documents ObjectId)
        if not req:
            try:
                req = await db.improvement_requests.find_one({"_id": ObjectId(request_id)})
                query_field = "_id"
            except:
                pass
        
        if not req:
            raise HTTPException(status_code=404, detail="Demande non trouvée")
        
        if req.get("improvement_id"):
            raise HTTPException(status_code=400, detail="Cette demande a déjà été convertie")
        
        improvement_id = str(uuid.uuid4())
        count = await db.improvements.count_documents({})
        numero = str(7000 + count + 1)
        
        date_limite_imp = None
        if date_limite:
            date_limite_imp = datetime.fromisoformat(date_limite.replace('Z', '+00:00'))
        elif req.get("date_limite_desiree"):
            date_limite_imp = req.get("date_limite_desiree")
        
        improvement_data = {
            "id": improvement_id,
            "numero": numero,
            "titre": req["titre"],
            "description": req["description"],
            "statut": "OUVERT",
            "priorite": req["priorite"],
            "equipement_id": req.get("equipement_id"),
            "equipement": req.get("equipement"),
            "emplacement_id": req.get("emplacement_id"),
            "emplacement": req.get("emplacement"),
            "assigne_a_id": assignee_id,
            "assigneA": None,
            "dateLimite": date_limite_imp,
            "tempsEstime": None,
            "dateCreation": datetime.utcnow(),
            "createdBy": req["created_by"],
            "createdByName": req.get("created_by_name"),
            "tempsReel": None,
            "dateTermine": None,
            "attachments": [],
            "comments": []
        }
        
        if assignee_id:
            assignee = await db.users.find_one({"id": assignee_id})
            if assignee:
                improvement_data["assigneA"] = {
                    "id": assignee["id"],
                    "nom": assignee["nom"],
                    "prenom": assignee["prenom"]
                }
        
        await db.improvements.insert_one(improvement_data)
        
        # Utiliser le bon champ pour la mise à jour
        if query_field == "id":
            update_query = {"id": request_id}
        else:
            update_query = {"_id": req["_id"]}
        
        await db.improvement_requests.update_one(
            update_query,
            {"$set": {
                "improvement_id": improvement_id,
                "improvement_numero": numero,
                "improvement_date_limite": date_limite_imp,
                "converted_at": datetime.utcnow(),
                "converted_by": current_user["id"]
            }}
        )
        
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}",
            user_email=current_user.get("email", ""),
            action=ActionType.CREATE,
            entity_type=EntityType.IMPROVEMENT,
            entity_id=improvement_id,
            entity_name=f"Amélioration #{numero}",
            details=f"Converti depuis demande: {req['titre']}"
        )
        
        return {
            "message": "Demande convertie en amélioration avec succès",
            "improvement_id": improvement_id,
            "improvement_numero": numero,
            "request_id": request_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur conversion demande: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Attachments et Comments pour Improvement Requests
@api_router.post("/improvement-requests/{request_id}/attachments", tags=["Demandes Amelioration"])
async def upload_improvement_request_attachment(
    request_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission("improvementRequests", "edit"))
):
    """Upload fichier pour une demande d'amélioration"""
    req = await db.improvement_requests.find_one({"id": request_id})
    if not req:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    return await upload_attachment_generic(request_id, file, "improvement_requests", current_user)

@api_router.get("/improvement-requests/{request_id}/attachments/{filename}", tags=["Demandes Amelioration"])
async def download_improvement_request_attachment(request_id: str, filename: str, current_user: dict = Depends(require_permission("improvementRequests", "view"))):
    """Télécharger un fichier d'une demande d'amélioration"""
    return await download_attachment_generic(request_id, filename, "improvement_requests")

@api_router.post("/improvement-requests/{request_id}/comments", tags=["Demandes Amelioration"])
async def add_improvement_request_comment(
    request_id: str,
    comment_data: dict,
    current_user: dict = Depends(require_permission("improvementRequests", "edit"))
):
    """Ajouter un commentaire à une demande d'amélioration"""
    req = await db.improvement_requests.find_one({"id": request_id})
    if not req:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    comment = {
        "id": str(uuid.uuid4()),
        "text": comment_data.get("text", ""),
        "user_id": current_user["id"],
        "user_name": f"{current_user.get('prenom', '')} {current_user.get('nom', '')}",
        "timestamp": datetime.utcnow().isoformat()
    }
    
    await db.improvement_requests.update_one(
        {"id": request_id},
        {"$push": {"comments": comment}}
    )
    
    return comment

@api_router.get("/improvement-requests/{request_id}/comments", tags=["Demandes Amelioration"])
async def get_improvement_request_comments(request_id: str, current_user: dict = Depends(require_permission("improvementRequests", "view"))):
    """Récupérer les commentaires d'une demande d'amélioration"""
    req = await db.improvement_requests.find_one({"id": request_id})
    if not req:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    return req.get("comments", [])

# Attachments pour Improvements
@api_router.post("/improvements/{imp_id}/attachments", tags=["Ameliorations"])
async def upload_improvement_attachment(
    imp_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission("improvements", "edit"))
):
    """Upload fichier pour une amélioration"""
    imp = await db.improvements.find_one({"id": imp_id})
    if not imp:
        raise HTTPException(status_code=404, detail="Amélioration non trouvée")
    
    return await upload_attachment_generic(imp_id, file, "improvements", current_user)


@api_router.get("/improvements/{imp_id}/attachments", tags=["Ameliorations"])
async def get_improvement_attachments(
    imp_id: str,
    current_user: dict = Depends(require_permission("improvements", "view"))
):
    """Récupérer la liste des pièces jointes d'une amélioration"""
    imp = await db.improvements.find_one({"id": imp_id})
    if not imp:
        raise HTTPException(status_code=404, detail="Amélioration non trouvée")
    
    attachments = imp.get("attachments", [])
    result = []
    for att in attachments:
        result.append({
            "id": att.get("id", ""),
            "filename": att.get("filename", ""),
            "original_filename": att.get("filename", att.get("original_filename", "")),
            "size": att.get("size", 0),
            "mime_type": att.get("type", att.get("mime_type", "application/octet-stream")),
            "uploaded_at": att.get("uploadedAt", att.get("uploaded_at", ""))
        })
    
    return result


@api_router.get("/improvements/{imp_id}/attachments/{attachment_id}", tags=["Ameliorations"])
async def download_improvement_attachment(
    imp_id: str, 
    attachment_id: str, 
    preview: bool = False,
    current_user: dict = Depends(require_permission("improvements", "view"))
):
    """Télécharger ou prévisualiser un fichier d'une amélioration"""
    imp = await db.improvements.find_one({"id": imp_id})
    if not imp:
        raise HTTPException(status_code=404, detail="Amélioration non trouvée")
    
    # Trouver l'attachment par son id
    attachment = None
    for att in imp.get("attachments", []):
        if att.get("id") == attachment_id:
            attachment = att
            break
    
    if not attachment:
        # Fallback: chercher par filename (ancien format)
        return await download_attachment_generic(imp_id, attachment_id, "improvements")
    
    file_path = attachment.get("path")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Fichier non trouvé sur le serveur")
    
    disposition = "inline" if preview else "attachment"
    return FileResponse(
        path=file_path,
        filename=attachment.get("filename", attachment.get("original_filename", "file")),
        media_type=attachment.get("type", attachment.get("mime_type", "application/octet-stream")),
        content_disposition_type=disposition
    )


@api_router.delete("/improvements/{imp_id}/attachments/{attachment_id}", tags=["Ameliorations"])
async def delete_improvement_attachment(
    imp_id: str,
    attachment_id: str,
    current_user: dict = Depends(require_permission("improvements", "edit"))
):
    """Supprimer une pièce jointe d'une amélioration"""
    imp = await db.improvements.find_one({"id": imp_id})
    if not imp:
        raise HTTPException(status_code=404, detail="Amélioration non trouvée")
    
    # Trouver l'attachment
    attachment = None
    for att in imp.get("attachments", []):
        if att.get("id") == attachment_id:
            attachment = att
            break
    
    if not attachment:
        raise HTTPException(status_code=404, detail="Pièce jointe non trouvée")
    
    # Supprimer le fichier physique
    file_path = attachment.get("path")
    if file_path and os.path.exists(file_path):
        os.remove(file_path)
    
    # Retirer de la base de données
    await db.improvements.update_one(
        {"id": imp_id},
        {"$pull": {"attachments": {"id": attachment_id}}}
    )
    
    return {"success": True, "message": "Pièce jointe supprimée"}


# Comments pour Improvements
@api_router.post("/improvements/{imp_id}/comments", tags=["Ameliorations"])
async def add_improvement_comment(
    imp_id: str,
    comment_data: dict,
    current_user: dict = Depends(require_permission("improvements", "edit"))
):
    """Ajouter un commentaire à une amélioration"""
    imp = await db.improvements.find_one({"id": imp_id})
    if not imp:
        raise HTTPException(status_code=404, detail="Amélioration non trouvée")
    
    comment = {
        "id": str(uuid.uuid4()),
        "text": comment_data.get("text", ""),
        "user_id": current_user["id"],
        "user_name": f"{current_user.get('prenom', '')} {current_user.get('nom', '')}",
        "timestamp": datetime.utcnow().isoformat()
    }
    
    await db.improvements.update_one(
        {"id": imp_id},
        {"$push": {"comments": comment}}
    )
    
    return comment

@api_router.get("/improvements/{imp_id}/comments", tags=["Ameliorations"])
async def get_improvement_comments(imp_id: str, current_user: dict = Depends(require_permission("improvements", "view"))):
    """Récupérer les commentaires d'une amélioration"""
    imp = await db.improvements.find_one({"id": imp_id})
    if not imp:
        raise HTTPException(status_code=404, detail="Amélioration non trouvée")
    
    return imp.get("comments", [])

# ==================== NOTIFICATIONS ENDPOINTS ====================

# --- Push Notifications (routes pour l'app mobile Expo) ---
# Ces routes DOIVENT etre definies AVANT les routes /notifications/{notification_id}
# pour eviter que "register", "unregister", "test" soient captures comme des IDs
from notifications import (
    send_expo_push_notification,
    DeviceTokenCreate
)
from web_push import (
    send_web_push_to_user,
    notify_work_order_assigned_web,
    notify_work_order_status_changed_web,
    notify_equipment_alert_web,
    notify_chat_message_web
)

# ============================================================
# WEB PUSH (PWA) - Endpoints
# ============================================================

@api_router.get("/web-push/vapid-key", tags=["Web Push PWA"])
async def get_vapid_public_key():
    """Retourne la cle publique VAPID pour l'abonnement push."""
    key = os.environ.get("VAPID_PUBLIC_KEY")
    if not key:
        raise HTTPException(status_code=500, detail="VAPID key not configured")
    return {"publicKey": key}

@api_router.post("/web-push/subscribe", tags=["Web Push PWA"])
async def web_push_subscribe(
    request: Request,
    current_user: dict = Depends(get_current_user),
):
    """Enregistre un abonnement web push pour l'utilisateur connecte."""
    body = await request.json()
    subscription = body.get("subscription")
    browser = body.get("browser", "unknown")
    
    if not subscription or not subscription.get("endpoint"):
        raise HTTPException(status_code=400, detail="Subscription invalide")
    
    user_id = str(current_user["id"])
    now = datetime.now(timezone.utc)
    
    # Desactiver les anciens abonnements du meme navigateur
    await db.web_push_subscriptions.update_many(
        {"user_id": user_id, "browser": browser, "subscription.endpoint": {"$ne": subscription["endpoint"]}},
        {"$set": {"is_active": False, "updated_at": now}}
    )
    
    # Upsert l'abonnement
    await db.web_push_subscriptions.update_one(
        {"subscription.endpoint": subscription["endpoint"]},
        {"$set": {
            "user_id": user_id,
            "subscription": subscription,
            "browser": browser,
            "is_active": True,
            "updated_at": now
        },
        "$setOnInsert": {"created_at": now}},
        upsert=True
    )
    
    logger.info(f"[WEB PUSH] Abonnement enregistre pour user {user_id} ({browser})")
    return {"message": "Abonnement enregistre", "status": "ok"}

@api_router.post("/web-push/unsubscribe", tags=["Web Push PWA"])
async def web_push_unsubscribe(
    request: Request,
    current_user: dict = Depends(get_current_user),
):
    """Desactive un abonnement web push."""
    body = await request.json()
    endpoint = body.get("endpoint")
    
    if endpoint:
        await db.web_push_subscriptions.update_one(
            {"subscription.endpoint": endpoint},
            {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
        )
    else:
        await db.web_push_subscriptions.update_many(
            {"user_id": str(current_user["id"])},
            {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
        )
    
    return {"message": "Desabonne"}

@api_router.post("/web-push/test", tags=["Web Push PWA"])
async def web_push_test(
    current_user: dict = Depends(get_current_user),
):
    """Envoie une notification de test a l'utilisateur connecte."""
    result = await send_web_push_to_user(
        db, str(current_user["id"]),
        title="Test de notification FSAO",
        body="Les notifications PWA fonctionnent correctement !",
        data={"type": "test"},
        tag="test-notification"
    )
    return result

@api_router.get("/web-push/subscriptions", tags=["Web Push PWA"])
async def web_push_list_subscriptions(
    current_user: dict = Depends(get_current_user),
):
    """Liste les abonnements web push."""
    is_admin = current_user.get("role") == "ADMIN"
    query = {} if is_admin else {"user_id": str(current_user["id"])}
    
    subs = []
    async for doc in db.web_push_subscriptions.find(query, {"_id": 0}):
        for k in ["created_at", "updated_at", "deactivated_at"]:
            if k in doc and hasattr(doc[k], 'isoformat'):
                doc[k] = doc[k].isoformat()
        # Masquer les details de la subscription pour la securite
        if "subscription" in doc:
            doc["endpoint_preview"] = doc["subscription"].get("endpoint", "")[:60] + "..."
            del doc["subscription"]
        subs.append(doc)
    
    return {"total": len(subs), "subscriptions": subs}

# ============================================================
# EXPO PUSH (Mobile) - Endpoints existants
# ============================================================

@api_router.post("/notifications/register", tags=["Push Notifications"])
async def mobile_register_device_token(
    token_data: DeviceTokenCreate,
    current_user: dict = Depends(get_current_user),
):
    """Register a device push token (mobile app endpoint).
    Un utilisateur + un appareil = un seul token actif.
    Les anciens tokens du meme device_name sont desactives."""
    try:
        user_id = str(current_user["id"])
        logger.info(f"[PUSH REGISTER] user_id={user_id}, token={token_data.push_token[:30]}..., platform={token_data.platform}, device={token_data.device_name}")
        now = datetime.now(timezone.utc)

        # Desactiver les anciens tokens du meme utilisateur + meme appareil
        if token_data.device_name:
            old_tokens = await db.device_tokens.update_many(
                {
                    "user_id": user_id,
                    "device_name": token_data.device_name,
                    "push_token": {"$ne": token_data.push_token}
                },
                {"$set": {"is_active": False, "updated_at": now}}
            )
            if old_tokens.modified_count > 0:
                logger.info(f"[PUSH REGISTER] Desactive {old_tokens.modified_count} ancien(s) token(s) pour {token_data.device_name}")

        # Upsert le nouveau token
        result = await db.device_tokens.update_one(
            {"push_token": token_data.push_token},
            {"$set": {
                "user_id": user_id,
                "platform": token_data.platform,
                "device_name": token_data.device_name,
                "updated_at": now,
                "is_active": True
            },
            "$setOnInsert": {
                "created_at": now
            }},
            upsert=True
        )

        if result.upserted_id:
            logger.info(f"[PUSH REGISTER] New token registered for user {user_id}: {str(result.upserted_id)}")
            return {"message": "Token registered", "token_id": str(result.upserted_id)}
        else:
            logger.info(f"[PUSH REGISTER] Token updated for user {user_id}")
            return {"message": "Token updated"}
    except Exception as e:
        logger.error(f"[PUSH REGISTER] ERROR: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur enregistrement token: {str(e)}")

@api_router.delete("/notifications/unregister", tags=["Push Notifications"])
async def mobile_unregister_device_token(
    push_token: str,
    current_user: dict = Depends(get_current_user),
):
    """Unregister a device push token (mobile app endpoint)."""
    user_id = str(current_user["id"])
    result = await db.device_tokens.update_one(
        {"user_id": user_id, "push_token": push_token},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Token not found")
    return {"message": "Token unregistered"}

@api_router.post("/notifications/test", tags=["Push Notifications"])
async def mobile_test_notification(
    current_user: dict = Depends(get_current_user),
):
    """Send a test notification to the current user (mobile app endpoint)."""
    user_id = str(current_user["id"])
    tokens_cursor = db.device_tokens.find({"user_id": user_id, "is_active": True})
    tokens = [doc["push_token"] async for doc in tokens_cursor]
    if not tokens:
        raise HTTPException(status_code=404, detail="No registered devices")
    result = await send_expo_push_notification(
        push_tokens=tokens,
        title="Test de notification",
        body="Les notifications fonctionnent correctement !",
        data={"type": "test"},
        db=db
    )
    return result

@api_router.get("/notifications/devices", tags=["Push Notifications"])
async def get_registered_devices(
    current_user: dict = Depends(get_current_user),
):
    """Diagnostic: list all registered device tokens (admin: all users, other: own only)."""
    if current_user.get("role") == "ADMIN":
        cursor = db.device_tokens.find({}, {"_id": 0})
    else:
        cursor = db.device_tokens.find(
            {"user_id": str(current_user["id"])}, {"_id": 0}
        )
    devices = []
    async for doc in cursor:
        if "created_at" in doc:
            doc["created_at"] = doc["created_at"].isoformat()
        if "updated_at" in doc:
            doc["updated_at"] = doc["updated_at"].isoformat()
        devices.append(doc)
    return {"total": len(devices), "devices": devices}


@api_router.get("/notifications/diagnostic", tags=["Push Notifications"])
async def push_notification_diagnostic(
    current_user: dict = Depends(get_current_user),
):
    """Diagnostic complet des notifications push.
    Teste chaque token individuellement et retourne le resultat brut d'Expo."""
    if current_user.get("role") != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin only")
    import httpx
    
    report = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "etapes": [],
        "tokens_db": [],
        "tokens_inactifs": [],
        "receipts_recents": [],
        "test_envoi": [],
        "conclusion": "",
        "actions_recommandees": []
    }
    
    # ETAPE 1: Lister TOUS les tokens (actifs + inactifs)
    all_tokens = []
    async for doc in db.device_tokens.find({}, {"_id": 0}):
        for k in ["created_at", "updated_at"]:
            if k in doc and hasattr(doc[k], 'isoformat'):
                doc[k] = doc[k].isoformat()
        all_tokens.append(doc)
    
    actifs = [t for t in all_tokens if t.get("is_active")]
    inactifs = [t for t in all_tokens if not t.get("is_active")]
    
    report["tokens_db"] = actifs
    report["tokens_inactifs"] = inactifs
    report["etapes"].append({
        "etape": "1. Tokens en base",
        "resultat": f"{len(actifs)} actif(s), {len(inactifs)} inactif(s), {len(all_tokens)} total",
        "statut": "OK" if actifs else "ERREUR - Aucun token actif"
    })
    
    if not actifs:
        report["conclusion"] = "ECHEC: Aucun token actif en base. L'application mobile n'a pas enregistre de token, ou le systeme de nettoyage les a supprimes."
        report["actions_recommandees"] = [
            "1. Ouvrir l'application mobile et se connecter",
            "2. Verifier que l'app appelle POST /api/notifications/register au demarrage",
            "3. Revenir ici et relancer le diagnostic"
        ]
        return report
    
    # ETAPE 2: Tester chaque token actif individuellement avec Expo
    for token_doc in actifs:
        push_token = token_doc["push_token"]
        test_result = {
            "token": push_token[:40] + "..." if len(push_token) > 40 else push_token,
            "user_id": token_doc.get("user_id"),
            "device_name": token_doc.get("device_name"),
            "platform": token_doc.get("platform"),
        }
        
        try:
            test_message = {
                "to": push_token,
                "title": "Diagnostic FSAO",
                "body": "Test de diagnostic automatique",
                "sound": "default",
                "priority": "high",
                "data": {"type": "diagnostic_test"}
            }
            
            async with httpx.AsyncClient() as client:
                resp = await client.post(
                    "https://exp.host/--/api/v2/push/send",
                    json=[test_message],
                    headers={
                        "Accept": "application/json",
                        "Content-Type": "application/json",
                    },
                    timeout=15.0
                )
                expo_response = resp.json()
            
            test_result["expo_http_status"] = resp.status_code
            test_result["expo_reponse_brute"] = expo_response
            
            # Analyser la reponse
            tickets = expo_response.get("data", [])
            if tickets:
                ticket = tickets[0]
                ticket_status = ticket.get("status")
                ticket_id = ticket.get("id")
                
                test_result["ticket_status"] = ticket_status
                test_result["ticket_id"] = ticket_id
                
                if ticket_status == "ok":
                    test_result["verdict"] = "OK - Expo a accepte le message. Notification en cours de livraison."
                    
                    # Verifier le receipt immediatement (attendre 5s)
                    if ticket_id:
                        await asyncio.sleep(5)
                        try:
                            receipt_resp = await httpx.AsyncClient().post(
                                "https://exp.host/--/api/v2/push/getReceipts",
                                json={"ids": [ticket_id]},
                                headers={"Content-Type": "application/json"},
                                timeout=10.0
                            )
                            receipt_data = receipt_resp.json()
                            receipt_info = receipt_data.get("data", {}).get(ticket_id, {})
                            test_result["receipt_verification"] = receipt_info if receipt_info else "Pas encore disponible (normal, reessayer dans 15min)"
                            
                            if receipt_info.get("status") == "error":
                                error_detail = receipt_info.get("details", {}).get("error", "")
                                test_result["verdict"] = f"ECHEC LIVRAISON - Expo a accepte puis refuse: {error_detail}"
                                if error_detail == "DeviceNotRegistered":
                                    test_result["explication"] = "Le token n'est pas reconnu par FCM/APNs. L'app mobile doit etre recompiee avec les credentials Firebase."
                        except Exception as e:
                            test_result["receipt_verification"] = f"Erreur verification: {str(e)}"
                    
                elif ticket_status == "error":
                    error_detail = ticket.get("details", {}).get("error", "")
                    error_message = ticket.get("message", "")
                    test_result["verdict"] = f"REFUSE PAR EXPO - {error_detail}: {error_message}"
                    
                    if error_detail == "DeviceNotRegistered":
                        test_result["explication"] = "Token invalide. L'appareil n'est plus enregistre aupres de FCM."
                    elif error_detail == "InvalidCredentials":
                        test_result["explication"] = "Les credentials FCM du projet Expo sont invalides."
                    elif error_detail == "MessageTooBig":
                        test_result["explication"] = "Message trop gros (ne devrait pas arriver en diagnostic)."
                    else:
                        test_result["explication"] = f"Erreur Expo: {error_detail}"
            else:
                test_result["verdict"] = "REPONSE VIDE - Expo n'a retourne aucun ticket"
                
        except Exception as e:
            test_result["verdict"] = f"ERREUR RESEAU - Impossible de contacter Expo: {str(e)}"
        
        report["test_envoi"].append(test_result)
    
    report["etapes"].append({
        "etape": "2. Test envoi Expo",
        "resultat": f"{len(report['test_envoi'])} token(s) teste(s)",
        "statut": "Voir details dans test_envoi"
    })
    
    # ETAPE 3: Verifier les receipts recents
    recent_receipts = []
    async for doc in db.push_receipts.find({}).sort("created_at", -1).limit(20):
        for k in ["created_at", "checked_at"]:
            if k in doc and hasattr(doc[k], 'isoformat'):
                doc[k] = doc[k].isoformat()
        doc.pop("_id", None)
        recent_receipts.append(doc)
    report["receipts_recents"] = recent_receipts
    report["etapes"].append({
        "etape": "3. Receipts recents",
        "resultat": f"{len(recent_receipts)} receipt(s) en base",
        "statut": "OK" if recent_receipts else "INFO - Aucun receipt (normal si aucune notif n'a ete envoyee)"
    })
    
    # CONCLUSION
    verdicts = [t.get("verdict", "") for t in report["test_envoi"]]
    all_ok = all("OK" in v for v in verdicts)
    all_device_not_registered = all("DeviceNotRegistered" in v for v in verdicts)
    all_refused = all("REFUSE" in v or "ECHEC" in v for v in verdicts)
    
    if all_ok:
        report["conclusion"] = "SUCCES: Toutes les notifications ont ete acceptees par Expo. Si vous ne les recevez toujours pas, le probleme est cote appareil (mode silencieux, batterie, permissions)."
        report["actions_recommandees"] = [
            "1. Verifier les permissions de notification sur l'appareil",
            "2. Verifier que le mode 'Ne pas deranger' est desactive",
            "3. Verifier l'optimisation batterie pour l'app FSAO",
            "4. Forcer l'arret de l'app et la relancer"
        ]
    elif all_device_not_registered:
        report["conclusion"] = "ECHEC: Tous les tokens sont rejetes avec DeviceNotRegistered. Les credentials Firebase (FCM) ne sont pas configurees dans le projet Expo."
        report["actions_recommandees"] = [
            "1. Verifier que google-services.json est present dans le projet mobile",
            "2. Configurer la Server Key FCM dans Expo (expo push:android:upload)",
            "3. Recompiler l'app avec eas build (pas juste expo start)",
            "4. Reinstaller l'app sur l'appareil et se reconnecter",
            "5. Relancer ce diagnostic"
        ]
    elif all_refused:
        report["conclusion"] = "ECHEC: Toutes les notifications sont refusees par Expo. Voir les details dans test_envoi."
        report["actions_recommandees"] = [
            "Transmettre ce rapport complet au support technique"
        ]
    else:
        report["conclusion"] = "RESULTAT MIXTE: Certains tokens fonctionnent, d'autres non. Voir les details."
        report["actions_recommandees"] = [
            "Purger les tokens invalides et reconnecter les appareils"
        ]
    
    return report

@api_router.post("/notifications/send-raw-test", tags=["Push Notifications"])
async def send_raw_test_notification(
    push_token: str = None,
    user_id: str = None,
    current_user: dict = Depends(get_current_user),
):
    """Envoie un test brut a un token specifique ou a un utilisateur.
    Retourne la reponse Expo complete sans filtrage."""
    if current_user.get("role") != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin only")
    import httpx
    
    tokens = []
    if push_token:
        tokens = [push_token]
    elif user_id:
        async for doc in db.device_tokens.find({"user_id": user_id, "is_active": True}):
            tokens.append(doc["push_token"])
    else:
        raise HTTPException(status_code=400, detail="Fournir push_token ou user_id")
    
    if not tokens:
        return {"error": "Aucun token trouve", "tokens_trouves": 0}
    
    results = []
    for token in tokens:
        message = {
            "to": token,
            "title": "Test direct FSAO",
            "body": f"Test envoye a {datetime.now(timezone.utc).strftime('%H:%M:%S')} UTC",
            "sound": "default",
            "priority": "high",
            "data": {"type": "test"}
        }
        
        try:
            async with httpx.AsyncClient() as client:
                resp = await client.post(
                    "https://exp.host/--/api/v2/push/send",
                    json=[message],
                    headers={"Accept": "application/json", "Content-Type": "application/json"},
                    timeout=15.0
                )
            results.append({
                "token": token[:40] + "...",
                "http_status": resp.status_code,
                "expo_response": resp.json()
            })
        except Exception as e:
            results.append({
                "token": token[:40] + "...",
                "error": str(e)
            })
    
    return {"tokens_testes": len(tokens), "resultats": results}


# --- Notifications in-app (routes existantes) ---

@api_router.get("/notifications",
    summary="Lister les notifications", tags=["Notifications"])
async def get_notifications(
    current_user: dict = Depends(get_current_user),
    unread_only: bool = False,
    limit: int = 50
):
    """Récupère les notifications de l'utilisateur connecté"""
    try:
        query = {"user_id": current_user.get("id")}
        if unread_only:
            query["read"] = False
        
        notifications = await db.notifications.find(query).sort("created_at", -1).limit(limit).to_list(limit)
        
        for notif in notifications:
            notif["id"] = str(notif.get("_id", notif.get("id", "")))
            if "_id" in notif:
                del notif["_id"]
        
        return notifications
    except Exception as e:
        logger.error(f"Erreur récupération notifications: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/notifications/count",
    summary="Compteur de notifications non lues", response_model=NotificationCountResponse, tags=["Notifications"])
async def get_notifications_count(current_user: dict = Depends(get_current_user)):
    """Compte les notifications non lues"""
    try:
        count = await db.notifications.count_documents({
            "user_id": current_user.get("id"),
            "read": False
        })
        return {"unread_count": count}
    except Exception as e:
        logger.error(f"Erreur comptage notifications: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/notifications/{notification_id}/read", tags=["Notifications"])
async def mark_notification_read(
    notification_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Marque une notification comme lue"""
    try:
        result = await db.notifications.update_one(
            {"id": notification_id, "user_id": current_user.get("id")},
            {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
        )
        if result.matched_count == 0:
            # Essayer avec _id
            result = await db.notifications.update_one(
                {"_id": ObjectId(notification_id), "user_id": current_user.get("id")},
                {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
            )
        return {"success": True}
    except Exception as e:
        logger.error(f"Erreur marquage notification lue: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/notifications/read-all", tags=["Notifications"])
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    """Marque toutes les notifications comme lues"""
    try:
        await db.notifications.update_many(
            {"user_id": current_user.get("id"), "read": False},
            {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"success": True}
    except Exception as e:
        logger.error(f"Erreur marquage toutes notifications lues: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/notifications/{notification_id}", tags=["Notifications"])
async def delete_notification(
    notification_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprime une notification"""
    try:
        result = await db.notifications.delete_one({
            "id": notification_id,
            "user_id": current_user.get("id")
        })
        if result.deleted_count == 0:
            result = await db.notifications.delete_one({
                "_id": ObjectId(notification_id),
                "user_id": current_user.get("id")
            })
        return {"success": True}
    except Exception as e:
        logger.error(f"Erreur suppression notification: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/notifications/create-rp", tags=["Notifications"])
async def create_rp_notification(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Crée une notification pour un OT 'Réparation à Planifier' (RP)"""
    try:
        # Récupérer tous les utilisateurs avec permission sur les OT (pour notifier les responsables)
        # Pour l'instant, notifier tous les admins et superviseurs
        users = await db.users.find({
            "$or": [
                {"role": "admin"},
                {"role": "supervisor"},
                {"permissions.workOrders.edit": True}
            ]
        }).to_list(100)
        
        notifications_created = 0
        for user in users:
            user_id = str(user.get("_id", user.get("id", "")))
            notification = {
                "id": str(uuid.uuid4()),
                "type": "rp_created",
                "title": f"Nouvel OT: {data.get('rp_ot_titre', 'RP-...')}",
                "message": f"Réparation à Planifier créé suite à {data.get('non_conformities_count', 0)} non-conformité(s) détectée(s) sur \"{data.get('original_ot_titre', 'OT')}\".",
                "priority": "high",
                "user_id": user_id,
                "link": "/work-orders",
                "metadata": {
                    "rp_ot_id": data.get("rp_ot_id"),
                    "rp_ot_titre": data.get("rp_ot_titre"),
                    "non_conformities_count": data.get("non_conformities_count"),
                    "is_rp_notification": True
                },
                "read": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "read_at": None
            }
            await db.notifications.insert_one(notification)
            
            # Émettre via WebSocket
            await realtime_manager.emit_event(
                "notification",
                "created",
                notification,
                user_id=user_id
            )
            notifications_created += 1
        
        return {"success": True, "notifications_created": notifications_created}
    except Exception as e:
        logger.error(f"Erreur création notification RP: {e}")
        raise HTTPException(status_code=500, detail=str(e))

async def create_notification(
    user_id: str,
    notif_type: str,
    title: str,
    message: str,
    priority: str = "medium",
    link: str = None,
    metadata: dict = None
):
    """Crée une notification pour un utilisateur"""
    try:
        notification = {
            "id": str(uuid.uuid4()),
            "type": notif_type,
            "title": title,
            "message": message,
            "priority": priority,
            "user_id": user_id,
            "link": link,
            "metadata": metadata or {},
            "read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "read_at": None
        }
        await db.notifications.insert_one(notification)
        
        # Émettre via WebSocket pour notification temps réel
        await realtime_manager.emit_event(
            "notification",
            "created",
            notification,
            user_id=user_id
        )
        
        return notification
    except Exception as e:
        logger.error(f"Erreur création notification: {e}")
        return None

async def check_pm_notifications():
    """
    Vérifie les maintenances préventives à venir et crée des notifications.
    Appelé par le scheduler quotidiennement.
    """
    try:
        logger.info("🔔 Vérification des notifications PM...")
        now = datetime.now(timezone.utc)
        
        # Récupérer toutes les PM actives
        pm_list = await db.preventive_maintenances.find({"statut": "ACTIF"}).to_list(1000)
        
        notifications_created = 0
        
        for pm in pm_list:
            pm_id = str(pm.get("_id", pm.get("id", "")))
            prochaine = pm.get("prochaineMaintenance")
            
            if not prochaine:
                continue
            
            # Convertir en datetime si nécessaire
            if isinstance(prochaine, str):
                prochaine = datetime.fromisoformat(prochaine.replace('Z', '+00:00'))
            
            # Rendre timezone-aware si nécessaire
            if prochaine.tzinfo is None:
                prochaine = prochaine.replace(tzinfo=timezone.utc)
            
            days_until = (prochaine - now).days
            
            # Récupérer l'utilisateur assigné
            assigne_a_id = pm.get("assigne_a_id")
            if not assigne_a_id:
                continue
            
            # Vérifier si une notification existe déjà pour aujourd'hui
            existing = await db.notifications.find_one({
                "metadata.pm_id": pm_id,
                "created_at": {"$gte": now.replace(hour=0, minute=0, second=0).isoformat()}
            })
            
            if existing:
                continue
            
            titre = pm.get("titre", "Maintenance préventive")
            
            # Notification si maintenance dans 3 jours
            if days_until == 3:
                await create_notification(
                    user_id=assigne_a_id,
                    notif_type="pm_upcoming",
                    title="Maintenance préventive dans 3 jours",
                    message=f"La maintenance \"{titre}\" est prévue dans 3 jours.",
                    priority="medium",
                    link="/preventive-maintenance",
                    metadata={"pm_id": pm_id, "days_until": 3}
                )
                notifications_created += 1
            
            # Notification si maintenance demain
            elif days_until == 1:
                await create_notification(
                    user_id=assigne_a_id,
                    notif_type="pm_upcoming",
                    title="Maintenance préventive demain",
                    message=f"La maintenance \"{titre}\" est prévue pour demain.",
                    priority="high",
                    link="/preventive-maintenance",
                    metadata={"pm_id": pm_id, "days_until": 1}
                )
                notifications_created += 1
            
            # Notification si maintenance aujourd'hui
            elif days_until == 0:
                await create_notification(
                    user_id=assigne_a_id,
                    notif_type="pm_upcoming",
                    title="Maintenance préventive aujourd'hui",
                    message=f"La maintenance \"{titre}\" est prévue pour aujourd'hui !",
                    priority="urgent",
                    link="/preventive-maintenance",
                    metadata={"pm_id": pm_id, "days_until": 0}
                )
                notifications_created += 1
            
            # Notification si maintenance en retard
            elif days_until < 0:
                await create_notification(
                    user_id=assigne_a_id,
                    notif_type="pm_overdue",
                    title="Maintenance préventive en retard",
                    message=f"La maintenance \"{titre}\" est en retard de {abs(days_until)} jour(s) !",
                    priority="urgent",
                    link="/preventive-maintenance",
                    metadata={"pm_id": pm_id, "days_overdue": abs(days_until)}
                )
                notifications_created += 1
        
        logger.info(f"🔔 {notifications_created} notifications PM créées")
        return notifications_created
        
    except Exception as e:
        logger.error(f"❌ Erreur vérification notifications PM: {e}")
        return 0

# ==================== IMPROVEMENTS (AMÉLIORATIONS) ENDPOINTS ====================

@api_router.get("/improvements", response_model=List[Improvement], tags=["Ameliorations"])
async def get_improvements(
    current_user: dict = Depends(require_permission("improvements", "view")),
    statut: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    date_type: str = "creation"
):
    """Récupérer toutes les améliorations avec filtres"""
    try:
        query = {}
        
        if statut:
            query["statut"] = statut
        
        if start_date or end_date:
            date_field = "dateCreation" if date_type == "creation" else "dateLimite"
            date_filter = {}
            if start_date:
                date_filter["$gte"] = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            if end_date:
                date_filter["$lte"] = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query[date_field] = date_filter
        
        improvements = []
        async for imp in db.improvements.find(query).sort("dateCreation", -1):
            if imp.get("assigne_a_id"):
                imp["assigneA"] = await get_user_by_id(imp["assigne_a_id"])
            if imp.get("emplacement_id"):
                imp["emplacement"] = await get_location_by_id(imp["emplacement_id"])
            if imp.get("equipement_id"):
                imp["equipement"] = await get_equipment_by_id(imp["equipement_id"])
            
            if imp.get("createdBy"):
                try:
                    creator = await db.users.find_one({"id": imp["createdBy"]})
                    if creator:
                        imp["createdByName"] = f"{creator.get('prenom', '')} {creator.get('nom', '')}".strip()
                except Exception as e:
                    logger.error(f"Erreur recherche créateur: {e}")
            
            if "numero" not in imp or not imp["numero"]:
                imp["numero"] = "N/A"
            
            improvements.append(Improvement(**imp))
        
        return improvements
    except Exception as e:
        logger.error(f"Erreur récupération améliorations: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/improvements/{imp_id}", response_model=Improvement, tags=["Ameliorations"])
async def get_improvement(imp_id: str, current_user: dict = Depends(require_permission("improvements", "view"))):
    """Détails d'une amélioration"""
    try:
        imp = await db.improvements.find_one({"id": imp_id})
        if not imp:
            raise HTTPException(status_code=404, detail="Amélioration non trouvée")
        
        imp = serialize_doc(imp)
        if imp.get("assigne_a_id"):
            imp["assigneA"] = await get_user_by_id(imp["assigne_a_id"])
        if imp.get("emplacement_id"):
            imp["emplacement"] = await get_location_by_id(imp["emplacement_id"])
        if imp.get("equipement_id"):
            imp["equipement"] = await get_equipment_by_id(imp["equipement_id"])
        
        if imp.get("createdBy"):
            try:
                creator = await db.users.find_one({"id": imp["createdBy"]})
                if creator:
                    imp["createdByName"] = f"{creator.get('prenom', '')} {creator.get('nom', '')}".strip()
            except Exception as e:
                logger.error(f"Erreur recherche créateur: {e}")
        
        if "numero" not in imp or not imp["numero"]:
            imp["numero"] = "N/A"
        
        return Improvement(**imp)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/improvements", response_model=Improvement, tags=["Ameliorations"])
async def create_improvement(imp_create: ImprovementCreate, current_user: dict = Depends(require_permission("improvements", "edit"))):
    """Créer une nouvelle amélioration"""
    count = await db.improvements.count_documents({})
    numero = str(7000 + count + 1)
    
    improvement_id = str(uuid.uuid4())
    improvement_data = imp_create.model_dump()
    improvement_data["id"] = improvement_id
    improvement_data["numero"] = numero
    improvement_data["statut"] = "OUVERT"
    improvement_data["dateCreation"] = datetime.utcnow()
    improvement_data["createdBy"] = current_user["id"]
    improvement_data["createdByName"] = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}"
    improvement_data["tempsReel"] = None
    improvement_data["dateTermine"] = None
    improvement_data["attachments"] = []
    improvement_data["comments"] = []
    
    if improvement_data.get("assigne_a_id"):
        assignee = await db.users.find_one({"id": improvement_data["assigne_a_id"]})
        if assignee:
            improvement_data["assigneA"] = {
                "id": assignee["id"],
                "nom": assignee["nom"],
                "prenom": assignee["prenom"]
            }
    
    if improvement_data.get("equipement_id"):
        equipment = await db.equipments.find_one({"id": improvement_data["equipement_id"]})
        if equipment:
            improvement_data["equipement"] = {"id": equipment["id"], "nom": equipment["nom"]}
    
    if improvement_data.get("emplacement_id"):
        location = await db.locations.find_one({"id": improvement_data["emplacement_id"]})
        if location:
            improvement_data["emplacement"] = {"id": location["id"], "nom": location["nom"]}
    
    await db.improvements.insert_one(improvement_data)
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "improvements",
        "created",
        improvement_data,
        user_id=current_user["id"]
    )
    
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=f"{current_user.get('nom', '')} {current_user.get('prenom', '')}",
        user_email=current_user["email"],
        action=ActionType.CREATE,
        entity_type=EntityType.IMPROVEMENT,
        entity_id=improvement_id,
        entity_name=imp_create.titre,
        details="Création amélioration"
    )
    
    return Improvement(**improvement_data)

@api_router.put("/improvements/{imp_id}", response_model=Improvement, tags=["Ameliorations"])
async def update_improvement(
    imp_id: str,
    imp_update: ImprovementUpdate,
    current_user: dict = Depends(require_permission("improvements", "edit"))
):
    """Mettre à jour une amélioration"""
    imp = await db.improvements.find_one({"id": imp_id})
    if not imp:
        raise HTTPException(status_code=404, detail="Amélioration non trouvée")
    
    # Gérer les champs qui peuvent être explicitement vidés (mis à null)
    update_data = {}
    sent_data = imp_update.model_dump(exclude_unset=True)
    for k, v in imp_update.model_dump().items():
        if v is not None:
            update_data[k] = v
        elif k in sent_data and k in ['assigne_a_id', 'equipement_id', 'emplacement_id', 'dateLimite']:
            update_data[k] = None
    
    if update_data.get("statut") == "TERMINE" and "dateTermine" not in update_data:
        update_data["dateTermine"] = datetime.utcnow()
    
    if "assigne_a_id" in update_data:
        if update_data["assigne_a_id"]:
            assignee = await db.users.find_one({"id": update_data["assigne_a_id"]})
            if assignee:
                update_data["assigneA"] = {
                    "id": assignee["id"],
                    "nom": assignee["nom"],
                    "prenom": assignee["prenom"]
                }
        else:
            update_data["assigneA"] = None
    
    if "equipement_id" in update_data:
        if update_data["equipement_id"]:
            equipment = await db.equipments.find_one({"id": update_data["equipement_id"]})
            if equipment:
                update_data["equipement"] = {"id": equipment["id"], "nom": equipment["nom"]}
        else:
            update_data["equipement"] = None
    
    if "emplacement_id" in update_data:
        if update_data["emplacement_id"]:
            location = await db.locations.find_one({"id": update_data["emplacement_id"]})
            if location:
                update_data["emplacement"] = {"id": location["id"], "nom": location["nom"]}
        else:
            update_data["emplacement"] = None
    
    await db.improvements.update_one({"id": imp_id}, {"$set": update_data})
    updated_imp = await db.improvements.find_one({"id": imp_id})
    
    updated_imp = serialize_doc(updated_imp)
    if updated_imp.get("assigne_a_id"):
        updated_imp["assigneA"] = await get_user_by_id(updated_imp["assigne_a_id"])
    if updated_imp.get("emplacement_id"):
        updated_imp["emplacement"] = await get_location_by_id(updated_imp["emplacement_id"])
    if updated_imp.get("equipement_id"):
        updated_imp["equipement"] = await get_equipment_by_id(updated_imp["equipement_id"])
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "improvements",
        "updated",
        updated_imp,
        user_id=current_user["id"]
    )
    
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=f"{current_user.get('nom', '')} {current_user.get('prenom', '')}",
        user_email=current_user["email"],
        action=ActionType.UPDATE,
        entity_type=EntityType.IMPROVEMENT,
        entity_id=imp_id,
        entity_name=updated_imp["titre"],
        details="Modification amélioration"
    )
    
    return Improvement(**updated_imp)

@api_router.delete("/improvements/{imp_id}", response_model=MessageResponse, tags=["Ameliorations"])
async def delete_improvement(imp_id: str, current_user: dict = Depends(require_permission("improvements", "delete"))):
    """Supprimer une amélioration"""
    imp = await db.improvements.find_one({"id": imp_id})
    if not imp:
        raise HTTPException(status_code=404, detail="Amélioration non trouvée")
    
    imp_titre = imp.get("titre", "Sans titre")
    
    await db.improvements.delete_one({"id": imp_id})
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "improvements",
        "deleted",
        {"id": imp_id, "titre": imp_titre},
        user_id=current_user["id"]
    )
    
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=f"{current_user.get('nom', '')} {current_user.get('prenom', '')}",
        user_email=current_user["email"],
        action=ActionType.DELETE,
        entity_type=EntityType.IMPROVEMENT,
        entity_id=imp_id,
        entity_name=imp_titre,
        details="Suppression amélioration"
    )
    
    return {"message": "Amélioration supprimée"}

@api_router.post("/improvements/{imp_id}/add-time", tags=["Ameliorations"])
async def add_time_to_improvement(imp_id: str, time_data: AddTimeSpent, current_user: dict = Depends(require_permission("improvements", "edit"))):
    """Ajouter du temps passé à une amélioration"""
    try:
        # Récupérer l'amélioration existante
        existing_imp = await db.improvements.find_one({"id": imp_id})
        if not existing_imp:
            raise HTTPException(status_code=404, detail="Amélioration non trouvée")
        
        # Convertir le temps en heures décimales
        time_to_add = time_data.hours + (time_data.minutes / 60.0)
        
        # Récupérer le temps réel actuel (0 si None)
        current_time = existing_imp.get("tempsReel", 0) or 0
        
        # Calculer le nouveau temps réel
        new_time = current_time + time_to_add
        
        # Créer une entrée d'historique de temps avec l'utilisateur qui l'a saisi
        time_entry = {
            "id": str(uuid.uuid4()),
            "user_id": current_user["id"],
            "user_name": f"{current_user['prenom']} {current_user['nom']}",
            "hours": time_to_add,
            "timestamp": datetime.now(timezone.utc)
        }
        
        # Mettre à jour l'amélioration avec le temps total ET l'entrée d'historique
        await db.improvements.update_one(
            {"id": imp_id},
            {
                "$set": {"tempsReel": new_time},
                "$push": {"time_entries": time_entry}
            }
        )
        
        # Log dans l'audit
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType.IMPROVEMENT,
            entity_id=str(existing_imp["id"]),
            entity_name=existing_imp["titre"],
            details=f"Ajout de temps passé: {time_data.hours}h{time_data.minutes:02d}min",
            changes={"tempsReel_old": current_time, "tempsReel_new": new_time, "time_added": time_to_add}
        )
        
        # Récupérer l'amélioration mise à jour
        imp = await db.improvements.find_one({"id": imp_id})
        imp = serialize_doc(imp)
        
        if imp.get("assigne_a_id"):
            imp["assigneA"] = await get_user_by_id(imp["assigne_a_id"])
        if imp.get("emplacement_id"):
            imp["emplacement"] = await get_location_by_id(imp["emplacement_id"])
        if imp.get("equipement_id"):
            imp["equipement"] = await get_equipment_by_id(imp["equipement_id"])
        
        return Improvement(**imp)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur lors de l'ajout de temps : {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

    if not imp:
        raise HTTPException(status_code=404, detail="Amélioration non trouvée")
    
    await db.improvements.delete_one({"id": imp_id})
    
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=f"{current_user.get('nom', '')} {current_user.get('prenom', '')}",
        user_email=current_user["email"],
        action=ActionType.DELETE,
        entity_type=EntityType.IMPROVEMENT,
        entity_id=imp_id,
        entity_name=imp["titre"],
        details="Suppression amélioration"
    )
    
    return {"message": "Amélioration supprimée"}


# ==================== UPDATE MANAGEMENT ENDPOINTS ====================
from update_service import UpdateService, MaintenanceMode

# Initialiser le service de mise à jour
update_service = UpdateService(db)

@api_router.get("/updates/check-version")
async def check_updates_version(current_user: dict = Depends(get_current_admin_user)):
    """
    Vérifie si une mise à jour est disponible via version.json (Admin uniquement)
    """
    try:
        update_info = await update_service.check_for_updates()
        return update_info if update_info else {"available": False, "current_version": update_service.current_version}
    except Exception as e:
        logger.error(f"❌ Erreur vérification mises à jour: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/updates/status")
async def get_update_status(current_user: dict = Depends(get_current_admin_user)):
    """
    Récupère le statut actuel du système de mise à jour
    """
    try:
        return {
            "current_version": update_service.current_version,
            "status": "ready"
        }
    except Exception as e:
        logger.error(f"❌ Erreur statut mise à jour: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/updates/check-conflicts")
async def check_git_conflicts(current_user: dict = Depends(get_current_admin_user)):
    """
    Vérifie s'il y a des conflits Git avant une mise à jour (Admin uniquement)
    Retourne la liste des fichiers modifiés localement
    """
    try:
        result = update_service.check_git_conflicts()
        return result
    except Exception as e:
        logger.error(f"Erreur lors de la vérification des conflits: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/updates/resolve-conflicts")
async def resolve_git_conflicts(
    strategy: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Résout les conflits Git selon la stratégie choisie (Admin uniquement)
    strategy: "reset" (écraser), "stash" (sauvegarder), ou "abort" (annuler)
    """
    try:
        result = update_service.resolve_git_conflicts(strategy)
        
        # Journaliser l'action
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType.SETTINGS,
            entity_id="git_conflicts",
            entity_name=f"Résolution conflits Git ({strategy})",
            details=result.get("message", "")
        )
        
        return result
    except Exception as e:
        logger.error(f"Erreur lors de la résolution des conflits: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

async def get_update_status(current_user: dict = Depends(get_current_admin_user)):
    """
    Récupère le statut actuel des mises à jour (Admin uniquement)
    """
    try:
        status = await update_service.get_update_status()
        return status
    except Exception as e:
        logger.error(f"❌ Erreur récupération statut: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/updates/dismiss/{version}")
async def dismiss_update(version: str, current_user: dict = Depends(get_current_admin_user)):
    """
    Marque une notification de mise à jour comme dismissée (Admin uniquement)
    """
    try:
        await update_service.dismiss_update_notification(version)
        return {"message": "Notification dismissée"}
    except Exception as e:
        logger.error(f"❌ Erreur dismiss notification: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/updates/broadcast-warning")
async def broadcast_update_warning(
    version: str = "",
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Diffuse un avertissement de mise à jour à TOUS les utilisateurs connectés via WebSocket.
    Après 30 secondes, les utilisateurs seront automatiquement déconnectés côté frontend.
    """
    try:
        admin_name = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip()
        connected_count = len(chat_manager.active_connections)
        
        logger.info(f"📢 Diffusion avertissement MAJ par {admin_name} - {connected_count} utilisateur(s) connecté(s)")
        
        # Broadcast via le WebSocket du chat (tous les utilisateurs connectés)
        await chat_manager.broadcast({
            "type": "update_warning",
            "message": "Une mise à jour va être effectuée. Vous serez déconnecté dans 30 secondes. Vous pourrez vous reconnecter dans 5 minutes.",
            "admin_name": admin_name,
            "version": version,
            "countdown_seconds": 30,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        
        # Broadcast via les WebSocket consignes aussi
        from consignes_routes import consigne_connections
        for uid, ws in list(consigne_connections.items()):
            try:
                await ws.send_json({
                    "type": "update_warning",
                    "message": "Une mise à jour va être effectuée. Vous serez déconnecté dans 30 secondes.",
                    "countdown_seconds": 30
                })
            except Exception:
                pass
        
        # Log dans l'audit
        await audit_service.log_action(
            user_id=current_user.get("id"),
            user_name=admin_name,
            user_email=current_user.get("email"),
            action=ActionType.UPDATE,
            entity_type=EntityType.SETTINGS,
            entity_id="update_warning_broadcast",
            entity_name=f"Avertissement MAJ diffusé ({connected_count} utilisateurs)"
        )
        
        return {
            "success": True,
            "connected_users": connected_count,
            "message": f"Avertissement envoyé à {connected_count} utilisateur(s)"
        }
    except Exception as e:
        logger.error(f"❌ Erreur broadcast avertissement MAJ: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/updates/apply")
async def apply_update_endpoint(
    version: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """Lance une mise a jour in-process (Admin uniquement)."""
    try:
        logger.info(f"[MAJ] Demande MAJ vers {version} par {current_user.get('email')}")
        await audit_service.log_action(
            user_id=current_user.get("id"),
            user_name=f"{current_user.get('prenom')} {current_user.get('nom')}",
            user_email=current_user.get("email"),
            action=ActionType.UPDATE,
            entity_type=EntityType.SETTINGS,
            entity_id="system_update",
            entity_name=f"Mise a jour vers {version}"
        )
        result = await update_service.apply_update(version)
        if result.get("accepted") or result.get("success"):
            return {
                "accepted": True,
                "success": True,
                "message": result.get("message", "Mise a jour lancee"),
                "update_id": result.get("update_id"),
                "version": version,
                "code_updated": result.get("code_updated", False),
                "errors": result.get("errors", []),
                "diagnostic": result.get("diagnostic", {})
            }
        else:
            raise HTTPException(status_code=500, detail=result.get("message", "Erreur"))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[MAJ] Erreur: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/updates/log")
async def get_update_log(current_user: dict = Depends(get_current_admin_user)):
    """
    Retourne le log de la derniere mise a jour.
    Source PRINCIPALE: MongoDB (fiable, survit au reboot).
    """
    try:
        last_result = await db.system_settings.find_one({"key": "last_update_result"}, {"_id": 0})
        if last_result and last_result.get("log_output"):
            return {
                "found": True,
                "path": "MongoDB",
                "content": last_result["log_output"],
                "in_progress": last_result.get("in_progress", False),
                "current_step": last_result.get("current_step", ""),
                "errors": last_result.get("errors", []),
                "status": last_result.get("status", ""),
                "success": last_result.get("success", False)
            }
        
        import glob as glob_mod
        log_candidates = ["/var/log/gmao-iris-update.log", "/var/log/gmao-iris-worker.log",
                          "/tmp/gmao-iris-update.log", "/tmp/gmao-iris-worker.log"]
        for path in log_candidates:
            if path and os.path.exists(path) and os.path.getsize(path) > 10:
                with open(path, 'r', errors='replace') as f:
                    content = f.read()
                return {
                    "found": True,
                    "path": path,
                    "content": content[-50000:],
                    "in_progress": last_result.get("in_progress", False) if last_result else False
                }

        return {
            "found": False,
            "content": "",
            "message": "Aucun log disponible."
        }
    except Exception as e:
        logger.error(f"[MAJ] Erreur lecture log: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/updates/last-result")
async def get_last_update_result(current_user: dict = Depends(get_current_admin_user)):
    """
    Récupère le résultat de la dernière mise à jour depuis la base de données.
    Permet au frontend de vérifier si la mise à jour a réellement réussi après un redémarrage.
    """
    try:
        result = await db.system_settings.find_one({"key": "last_update_result"}, {"_id": 0})
        if result:
            return {
                "has_result": True,
                "success": result.get("success", False),
                "code_updated": result.get("code_updated", False),
                "in_progress": result.get("in_progress", False),
                "version_before": result.get("version_before"),
                "version_after": result.get("version_after"),
                "errors": result.get("errors", []),
                "warnings": result.get("warnings", []),
                "completed_at": result.get("completed_at")
            }
        return {"has_result": False, "in_progress": False}
    except Exception as e:
        logger.error(f"❌ Erreur récupération résultat MAJ: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/maintenance/activate")
async def activate_maintenance_mode(current_user: dict = Depends(get_current_admin_user)):
    """Active la page de maintenance NGINX (Admin uniquement)."""
    try:
        maintenance = MaintenanceMode(Path(update_service.app_root))
        success = maintenance.activate()
        if success:
            return {"status": "ok", "message": "Page de maintenance activée", "maintenance_active": True}
        raise HTTPException(status_code=500, detail="Échec de l'activation de la maintenance")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/maintenance/deactivate")
async def deactivate_maintenance_mode(current_user: dict = Depends(get_current_admin_user)):
    """Désactive la page de maintenance NGINX (Admin uniquement)."""
    try:
        maintenance = MaintenanceMode(Path(update_service.app_root))
        success = maintenance.deactivate()
        if success:
            return {"status": "ok", "message": "Page de maintenance désactivée", "maintenance_active": False}
        raise HTTPException(status_code=500, detail="Échec de la désactivation de la maintenance")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/maintenance/status")
async def get_maintenance_status(current_user: dict = Depends(get_current_admin_user)):
    """Vérifie si la page de maintenance est active (Admin uniquement)."""
    try:
        flag_path = Path(update_service.app_root) / "maintenance.flag"
        state_path = Path(update_service.app_root) / "health_state.json"
        history_path = Path(update_service.app_root) / "health_recovery_history.json"
        result = {
            "maintenance_active": flag_path.exists(),
            "health_state": None,
            "recovery_history": [],
        }
        if state_path.exists():
            import json as json_mod
            with open(state_path) as f:
                health_state = json_mod.load(f)
            result["health_state"] = {
                "consecutive_failures": health_state.get("consecutive_failures", 0),
                "last_check": health_state.get("last_check"),
                "last_success": health_state.get("last_success"),
                "last_failure": health_state.get("last_failure"),
                "last_recovery_level": health_state.get("last_recovery_level", 0),
                "total_recoveries": health_state.get("total_recoveries", 0),
            }
        if history_path.exists():
            import json as json_mod
            with open(history_path) as f:
                result["recovery_history"] = json_mod.load(f)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/health/recovery-history")
async def get_recovery_history(current_user: dict = Depends(get_current_admin_user)):
    """Historique des récupérations automatiques (Admin uniquement)."""
    try:
        history_path = Path(update_service.app_root) / "health_recovery_history.json"
        if history_path.exists():
            import json as json_mod
            with open(history_path) as f:
                return json_mod.load(f)
        return []
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/health/force-check")
async def force_health_check(current_user: dict = Depends(get_current_admin_user)):
    """Lance un health check immédiat (Admin uniquement)."""
    try:
        import urllib.request
        checks = {}
        # Backend self-check
        checks["backend"] = {"status": "ok", "message": "API opérationnelle"}
        # MongoDB check
        try:
            await db.command("ping")
            checks["mongodb"] = {"status": "ok", "message": "MongoDB connecté"}
        except Exception as e:
            checks["mongodb"] = {"status": "error", "message": str(e)}
        # Disk usage
        try:
            import shutil
            usage = shutil.disk_usage("/")
            used_pct = round((usage.used / usage.total) * 100, 1)
            free_gb = round(usage.free / (1024**3), 1)
            checks["disk"] = {
                "status": "ok" if used_pct < 90 else "warning",
                "message": f"{used_pct}% utilisé, {free_gb} Go libre"
            }
        except Exception as e:
            checks["disk"] = {"status": "error", "message": str(e)}
        # Memory
        try:
            with open("/proc/meminfo") as f:
                meminfo = f.read()
            total = int([l for l in meminfo.split('\n') if 'MemTotal' in l][0].split()[1])
            available = int([l for l in meminfo.split('\n') if 'MemAvailable' in l][0].split()[1])
            used_pct = round(((total - available) / total) * 100, 1)
            checks["memory"] = {
                "status": "ok" if used_pct < 85 else "warning",
                "message": f"{used_pct}% utilisé"
            }
        except Exception:
            checks["memory"] = {"status": "unknown", "message": "Impossible de lire /proc/meminfo"}

        overall = "ok"
        for c in checks.values():
            if c["status"] == "error":
                overall = "error"
                break
            if c["status"] == "warning":
                overall = "warning"
        return {"overall": overall, "checks": checks, "timestamp": datetime.now(timezone.utc).isoformat()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/health/reset-failures")
async def reset_failure_counter(current_user: dict = Depends(get_current_admin_user)):
    """Remet à zéro le compteur d'échecs consécutifs (Admin uniquement)."""
    try:
        state_path = Path(update_service.app_root) / "health_state.json"
        import json as json_mod
        state = {}
        if state_path.exists():
            with open(state_path) as f:
                state = json_mod.load(f)
        state["consecutive_failures"] = 0
        state["last_recovery_level"] = 0
        with open(state_path, "w") as f:
            json_mod.dump(state, f, indent=2)
        return {"status": "ok", "message": "Compteur d'échecs remis à zéro"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ──────── HEALTH ALERTS CONFIG ────────

@api_router.get("/health/alerts-config")
async def get_health_alerts_config(current_user: dict = Depends(get_current_admin_user)):
    """Récupère la configuration des alertes santé système."""
    config = await db.health_alerts_config.find_one({}, {"_id": 0})
    if not config:
        config = {
            "enabled": False,
            "recipients": [],
            "cooldown_hours": 24,
            "alerts": {
                "app_down": {"enabled": True, "threshold": 1},
                "recovery_success": {"enabled": True},
                "recovery_failed": {"enabled": True},
                "disk_warning": {"enabled": True, "threshold": 80},
                "memory_warning": {"enabled": True, "threshold": 85},
                "maintenance_changed": {"enabled": False},
            },
            "last_test_sent": None,
        }
    return config


@api_router.put("/health/alerts-config")
async def update_health_alerts_config(
    config: dict,
    current_user: dict = Depends(get_current_admin_user)
):
    """Met à jour la configuration des alertes santé système."""
    try:
        allowed_fields = ["enabled", "recipients", "cooldown_hours", "alerts"]
        update_data = {k: v for k, v in config.items() if k in allowed_fields}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        update_data["updated_by"] = current_user.get("name", current_user.get("email", ""))

        await db.health_alerts_config.update_one(
            {}, {"$set": update_data}, upsert=True
        )
        return {"status": "ok", "message": "Configuration des alertes mise à jour"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/health/alerts-test")
async def test_health_alert(current_user: dict = Depends(get_current_admin_user)):
    """Envoie un email de test pour vérifier la configuration des alertes."""
    try:
        config = await db.health_alerts_config.find_one({}, {"_id": 0})
        if not config or not config.get("recipients"):
            raise HTTPException(status_code=400, detail="Aucun destinataire configuré")

        from health_alert_service import send_email, _build_html_email
        recipients = config["recipients"]
        admin_name = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip() or current_user.get("email", "")
        now_str = datetime.now().strftime("%d/%m/%Y à %H:%M")

        details_html = f"""
        <div style="background: #EFF6FF; border-left: 4px solid #2563EB; padding: 12px 16px; border-radius: 4px; margin-bottom: 16px;">
          <p style="margin: 0; color: #1E40AF; font-weight: 600;">Test de configuration réussi</p>
          <p style="margin: 4px 0 0; color: #3B82F6; font-size: 13px;">Déclenché par : {admin_name}</p>
        </div>
        <p style="font-size: 14px; color: #334155;">
          Si vous recevez cet email, les alertes de santé système sont correctement configurées.
        </p>
        <p style="font-size: 13px; color: #64748b;">
          Les alertes actives vous notifieront automatiquement en cas de problème.
        </p>
        """
        html = _build_html_email("[TEST] FSAO Iris - Alerte Système", "info", details_html, f"Date : {now_str}<br>")

        sent = 0
        errors = []
        for email in recipients:
            try:
                ok = send_email(email.strip(), "[TEST] FSAO Iris - Test Alerte Système", html)
                if ok:
                    sent += 1
                else:
                    errors.append(email)
            except Exception as e:
                errors.append(f"{email}: {str(e)}")

        # Update last test timestamp
        await db.health_alerts_config.update_one(
            {}, {"$set": {"last_test_sent": datetime.now(timezone.utc).isoformat()}}, upsert=True
        )

        if sent > 0:
            return {"status": "ok", "message": f"Email de test envoyé à {sent} destinataire(s)", "sent": sent, "errors": errors}
        raise HTTPException(status_code=500, detail=f"Échec d'envoi : {', '.join(errors)}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/health/alerts-history")
async def get_health_alerts_history(current_user: dict = Depends(get_current_admin_user)):
    """Récupère l'historique des alertes envoyées."""
    try:
        history_path = Path(update_service.app_root) / "health_alert_history.json"
        if history_path.exists():
            import json as json_mod
            with open(history_path) as f:
                return json_mod.load(f)
        return {}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/updates/recent-info")
async def get_recent_update_info(current_user: dict = Depends(get_current_user)):
    """
    Récupère les informations des mises à jour récentes (pour le popup utilisateur)
    Disponible pour tous les utilisateurs connectés
    """
    try:
        info = await update_service.get_recent_updates_info(days=3)
        return info
    except Exception as e:
        logger.error(f"❌ Erreur récupération info MAJ récente: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/updates/history-list")
async def get_update_history_list(
    limit: int = 50,
    current_user: dict = Depends(get_current_admin_user)
):
    """
    Récupère l'historique des mises à jour depuis la BDD (admin uniquement)
    Compatible avec le frontend Updates.jsx
    """
    try:
        # Récupérer depuis la collection system_update_history
        history = await db.system_update_history.find(
            {},
            {"_id": 0}
        ).sort("started_at", -1).limit(limit).to_list(limit)
        
        return {"data": history, "total": len(history)}
    except Exception as e:
        logger.error(f"❌ Erreur récupération historique mises à jour: {str(e)}")
        # Retourner une liste vide en cas d'erreur plutôt qu'une exception
        return {"data": [], "total": 0}


@api_router.get("/changelog")
async def get_changelog(current_user: dict = Depends(get_current_user)):
    """Récupère le changelog des mises à jour pour l'utilisateur"""
    try:
        user_id = current_user.get("id")
        
        # Récupérer les entrées de changelog récentes
        entries = await db.system_update_history.find(
            {},
            {"_id": 0}
        ).sort("started_at", -1).limit(20).to_list(20)
        
        # Générer un identifiant unique pour chaque entrée (version ou started_at)
        for entry in entries:
            if not entry.get("version"):
                entry["version"] = entry.get("started_at", "unknown")
        
        # Récupérer les versions lues par cet utilisateur
        user_seen = await db.changelog_seen.find_one({"user_id": user_id}, {"_id": 0})
        seen_versions = set(user_seen.get("versions", [])) if user_seen else set()
        
        for entry in entries:
            entry["seen"] = entry.get("version", "") in seen_versions
        
        return {"entries": entries, "unseen_count": sum(1 for e in entries if not e.get("seen"))}
    except Exception as e:
        logger.error(f"❌ Erreur récupération changelog: {str(e)}")
        return {"entries": [], "unseen_count": 0}


@api_router.post("/changelog/mark-seen")
async def mark_changelog_seen(current_user: dict = Depends(get_current_user)):
    """Marque toutes les entrées du changelog comme lues"""
    try:
        user_id = current_user.get("id")
        
        entries = await db.system_update_history.find({}, {"_id": 0, "version": 1, "started_at": 1}).to_list(None)
        # Utiliser version ou started_at comme identifiant unique
        all_versions = [e.get("version") or e.get("started_at", "") for e in entries if e.get("version") or e.get("started_at")]
        
        await db.changelog_seen.update_one(
            {"user_id": user_id},
            {"$set": {"user_id": user_id, "versions": all_versions, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        
        return {"success": True}
    except Exception as e:
        logger.error(f"❌ Erreur marquage changelog: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/menu-badges")
async def get_menu_badges(current_user: dict = Depends(get_current_user)):
    """Récupère les badges 'Nouveau' pour les menus récemment ajoutés"""
    try:
        user_id = current_user.get("id")
        
        # Récupérer la date de dernière consultation des badges
        user_badge = await db.menu_badge_seen.find_one({"user_id": user_id}, {"_id": 0})
        last_seen = user_badge.get("last_seen_at") if user_badge else None
        
        # Menus ajoutés récemment (depuis le dernier check ou depuis 7 jours)
        new_menu_ids = []
        if not last_seen:
            # Première connexion ou pas encore de données - montrer les menus les plus récents
            new_menu_ids = ["mes", "mes-reports", "analytics-checklists", "service-dashboard", "cameras", "weekly-reports"]
        
        return {"new_menu_ids": new_menu_ids}
    except Exception as e:
        return {"new_menu_ids": []}


@api_router.post("/menu-badges/dismiss")
async def dismiss_menu_badges(current_user: dict = Depends(get_current_user)):
    """Marque les badges de menus comme vus"""
    try:
        user_id = current_user.get("id")
        await db.menu_badge_seen.update_one(
            {"user_id": user_id},
            {"$set": {"user_id": user_id, "last_seen_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        return {"success": True}
    except Exception as e:
        return {"success": False}


# Import surveillance routes
from surveillance_routes import router as surveillance_router, init_surveillance_routes
from realtime_manager import realtime_manager

# Initialize surveillance routes with database, audit service and realtime manager
init_surveillance_routes(db, audit_service, realtime_manager)

# Include surveillance routes
api_router.include_router(surveillance_router)

# Import and initialize AI maintenance routes (checklists + maintenance préventive)
from ai_maintenance_routes import router as ai_maintenance_router, init_ai_maintenance_routes
init_ai_maintenance_routes(db, audit_service)
api_router.include_router(ai_maintenance_router)

from ai_presqu_accident_routes import router as ai_pa_router, init_ai_pa_routes
init_ai_pa_routes(db, audit_service)
api_router.include_router(ai_pa_router)

from ai_work_order_routes import router as ai_wo_router, init_ai_wo_routes
init_ai_wo_routes(db, audit_service)
api_router.include_router(ai_wo_router)

from ai_weekly_report_routes import router as ai_report_router, init_ai_report_routes
init_ai_report_routes(db)
api_router.include_router(ai_report_router)

from ai_sensor_routes import router as ai_sensor_router, init_ai_sensor_routes
init_ai_sensor_routes(db)
api_router.include_router(ai_sensor_router)

from ai_purchase_history_routes import router as ai_purchase_history_router, init_ai_purchase_history_routes
init_ai_purchase_history_routes(db, audit_service)
api_router.include_router(ai_purchase_history_router)

from automation_routes import router as automation_router, init_automation_routes
init_automation_routes(db)
api_router.include_router(automation_router)




# Import presqu'accident routes
from presqu_accident_routes import router as presqu_accident_router, init_presqu_accident_routes

# Initialize presqu'accident routes with database, audit service and realtime manager
init_presqu_accident_routes(db, audit_service, realtime_manager)

# Include presqu'accident routes
api_router.include_router(presqu_accident_router)

# Import documentations routes
from documentations_routes import router as documentations_router, init_documentations_routes
from ssh_routes import router as ssh_router
from user_preferences_routes import router as user_preferences_router
from surveillance_history_routes import router as surveillance_history_router
from tailscale_routes import router as tailscale_router
from autorisation_routes import router as autorisation_router

# Initialize documentations routes with database, audit service and realtime manager
init_documentations_routes(db, audit_service, realtime_manager)

# Include documentations routes
api_router.include_router(documentations_router)
api_router.include_router(ssh_router)
api_router.include_router(user_preferences_router)
api_router.include_router(surveillance_history_router)
api_router.include_router(tailscale_router)
api_router.include_router(autorisation_router)

# Demandes d'arrêt pour maintenance (refactorisé en modules)
from demande_arret_routes import router as demande_arret_router
from demande_arret_reports_routes import router as demande_arret_reports_router
from demande_arret_attachments_routes import router as demande_arret_attachments_router
api_router.include_router(demande_arret_reports_router)  # Routes reports EN PREMIER (avant routes avec {demande_id})
api_router.include_router(demande_arret_attachments_router)  # Routes attachments
api_router.include_router(demande_arret_router)  # Routes principales EN DERNIER

# Import/Export routes (modularisé)
from import_export_routes import router as import_export_router, init_db as init_import_export_db
init_import_export_db(db)
api_router.include_router(import_export_router)

# Backup routes (sauvegardes automatiques)
from backup_routes import router as backup_router, init_db as init_backup_db, set_scheduler as set_backup_scheduler
from backup_service import init_db as init_backup_service_db
init_backup_db(db)
init_backup_service_db(db)
api_router.include_router(backup_router)

# Chat Live
from chat_routes import router as chat_router, init_chat_routes
init_chat_routes(db)
api_router.include_router(chat_router)

# Chat Cleanup Service
from chat_cleanup_service import init_chat_cleanup_service
chat_cleanup_service = init_chat_cleanup_service(db)

# Manuel utilisateur
from manual_routes import router as manual_router
api_router.include_router(manual_router)

# Changelog "Quoi de neuf ?" (releases)
from changelog_routes import router as releases_router
api_router.include_router(releases_router)

# QR Codes équipements
from qr_routes import router as qr_router
api_router.include_router(qr_router)

# QR Codes inventaire
from qr_inventory_routes import router as qr_inventory_router
api_router.include_router(qr_inventory_router)

# Purchase Request routes
from purchase_request_routes import router as purchase_request_router
api_router.include_router(purchase_request_router)

# MQTT routes
from mqtt_routes import router as mqtt_router, init_mqtt_routes
init_mqtt_routes(db)
api_router.include_router(mqtt_router)

# MQTT Manager - pour connexion automatique au démarrage
from mqtt_manager import mqtt_manager

# MQTT Meter Collector
from mqtt_meter_collector import mqtt_meter_collector

# Sensor routes
from sensor_routes import router as sensor_router, init_sensor_routes
init_sensor_routes(db, realtime_manager)
api_router.include_router(sensor_router)

# MQTT Sensor Collector
from mqtt_sensor_collector import mqtt_sensor_collector

# Alert routes and service
from alert_routes import router as alert_router, init_alert_routes
init_alert_routes(db)
api_router.include_router(alert_router)

from alert_service import alert_service

# MQTT Logger
from mqtt_logger import init_mqtt_logger
mqtt_logger = init_mqtt_logger(db)

# MQTT Logs routes
from mqtt_logs_routes import router as mqtt_logs_router, init_mqtt_logs_routes
init_mqtt_logs_routes(db, mqtt_logger)
api_router.include_router(mqtt_logs_router)

# M.E.S (Manufacturing Execution System) routes
from mes_routes import router as mes_router, init_mes_routes, mes_service as _mes_svc_ref
init_mes_routes(db, mqtt_manager)
api_router.include_router(mes_router)

# M.E.S Report Scheduler (envoi automatique des rapports)
from mes_report_scheduler import init_mes_report_scheduler
import email_service as email_service_module

@app.on_event("startup")
async def start_mes_report_scheduler():
    try:
        await init_mes_report_scheduler(db, _mes_svc_ref, email_service_module)
        logger.info("Scheduler rapports M.E.S. demarre")
    except Exception as e:
        logger.warning(f"Erreur demarrage scheduler rapports M.E.S.: {e}")


# AI Chatbot routes
from ai_chat_routes import router as ai_router, init_ai_routes
init_ai_routes(db)
api_router.include_router(ai_router)

# Roles Management routes
from roles_routes import router as roles_router, init_system_roles, init_roles_routes
init_roles_routes(db)
api_router.include_router(roles_router)

# Timezone Configuration routes
from timezone_routes import router as timezone_router, init_timezone_routes
init_timezone_routes(db)
api_router.include_router(timezone_router)

# Consignes routes (notifications MQTT)
from consignes_routes import router as consignes_router, init_consignes_routes, consignes_websocket_endpoint
init_consignes_routes(db, get_current_user, mqtt_manager, audit_service)
api_router.include_router(consignes_router)

# Work Order Templates routes (Ordres Type)
from work_order_templates_routes import router as wo_templates_router
api_router.include_router(wo_templates_router)

# Custom Widgets routes (Widgets personnalisés pour responsables de service)
from custom_widgets_routes import router as custom_widgets_router, init_custom_widgets_routes
init_custom_widgets_routes(db, audit_service)
api_router.include_router(custom_widgets_router)

from ai_widget_routes import router as ai_widget_router, init_ai_widget_routes
init_ai_widget_routes(db)
api_router.include_router(ai_widget_router)

# Service de filtrage par service
from service_filter import init_service_filter
init_service_filter(db)

# Service d'email pour les demandes d'amélioration
from improvement_request_email_service import init_improvement_request_email_service
init_improvement_request_email_service(db)

# Whiteboard (Tableau d'affichage) routes
from whiteboard_routes import router as whiteboard_router, init_whiteboards, init_whiteboard_audit
from whiteboard_object_routes import router as whiteboard_object_router
from whiteboard_manager import whiteboard_manager, handle_whiteboard_message
init_whiteboard_audit(audit_service)  # Initialiser le service d'audit pour le whiteboard
api_router.include_router(whiteboard_router)
api_router.include_router(whiteboard_object_router)  # Nouvelles routes API granulaires

# Routes des rapports hebdomadaires/mensuels/annuels
from weekly_report_routes import router as weekly_report_router, set_database as set_weekly_report_db
set_weekly_report_db(db)
api_router.include_router(weekly_report_router)

# Routes de gestion d'équipe et pointage
from team_management_routes import router as team_router, set_database as set_team_db
set_team_db(db)
api_router.include_router(team_router)

from time_tracking_routes import router as time_tracking_router, set_database as set_time_tracking_db
set_time_tracking_db(db)
api_router.include_router(time_tracking_router)

# Routes de gestion des caméras RTSP/ONVIF
# IMPORTANT: Les routes Frigate doivent être incluses AVANT camera_router car
# celui-ci a des routes dynamiques /{camera_id} qui capturent tout
from frigate_routes import router as frigate_router, set_database as set_frigate_db, init_frigate_from_db
set_frigate_db(db)
api_router.include_router(frigate_router, prefix="/cameras")

from camera_routes import router as camera_router, set_database as set_camera_db
from camera_snapshot_scheduler import set_database as set_camera_scheduler_db, start_snapshot_scheduler
set_camera_db(db)
set_camera_scheduler_db(db)
api_router.include_router(camera_router)

# Initialiser Frigate depuis la DB au démarrage
@app.on_event("startup")
async def init_frigate():
    await init_frigate_from_db()

# Routes Analytics Checklists
from analytics_routes import router as analytics_router, set_database as set_analytics_db
set_analytics_db(db)
api_router.include_router(analytics_router)

# Routes Contrats
from contract_routes import router as contract_router, init_db as init_contract_db
init_contract_db(db, audit_service)
api_router.include_router(contract_router)

# Routes LOTO (Lockout/Tagout - Consignations de sécurité)
from loto_routes import router as loto_router, init_loto_routes
init_loto_routes(db, audit_service)
api_router.include_router(loto_router)

# Push Notifications routes
from notifications import router as push_notifications_router, set_db as set_notifications_db, check_push_receipts
set_notifications_db(db)
api_router.include_router(push_notifications_router)

# Routes Formation (Training)
from training_routes import router as training_router, init_training_routes
init_training_routes(db)
api_router.include_router(training_router)


# WebSocket pour le tableau d'affichage
from fastapi import WebSocket, WebSocketDisconnect

@app.websocket("/api/ws/whiteboard/{board_id}")
async def whiteboard_websocket(websocket: WebSocket, board_id: str):
    """WebSocket pour la synchronisation temps réel du tableau d'affichage"""
    # Récupérer les paramètres de l'utilisateur depuis la query string
    user_id = websocket.query_params.get("user_id", "anonymous")
    user_name = websocket.query_params.get("user_name", "Anonyme")
    
    await whiteboard_manager.connect(websocket, board_id, user_id, user_name)
    
    try:
        # Envoyer l'état initial du tableau
        board = await db.whiteboards.find_one({"board_id": board_id}, {"_id": 0})
        if board:
            await websocket.send_json({
                "type": "sync_response",
                "board": board
            })
        
        # Écouter les messages
        while True:
            data = await websocket.receive_json()
            await handle_whiteboard_message(websocket, board_id, user_id, user_name, data, db)
            
    except WebSocketDisconnect:
        await whiteboard_manager.disconnect(board_id, user_id)
    except Exception as e:
        logger.error(f"Erreur WebSocket whiteboard: {e}")
        await whiteboard_manager.disconnect(board_id, user_id)

# WebSocket Centralisé pour toutes les entités temps réel
from realtime_manager import realtime_manager
from realtime_events import EntityType as RealtimeEntityType

@app.websocket("/api/ws/realtime/{entity_type}")
async def realtime_websocket(websocket: WebSocket, entity_type: str, user_id: str = None):
    """
    WebSocket centralisé pour la synchronisation temps réel de toutes les entités
    
    Args:
        entity_type: Type d'entité (work_orders, equipments, etc.)
        user_id: ID de l'utilisateur connecté
    """
    try:
        logger.info(f"[Realtime] Nouvelle connexion WebSocket demandée: entity_type={entity_type}, user_id={user_id}")
        
        # Valider le type d'entité
        valid_types = [e.value for e in RealtimeEntityType]
        if entity_type not in valid_types:
            logger.warning(f"[Realtime] Type d'entité invalide: {entity_type}. Types valides: {valid_types}")
            await websocket.close(code=1008, reason=f"Invalid entity type: {entity_type}")
            return
        
        # Valider user_id
        if not user_id:
            logger.warning(f"[Realtime] user_id manquant pour {entity_type}")
            await websocket.close(code=1008, reason="user_id is required")
            return
        
        # Accepter la connexion WebSocket
        await websocket.accept()
        logger.info(f"[Realtime] WebSocket accepté: {entity_type}/{user_id}")
        
        # Connecter l'utilisateur au manager (connexion déjà acceptée)
        await realtime_manager.connect(entity_type, user_id, websocket, already_accepted=True)
        logger.info(f"[Realtime] Utilisateur {user_id} connecté au room {entity_type}. Total: {realtime_manager.get_connection_count(entity_type)}")
        
        # Garder la connexion ouverte
        while True:
            # Recevoir les messages du client (pour ping/pong ou autres commandes)
            data = await websocket.receive_json()
            
            # Gérer les commandes spéciales si nécessaire
            if data.get("type") == "ping":
                await realtime_manager.send_to_user(entity_type, user_id, {"type": "pong"})
            
    except WebSocketDisconnect:
        realtime_manager.disconnect(entity_type, user_id)
        logger.info(f"[Realtime] WebSocket déconnecté: {entity_type}/{user_id}")
    except Exception as e:
        logger.error(f"[Realtime] Erreur WebSocket {entity_type}/{user_id}: {e}")
        realtime_manager.disconnect(entity_type, user_id)

# WebSocket pour le Chat Live
from websocket_manager import manager as chat_manager

@app.websocket("/api/ws/chat")
async def chat_live_websocket(websocket: WebSocket, token: str = None, user_id: str = None):
    """WebSocket pour le chat en temps réel"""
    ws_user_id = None
    user_name = "Unknown"
    
    try:
        # Support: user_id direct (préféré pour la compatibilité proxy) ou token JWT
        if user_id:
            ws_user_id = user_id
            user_data = await db.users.find_one({"_id": ObjectId(ws_user_id)})
            if not user_data:
                await websocket.close(code=1008, reason="User not found")
                return
            user_name = f"{user_data.get('prenom', '')} {user_data.get('nom', '')}".strip()
        elif token:
            payload = decode_access_token(token)
            if not payload:
                await websocket.close(code=1008, reason="Invalid token")
                return
            ws_user_id = payload.get("sub")
            if not ws_user_id:
                await websocket.close(code=1008, reason="Invalid token - no user_id")
                return
            user_data = await db.users.find_one({"_id": ObjectId(ws_user_id)})
            if not user_data:
                await websocket.close(code=1008, reason="User not found")
                return
            user_name = f"{user_data.get('prenom', '')} {user_data.get('nom', '')}".strip()
        else:
            await websocket.close(code=1008, reason="user_id or token required")
            return
        
        # Connecter l'utilisateur
        await chat_manager.connect(websocket, ws_user_id, user_name)
        
        # Marquer l'utilisateur comme en ligne
        await db.user_chat_activity.update_one(
            {"user_id": ws_user_id},
            {
                "$set": {
                    "user_id": ws_user_id,
                    "is_online": True,
                    "last_activity": datetime.now(timezone.utc).isoformat()
                }
            },
            upsert=True
        )
        
        try:
            while True:
                # Recevoir les messages du client
                data = await websocket.receive_json()
                message_type = data.get("type")
                
                if message_type == "heartbeat":
                    await db.user_chat_activity.update_one(
                        {"user_id": ws_user_id},
                        {"$set": {"last_activity": datetime.now(timezone.utc).isoformat()}}
                    )
                    await websocket.send_json({"type": "heartbeat_ack"})
                
                elif message_type == "message":
                    message_content = data.get("message", "")
                    recipient_ids = data.get("recipient_ids", [])
                    reply_to_id = data.get("reply_to_id")
                    
                    chat_message = {
                        "id": str(uuid.uuid4()),
                        "user_id": ws_user_id,
                        "user_name": user_name,
                        "user_role": user_data.get("role", ""),
                        "message": message_content,
                        "recipient_ids": recipient_ids,
                        "recipient_names": [],
                        "timestamp": datetime.now(timezone.utc).isoformat(),
                        "is_deleted": False,
                        "deleted_at": None,
                        "reply_to_id": reply_to_id,
                        "reply_to_preview": None,
                        "reactions": [],
                        "attachments": [],
                        "deletable_until": (datetime.now(timezone.utc) + timedelta(seconds=10)).isoformat(),
                        "is_private": len(recipient_ids) > 0
                    }
                    
                    if reply_to_id:
                        original_msg = await db.chat_messages.find_one({"id": reply_to_id})
                        if original_msg:
                            chat_message["reply_to_preview"] = original_msg.get("message", "")[:100]
                    
                    if recipient_ids:
                        recipient_object_ids = [ObjectId(rid) for rid in recipient_ids if ObjectId.is_valid(rid)]
                        recipients = await db.users.find({"_id": {"$in": recipient_object_ids}}).to_list(length=None)
                        chat_message["recipient_names"] = [
                            f"{r.get('prenom', '')} {r.get('nom', '')}".strip()
                            for r in recipients
                        ]
                    
                    await db.chat_messages.insert_one(chat_message)
                    
                    broadcast_data = {
                        "type": "new_message",
                        "message": {k: v for k, v in chat_message.items() if k != "_id"}
                    }
                    
                    if recipient_ids:
                        await chat_manager.send_to_users(broadcast_data, recipient_ids + [ws_user_id])
                    else:
                        await chat_manager.broadcast(broadcast_data)
                
                elif message_type == "typing":
                    await chat_manager.broadcast({
                        "type": "user_typing",
                        "user_id": ws_user_id,
                        "user_name": user_name
                    }, exclude_user_id=ws_user_id)
        
        except WebSocketDisconnect:
            logger.info(f"Chat WebSocket déconnecté: {user_name}")
    
    except Exception as e:
        logger.error(f"Erreur Chat WebSocket: {e}")
    
    finally:
        if ws_user_id:
            chat_manager.disconnect(ws_user_id, user_name, websocket=websocket)
            # Vérifier s'il reste des connexions pour cet utilisateur
            if not chat_manager.is_user_online(ws_user_id):
                await db.user_chat_activity.update_one(
                    {"user_id": ws_user_id},
                    {"$set": {"is_online": False, "last_activity": datetime.now(timezone.utc).isoformat()}}
                )
                await chat_manager.broadcast_user_status(ws_user_id, user_name, "offline")

# WebSocket pour les consignes (notifications temps réel)
@app.websocket("/api/ws/consignes")
async def consignes_websocket(websocket: WebSocket, token: str = None, user_id: str = None):
    """WebSocket pour recevoir les consignes en temps réel"""
    if user_id:
        # Connexion par user_id (compatible proxy)
        try:
            user_data = await db.users.find_one({"_id": ObjectId(user_id)})
            if not user_data:
                await websocket.close(code=1008, reason="User not found")
                return
            await websocket.accept()
            try:
                while True:
                    data = await websocket.receive_text()
            except WebSocketDisconnect:
                pass
        except Exception as e:
            logger.error(f"Erreur consignes WS: {e}")
    elif token:
        await consignes_websocket_endpoint(websocket, token)
    else:
        await websocket.close(code=1008, reason="user_id or token required")


# ==================== ADMIN RESET ROUTES ====================

RESET_COLLECTIONS = {
    "work_orders": "Ordres de travail",
    "intervention_requests": "Demandes d'intervention",
    "improvement_requests": "Demandes d'amélioration",
    "improvements": "Améliorations",
    "equipments": "Équipements",
    "inventory": "Inventaire",
    "locations": "Zones / Emplacements",
    "preventive_maintenance": "Maintenance préventive",
    "vendors": "Fournisseurs",
    "purchase_history": "Historique d'achat",
    "purchase_requests": "Demandes d'achat",
    "sensors": "Capteurs MQTT",
    "chat_messages": "Messages Chat Live",
    "users": "Utilisateurs",
    "surveillance_items": "Plan de surveillance",
    "presqu_accident_items": "Presqu'accidents",
}

@api_router.delete("/admin/reset/{section}",
    summary="Reinitialiser une section", response_model=ResetSectionResponse, tags=["Administration"])
async def reset_section(section: str, current_user: dict = Depends(get_current_admin_user)):
    """Réinitialiser une section (admin uniquement)"""
    if section not in RESET_COLLECTIONS:
        raise HTTPException(status_code=400, detail=f"Section inconnue: {section}")
    
    query = {}
    if section == "users":
        query = {"_id": {"$ne": ObjectId(current_user["id"])}}
    
    result = await db[section].delete_many(query)
    
    # Log d'audit via le service centralisé
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip(),
        user_email=current_user["email"],
        action=ActionType.DELETE,
        entity_type=EntityType.USER,
        entity_name=RESET_COLLECTIONS[section],
        details=f"Réinitialisation de {RESET_COLLECTIONS[section]}: {result.deleted_count} éléments supprimés"
    )
    
    return {
        "success": True,
        "section": RESET_COLLECTIONS[section],
        "deleted_count": result.deleted_count
    }

@api_router.delete("/admin/reset-all",
    summary="Reinitialiser toutes les donnees", response_model=ResetAllResponse, tags=["Administration"])
async def reset_all(current_user: dict = Depends(get_current_admin_user)):
    """Réinitialiser toutes les données (admin uniquement)"""
    details = {}
    total = 0
    
    for section, label in RESET_COLLECTIONS.items():
        query = {}
        if section == "users":
            query = {"_id": {"$ne": ObjectId(current_user["id"])}}
        
        result = await db[section].delete_many(query)
        details[label] = result.deleted_count
        total += result.deleted_count
    
    # Log d'audit via le service centralisé
    await audit_service.log_action(
        user_id=current_user["id"],
        user_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip(),
        user_email=current_user["email"],
        action=ActionType.DELETE,
        entity_type=EntityType.USER,
        entity_name="Toutes les données",
        details=f"Réinitialisation complète: {total} éléments supprimés"
    )
    
    return {
        "success": True,
        "total_deleted": total,
        "details": details
    }


# Include the router in the main app (MUST be after all endpoint definitions)
app.include_router(api_router)

# Fonction de vérification des versions LLM (appelée par le scheduler)
async def check_llm_versions_job():
    """Vérifie les nouvelles versions des modèles LLM et notifie les admins"""
    try:
        from datetime import datetime, timezone, timedelta
        logger.info("🔍 Vérification automatique des versions LLM...")
        
        # Mettre à jour la date de dernière vérification
        now = datetime.now(timezone.utc)
        next_monday = now + timedelta(days=7)
        
        await db.llm_versions.update_one(
            {"id": "current"},
            {"$set": {
                "last_check": now.isoformat(),
                "next_check": next_monday.isoformat(),
                "checked_by": "scheduler"
            }},
            upsert=True
        )
        
        logger.info("✅ Vérification des versions LLM terminée")
        
    except Exception as e:
        logger.error(f"❌ Erreur vérification versions LLM: {e}")


@app.on_event("startup")
async def fix_surveillance_ecart_data():
    """Correction de données : reset ecart_jours pour les items non réalisés."""
    try:
        result = await db.surveillance_items.update_many(
            {"status": {"$in": ["PLANIFIER", "PLANIFIE"]}, "ecart_jours": {"$ne": None}},
            {"$set": {"ecart_jours": None}}
        )
        if result.modified_count > 0:
            logger.info(f"🔧 Correction données: ecart_jours réinitialisé pour {result.modified_count} item(s) non réalisé(s)")
    except Exception as e:
        logger.warning(f"Correction ecart_jours ignorée: {e}")


@app.on_event("startup")
async def create_notification_indexes():
    """Create MongoDB indexes for push notifications collections."""
    try:
        # Supprimer l'ancien index unique s'il existe (cause des erreurs 500)
        try:
            await db.device_tokens.drop_index("push_token_1")
        except Exception:
            pass
        await db.device_tokens.create_index([("user_id", 1), ("is_active", 1)])
        await db.device_tokens.create_index([("push_token", 1)])
        await db.push_receipts.create_index([("checked", 1), ("created_at", 1)])
        await db.push_receipts.create_index([("ticket_id", 1)])
        logger.info("Indexes device_tokens et push_receipts crees avec succes")
    except Exception as e:
        logger.warning(f"Erreur creation indexes: {e}")



@app.on_event("startup")
async def check_update_results_on_startup():
    """Vérifie s'il existe des résultats de mise à jour non traités au démarrage."""
    try:
        await update_service.check_and_save_update_result()
    except Exception as e:
        logger.error(f"[MAJ] Erreur vérification résultats MAJ au démarrage: {e}")


@app.on_event("startup")
async def migrate_inventory_services():
    """Migration: créer le service 'Non classé' et assigner les articles existants sans service_id."""
    try:
        # Créer 'Non classé' s'il n'existe pas
        non_classe = await db.inventory_services.find_one({"name": "Non classé"})
        if not non_classe:
            nc_doc = {
                "id": str(uuid.uuid4()),
                "name": "Non classé",
                "created_by": "system",
                "created_by_name": "Système",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.inventory_services.insert_one(nc_doc)
            nc_id = nc_doc["id"]
            logger.info("[Inventaire] Service 'Non classé' créé")
        else:
            nc_id = non_classe["id"]
        
        # Migrer les articles sans service_id
        result = await db.inventory.update_many(
            {"$or": [{"service_id": None}, {"service_id": {"$exists": False}}]},
            {"$set": {"service_id": nc_id, "shared_service_ids": []}}
        )
        if result.modified_count > 0:
            logger.info(f"[Inventaire] {result.modified_count} articles migrés vers 'Non classé'")
    except Exception as e:
        logger.error(f"[Inventaire] Erreur migration services: {e}")



@app.on_event("startup")
async def startup_scheduler():
    """Démarre le scheduler au démarrage de l'application"""
    try:
        # Configurer le scheduler pour s'exécuter chaque jour à minuit (heure locale)
        scheduler.add_job(
            auto_check_preventive_maintenance,
            CronTrigger(hour=0, minute=0),  # Tous les jours à minuit
            id='check_preventive_maintenance',
            name='Vérification automatique maintenances préventives',
            replace_existing=True
        )
        
        # Configurer la vérification automatique des mises à jour à 1h00 du matin
        scheduler.add_job(
            update_service.check_for_updates,
            CronTrigger(hour=1, minute=0),  # Tous les jours à 1h00
            id='check_updates',
            name='Vérification automatique des mises à jour',
            replace_existing=True
        )
        
        # Configurer la vérification des demandes d'arrêt expirées à 2h00 du matin
        from demande_arret_routes import check_expired_demandes_cron
        scheduler.add_job(
            check_expired_demandes_cron,
            CronTrigger(hour=2, minute=0),  # Tous les jours à 2h00
            id='check_expired_demandes',
            name='Vérification demandes arrêt expirées (7 jours)',
            replace_existing=True
        )
        
        # Configurer la gestion des maintenances planifiées à 0h05 et 12h00
        scheduler.add_job(
            manage_planned_maintenance_status,
            CronTrigger(hour=0, minute=5),  # Tous les jours à 0h05
            id='manage_planned_maintenance_morning',
            name='Gestion maintenances planifiées (matin)',
            replace_existing=True
        )
        scheduler.add_job(
            manage_planned_maintenance_status,
            CronTrigger(hour=12, minute=0),  # Tous les jours à 12h00
            id='manage_planned_maintenance_noon',
            name='Gestion maintenances planifiées (midi)',
            replace_existing=True
        )
        
        # Configurer le nettoyage automatique des messages du chat à 3h00 du matin
        scheduler.add_job(
            chat_cleanup_service.cleanup_old_messages,
            CronTrigger(hour=3, minute=0),  # Tous les jours à 3h00
            id='chat_cleanup',
            name='Nettoyage messages chat > 60 jours',
            replace_existing=True
        )
        
        # Configurer la vérification des versions LLM chaque lundi à 3h00 GMT
        scheduler.add_job(
            check_llm_versions_job,
            CronTrigger(day_of_week='mon', hour=3, minute=0),  # Chaque lundi à 3h00
            id='llm_version_check',
            name='Vérification versions LLM',
            replace_existing=True
        )
        
        # Configurer la vérification des notifications PM tous les jours à 7h00 GMT
        scheduler.add_job(
            check_pm_notifications,
            CronTrigger(hour=7, minute=0),  # Tous les jours à 7h00
            id='pm_notifications_check',
            name='Vérification notifications PM',
            replace_existing=True
        )
        
        # Configurer la vérification des rappels de surveillance tous les jours à 7h30 GMT
        from surveillance_routes import check_surveillance_reminders
        scheduler.add_job(
            check_surveillance_reminders,
            CronTrigger(hour=7, minute=30),  # Tous les jours à 7h30
            id='surveillance_reminders_check',
            name='Rappels surveillance',
            replace_existing=True
        )
        
        # Configurer la vérification des alertes de contrats tous les jours à 8h00 GMT
        from contract_routes import check_contract_alerts
        scheduler.add_job(
            check_contract_alerts,
            CronTrigger(hour=8, minute=0),  # Tous les jours à 8h00
            id='contract_alerts_check',
            name='Alertes contrats',
            replace_existing=True
        )
        
        # Configurer le nettoyage des push tokens invalides toutes les 20 minutes
        scheduler.add_job(
            check_push_receipts,
            CronTrigger(minute='*/20'),  # Toutes les 20 minutes
            id='push_receipts_check',
            name='Verification recus push et nettoyage tokens invalides',
            replace_existing=True
        )
        
        scheduler.start()
        logger.info("✅ Scheduler démarré:")
        logger.info("   - Vérification maintenances préventives: tous les jours à 00h00")
        logger.info("   - Gestion maintenances planifiées: tous les jours à 00h05 et 12h00")
        logger.info("   - Vérification mises à jour: tous les jours à 01h00")
        logger.info("   - Vérification demandes expirées: tous les jours à 02h00")
        logger.info("   - Nettoyage messages chat (60j): tous les jours à 03h00")
        logger.info("   - Vérification versions LLM: chaque lundi à 03h00")
        logger.info("   - Notifications PM: tous les jours à 07h00")
        logger.info("   - Rappels surveillance: tous les jours à 07h30")
        logger.info("   - Alertes contrats: tous les jours à 08h00")
        logger.info("   - Nettoyage push tokens invalides: toutes les 20 min")

        # Charger les planifications de backup automatique
        try:
            set_backup_scheduler(scheduler)
            from backup_routes import _reload_scheduler as reload_backup_jobs
            await reload_backup_jobs()
            logger.info("   - Sauvegardes automatiques: planifications chargées")
        except Exception as e:
            logger.warning(f"   - Sauvegardes automatiques: erreur chargement ({e})")
        
        # M.E.S - Calcul cadence chaque minute (abonnement MQTT sera fait APRÈS connexion)
        from mes_routes import mes_service as _mes_ref
        if _mes_ref:
            from apscheduler.triggers.interval import IntervalTrigger
            scheduler.add_job(
                _mes_ref.calculate_minute_cadence,
                IntervalTrigger(minutes=1),
                id='mes_cadence_calc',
                name='M.E.S - Calcul cadence par minute',
                replace_existing=True
            )
            scheduler.add_job(
                _mes_ref.cleanup_old_data,
                CronTrigger(hour=4, minute=0),
                id='mes_cleanup',
                name='M.E.S - Nettoyage données > 1 an',
                replace_existing=True
            )
            logger.info("   - M.E.S cadence: chaque minute")
        
        # Auto-connexion MQTT si configuré (DOIT être fait AVANT les abonnements M.E.S.)
        mqtt_connected = False
        try:
            mqtt_config = await db.mqtt_config.find_one({"id": "default"})
            if mqtt_config and mqtt_config.get("host"):
                logger.info(f"🔌 Configuration MQTT trouvée: {mqtt_config.get('host')}:{mqtt_config.get('port', 1883)}")
                mqtt_manager.set_database(db)
                mqtt_manager.configure(
                    host=mqtt_config["host"],
                    port=mqtt_config.get("port", 1883),
                    username=mqtt_config.get("username"),
                    password=mqtt_config.get("password"),
                    use_ssl=mqtt_config.get("use_ssl", False),
                    client_id=mqtt_config.get("client_id", "gmao_iris")
                )
                success = mqtt_manager.connect()
                if success:
                    # Attendre que la connexion soit vraiment établie
                    import time
                    max_wait = 5  # secondes
                    waited = 0
                    while not mqtt_manager.is_connected and waited < max_wait:
                        await asyncio.sleep(0.1)
                        waited += 0.1
                    mqtt_connected = mqtt_manager.is_connected
                    if mqtt_connected:
                        logger.info("✅ Connexion MQTT automatique établie")
                    else:
                        logger.warning("⚠️ Connexion MQTT initiée mais pas encore établie")
                else:
                    logger.warning("⚠️ Échec de la connexion MQTT automatique")
            else:
                logger.info("ℹ️ Aucune configuration MQTT trouvée, connexion manuelle requise")
        except Exception as mqtt_err:
            logger.error(f"❌ Erreur lors de l'auto-connexion MQTT: {mqtt_err}")
        
        # M.E.S - Abonnement MQTT (APRÈS connexion MQTT)
        if _mes_ref:
            await _mes_ref.subscribe_all()
            if mqtt_connected:
                logger.info("✅ M.E.S. abonné aux topics MQTT")
            else:
                logger.info("ℹ️ M.E.S. topics en attente (MQTT non connecté)")
        
        # Initialiser et démarrer les collecteurs MQTT
        await mqtt_meter_collector.initialize(db)
        await mqtt_meter_collector.start()
        logger.info("✅ Collecteur MQTT compteurs démarré")
        
        await mqtt_sensor_collector.initialize(db)
        await mqtt_sensor_collector.start()
        logger.info("✅ Collecteur MQTT capteurs démarré")
        
        # Initialiser le service d'alertes
        await alert_service.initialize(db)
        logger.info("✅ Service d'alertes initialisé")
        
        # Initialiser les rôles système
        await init_system_roles()
        logger.info("✅ Rôles système initialisés")
        
        # Migrer les permissions des rôles (ajouter les modules manquants)
        try:
            from roles_routes import get_default_permissions_by_role
            NEW_PERMISSION_KEYS = ["mes", "mesReports", "serviceDashboard", "weeklyReports", "demandesArret", "consignes", "autorisationsParticulieres", "timeTracking", "cameras", "analyticsChecklists"]
            roles = await db.roles.find({}).to_list(length=None)
            perm_updated = 0
            for role in roles:
                perms = role.get("permissions", {})
                if not isinstance(perms, dict):
                    perms = {}
                needs_update = False
                for key in NEW_PERMISSION_KEYS:
                    if key not in perms:
                        needs_update = True
                        default_perms = get_default_permissions_by_role(role.get("code", ""))
                        default_dict = default_perms.model_dump()
                        perms[key] = default_dict.get(key, {"view": False, "edit": False, "delete": False})
                if needs_update:
                    await db.roles.update_one({"id": role["id"]}, {"$set": {"permissions": perms}})
                    perm_updated += 1
            if perm_updated > 0:
                logger.info(f"✅ Permissions migrées pour {perm_updated} rôle(s)")
        except Exception as role_mig_err:
            import traceback
            logger.error(f"❌ Erreur migration permissions rôles: {role_mig_err}\n{traceback.format_exc()}")
        
        # Migrer les permissions de TOUS les utilisateurs existants
        try:
            all_users = await db.users.find({}).to_list(length=None)
            users_perm_updated = 0
            for u in all_users:
                user_role = u.get("role", "VISUALISEUR")
                current_perms = u.get("permissions", {})
                if not isinstance(current_perms, dict):
                    current_perms = {}
                default_perms = get_default_permissions_by_role(user_role).model_dump()
                needs_update = False
                valid_module_keys = set(default_perms.keys())
                keys_to_remove = [k for k in current_perms if k not in valid_module_keys]
                if keys_to_remove:
                    for k in keys_to_remove:
                        del current_perms[k]
                    needs_update = True
                for module_key, module_val in default_perms.items():
                    if module_key not in current_perms:
                        current_perms[module_key] = module_val
                        needs_update = True
                    elif not isinstance(current_perms[module_key], dict):
                        current_perms[module_key] = module_val
                        needs_update = True
                    else:
                        valid_keys = {"view", "edit", "delete"}
                        existing = current_perms[module_key]
                        has_extra = any(k not in valid_keys for k in existing.keys())
                        if has_extra:
                            current_perms[module_key] = {
                                "view": existing.get("view", module_val.get("view", False)),
                                "edit": existing.get("edit", module_val.get("edit", False)),
                                "delete": existing.get("delete", module_val.get("delete", False))
                            }
                            needs_update = True
                if needs_update:
                    await db.users.update_one({"_id": u["_id"]}, {"$set": {"permissions": current_perms}})
                    users_perm_updated += 1
            if users_perm_updated > 0:
                logger.info(f"✅ Permissions utilisateurs migrées pour {users_perm_updated} utilisateur(s)")
        except Exception as user_mig_err:
            import traceback
            logger.error(f"❌ Erreur migration permissions utilisateurs: {user_mig_err}\n{traceback.format_exc()}")
        
        # Migrer le manuel utilisateur (ajouter chapitres manquants)
        import json as json_lib
        manual_json_path = os.path.join(os.path.dirname(__file__), "manual_default_content.json")
        if os.path.exists(manual_json_path):
            with open(manual_json_path, "r", encoding="utf-8") as f:
                manual_data = json_lib.load(f)
            now_utc = datetime.now(timezone.utc)
            manual_ch_added = 0
            manual_sec_added = 0
            for chapter in manual_data.get("chapters", []):
                if "id" not in chapter:
                    continue
                existing = await db.manual_chapters.find_one({"id": chapter["id"]})
                if not existing:
                    chapter.setdefault("created_at", now_utc.isoformat())
                    chapter.setdefault("updated_at", now_utc.isoformat())
                    await db.manual_chapters.insert_one(chapter)
                    manual_ch_added += 1
                else:
                    json_sections = chapter.get("sections", [])
                    db_sections = existing.get("sections", [])
                    new_secs = [s for s in json_sections if s not in db_sections]
                    if new_secs:
                        await db.manual_chapters.update_one(
                            {"id": chapter["id"]},
                            {"$addToSet": {"sections": {"$each": new_secs}}}
                        )
                        manual_ch_added += 1
            for section in manual_data.get("sections", []):
                if "id" not in section:
                    continue
                existing = await db.manual_sections.find_one({"id": section["id"]})
                if not existing:
                    section.setdefault("created_at", now_utc.isoformat())
                    section.setdefault("updated_at", now_utc.isoformat())
                    await db.manual_sections.insert_one(section)
                    manual_sec_added += 1
            if manual_ch_added > 0 or manual_sec_added > 0:
                existing_version = await db.manual_versions.find_one({"is_current": True})
                if not existing_version:
                    await db.manual_versions.insert_one({
                        "id": f"init-{now_utc.strftime('%Y%m%d%H%M%S')}",
                        "version": "2.3",
                        "release_date": now_utc.isoformat(),
                        "changes": ["Initialisation complete du manuel utilisateur"],
                        "author_id": "system",
                        "author_name": "Initialisation systeme",
                        "is_current": True
                    })
                logger.info(f"✅ Manuel utilisateur: {manual_ch_added} chapitre(s) et {manual_sec_added} section(s) ajouté(s)")
        
        # Migrer les menus utilisateurs (ajouter menus manquants)
        complete_menu_ref = {
            "dashboard": {"label": "Tableau de bord", "path": "/dashboard", "icon": "LayoutDashboard", "module": "dashboard"},
            "service-dashboard": {"label": "Dashboard Service", "path": "/service-dashboard", "icon": "Presentation", "module": "serviceDashboard"},
            "chat-live": {"label": "Chat Live", "path": "/chat-live", "icon": "Mail", "module": "chatLive"},
            "intervention-requests": {"label": "Demandes d'inter.", "path": "/intervention-requests", "icon": "MessageSquare", "module": "interventionRequests"},
            "work-orders": {"label": "Ordres de travail", "path": "/work-orders", "icon": "ClipboardList", "module": "workOrders"},
            "improvement-requests": {"label": "Demandes d'amél.", "path": "/improvement-requests", "icon": "Lightbulb", "module": "improvementRequests"},
            "improvements": {"label": "Améliorations", "path": "/improvements", "icon": "Sparkles", "module": "improvements"},
            "preventive-maintenance": {"label": "Maintenance prev.", "path": "/preventive-maintenance", "icon": "Calendar", "module": "preventiveMaintenance"},
            "planning-mprev": {"label": "Planning M.Prev.", "path": "/planning-mprev", "icon": "Calendar", "module": "planningMprev"},
            "assets": {"label": "Équipements", "path": "/assets", "icon": "Wrench", "module": "assets"},
            "inventory": {"label": "Inventaire", "path": "/inventory", "icon": "Package", "module": "inventory"},
            "purchase-requests": {"label": "Demandes d'Achat", "path": "/purchase-requests", "icon": "ShoppingCart", "module": "purchaseRequests"},
            "locations": {"label": "Zones", "path": "/locations", "icon": "MapPin", "module": "locations"},
            "meters": {"label": "Compteurs", "path": "/meters", "icon": "Gauge", "module": "meters"},
            "surveillance-plan": {"label": "Plan de Surveillance", "path": "/surveillance-plan", "icon": "Eye", "module": "surveillance"},
            "surveillance-rapport": {"label": "Rapport Surveillance", "path": "/surveillance-rapport", "icon": "FileText", "module": "surveillanceRapport"},
            "weekly-reports": {"label": "Rapports Hebdo.", "path": "/weekly-reports", "icon": "FileText", "module": "weeklyReports"},
            "presqu-accident": {"label": "Presqu'accident", "path": "/presqu-accident", "icon": "AlertTriangle", "module": "presquaccident"},
            "presqu-accident-rapport": {"label": "Rapport P.accident", "path": "/presqu-accident-rapport", "icon": "FileText", "module": "presquaccidentRapport"},
            "documentations": {"label": "Documentations", "path": "/documentations", "icon": "FolderOpen", "module": "documentations"},
            "reports": {"label": "Rapports", "path": "/reports", "icon": "BarChart3", "module": "reports"},
            "team-management": {"label": "Gestion d'équipe", "path": "/team-management", "icon": "UserCog", "module": "timeTracking"},
            "cameras": {"label": "Caméras", "path": "/cameras", "icon": "Camera", "module": "cameras"},
            "mes": {"label": "M.E.S.", "path": "/mes", "icon": "Zap", "module": "mes"},
            "mes-reports": {"label": "Rapports M.E.S.", "path": "/mes-reports", "icon": "FileBarChart", "module": "mesReports"},
            "analytics-checklists": {"label": "Analytics Checklists", "path": "/analytics/checklists", "icon": "BarChart3", "module": "analyticsChecklists"},
            "people": {"label": "Utilisateurs", "path": "/people", "icon": "Users", "module": "people"},
            "planning": {"label": "Planning", "path": "/planning", "icon": "Calendar", "module": "planning"},
            "vendors": {"label": "Fournisseurs", "path": "/vendors", "icon": "ShoppingCart", "module": "vendors"},
            "contrats": {"label": "Contrats", "path": "/contrats", "icon": "FileSignature", "module": "contrats"},
            "purchase-history": {"label": "Historique Achat", "path": "/purchase-history", "icon": "ShoppingBag", "module": "purchaseHistory"},
            "import-export": {"label": "Import / Export", "path": "/import-export", "icon": "Database", "module": "importExport"},
            "sensors": {"label": "Capteurs MQTT", "path": "/sensors", "icon": "Activity", "module": "sensors"},
            "iot-dashboard": {"label": "Dashboard IoT", "path": "/iot-dashboard", "icon": "BarChart3", "module": "iotDashboard"},
            "mqtt-logs": {"label": "Logs MQTT", "path": "/mqtt-logs", "icon": "Terminal", "module": "mqttLogs"},
            "whiteboard": {"label": "Tableau d'affichage", "path": "/whiteboard", "icon": "Presentation", "module": "whiteboard"},
        }
        user_prefs = await db.user_preferences.find({}).to_list(length=None)
        menus_migrated = 0
        for pref in user_prefs:
            menu_items = pref.get("menu_items", [])
            # Filtrer les entrées non-dict (corruption de données)
            menu_items = [item for item in menu_items if isinstance(item, dict)]
            existing_ids = {item.get("id") for item in menu_items}
            needs_update = False
            # Fix existing items that are missing required fields
            for item in menu_items:
                ref = complete_menu_ref.get(item.get("id"))
                if ref:
                    for field in ["label", "path", "icon", "module"]:
                        if field not in item or not item[field]:
                            item[field] = ref[field]
                            needs_update = True
            # Add missing menu items
            max_order = max((item.get("order", 0) for item in menu_items), default=0)
            for i, (mid, ref) in enumerate(complete_menu_ref.items()):
                if mid not in existing_ids:
                    menu_items.append({"id": mid, "label": ref["label"], "path": ref["path"], "icon": ref["icon"], "module": ref["module"], "visible": True, "favorite": False, "order": max_order + 1 + i, "category_id": None})
                    needs_update = True
            if needs_update:
                await db.user_preferences.update_one({"_id": pref["_id"]}, {"$set": {"menu_items": menu_items}})
                menus_migrated += 1
        if menus_migrated > 0:
            logger.info(f"✅ Menus migrés pour {menus_migrated} utilisateur(s)")
        
        # Initialiser le scheduler des rapports hebdomadaires
        from weekly_report_scheduler import init_report_scheduler, load_all_report_schedules
        init_report_scheduler(scheduler, db)
        scheduled_count = await load_all_report_schedules(db)
        logger.info(f"✅ Scheduler rapports initialisé ({scheduled_count} rapport(s) planifié(s))")
        
        # Initialiser le service d'alertes caméras
        from camera_alert_service import set_database as set_camera_alert_db, start_camera_alert_scheduler
        set_camera_alert_db(db)
        asyncio.create_task(start_camera_alert_scheduler(interval_seconds=60))
        logger.info("✅ Service d'alertes caméras démarré (vérification toutes les 60s)")
        
    except Exception as e:
        import traceback
        logger.error(f"❌ Erreur lors du démarrage du scheduler: {str(e)}\n{traceback.format_exc()}")

@app.on_event("shutdown")
async def shutdown_services():
    """Arrête les services lors de l'arrêt de l'application"""
    try:
        scheduler.shutdown()
        logger.info("✅ Scheduler arrêté")
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'arrêt du scheduler: {str(e)}")
    
