"""
Routes API pour les Presqu'accidents (Near Miss)
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from pathlib import Path
import uuid
import logging

from models import (
    PresquAccidentItem,
    PresquAccidentItemCreate,
    PresquAccidentItemUpdate,
    PresquAccidentStatus,
    PresquAccidentService,
    PresquAccidentSeverity,
    ActionType,
    EntityType,
    SuccessResponse
)
from dependencies import get_current_user, get_current_admin_user
from audit_service import AuditService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/presqu-accident", tags=["presqu-accident"])

# Variables globales (seront injectées depuis server.py)
db = None
audit_service = None
realtime_manager = None

def init_presqu_accident_routes(database, audit_svc, realtime_mgr=None):
    """Initialise les routes avec la connexion DB, audit service et realtime manager"""
    global db, audit_service, realtime_manager
    db = database
    audit_service = audit_svc
    realtime_manager = realtime_mgr


# ==================== CRUD Routes ====================

@router.get("/items", response_model=List[dict])
async def get_presqu_accident_items(
    service: Optional[str] = None,
    status: Optional[str] = None,
    severite: Optional[str] = None,
    lieu: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer tous les presqu'accidents avec filtres et filtrage par service"""
    from service_filter import apply_service_filter
    
    try:
        query = {}
        
        if service:
            query["service"] = service
        if status:
            query["status"] = status
        if severite:
            query["severite"] = severite
        if lieu:
            query["lieu"] = {"$regex": lieu, "$options": "i"}
        
        # Appliquer le filtre par service automatique si pas de filtre service explicite
        if not service:
            query = await apply_service_filter(query, current_user, "service")
        
        items = await db.presqu_accident_items.find(query).to_list(length=None)
        
        # Convertir _id en string
        for item in items:
            if "_id" in item:
                del item["_id"]
        
        return items
    except Exception as e:
        logger.error(f"Erreur récupération presqu'accidents: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/items/{item_id}")
async def get_presqu_accident_item(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer un presqu'accident spécifique"""
    try:
        item = await db.presqu_accident_items.find_one({"id": item_id})
        
        if not item:
            raise HTTPException(status_code=404, detail="Presqu'accident non trouvé")
        
        if "_id" in item:
            del item["_id"]
        
        return item
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur récupération presqu'accident {item_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/items")
async def create_presqu_accident_item(
    item_data: PresquAccidentItemCreate,
    current_user: dict = Depends(get_current_user)
):
    """Créer un nouveau presqu'accident"""
    try:
        # Générer le numéro au format [année]-[numéro incrémenté]
        current_year = datetime.now().year
        
        # Compter les presqu'accidents de l'année en cours
        count = await db.presqu_accident_items.count_documents({
            "numero": {"$regex": f"^{current_year}-"}
        })
        
        # Générer le numéro (commence à 001)
        numero = f"{current_year}-{str(count + 1).zfill(3)}"
        
        item = PresquAccidentItem(
            **item_data.model_dump(),
            numero=numero,
            status=PresquAccidentStatus.A_TRAITER,  # Statut initial obligatoire
            created_by=current_user.get("id"),
            updated_by=current_user.get("id")
        )
        
        item_dict = item.model_dump()
        await db.presqu_accident_items.insert_one(item_dict)
        
        # Audit
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.CREATE,
            entity_type=EntityType.PRESQU_ACCIDENT,
            entity_id=item.id,
            entity_name=f"Presqu'accident: {item.titre}"
        )
        
        if "_id" in item_dict:
            del item_dict["_id"]
        
        # Broadcast WebSocket pour la synchronisation temps réel
        if realtime_manager:
            await realtime_manager.emit_event(
                "near_miss",
                "created",
                item_dict,
                user_id=current_user["id"]
            )
        
        return item_dict
    except Exception as e:
        logger.error(f"Erreur création presqu'accident: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



@router.post("/import-bulk")
async def import_bulk_presqu_accidents(
    items_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Import en masse de presqu'accidents depuis l'extraction IA.
    Accepte TOUS les champs y compris status, commentaire_traitement, responsable_action.
    """
    try:
        items_list = items_data.get("items", [])
        if not items_list:
            raise HTTPException(status_code=400, detail="Aucun item a importer")

        current_year = datetime.now().year
        count = await db.presqu_accident_items.count_documents({
            "numero": {"$regex": f"^{current_year}-"}
        })

        # Mapping statut: valider contre l'enum
        valid_statuses = {s.value for s in PresquAccidentStatus}
        valid_services = {s.value for s in PresquAccidentService}

        created = 0
        errors_list = []

        for idx, raw in enumerate(items_list):
            try:
                count += 1
                numero = f"{current_year}-{str(count).zfill(3)}"

                # Valider le service
                service = raw.get("service", "AUTRE")
                if service not in valid_services:
                    service = "AUTRE"

                # Valider le statut (ou defaut A_TRAITER)
                status_val = raw.get("status", "A_TRAITER")
                if status_val not in valid_statuses:
                    status_val = "A_TRAITER"

                # Construire l'item complet avec TOUS les champs
                item = PresquAccidentItem(
                    numero=numero,
                    titre=raw.get("titre", "Sans titre")[:200],
                    description=raw.get("description") or "-",
                    date_incident=raw.get("date_incident") or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    lieu=raw.get("lieu") or "-",
                    service=PresquAccidentService(service),
                    categorie_incident=raw.get("categorie_incident") or None,
                    declarant=raw.get("declarant") or None,
                    personnes_impliquees=raw.get("personnes_impliquees") or None,
                    mesures_immediates=raw.get("mesures_immediates") or None,
                    actions_proposees=raw.get("actions_proposees") or None,
                    contexte_cause=raw.get("contexte_cause") or None,
                    conditions_incident=raw.get("conditions_incident") or None,
                    commentaire_traitement=raw.get("commentaire_traitement") or None,
                    responsable_action=raw.get("responsable_action") or None,
                    status=PresquAccidentStatus(status_val),
                    severite=PresquAccidentSeverity(raw.get("severite", "MOYEN")),
                    created_by=current_user.get("id"),
                    updated_by=current_user.get("id")
                )

                item_dict = item.model_dump()
                await db.presqu_accident_items.insert_one(item_dict)
                if "_id" in item_dict:
                    del item_dict["_id"]
                created += 1

            except Exception as e:
                errors_list.append(f"Item {idx+1}: {str(e)}")
                logger.error(f"Erreur import PA item {idx+1}: {e}")

        # Audit une seule fois pour l'import en masse
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.CREATE,
            entity_type=EntityType.PRESQU_ACCIDENT,
            entity_id="bulk_import",
            entity_name=f"Import en masse: {created} presqu'accidents"
        )

        return {
            "success": True,
            "created": created,
            "errors": len(errors_list),
            "error_details": errors_list[:10]
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur import en masse: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



@router.put("/items/{item_id}")
async def update_presqu_accident_item(
    item_id: str,
    item_update: PresquAccidentItemUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Mettre à jour un presqu'accident"""
    try:
        # Vérifier que l'item existe
        existing = await db.presqu_accident_items.find_one({"id": item_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Presqu'accident non trouvé")
        
        # Préparer les mises à jour
        update_data = {
            k: v for k, v in item_update.model_dump(exclude_unset=True).items()
            if v is not None
        }
        
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        update_data["updated_by"] = current_user.get("id")
        
        # Si le statut passe à TERMINE, ajouter la date de clôture
        if update_data.get("status") == PresquAccidentStatus.TERMINE.value and not existing.get("date_cloture"):
            update_data["date_cloture"] = datetime.now(timezone.utc).isoformat()
        
        # Calculer et stocker la priorité si sévérité et récurrence sont présentes
        severite_val = update_data.get("severite_traitement") or existing.get("severite_traitement")
        recurrence_val = update_data.get("recurrence") or existing.get("recurrence")
        if severite_val and recurrence_val:
            try:
                score = int(severite_val) * int(recurrence_val)
                if score <= 4:
                    update_data["priorite"] = "Faible"
                elif score <= 8:
                    update_data["priorite"] = "Moyenne"
                elif score <= 12:
                    update_data["priorite"] = "Élevée"
                else:
                    update_data["priorite"] = "Critique"
                update_data["priorite_score"] = score
            except:
                pass
        
        # Mettre à jour
        await db.presqu_accident_items.update_one(
            {"id": item_id},
            {"$set": update_data}
        )
        
        # Récupérer l'item mis à jour
        updated_item = await db.presqu_accident_items.find_one({"id": item_id})
        
        # Préparer les détails d'audit pour les changements de traitement
        audit_details = []
        if "severite_traitement" in update_data and update_data["severite_traitement"] != existing.get("severite_traitement"):
            severite_labels = {"1": "Mineur", "2": "Modéré", "3": "Grave", "4": "Très grave"}
            old_val = severite_labels.get(str(existing.get("severite_traitement", "")), "-")
            new_val = severite_labels.get(str(update_data["severite_traitement"]), "-")
            audit_details.append(f"Sévérité: {old_val} → {new_val}")
        
        if "recurrence" in update_data and update_data["recurrence"] != existing.get("recurrence"):
            recurrence_labels = {"1": "Très rare", "2": "Rare", "3": "Fréquent", "4": "Très fréquent"}
            old_val = recurrence_labels.get(str(existing.get("recurrence", "")), "-")
            new_val = recurrence_labels.get(str(update_data["recurrence"]), "-")
            audit_details.append(f"Récurrence: {old_val} → {new_val}")
        
        if "status" in update_data and update_data["status"] != existing.get("status"):
            status_labels = {"A_TRAITER": "À traiter", "EN_COURS": "En cours", "TERMINE": "Terminé", "RISQUE_RESIDUEL": "Risque résiduel"}
            old_val = status_labels.get(existing.get("status", ""), "-")
            new_val = status_labels.get(update_data["status"], "-")
            audit_details.append(f"Statut: {old_val} → {new_val}")
        
        # Audit avec détails
        audit_description = f"Presqu'accident: {existing.get('titre')}"
        if audit_details:
            audit_description += f" ({', '.join(audit_details)})"
        
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.UPDATE,
            entity_type=EntityType.PRESQU_ACCIDENT,
            entity_id=item_id,
            entity_name=audit_description
        )
        
        if "_id" in updated_item:
            del updated_item["_id"]
        
        # Broadcast WebSocket pour la synchronisation temps réel
        if realtime_manager:
            await realtime_manager.emit_event(
                "near_miss",
                "updated",
                updated_item,
                user_id=current_user["id"]
            )
        
        return updated_item
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur mise à jour presqu'accident {item_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/items/{item_id}", response_model=SuccessResponse)
async def delete_presqu_accident_item(
    item_id: str,
    current_user: dict = Depends(get_current_admin_user)
):
    """Supprimer un presqu'accident (Admin/QHSE uniquement)"""
    try:
        item = await db.presqu_accident_items.find_one({"id": item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Presqu'accident non trouvé")
        
        await db.presqu_accident_items.delete_one({"id": item_id})
        
        # Audit
        await audit_service.log_action(
            user_id=current_user["id"],
            user_name=f"{current_user['prenom']} {current_user['nom']}",
            user_email=current_user["email"],
            action=ActionType.DELETE,
            entity_type=EntityType.PRESQU_ACCIDENT,
            entity_id=item_id,
            entity_name=f"Presqu'accident: {item.get('titre')}"
        )
        
        # Broadcast WebSocket pour la synchronisation temps réel
        if realtime_manager:
            await realtime_manager.emit_event(
                "near_miss",
                "deleted",
                {"id": item_id, "titre": item.get('titre')},
                user_id=current_user["id"]
            )
        
        return {"success": True, "message": "Presqu'accident supprimé"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur suppression presqu'accident {item_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Statistiques et Indicateurs ====================

@router.get("/stats")
async def get_presqu_accident_stats(current_user: dict = Depends(get_current_user)):
    """Récupérer les statistiques globales des presqu'accidents"""
    try:
        items = await db.presqu_accident_items.find().to_list(length=None)
        
        total = len(items)
        a_traiter = len([i for i in items if i.get("status") == PresquAccidentStatus.A_TRAITER.value])
        en_cours = len([i for i in items if i.get("status") == PresquAccidentStatus.EN_COURS.value])
        termine = len([i for i in items if i.get("status") == PresquAccidentStatus.TERMINE.value])
        risque_residuel = len([i for i in items if i.get("status") == PresquAccidentStatus.RISQUE_RESIDUEL.value])
        
        # Par service
        by_service = {}
        for svc in PresquAccidentService:
            svc_items = [i for i in items if i.get("service") == svc.value]
            svc_termine = len([i for i in svc_items if i.get("status") == PresquAccidentStatus.TERMINE.value])
            by_service[svc.value] = {
                "total": len(svc_items),
                "termine": svc_termine,
                "pourcentage": round((svc_termine / len(svc_items) * 100) if svc_items else 0, 1)
            }
        
        # Par sévérité
        by_severite = {}
        for sev in PresquAccidentSeverity:
            sev_items = [i for i in items if i.get("severite") == sev.value]
            by_severite[sev.value] = len(sev_items)
        
        return {
            "global": {
                "total": total,
                "a_traiter": a_traiter,
                "en_cours": en_cours,
                "termine": termine,
                "risque_residuel": risque_residuel,
                "pourcentage_traitement": round((termine / total * 100) if total > 0 else 0, 1)
            },
            "by_service": by_service,
            "by_severite": by_severite
        }
    except Exception as e:
        logger.error(f"Erreur récupération statistiques: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rapport-stats")
async def get_rapport_stats(current_user: dict = Depends(get_current_user)):
    """
    Récupérer les statistiques complètes pour la page Rapport
    Inclut tous les KPIs : taux de traitement par service, sévérité, lieu, etc.
    """
    try:
        items = await db.presqu_accident_items.find().to_list(length=None)
        
        total = len(items)
        if total == 0:
            return {
                "global": {
                    "total": 0,
                    "a_traiter": 0,
                    "en_cours": 0,
                    "termine": 0,
                    "risque_residuel": 0,
                    "pourcentage_traitement": 0,
                    "delai_moyen_traitement": 0,
                    "en_retard": 0
                },
                "by_service": {},
                "by_severite": {},
                "by_lieu": {},
                "by_month": {}
            }
        
        today = datetime.now(timezone.utc).date()
        
        # Statistiques globales
        a_traiter = [i for i in items if i.get("status") == PresquAccidentStatus.A_TRAITER.value]
        en_cours = [i for i in items if i.get("status") == PresquAccidentStatus.EN_COURS.value]
        termine = [i for i in items if i.get("status") == PresquAccidentStatus.TERMINE.value]
        risque_residuel = [i for i in items if i.get("status") == PresquAccidentStatus.RISQUE_RESIDUEL.value]
        
        # Calculer le délai moyen de traitement (en jours)
        delais = []
        for item in termine:
            if item.get("date_incident") and item.get("date_cloture"):
                try:
                    date_incident = datetime.fromisoformat(item["date_incident"]).date()
                    date_cloture = datetime.fromisoformat(item["date_cloture"]).date()
                    delais.append((date_cloture - date_incident).days)
                except:
                    pass
        delai_moyen = round(sum(delais) / len(delais)) if delais else 0
        
        # Compter les items en retard (actions avec échéance dépassée et non terminés)
        en_retard = 0
        for item in items:
            if item.get("status") not in [PresquAccidentStatus.TERMINE.value, PresquAccidentStatus.RISQUE_RESIDUEL.value]:
                if item.get("date_echeance_action"):
                    try:
                        echeance = datetime.fromisoformat(item["date_echeance_action"]).date()
                        if echeance < today:
                            en_retard += 1
                    except:
                        pass
        
        # Par service
        by_service = {}
        for svc in PresquAccidentService:
            svc_items = [i for i in items if i.get("service") == svc.value]
            svc_termine = len([i for i in svc_items if i.get("status") == PresquAccidentStatus.TERMINE.value])
            by_service[svc.value] = {
                "total": len(svc_items),
                "termine": svc_termine,
                "pourcentage": round((svc_termine / len(svc_items) * 100) if svc_items else 0, 1)
            }
        
        # Par sévérité
        by_severite = {}
        for sev in PresquAccidentSeverity:
            sev_items = [i for i in items if i.get("severite") == sev.value]
            sev_termine = len([i for i in sev_items if i.get("status") == PresquAccidentStatus.TERMINE.value])
            by_severite[sev.value] = {
                "total": len(sev_items),
                "termine": sev_termine,
                "pourcentage": round((sev_termine / len(sev_items) * 100) if sev_items else 0, 1)
            }
        
        # Par lieu (top 10)
        by_lieu = {}
        lieux = set([i.get("lieu", "Non spécifié") for i in items])
        for lieu in lieux:
            lieu_items = [i for i in items if i.get("lieu") == lieu]
            lieu_termine = len([i for i in lieu_items if i.get("status") == PresquAccidentStatus.TERMINE.value])
            by_lieu[lieu] = {
                "total": len(lieu_items),
                "termine": lieu_termine,
                "pourcentage": round((lieu_termine / len(lieu_items) * 100) if lieu_items else 0, 1)
            }
        
        # Par mois (12 derniers mois)
        by_month = {}
        for i in range(12):
            month_start = (datetime.now(timezone.utc) - timedelta(days=30*i)).replace(day=1)
            month_key = month_start.strftime("%Y-%m")
            month_items = []
            for item in items:
                if item.get("date_incident"):
                    try:
                        incident_date = datetime.fromisoformat(item["date_incident"])
                        if incident_date.strftime("%Y-%m") == month_key:
                            month_items.append(item)
                    except:
                        pass
            by_month[month_key] = len(month_items)
        
        return {
            "global": {
                "total": total,
                "a_traiter": len(a_traiter),
                "en_cours": len(en_cours),
                "termine": len(termine),
                "risque_residuel": len(risque_residuel),
                "pourcentage_traitement": round((len(termine) / total * 100), 1),
                "delai_moyen_traitement": delai_moyen,
                "en_retard": en_retard
            },
            "by_service": by_service,
            "by_severite": by_severite,
            "by_lieu": dict(sorted(by_lieu.items(), key=lambda x: x[1]["total"], reverse=True)[:10]),
            "by_month": dict(sorted(by_month.items()))
        }
    except Exception as e:
        logger.error(f"Erreur récupération rapport stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/badge-stats")
async def get_badge_stats(current_user: dict = Depends(get_current_user)):
    """
    Récupérer les statistiques pour le badge de notification du header
    - Nombre de presqu'accidents à traiter
    - Nombre d'actions en retard
    """
    try:
        items = await db.presqu_accident_items.find().to_list(length=None)
        
        total = len(items)
        if total == 0:
            return {
                "a_traiter": 0,
                "en_retard": 0
            }
        
        # Compter les items à traiter
        a_traiter = len([i for i in items if i.get("status") == PresquAccidentStatus.A_TRAITER.value])
        
        # Compter les items en retard
        en_retard = 0
        today = datetime.now(timezone.utc).date()
        
        for item in items:
            if item.get("status") not in [PresquAccidentStatus.TERMINE.value, PresquAccidentStatus.RISQUE_RESIDUEL.value]:
                if item.get("date_echeance_action"):
                    try:
                        echeance = datetime.fromisoformat(item["date_echeance_action"]).date()
                        if echeance < today:
                            en_retard += 1
                    except:
                        pass
        
        return {
            "a_traiter": a_traiter,
            "en_retard": en_retard
        }
    except Exception as e:
        logger.error(f"Erreur récupération badge stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/alerts")
async def get_presqu_accident_alerts(current_user: dict = Depends(get_current_user)):
    """Récupérer les presqu'accidents nécessitant attention (à traiter, en retard)"""
    try:
        items = await db.presqu_accident_items.find().to_list(length=None)
        
        alerts = []
        today = datetime.now(timezone.utc).date()
        
        for item in items:
            alert_item = None
            urgence = "normal"
            
            # Items à traiter
            if item.get("status") == PresquAccidentStatus.A_TRAITER.value:
                alert_item = item
                urgence = "important"
            
            # Items en retard
            if item.get("status") not in [PresquAccidentStatus.TERMINE.value, PresquAccidentStatus.RISQUE_RESIDUEL.value]:
                if item.get("date_echeance_action"):
                    try:
                        echeance = datetime.fromisoformat(item["date_echeance_action"]).date()
                        days_until = (echeance - today).days
                        if days_until < 0:
                            alert_item = item
                            urgence = "critique"
                            item["days_overdue"] = abs(days_until)
                        elif days_until <= 7:
                            alert_item = item
                            urgence = "important"
                            item["days_until"] = days_until
                    except:
                        pass
            
            if alert_item:
                if "_id" in alert_item:
                    del alert_item["_id"]
                alert_item["urgence"] = urgence
                alerts.append(alert_item)
        
        # Trier par urgence (critique > important > normal)
        urgence_order = {"critique": 0, "important": 1, "normal": 2}
        alerts.sort(key=lambda x: urgence_order.get(x.get("urgence", "normal"), 2))
        
        return {
            "count": len(alerts),
            "alerts": alerts
        }
    except Exception as e:
        logger.error(f"Erreur récupération alertes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Upload de pièces jointes ====================

@router.post("/items/{item_id}/attachments")
async def upload_attachment(
    item_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload une pièce jointe pour un presqu'accident (nouveau format multi-fichiers)"""
    try:
        # Vérifier que l'item existe
        item = await db.presqu_accident_items.find_one({"id": item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Presqu'accident non trouvé")
        
        # Créer le répertoire uploads/presqu-accident si nécessaire
        upload_dir = Path("/app/backend/uploads/presqu-accident")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Générer un nom de fichier unique
        file_ext = Path(file.filename).suffix
        attachment_id = str(uuid.uuid4())
        unique_filename = f"{attachment_id}{file_ext}"
        file_path = upload_dir / unique_filename
        
        # Lire et sauvegarder le fichier
        content = await file.read()
        file_size = len(content)
        
        with open(file_path, "wb") as f:
            f.write(content)
        
        # Créer l'objet attachment
        new_attachment = {
            "id": attachment_id,
            "filename": unique_filename,
            "original_filename": file.filename,
            "path": str(file_path),
            "mime_type": file.content_type or "application/octet-stream",
            "size": file_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "uploaded_by": current_user.get("id")
        }
        
        # Ajouter au tableau attachments
        await db.presqu_accident_items.update_one(
            {"id": item_id},
            {
                "$push": {"attachments": new_attachment},
                "$set": {
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user.get("id")
                }
            }
        )
        
        logger.info(f"Pièce jointe ajoutée au presqu'accident {item_id}: {file.filename}")
        
        return {
            "success": True,
            "attachment": new_attachment
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur upload pièce jointe: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/items/{item_id}/attachments")
async def get_attachments(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Récupérer les pièces jointes d'un presqu'accident"""
    try:
        item = await db.presqu_accident_items.find_one({"id": item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Presqu'accident non trouvé")
        
        return item.get("attachments", [])
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur récupération pièces jointes: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/items/{item_id}/attachments/{attachment_id}")
async def download_attachment(
    item_id: str,
    attachment_id: str,
    preview: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Télécharger ou prévisualiser une pièce jointe"""
    from fastapi.responses import FileResponse
    import os
    
    try:
        item = await db.presqu_accident_items.find_one({"id": item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Presqu'accident non trouvé")
        
        # Trouver la pièce jointe
        attachment = None
        for att in item.get("attachments", []):
            if att.get("id") == attachment_id:
                attachment = att
                break
        
        if not attachment:
            raise HTTPException(status_code=404, detail="Pièce jointe non trouvée")
        
        file_path = attachment.get("path")
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Fichier non trouvé sur le serveur")
        
        disposition = "inline" if preview else "attachment"
        return FileResponse(
            path=file_path,
            filename=attachment.get("original_filename", attachment.get("filename")),
            media_type=attachment.get("mime_type", "application/octet-stream"),
            content_disposition_type=disposition
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur téléchargement pièce jointe: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/items/{item_id}/attachments/{attachment_id}", response_model=SuccessResponse)
async def delete_attachment(
    item_id: str,
    attachment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Supprimer une pièce jointe"""
    import os
    
    try:
        item = await db.presqu_accident_items.find_one({"id": item_id})
        if not item:
            raise HTTPException(status_code=404, detail="Presqu'accident non trouvé")
        
        # Trouver et supprimer la pièce jointe
        attachment = None
        for att in item.get("attachments", []):
            if att.get("id") == attachment_id:
                attachment = att
                break
        
        if not attachment:
            raise HTTPException(status_code=404, detail="Pièce jointe non trouvée")
        
        # Supprimer le fichier physique
        file_path = attachment.get("path")
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
        
        # Retirer du tableau attachments
        await db.presqu_accident_items.update_one(
            {"id": item_id},
            {
                "$pull": {"attachments": {"id": attachment_id}},
                "$set": {
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user.get("id")
                }
            }
        )
        
        logger.info(f"Pièce jointe supprimée du presqu'accident {item_id}: {attachment_id}")
        
        return {"success": True, "message": "Pièce jointe supprimée"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur suppression pièce jointe: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/items/{item_id}/upload")
async def upload_piece_jointe_legacy(
    item_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload une pièce jointe pour un presqu'accident (endpoint legacy - redirige vers nouveau)"""
    # Rediriger vers le nouveau endpoint
    return await upload_attachment(item_id, file, current_user)


# ==================== Import/Export ====================

@router.post("/import")
async def import_presqu_accident_data(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_admin_user)
):
    """Importer des données depuis un fichier CSV/Excel"""
    try:
        import pandas as pd
        from io import BytesIO
        
        content = await file.read()
        
        # Lire le fichier selon l'extension
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Format de fichier non supporté")
        
        # Mapper les colonnes
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                item = PresquAccidentItem(
                    titre=str(row.get('titre', '')),
                    description=str(row.get('description', '')),
                    date_incident=str(row.get('date_incident', '')),
                    lieu=str(row.get('lieu', '')),
                    service=str(row.get('service', 'AUTRE')),
                    personnes_impliquees=str(row.get('personnes_impliquees', '')) if pd.notna(row.get('personnes_impliquees')) else None,
                    declarant=str(row.get('declarant', '')) if pd.notna(row.get('declarant')) else None,
                    contexte_cause=str(row.get('contexte_cause', '')) if pd.notna(row.get('contexte_cause')) else None,
                    severite=str(row.get('severite', 'MOYEN')),
                    actions_proposees=str(row.get('actions_proposees', '')) if pd.notna(row.get('actions_proposees')) else None,
                    commentaire=str(row.get('commentaire', '')) if pd.notna(row.get('commentaire')) else None,
                    created_by=current_user.get("id"),
                    updated_by=current_user.get("id")
                )
                
                await db.presqu_accident_items.insert_one(item.model_dump())
                imported_count += 1
            except Exception as e:
                errors.append(f"Ligne {index + 2}: {str(e)}")
        
        return {
            "success": True,
            "imported_count": imported_count,
            "errors": errors[:10]  # Limiter à 10 erreurs
        }
    except Exception as e:
        logger.error(f"Erreur import données: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/template")
async def export_template(current_user: dict = Depends(get_current_user)):
    """Télécharger un template CSV pour l'import"""
    try:
        import pandas as pd
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        # Créer un DataFrame avec les colonnes attendues
        template_data = {
            "titre": ["Risque de chute", "Odeur suspecte"],
            "description": ["Risque de chute lors du chargement camion", "Odeur anormale lors préparation"],
            "date_incident": ["2025-03-26", "2025-03-27"],
            "lieu": ["Entrepôt", "Atelier B2"],
            "service": ["LOGISTIQUE", "PRODUCTION"],
            "personnes_impliquees": ["Jean DUPONT", "Marie MARTIN"],
            "declarant": ["Paul LEFEBVRE", "Sophie BERNARD"],
            "contexte_cause": ["Pas d'escalier adapté", "Ventilation insuffisante"],
            "severite": ["ELEVE", "MOYEN"],
            "actions_proposees": ["Installer escalier mobile", "Vérifier ventilation"],
            "commentaire": ["Urgent", "À surveiller"]
        }
        
        df = pd.DataFrame(template_data)
        
        # Créer un buffer
        buffer = BytesIO()
        df.to_csv(buffer, index=False, encoding='utf-8-sig')
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=template_presqu_accidents.csv"
            }
        )
    except Exception as e:
        logger.error(f"Erreur export template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export")
async def export_presqu_accidents(current_user: dict = Depends(get_current_user)):
    """Exporter tous les presqu'accidents en fichier Excel (.xlsx)"""
    try:
        import pandas as pd
        from io import BytesIO
        from fastapi.responses import StreamingResponse

        items = await db.presqu_accident_items.find(
            {}, {"_id": 0}
        ).sort("created_at", -1).to_list(length=10000)

        if not items:
            items = [{"titre": "Aucun presqu'accident"}]

        df = pd.DataFrame(items)

        col_order = [
            "numero", "titre", "description", "date_incident", "lieu",
            "service", "categorie_incident", "severite", "status",
            "personnes_impliquees", "declarant", "contexte_cause",
            "conditions_incident", "mesures_immediates", "actions_proposees",
            "commentaire_traitement", "responsable_action", "created_at"
        ]
        existing_cols = [c for c in col_order if c in df.columns]
        other_cols = [c for c in df.columns if c not in col_order]
        df = df[existing_cols + other_cols]

        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine='openpyxl', sheet_name='Presqu-accidents')
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=presqu_accidents.xlsx"
            }
        )
    except Exception as e:
        logger.error(f"Erreur export presqu'accidents: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))





async def _extract_with_gemini(content: bytes, filename: str) -> list:
    """Utilise Gemini pour extraire les presqu'accidents d'un PDF ou image."""
    import os
    import json
    import tempfile
    from dotenv import load_dotenv
    load_dotenv()

    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="Cle API Gemini non configuree")

    from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType

    # Determiner le mime type
    mime_map = {
        '.pdf': 'application/pdf',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.webp': 'image/webp',
    }
    ext = '.' + filename.rsplit('.', 1)[-1]
    mime_type = mime_map.get(ext, 'application/octet-stream')

    # Sauvegarder temporairement le fichier
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        chat = LlmChat(
            api_key=api_key,
            session_id=f"presqu-extract-{os.urandom(4).hex()}",
            system_message="""Tu es un assistant specialise dans l'extraction de donnees de securite industrielle.
Tu analyses des documents (PDF, images, tableaux) contenant des presqu'accidents (near-miss) et tu extrais les informations structurees.
Tu DOIS repondre UNIQUEMENT avec un JSON valide, sans texte avant ou apres. Pas de markdown, pas de ```json."""
        ).with_model("gemini", "gemini-2.5-flash")

        file_attachment = FileContentWithMimeType(
            file_path=tmp_path,
            mime_type=mime_type
        )

        prompt = """Analyse ce document et extrais TOUS les presqu'accidents (near-miss / incidents) qu'il contient.

Pour chaque presqu'accident, extrais ces champs:
- titre: titre court (ex: "PA 1 - chute d'objet")
- description: description detaillee / circonstances
- date_incident: date au format YYYY-MM-DD (si disponible)
- lieu: lieu de l'incident
- service: service concerne (PRODUCTION, MAINTENANCE, QUALITE, LOGISTIQUE, ADV, LABO, DIRECTION, FORMULATION, ou AUTRE)
- categorie_incident: type de danger (chute, brulure, projection, etc.)
- declarant: nom de la personne qui declare
- mesures_immediates: actions immediates prises
- actions_proposees: actions correctives proposees
- personnes_impliquees: personnes impliquees

Reponds avec un JSON au format: {"items": [{...}, {...}]}
Si le document ne contient pas de presqu'accidents, reponds: {"items": []}"""

        user_message = UserMessage(
            text=prompt,
            file_contents=[file_attachment]
        )

        response = await chat.send_message(user_message)

        # Parser le JSON de la reponse
        response_text = response.strip()
        # Nettoyer les balises markdown si presentes
        if response_text.startswith('```'):
            response_text = response_text.split('\n', 1)[1] if '\n' in response_text else response_text[3:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]
            response_text = response_text.strip()

        data = json.loads(response_text)
        items = data.get("items", [])

        # Normaliser les champs
        def clean(val):
            """Convertit en string et nettoie les valeurs None/null."""
            s = str(val or "").strip()
            return "" if s.lower() in ("none", "null", "n/a") else s

        result = []
        for idx, item in enumerate(items):
            service = clean(item.get("service")).upper() or "AUTRE"
            valid_services = ["PRODUCTION", "MAINTENANCE", "QUALITE", "LOGISTIQUE", "ADV", "LABO", "DIRECTION", "FORMULATION"]
            if service not in valid_services:
                service = "AUTRE"

            titre = clean(item.get("titre"))
            if not titre:
                desc = clean(item.get("description"))
                titre = f"PA {idx+1} - {desc[:80]}" if desc else f"PA {idx+1}"

            result.append({
                "titre": titre[:200],
                "description": clean(item.get("description")),
                "date_incident": clean(item.get("date_incident")),
                "lieu": clean(item.get("lieu")),
                "service": service,
                "categorie_incident": clean(item.get("categorie_incident")),
                "declarant": clean(item.get("declarant")),
                "mesures_immediates": clean(item.get("mesures_immediates")),
                "actions_proposees": clean(item.get("actions_proposees")),
                "personnes_impliquees": clean(item.get("personnes_impliquees")),
            })
        return result

    finally:
        os.unlink(tmp_path)



@router.post("/ai/extract")
async def ai_extract_presqu_accidents(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Extraction IA: analyse un fichier Excel (.xls/.xlsx), PDF ou image
    et retourne les presqu'accidents extraits, prets a etre importes.
    Pour Excel: parsing direct. Pour PDF/images: analyse par Gemini.
    """
    import os
    from io import BytesIO

    content = await file.read()
    fname = file.filename.lower()

    supported = ('.xls', '.xlsx', '.pdf', '.png', '.jpg', '.jpeg', '.webp')
    if not fname.endswith(supported):
        raise HTTPException(status_code=400, detail=f"Format non supporte. Formats acceptes: {', '.join(supported)}")

    try:
        extracted = []

        # === PDF / Images => Gemini ===
        if fname.endswith(('.pdf', '.png', '.jpg', '.jpeg', '.webp')):
            extracted = await _extract_with_gemini(content, fname)

        elif fname.endswith('.xls') or fname.endswith('.xlsx'):
            # --- Fonctions utilitaires partagees ---
            def _map_headers(headers):
                """Mapping intelligent des en-tetes Excel vers les champs de l'application."""
                col_map = {}
                for i, h in enumerate(headers):
                    hl = h.lower().replace('\n', ' ').replace('\r', '')
                    # N° (numero)
                    if hl.startswith('n°') or hl == 'n' or hl == 'no':
                        col_map.setdefault('numero', i)
                    # Service
                    elif 'service' in hl:
                        col_map.setdefault('service', i)
                    # Personnes impliquees
                    elif 'personnes' in hl or 'impliq' in hl:
                        col_map.setdefault('personnes_impliquees', i)
                    # Date presqu'accident (prioritaire sur date incident)
                    elif 'date' in hl and ('presqu' in hl or 'accident' in hl):
                        col_map['date_pa'] = i
                    # Date de l'incident
                    elif 'date' in hl and 'incident' in hl:
                        col_map.setdefault('date_incident', i)
                    # Lieu
                    elif 'lieu' in hl:
                        col_map.setdefault('lieu', i)
                    # Conditions au moment de l'incident / situation dangereuse / categorie
                    elif 'condition' in hl or 'situation' in hl:
                        col_map.setdefault('conditions_incident', i)
                    # Description: "que s'est-il passe", "circonstances", "description"
                    elif any(kw in hl for kw in ["pass", "circonstance", "détaillé", "detaille", "description"]):
                        col_map.setdefault('description', i)
                    # Mesures immediates: "mesures imm", "actions immediates"
                    elif any(kw in hl for kw in ["mesures imm", "actions imm"]):
                        col_map.setdefault('mesures_immediates', i)
                    # Actions suggerees / correctives / proposees
                    elif 'action' in hl and ('sugg' in hl or 'corr' in hl or 'propos' in hl):
                        col_map.setdefault('actions_proposees', i)
                    # Decision
                    elif 'décision' in hl or 'decision' in hl:
                        col_map.setdefault('decision', i)
                    # Statut
                    elif 'statut' in hl:
                        col_map.setdefault('statut', i)
                    # Plan action / commentaires
                    elif 'plan action' in hl or 'commentaire' in hl:
                        col_map.setdefault('commentaire', i)
                    # Acteur / responsable
                    elif 'acteur' in hl or 'responsable' in hl:
                        col_map.setdefault('acteur', i)
                    # Declarant (nom/prenom si pas deja mappe)
                    elif ('nom' in hl or 'prénom' in hl or 'prenom' in hl or 'déclarant' in hl or 'declarant' in hl):
                        col_map.setdefault('declarant', i)
                return col_map

            SERVICE_MAP = {
                'PRODUCTION': 'PRODUCTION', 'PROD': 'PRODUCTION',
                'MAINTENANCE': 'MAINTENANCE', 'MAINT': 'MAINTENANCE',
                'QUALITE': 'QHSE', 'QHSE': 'QHSE', 'QSE': 'QHSE', 'QUALITÉ': 'QHSE',
                'LOGISTIQUE': 'LOGISTIQUE', 'LOG': 'LOGISTIQUE',
                'ADV': 'ADV',
                'LABO': 'LABO', 'LABORATOIRE': 'LABO',
                'DIRECTION': 'AUTRE', 'FORMULATION': 'PRODUCTION',
                'INDUS': 'INDUS', 'INDUSTRIALISATION': 'INDUS',
            }

            STATUS_MAP = {
                'SOLDEE': 'TERMINE', 'SOLDE': 'TERMINE', 'SOLDÉE': 'TERMINE',
                'TERMINEE': 'TERMINE', 'TERMINE': 'TERMINE', 'TERMINÉ': 'TERMINE', 'TERMINÉE': 'TERMINE',
                'EN COURS': 'EN_COURS', 'ENCOURS': 'EN_COURS',
                'A TRAITER': 'A_TRAITER', 'ATRAITER': 'A_TRAITER',
                'SUPPRIMEE': 'TERMINE', 'SUPPRIMÉE': 'TERMINE',
            }

            def _map_service(raw):
                raw_up = raw.upper().strip()
                if raw_up in SERVICE_MAP:
                    return SERVICE_MAP[raw_up]
                for k, v in SERVICE_MAP.items():
                    if k in raw_up:
                        return v
                return 'AUTRE'

            def _map_status(raw):
                raw_up = raw.upper().strip()
                if raw_up in STATUS_MAP:
                    return STATUS_MAP[raw_up]
                for k, v in STATUS_MAP.items():
                    if k in raw_up:
                        return v
                return 'A_TRAITER'

            def _build_item(num, service_raw, personnes, date_val, lieu, conditions,
                            description, mesures, date_incident, actions, decision,
                            statut, commentaire, acteur):
                service = _map_service(service_raw)
                status = _map_status(statut) if statut else 'A_TRAITER'
                # Nettoyer le numero (enlever .0 des floats Excel)
                num_clean = num.replace('.0', '') if num.endswith('.0') else num
                titre_suffix = conditions if conditions else (lieu if lieu else description[:60] if description else "")
                titre = f"PA {num_clean} - {titre_suffix}" if num_clean else f"PA - {titre_suffix}"
                # Utiliser date_pa en priorite, sinon date_incident
                final_date = date_val if date_val else date_incident
                return {
                    "titre": titre[:200],
                    "description": description,
                    "date_incident": final_date,
                    "lieu": lieu,
                    "service": service,
                    "categorie_incident": conditions,
                    "declarant": "",
                    "personnes_impliquees": personnes,
                    "mesures_immediates": mesures,
                    "actions_proposees": actions,
                    "contexte_cause": decision,
                    "conditions_incident": conditions,
                    "commentaire_traitement": commentaire,
                    "responsable_action": acteur,
                    "status": status,
                }

            # --- Extraction XLS ---
            if fname.endswith('.xls'):
                import xlrd
                wb = xlrd.open_workbook(file_contents=content)
                sh = wb.sheet_by_index(0)

                header_row = 0
                for r in range(min(10, sh.nrows)):
                    for c in range(min(5, sh.ncols)):
                        if str(sh.cell_value(r, c)).strip().lower().startswith('n°'):
                            header_row = r
                            break

                headers = []
                for c in range(sh.ncols):
                    headers.append(str(sh.cell_value(header_row, c)).strip())

                def excel_date(val, book):
                    if not val:
                        return ""
                    try:
                        if isinstance(val, (int, float)) and val > 40000:
                            dt = xlrd.xldate_as_tuple(val, book.datemode)
                            return f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d}"
                        return str(val).strip()
                    except Exception:
                        return str(val).strip()

                col_map = _map_headers(headers)
                logger.info(f"XLS col_map: {col_map}")

                for r in range(header_row + 1, sh.nrows):
                    def cell(col_key):
                        ci = col_map.get(col_key)
                        if ci is None:
                            return ""
                        return str(sh.cell_value(r, ci)).strip()

                    num_val = cell('numero')
                    if not num_val and not cell('personnes_impliquees') and not cell('description'):
                        continue
                    try:
                        float(num_val) if num_val else None
                    except ValueError:
                        continue

                    date_pa = ""
                    di_pa = col_map.get('date_pa')
                    if di_pa is not None:
                        date_pa = excel_date(sh.cell_value(r, di_pa), wb)
                    date_inc = ""
                    di_inc = col_map.get('date_incident')
                    if di_inc is not None:
                        date_inc = excel_date(sh.cell_value(r, di_inc), wb)

                    extracted.append(_build_item(
                        num=cell('numero'),
                        service_raw=cell('service'),
                        personnes=cell('personnes_impliquees') or cell('declarant'),
                        date_val=date_pa,
                        lieu=cell('lieu'),
                        conditions=cell('conditions_incident'),
                        description=cell('description'),
                        mesures=cell('mesures_immediates'),
                        date_incident=date_inc,
                        actions=cell('actions_proposees'),
                        decision=cell('decision'),
                        statut=cell('statut'),
                        commentaire=cell('commentaire'),
                        acteur=cell('acteur'),
                    ))

            # --- Extraction XLSX ---
            else:
                import openpyxl
                wbx = openpyxl.load_workbook(BytesIO(content), data_only=True)
                ws = wbx.active

                header_row_idx = 1
                for r in range(1, min(10, ws.max_row + 1)):
                    for c in range(1, min(5, ws.max_column + 1)):
                        v = str(ws.cell(row=r, column=c).value or '').strip().lower()
                        if v.startswith('n°'):
                            header_row_idx = r
                            break

                headers = []
                for c in range(1, ws.max_column + 1):
                    headers.append(str(ws.cell(row=header_row_idx, column=c).value or '').strip())

                col_map = _map_headers(headers)
                logger.info(f"XLSX col_map: {col_map}")

                for r in range(header_row_idx + 1, ws.max_row + 1):
                    def cellv(col_key):
                        ci = col_map.get(col_key)
                        if ci is None:
                            return ""
                        return str(ws.cell(row=r, column=ci + 1).value or '').strip()

                    num_val = cellv('numero')
                    if not num_val and not cellv('personnes_impliquees') and not cellv('description'):
                        continue

                    date_pa = ""
                    di_pa = col_map.get('date_pa')
                    if di_pa is not None:
                        v = ws.cell(row=r, column=di_pa + 1).value
                        if isinstance(v, datetime):
                            date_pa = v.strftime('%Y-%m-%d')
                        elif v:
                            date_pa = str(v).strip()
                    date_inc = ""
                    di_inc = col_map.get('date_incident')
                    if di_inc is not None:
                        v = ws.cell(row=r, column=di_inc + 1).value
                        if isinstance(v, datetime):
                            date_inc = v.strftime('%Y-%m-%d')
                        elif v:
                            date_inc = str(v).strip()

                    extracted.append(_build_item(
                        num=cellv('numero'),
                        service_raw=cellv('service'),
                        personnes=cellv('personnes_impliquees') or cellv('declarant'),
                        date_val=date_pa,
                        lieu=cellv('lieu'),
                        conditions=cellv('conditions_incident'),
                        description=cellv('description'),
                        mesures=cellv('mesures_immediates'),
                        date_incident=date_inc,
                        actions=cellv('actions_proposees'),
                        decision=cellv('decision'),
                        statut=cellv('statut'),
                        commentaire=cellv('commentaire'),
                        acteur=cellv('acteur'),
                    ))

        return {
            "success": True,
            "count": len(extracted),
            "items": extracted
        }

    except Exception as e:
        logger.error(f"Erreur extraction IA presqu'accident: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erreur extraction: {str(e)}")
