"""
Routes API M.E.S (Manufacturing Execution System)
"""
from fastapi import APIRouter, Depends, HTTPException
from dependencies import get_current_user, get_current_admin_user, get_database
from models import SuccessResponse, MessageResponse
from openapi_config import CRUD_ERRORS, STANDARD_ERRORS

router = APIRouter(prefix="/mes", tags=["MES"])

# Service will be initialized from server.py
mes_service = None
audit_service_ref = None

def init_mes_routes(db, mqtt_manager=None):
    global mes_service, audit_service_ref
    from mes_service import MESService
    from audit_service import AuditService
    mes_service = MESService(db, mqtt_manager)
    audit_service_ref = AuditService(db)


# ==================== MACHINES ====================

@router.get("/machines",
    summary="Lister les machines M.E.S.",
    description="Retourne la liste de toutes les machines configurees dans le module M.E.S. avec leurs parametres de production.",
    responses={**STANDARD_ERRORS}
)
async def list_machines(current_user: dict = Depends(get_current_user)):
    return await mes_service.get_machines()

@router.get("/machines/{machine_id}",
    summary="Detail d'une machine",
    description="Retourne les informations detaillees d'une machine M.E.S. incluant sa configuration, ses parametres de production et sa reference produit active.",
    responses={**CRUD_ERRORS}
)
async def get_machine(machine_id: str, current_user: dict = Depends(get_current_user)):
    m = await mes_service.get_machine(machine_id)
    if not m:
        raise HTTPException(404, "Machine non trouvée")
    return m

@router.post("/machines",
    summary="Creer une machine",
    description="Ajoute une nouvelle machine au suivi M.E.S. avec sa configuration initiale (cadence objectif, seuils d'alerte, planning de production).",
    responses={**STANDARD_ERRORS}
)
async def create_machine(data: dict, current_user: dict = Depends(get_current_user)):
    return await mes_service.create_machine(data)

@router.put("/machines/{machine_id}",
    summary="Modifier une machine",
    description="Met a jour la configuration d'une machine M.E.S. (parametres de production, seuils, planning, objectif TRS).",
    responses={**CRUD_ERRORS}
)
async def update_machine(machine_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    m = await mes_service.update_machine(machine_id, data)
    if not m:
        raise HTTPException(404, "Machine non trouvée")
    return m

@router.delete("/machines/{machine_id}", response_model=SuccessResponse,
    summary="Supprimer une machine",
    description="Supprime definitivement une machine et toutes ses donnees associees (metriques, alertes, rebuts).",
    responses={**CRUD_ERRORS}
)
async def delete_machine(machine_id: str, current_user: dict = Depends(get_current_user)):
    await mes_service.delete_machine(machine_id)
    return {"success": True, "message": "Machine supprimée"}


# ==================== METRICS ====================

@router.get("/machines/{machine_id}/metrics",
    summary="Metriques temps reel d'une machine",
    description="Retourne les metriques de production en temps reel : pieces produites, cadence instantanee, TRS actuel (Disponibilite x Performance x Qualite), temps d'arret et statut de production.",
    responses={**CRUD_ERRORS}
)
async def get_metrics(machine_id: str, current_user: dict = Depends(get_current_user)):
    return await mes_service.get_realtime_metrics(machine_id)

@router.get("/machines/{machine_id}/history",
    summary="Historique de cadence",
    description="Retourne l'historique de la cadence de production sur une periode donnee (6h, 12h, 24h, 7d ou dates personnalisees).",
    responses={**CRUD_ERRORS}
)
async def get_history(machine_id: str, period: str = "6h",
                      date_from: str = None, date_to: str = None,
                      current_user: dict = Depends(get_current_user)):
    return await mes_service.get_cadence_history(machine_id, period, date_from, date_to)


# ==================== ALERTS ====================

@router.get("/alerts",
    summary="Lister les alertes M.E.S.",
    description="Retourne les alertes de production (arret machine, TRS bas, panne capteur). Filtrable par statut lu/non-lu.",
    responses={**STANDARD_ERRORS}
)
async def list_alerts(unread_only: bool = False, limit: int = 50,
                      current_user: dict = Depends(get_current_user)):
    return await mes_service.get_alerts(unread_only, limit)

@router.get("/alerts/count",
    summary="Compteur d'alertes non lues",
    description="Retourne le nombre d'alertes M.E.S. non lues pour affichage dans les badges de notification.",
    responses={**STANDARD_ERRORS}
)
async def alert_count(current_user: dict = Depends(get_current_user)):
    count = await mes_service.get_unread_alert_count()
    return {"count": count}

@router.put("/alerts/{alert_id}/read", response_model=SuccessResponse,
    summary="Marquer une alerte comme lue",
    responses={**CRUD_ERRORS}
)
async def mark_read(alert_id: str, current_user: dict = Depends(get_current_user)):
    await mes_service.mark_alert_read(alert_id)
    return {"success": True, "message": "Alerte marquée comme lue"}

@router.put("/alerts/read-all", response_model=SuccessResponse,
    summary="Marquer toutes les alertes comme lues",
    responses={**STANDARD_ERRORS}
)
async def mark_all_read(current_user: dict = Depends(get_current_user)):
    await mes_service.mark_all_alerts_read()
    return {"success": True, "message": "Toutes les alertes marquées comme lues"}

@router.delete("/alerts/all", response_model=SuccessResponse,
    summary="Supprimer toutes les alertes",
    description="Supprime definitivement toutes les alertes M.E.S. Cette action est irreversible.",
    responses={**STANDARD_ERRORS}
)
async def delete_all_alerts(current_user: dict = Depends(get_current_user)):
    await mes_service.delete_all_alerts()
    return {"success": True, "message": "Toutes les alertes supprimées"}


# ==================== TOOLS ====================

@router.post("/machines/{machine_id}/ping",
    summary="Ping capteur machine",
    description="Envoie un ping au capteur MQTT de la machine pour verifier la connectivite.",
    responses={**CRUD_ERRORS}
)
async def ping_sensor(machine_id: str, current_user: dict = Depends(get_current_user)):
    return await mes_service.ping_sensor(machine_id)

@router.post("/machines/{machine_id}/simulate-pulse", response_model=SuccessResponse,
    summary="Simuler une impulsion",
    description="Simule une impulsion de production pour tester la configuration de la machine (increment du compteur de pieces).",
    responses={**CRUD_ERRORS}
)
async def simulate_pulse(machine_id: str, current_user: dict = Depends(get_current_user)):
    await mes_service.record_pulse(machine_id, 1)
    return {"success": True, "message": "Impulsion simulée"}


# ==================== CONFIGURATION (Rétention) ====================

@router.get("/config/retention",
    summary="Obtenir la configuration de rétention M.E.S",
    description="Récupère la durée de rétention en jours et les statistiques de volumétrie.",
    responses={**STANDARD_ERRORS}
)
async def get_mes_retention(current_user: dict = Depends(get_current_user)):
    stats = await mes_service.get_storage_stats()
    return {
        "retention_days": stats["retention_days"],
        "min_days": mes_service.MIN_RETENTION_DAYS,
        "max_days": mes_service.MAX_RETENTION_DAYS,
        "pulses_count": stats["pulses_count"],
        "cadence_count": stats["cadence_count"],
        "oldest_pulse": stats["oldest_pulse"],
    }


@router.put("/config/retention",
    summary="Modifier la durée de rétention M.E.S (admin)",
    description="Définit la durée de rétention en jours. Le prochain nettoyage automatique appliquera cette nouvelle valeur.",
    responses={**STANDARD_ERRORS}
)
async def set_mes_retention(payload: dict, current_user: dict = Depends(get_current_admin_user)):
    days = payload.get("retention_days")
    if days is None:
        raise HTTPException(status_code=400, detail="retention_days requis")
    try:
        days = int(days)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="retention_days doit être un entier")
    effective = await mes_service.set_retention_days(days)
    return {"success": True, "retention_days": effective}


@router.post("/config/cleanup-now",
    summary="Lancer le nettoyage immédiat (admin)",
    description="Force l'exécution immédiate du nettoyage des données M.E.S anciennes selon la rétention configurée.",
    responses={**STANDARD_ERRORS}
)
async def cleanup_now(current_user: dict = Depends(get_current_admin_user)):
    result = await mes_service.cleanup_old_data()
    return {"success": True, **(result or {})}


# ==================== RAPPORTS DE POSTE (3×8) ====================

@router.get("/machines/{machine_id}/shifts",
    summary="Rapports de poste 3×8 d'une machine",
    description="Liste les derniers rapports de poste reçus depuis l'ESP32 via topic shift_end.",
    responses={**STANDARD_ERRORS}
)
async def get_machine_shifts(machine_id: str, limit: int = 30, current_user: dict = Depends(get_current_user)):
    return await mes_service.get_shift_summaries(machine_id, limit=limit)


# ==================== REJECT REASONS (Admin) ====================

@router.get("/reject-reasons",
    summary="Lister les motifs de rebut",
    description="Retourne la liste des motifs de rebut predefinis utilisables lors de la declaration des pieces defectueuses.",
    responses={**STANDARD_ERRORS}
)
async def list_reject_reasons(current_user: dict = Depends(get_current_user)):
    return await mes_service.get_reject_reasons()

@router.post("/reject-reasons",
    summary="Creer un motif de rebut",
    description="Ajoute un nouveau motif de rebut a la liste des motifs predefinis (ex: defaut visuel, casse, hors tolerance).",
    responses={**STANDARD_ERRORS, 400: {"description": "Libelle requis"}}
)
async def create_reject_reason(data: dict, current_user: dict = Depends(get_current_user)):
    if not data.get("label"):
        raise HTTPException(400, "Le libellé est requis")
    return await mes_service.create_reject_reason(data)

@router.put("/reject-reasons/{reason_id}",
    summary="Modifier un motif de rebut",
    responses={**CRUD_ERRORS}
)
async def update_reject_reason(reason_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    result = await mes_service.update_reject_reason(reason_id, data)
    if not result:
        raise HTTPException(404, "Motif non trouvé")
    return result

@router.delete("/reject-reasons/{reason_id}", response_model=SuccessResponse,
    summary="Supprimer un motif de rebut",
    responses={**CRUD_ERRORS}
)
async def delete_reject_reason(reason_id: str, current_user: dict = Depends(get_current_user)):
    await mes_service.delete_reject_reason(reason_id)
    return {"success": True, "message": "Motif de rebut supprimé"}


# ==================== REJECTS (Operator) ====================

@router.post("/machines/{machine_id}/rejects",
    summary="Declarer un rebut",
    description="Declare des pieces defectueuses sur une machine avec quantite, motif et commentaire optionnel. Impacte le calcul du TRS Qualite.",
    responses={**CRUD_ERRORS, 400: {"description": "Quantite invalide"}}
)
async def declare_reject(machine_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    if not data.get("quantity") or int(data["quantity"]) <= 0:
        raise HTTPException(400, "La quantité doit être supérieure à 0")
    data["operator"] = current_user.get("name", current_user.get("email", ""))
    return await mes_service.declare_reject(machine_id, data)

@router.get("/machines/{machine_id}/rejects",
    summary="Historique des rebuts d'une machine",
    description="Retourne l'historique des rebuts declares sur une machine avec filtres de date optionnels.",
    responses={**CRUD_ERRORS}
)
async def list_rejects(machine_id: str, date_from: str = None, date_to: str = None,
                       current_user: dict = Depends(get_current_user)):
    return await mes_service.get_rejects(machine_id, date_from, date_to)

@router.delete("/rejects/{reject_id}", response_model=SuccessResponse,
    summary="Supprimer un rebut",
    responses={**CRUD_ERRORS}
)
async def delete_reject(reject_id: str, current_user: dict = Depends(get_current_user)):
    await mes_service.delete_reject(reject_id)
    return {"success": True, "message": "Rebut supprimé"}


# ==================== PRODUCT REFERENCES ====================

@router.get("/product-references",
    summary="Lister les references produit",
    description="Retourne toutes les references produit configurees. Une reference contient les parametres de production pre-configures (cadence, seuils, objectif TRS) applicables a une machine.",
    responses={**STANDARD_ERRORS}
)
async def list_product_references(current_user: dict = Depends(get_current_user)):
    return await mes_service.get_product_references()

@router.post("/product-references",
    summary="Creer une reference produit",
    description="Cree un nouveau template de parametres de production. Accessible uniquement aux administrateurs. L'action est tracee dans le journal d'audit.",
    responses={**STANDARD_ERRORS, 400: {"description": "Nom requis"}}
)
async def create_product_reference(data: dict, current_user: dict = Depends(get_current_admin_user)):
    if not data.get("name"):
        raise HTTPException(400, "Le nom de la reference est requis")
    ref = await mes_service.create_product_reference(data)
    from models import ActionType, EntityType
    await audit_service_ref.log_action(
        user_id=str(current_user.get("id", "")),
        user_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}",
        user_email=current_user.get("email", ""),
        action=ActionType.CREATE,
        entity_type=EntityType.MES_PRODUCT_REFERENCE,
        entity_id=ref.get("id"),
        entity_name=ref.get("name"),
        details=f"Creation reference produite: {ref.get('name')}"
    )
    return ref

@router.put("/product-references/{ref_id}",
    summary="Modifier une reference produit",
    description="Met a jour les parametres d'une reference produit existante. Admin uniquement. Tracee dans l'audit.",
    responses={**CRUD_ERRORS}
)
async def update_product_reference(ref_id: str, data: dict, current_user: dict = Depends(get_current_admin_user)):
    result = await mes_service.update_product_reference(ref_id, data)
    if not result:
        raise HTTPException(404, "Reference non trouvee")
    from models import ActionType, EntityType
    await audit_service_ref.log_action(
        user_id=str(current_user.get("id", "")),
        user_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}",
        user_email=current_user.get("email", ""),
        action=ActionType.UPDATE,
        entity_type=EntityType.MES_PRODUCT_REFERENCE,
        entity_id=ref_id,
        entity_name=result.get("name"),
        details=f"Modification reference produite: {result.get('name')}"
    )
    return result

@router.delete("/product-references/{ref_id}", response_model=SuccessResponse,
    summary="Supprimer une reference produit",
    description="Supprime definitivement une reference produit. Admin uniquement. Les machines utilisant cette reference ne sont pas affectees.",
    responses={**CRUD_ERRORS}
)
async def delete_product_reference(ref_id: str, current_user: dict = Depends(get_current_admin_user)):
    await mes_service.delete_product_reference(ref_id)
    from models import ActionType, EntityType
    await audit_service_ref.log_action(
        user_id=str(current_user.get("id", "")),
        user_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}",
        user_email=current_user.get("email", ""),
        action=ActionType.DELETE,
        entity_type=EntityType.MES_PRODUCT_REFERENCE,
        entity_id=ref_id,
        details="Suppression reference produite"
    )
    return {"success": True, "message": "Référence produite supprimée"}

@router.post("/machines/{machine_id}/select-reference",
    summary="Selectionner une reference produit pour une machine",
    description="Applique les parametres d'une reference produit a une machine. Met a jour automatiquement la cadence objectif, les seuils et le planning de la machine.",
    responses={**CRUD_ERRORS, 400: {"description": "reference_id requis"}}
)
async def select_reference(machine_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    ref_id = data.get("reference_id")
    if not ref_id:
        raise HTTPException(400, "reference_id requis")
    result = await mes_service.select_reference_for_machine(machine_id, ref_id)
    if not result:
        raise HTTPException(404, "Reference non trouvee")
    from models import ActionType, EntityType
    await audit_service_ref.log_action(
        user_id=str(current_user.get("id", "")),
        user_name=f"{current_user.get('prenom', '')} {current_user.get('nom', '')}",
        user_email=current_user.get("email", ""),
        action=ActionType.UPDATE,
        entity_type=EntityType.MES_PRODUCT_REFERENCE,
        entity_id=ref_id,
        entity_name=result.get("equipment_name"),
        details=f"Changement reference produite sur machine {result.get('equipment_name')}"
    )
    return result


# ==================== TRS HISTORY ====================

@router.get("/machines/{machine_id}/trs-history",
    summary="Historique TRS quotidien",
    description="Retourne l'historique du TRS jour par jour pour une machine sur les N derniers jours. Inclut Disponibilite, Performance et Qualite.",
    responses={**CRUD_ERRORS}
)
async def get_trs_history(machine_id: str, days: int = 7,
                          current_user: dict = Depends(get_current_user)):
    return await mes_service.get_trs_daily_history(machine_id, days)


# ==================== REPORTING ====================

@router.post("/reports/data",
    summary="Donnees de rapport M.E.S.",
    description="Genere les donnees agregees pour les rapports M.E.S. Supporte plusieurs types : TRS, production, arrets, rebuts, alertes ou rapport complet. Filtrable par machine(s) et periode.",
    responses={**STANDARD_ERRORS, 400: {"description": "Dates requises"}}
)
async def get_report_data(data: dict, current_user: dict = Depends(get_current_user)):
    """Get report data for specified machines and period"""
    machine_ids = data.get("machine_ids", ["all"])
    report_type = data.get("report_type", "all")  # trs, production, stops, rejects, alerts, all
    date_from = data.get("date_from")
    date_to = data.get("date_to")
    
    if not date_from or not date_to:
        raise HTTPException(400, "date_from et date_to sont requis")
    
    return await mes_service.get_report_data(machine_ids, report_type, date_from, date_to)


@router.post("/reports/export/excel",
    summary="Exporter rapport M.E.S. en Excel",
    description="Genere et telecharge un fichier Excel multi-feuilles avec les donnees M.E.S. agregees (TRS, production, arrets, rebuts).",
    responses={**STANDARD_ERRORS, 200: {"description": "Fichier Excel genere", "content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}}}}
)
async def export_excel_report(data: dict, current_user: dict = Depends(get_current_user)):
    """Export report data to Excel file"""
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    machine_ids = data.get("machine_ids", ["all"])
    report_type = data.get("report_type", "all")
    date_from = data.get("date_from")
    date_to = data.get("date_to")
    
    if not date_from or not date_to:
        raise HTTPException(400, "date_from et date_to sont requis")
    
    report_data = await mes_service.get_report_data(machine_ids, report_type, date_from, date_to)
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Resume"
    summary = report_data.get("summary", {})
    ws_summary.append(["RAPPORT M.E.S."])
    ws_summary.append(["Periode", f"{date_from} - {date_to}"])
    ws_summary.append(["Genere le", report_data.get("generated_at", "")[:19]])
    ws_summary.append([])
    ws_summary.append(["RESUME GLOBAL"])
    ws_summary.append(["Nombre de machines", summary.get("total_machines", 0)])
    ws_summary.append(["Production totale", summary.get("total_production", 0)])
    ws_summary.append(["Rebuts totaux", summary.get("total_rejects", 0)])
    ws_summary.append(["Temps d'arret total (h)", summary.get("total_downtime_hours", 0)])
    ws_summary.append(["TRS moyen (%)", summary.get("average_trs", 0)])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A5"].font = Font(bold=True)
    auto_width(ws_summary)
    
    # TRS Sheet
    if report_type in ["trs", "all"]:
        ws_trs = wb.create_sheet("TRS")
        ws_trs.append(["Machine", "Date", "TRS (%)", "Disponibilite (%)", "Performance (%)", "Qualite (%)", "Production", "Rebuts"])
        style_header(ws_trs)
        for machine in report_data.get("machines", []):
            trs_data = machine.get("trs", {})
            for v in trs_data.get("trs_values", []):
                if v.get("is_production_day"):
                    ws_trs.append([
                        machine.get("name", ""),
                        v.get("date", ""),
                        v.get("trs", 0),
                        v.get("availability", 0),
                        v.get("performance", 0),
                        v.get("quality", 0),
                        v.get("production", 0),
                        v.get("rejects", 0),
                    ])
        auto_width(ws_trs)
    
    # Production Sheet
    if report_type in ["production", "all"]:
        ws_prod = wb.create_sheet("Production")
        ws_prod.append(["Machine", "Date", "Production"])
        style_header(ws_prod)
        for machine in report_data.get("machines", []):
            prod_data = machine.get("production", {})
            for v in prod_data.get("daily_values", []):
                ws_prod.append([machine.get("name", ""), v.get("date", ""), v.get("production", 0)])
        auto_width(ws_prod)
    
    # Rejects Sheet
    if report_type in ["rejects", "all"]:
        ws_rej = wb.create_sheet("Rebuts")
        ws_rej.append(["Machine", "Total Rebuts", "Motif", "Quantite"])
        style_header(ws_rej)
        for machine in report_data.get("machines", []):
            rej_data = machine.get("rejects", {})
            first = True
            for r in rej_data.get("by_reason", []):
                if first:
                    ws_rej.append([machine.get("name", ""), rej_data.get("total", 0), r.get("reason", ""), r.get("quantity", 0)])
                    first = False
                else:
                    ws_rej.append(["", "", r.get("reason", ""), r.get("quantity", 0)])
            if not rej_data.get("by_reason"):
                ws_rej.append([machine.get("name", ""), 0, "", ""])
        auto_width(ws_rej)
    
    # Stops Sheet
    if report_type in ["stops", "all"]:
        ws_stops = wb.create_sheet("Arrets")
        ws_stops.append(["Machine", "Temps arret total (h)", "Nb evenements", "Date/Heure", "Type", "Message"])
        style_header(ws_stops)
        for machine in report_data.get("machines", []):
            stops_data = machine.get("stops", {})
            first = True
            for s in stops_data.get("stop_events", []):
                if first:
                    ws_stops.append([
                        machine.get("name", ""),
                        stops_data.get("total_downtime_hours", 0),
                        stops_data.get("stop_count", 0),
                        s.get("timestamp", "")[:19],
                        s.get("type", ""),
                        s.get("message", ""),
                    ])
                    first = False
                else:
                    ws_stops.append(["", "", "", s.get("timestamp", "")[:19], s.get("type", ""), s.get("message", "")])
            if not stops_data.get("stop_events"):
                ws_stops.append([machine.get("name", ""), stops_data.get("total_downtime_hours", 0), 0, "", "", ""])
        auto_width(ws_stops)
    
    # Alerts Sheet
    if report_type in ["alerts", "all"]:
        ws_alerts = wb.create_sheet("Alertes")
        ws_alerts.append(["Machine", "Total Alertes", "Type", "Nombre", "Date/Heure", "Message"])
        style_header(ws_alerts)
        for machine in report_data.get("machines", []):
            alerts_data = machine.get("alerts", {})
            first = True
            # By type summary
            for a in alerts_data.get("by_type", []):
                if first:
                    ws_alerts.append([machine.get("name", ""), alerts_data.get("total", 0), a.get("type", ""), a.get("count", 0), "", ""])
                    first = False
                else:
                    ws_alerts.append(["", "", a.get("type", ""), a.get("count", 0), "", ""])
            if not alerts_data.get("by_type"):
                ws_alerts.append([machine.get("name", ""), 0, "", "", "", ""])
        auto_width(ws_alerts)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"rapport_mes_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/reports/export/pdf")
async def export_pdf_report(data: dict, current_user: dict = Depends(get_current_user)):
    """Export report data to PDF file"""
    from fastapi.responses import StreamingResponse
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    machine_ids = data.get("machine_ids", ["all"])
    report_type = data.get("report_type", "all")
    date_from = data.get("date_from")
    date_to = data.get("date_to")
    
    if not date_from or not date_to:
        raise HTTPException(400, "date_from et date_to sont requis")
    
    report_data = await mes_service.get_report_data(machine_ids, report_type, date_from, date_to)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=20)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, alignment=TA_LEFT, spaceAfter=10)
    normal_style = styles['Normal']
    
    elements = []
    
    # Title
    elements.append(Paragraph("RAPPORT M.E.S.", title_style))
    elements.append(Paragraph(f"Periode: {date_from} - {date_to}", normal_style))
    elements.append(Paragraph(f"Genere le: {report_data.get('generated_at', '')[:19]}", normal_style))
    elements.append(Spacer(1, 20))
    
    # Summary
    summary = report_data.get("summary", {})
    elements.append(Paragraph("RESUME GLOBAL", subtitle_style))
    summary_data = [
        ["Indicateur", "Valeur"],
        ["Nombre de machines", str(summary.get("total_machines", 0))],
        ["Production totale", str(summary.get("total_production", 0))],
        ["Rebuts totaux", str(summary.get("total_rejects", 0))],
        ["Temps d'arret total", f"{summary.get('total_downtime_hours', 0)} h"],
        ["TRS moyen", f"{summary.get('average_trs', 0)} %"],
    ]
    summary_table = Table(summary_data, colWidths=[8*cm, 5*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # TRS Table
    if report_type in ["trs", "all"]:
        elements.append(Paragraph("TRS PAR MACHINE", subtitle_style))
        trs_header = ["Machine", "TRS Moy.", "Dispo.", "Perf.", "Qualite"]
        trs_rows = [trs_header]
        for machine in report_data.get("machines", []):
            trs_data = machine.get("trs", {})
            trs_rows.append([
                machine.get("name", "")[:30],
                f"{trs_data.get('average_trs', 0)} %",
                f"{trs_data.get('average_availability', 0)} %",
                f"{trs_data.get('average_performance', 0)} %",
                f"{trs_data.get('average_quality', 0)} %",
            ])
        if len(trs_rows) > 1:
            trs_table = Table(trs_rows, colWidths=[6*cm, 3*cm, 3*cm, 3*cm, 3*cm])
            trs_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ]))
            elements.append(trs_table)
        elements.append(Spacer(1, 20))
    
    # Production Table
    if report_type in ["production", "all"]:
        elements.append(Paragraph("PRODUCTION PAR MACHINE", subtitle_style))
        prod_header = ["Machine", "Production Totale", "Moyenne Journaliere"]
        prod_rows = [prod_header]
        for machine in report_data.get("machines", []):
            prod_data = machine.get("production", {})
            prod_rows.append([
                machine.get("name", "")[:30],
                str(prod_data.get("total", 0)),
                str(prod_data.get("average_daily", 0)),
            ])
        if len(prod_rows) > 1:
            prod_table = Table(prod_rows, colWidths=[8*cm, 5*cm, 5*cm])
            prod_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ]))
            elements.append(prod_table)
        elements.append(Spacer(1, 20))
    
    # Rejects Table
    if report_type in ["rejects", "all"]:
        elements.append(Paragraph("REBUTS PAR MACHINE", subtitle_style))
        rej_header = ["Machine", "Total Rebuts", "Principal Motif"]
        rej_rows = [rej_header]
        for machine in report_data.get("machines", []):
            rej_data = machine.get("rejects", {})
            top_reason = rej_data.get("by_reason", [{}])[0].get("reason", "-") if rej_data.get("by_reason") else "-"
            rej_rows.append([
                machine.get("name", "")[:30],
                str(rej_data.get("total", 0)),
                top_reason[:40],
            ])
        if len(rej_rows) > 1:
            rej_table = Table(rej_rows, colWidths=[8*cm, 4*cm, 6*cm])
            rej_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ]))
            elements.append(rej_table)
        elements.append(Spacer(1, 20))
    
    # Stops Table
    if report_type in ["stops", "all"]:
        elements.append(Paragraph("ARRETS PAR MACHINE", subtitle_style))
        stops_header = ["Machine", "Temps Arret (h)", "Nb Evenements"]
        stops_rows = [stops_header]
        for machine in report_data.get("machines", []):
            stops_data = machine.get("stops", {})
            stops_rows.append([
                machine.get("name", "")[:30],
                str(stops_data.get("total_downtime_hours", 0)),
                str(stops_data.get("stop_count", 0)),
            ])
        if len(stops_rows) > 1:
            stops_table = Table(stops_rows, colWidths=[8*cm, 5*cm, 5*cm])
            stops_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ]))
            elements.append(stops_table)
        elements.append(Spacer(1, 20))
    
    # Alerts Table
    if report_type in ["alerts", "all"]:
        elements.append(Paragraph("ALERTES PAR MACHINE", subtitle_style))
        alerts_header = ["Machine", "Total Alertes", "Principal Type"]
        alerts_rows = [alerts_header]
        for machine in report_data.get("machines", []):
            alerts_data = machine.get("alerts", {})
            top_type = alerts_data.get("by_type", [{}])[0].get("type", "-") if alerts_data.get("by_type") else "-"
            alerts_rows.append([
                machine.get("name", "")[:30],
                str(alerts_data.get("total", 0)),
                top_type,
            ])
        if len(alerts_rows) > 1:
            alerts_table = Table(alerts_rows, colWidths=[8*cm, 5*cm, 5*cm])
            alerts_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ]))
            elements.append(alerts_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"rapport_mes_{date_from}_{date_to}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== SCHEDULED REPORTS ====================

@router.get("/scheduled-reports",
    summary="Lister les rapports planifies",
    description="Retourne tous les rapports M.E.S. planifies avec leur frequence, destinataires et prochaine execution.",
    responses={**STANDARD_ERRORS}
)
async def list_scheduled_reports(current_user: dict = Depends(get_current_user)):
    return await mes_service.get_scheduled_reports()

@router.get("/scheduled-reports/{report_id}",
    summary="Detail d'un rapport planifie",
    responses={**CRUD_ERRORS}
)
async def get_scheduled_report(report_id: str, current_user: dict = Depends(get_current_user)):
    report = await mes_service.get_scheduled_report(report_id)
    if not report:
        raise HTTPException(404, "Rapport planifie non trouve")
    return report

@router.post("/scheduled-reports",
    summary="Creer un rapport planifie",
    description="Configure un nouveau rapport automatique. Supporte les frequences quotidienne, hebdomadaire (jour choisi) et mensuelle (jour du mois). Les rapports sont envoyes par email en PDF ou Excel.",
    responses={**STANDARD_ERRORS, 400: {"description": "Nom ou destinataire(s) manquant(s)"}}
)
async def create_scheduled_report(data: dict, current_user: dict = Depends(get_current_user)):
    """Create a new scheduled report"""
    if not data.get("name"):
        raise HTTPException(400, "Le nom du rapport est requis")
    if not data.get("recipients"):
        raise HTTPException(400, "Au moins un destinataire est requis")
    data["created_by"] = current_user.get("email", "")
    return await mes_service.create_scheduled_report(data)

@router.put("/scheduled-reports/{report_id}",
    summary="Modifier un rapport planifie",
    description="Met a jour la configuration d'un rapport planifie (frequence, destinataires, format).",
    responses={**CRUD_ERRORS}
)
async def update_scheduled_report(report_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    report = await mes_service.update_scheduled_report(report_id, data)
    if not report:
        raise HTTPException(404, "Rapport planifie non trouve")
    return report

@router.delete("/scheduled-reports/{report_id}", response_model=SuccessResponse,
    summary="Supprimer un rapport planifie",
    responses={**CRUD_ERRORS}
)
async def delete_scheduled_report(report_id: str, current_user: dict = Depends(get_current_user)):
    await mes_service.delete_scheduled_report(report_id)
    return {"success": True, "message": "Rapport planifié supprimé"}

@router.post("/scheduled-reports/{report_id}/send-now", response_model=SuccessResponse,
    summary="Envoyer un rapport immediatement",
    description="Declenche l'envoi immediat d'un rapport planifie aux destinataires configures. Utile pour tester la configuration.",
    responses={**CRUD_ERRORS, 500: {"description": "Scheduler non disponible"}}
)
async def send_scheduled_report_now(report_id: str, current_user: dict = Depends(get_current_user)):
    report = await mes_service.get_scheduled_report(report_id)
    if not report:
        raise HTTPException(404, "Rapport planifie non trouve")
    from mes_report_scheduler import mes_report_scheduler
    if mes_report_scheduler:
        await mes_report_scheduler.send_report(report_id)
        return {"success": True, "message": "Rapport envoye"}
    else:
        raise HTTPException(500, "Scheduler non disponible")

