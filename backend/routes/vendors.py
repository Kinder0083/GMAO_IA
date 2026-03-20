"""
Routes des Fournisseurs et Historique d'Achats - CRUD, AI Extract
Extrait de server.py.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from bson import ObjectId
from datetime import datetime, timezone
from typing import List, Optional
import uuid
import logging

from models import (
    ActionType, EntityType, MessageResponse,
    PurchaseHistory, PurchaseHistoryCreate, PurchaseHistoryUpdate,
    Vendor, VendorCreate, VendorUpdate
)
from dependencies import get_current_user, get_current_admin_user, require_permission
from routes.shared import db, audit_service, serialize_doc

EntityType_Audit = EntityType
logger = logging.getLogger(__name__)

router = APIRouter(tags=["Fournisseurs"])

@router.get("/vendors",
    summary="Lister les fournisseurs", response_model=List[Vendor], tags=["Fournisseurs"])
async def get_vendors(current_user: dict = Depends(require_permission("vendors", "view"))):
    """Liste tous les fournisseurs"""
    vendors = await db.vendors.find().to_list(1000)
    return [Vendor(**serialize_doc(vendor)) for vendor in vendors]

@router.post("/vendors",
    summary="Creer un fournisseur", response_model=Vendor, tags=["Fournisseurs"])
async def create_vendor(vendor_create: VendorCreate, current_user: dict = Depends(require_permission("vendors", "edit"))):
    """Créer un nouveau fournisseur"""
    vendor_dict = vendor_create.model_dump()
    vendor_dict["dateCreation"] = datetime.utcnow()
    vendor_dict["_id"] = ObjectId()
    
    await db.vendors.insert_one(vendor_dict)
    
    vendor_data = serialize_doc(vendor_dict)
    
    # Broadcast WebSocket pour la synchronisation temps réel
    await realtime_manager.emit_event(
        "suppliers",
        "created",
        vendor_data,
        user_id=current_user.get("id")
    )
    
    return Vendor(**vendor_data)

@router.put("/vendors/{vendor_id}",
    summary="Modifier un fournisseur", response_model=Vendor, tags=["Fournisseurs"])
async def update_vendor(vendor_id: str, vendor_update: VendorUpdate, current_user: dict = Depends(require_permission("vendors", "edit"))):
    """Modifier un fournisseur"""
    try:
        update_data = {k: v for k, v in vendor_update.model_dump().items() if v is not None}
        
        await db.vendors.update_one(
            {"_id": ObjectId(vendor_id)},
            {"$set": update_data}
        )
        
        vendor = await db.vendors.find_one({"_id": ObjectId(vendor_id)})
        vendor_data = serialize_doc(vendor)
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "suppliers",
            "updated",
            vendor_data,
            user_id=current_user.get("id")
        )
        
        return Vendor(**vendor_data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/vendors/{vendor_id}", response_model=MessageResponse,
    summary="Supprimer un fournisseur", tags=["Fournisseurs"])
async def delete_vendor(vendor_id: str, current_user: dict = Depends(require_permission("vendors", "delete"))):
    """Supprimer un fournisseur"""
    try:
        # Récupérer le fournisseur avant suppression pour le broadcast
        vendor = await db.vendors.find_one({"_id": ObjectId(vendor_id)})
        vendor_name = vendor.get("nom", "Inconnu") if vendor else "Inconnu"
        
        result = await db.vendors.delete_one({"_id": ObjectId(vendor_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Fournisseur non trouvé")
        
        # Broadcast WebSocket pour la synchronisation temps réel
        await realtime_manager.emit_event(
            "suppliers",
            "deleted",
            {"id": vendor_id, "nom": vendor_name},
            user_id=current_user.get("id")
        )
        
        return {"message": "Fournisseur supprimé"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ==================== VENDOR AI EXTRACT ====================
@router.post("/vendors/ai/extract",
    summary="Extraire les informations fournisseur d'un document via IA", tags=["Fournisseurs"])
async def extract_vendor_from_document(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_permission("vendors", "edit"))
):
    """
    Analyse un document (Excel, PDF, image) via IA et extrait les informations
    pour créer une fiche fournisseur.
    Supporte: PDF, images, Excel (converti en texte avant envoi à l'IA)
    """
    import tempfile
    import json as json_mod

    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType

        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Clé LLM non configurée")

        ext = os.path.splitext(file.filename)[1].lower()

        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        # Formats supportés nativement par Gemini (fichier binaire)
        native_formats = {".pdf", ".png", ".jpg", ".jpeg", ".webp"}
        # Formats à convertir en texte
        spreadsheet_formats = {".xlsx", ".xls", ".csv"}
        
        system_prompt = """Tu es un assistant spécialisé dans l'extraction d'informations fournisseurs à partir de documents administratifs et commerciaux.

Analyse le document fourni et extrais TOUTES les informations relatives au fournisseur.
Le document peut être un formulaire de création fournisseur, un devis, une facture, un bon de commande, un contrat, ou tout autre document commercial.

Réponds UNIQUEMENT avec un JSON valide, sans texte autour ni backticks.

Format attendu:
{
  "nom": "Nom de la société/entreprise",
  "contact": "Nom du contact principal (Prénom Nom) ou null",
  "contact_fonction": "Fonction/poste du contact ou null",
  "email": "Email du contact ou de l'entreprise ou null",
  "telephone": "Numéro de téléphone ou null",
  "adresse": "Adresse complète (rue) ou null",
  "code_postal": "Code postal ou null",
  "ville": "Ville ou null",
  "pays": "Code pays (FR, DE, LU, etc.) ou null",
  "specialite": "Domaine d'activité/spécialité déduit du document",
  "tva_intra": "N° TVA intracommunautaire ou null",
  "siret": "N° SIRET/SIREN ou numéro d'enregistrement ou null",
  "conditions_paiement": "valeur parmi: 30J_NET, 30J_FDM, 45J_FDM, 60J_FDM, 90J_FDM ou null",
  "devise": "EUR, USD, GBP, etc. ou null",
  "categorie": "valeur parmi: MAINTENANCE, FOURNITURES, SERVICES, EQUIPEMENTS, SOUS_TRAITANCE, ENERGIE, INFORMATIQUE, LOGISTIQUE, NETTOYAGE, SECURITE, AUTRE ou null",
  "sous_traitant": false,
  "site_web": "URL du site web ou null",
  "notes": "Informations complémentaires utiles extraites du document ou null",
  "confidence": 0.8
}

RÈGLES:
- Si une information n'est pas trouvée, mets null
- Pour le nom de société, cherche: raison sociale, nom commercial, dénomination
- Pour le contact, cherche: interlocuteur, responsable, signataire
- Déduis la spécialité et la catégorie à partir du contenu du document
- Le champ conditions_paiement doit correspondre EXACTEMENT à une des valeurs listées
- Le champ categorie doit correspondre EXACTEMENT à une des valeurs listées
- Extrais le maximum d'informations possibles"""

        chat = LlmChat(
            api_key=api_key,
            session_id=f"vendor_extract_{uuid.uuid4().hex[:8]}",
            system_message=system_prompt
        ).with_model("gemini", "gemini-2.5-flash")

        if ext in spreadsheet_formats:
            # Convertir Excel/CSV en texte pour l'envoyer à Gemini
            text_content = ""
            try:
                import openpyxl
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text_content += f"=== Feuille: {sheet_name} ===\n"
                    for row in ws.iter_rows(values_only=False):
                        row_texts = []
                        for cell in row:
                            if cell.value is not None:
                                row_texts.append(f"{str(cell.value).strip()}")
                        if any(t for t in row_texts):
                            text_content += " | ".join(row_texts) + "\n"
                    text_content += "\n"
            except Exception:
                # Fallback: lire comme CSV
                try:
                    with open(tmp_path, 'r', encoding='utf-8', errors='replace') as f:
                        text_content = f.read()
                except Exception:
                    text_content = "Impossible de lire le fichier"

            response = await chat.send_message(
                UserMessage(
                    text=f"Voici le contenu extrait d'un document fournisseur ({file.filename}). Analyse-le et extrais les informations du fournisseur. Réponds uniquement en JSON.\n\n---\n{text_content[:15000]}"
                )
            )
        else:
            # Formats natifs (PDF, images) — envoi direct du fichier
            mime_map = {
                ".pdf": "application/pdf",
                ".png": "image/png",
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".webp": "image/webp",
            }
            mime_type = mime_map.get(ext, "application/octet-stream")
            
            response = await chat.send_message(
                UserMessage(
                    text="Analyse ce document et extrais les informations du fournisseur. Réponds uniquement en JSON.",
                    file_contents=[FileContentWithMimeType(file_path=tmp_path, mime_type=mime_type)]
                )
            )

        # Nettoyer le fichier temporaire
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

        # Parser la réponse JSON (response est un string directement)
        response_text = response.strip()
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        response_text = response_text.strip()

        extracted_data = json_mod.loads(response_text)

        return {
            "success": True,
            "extracted_data": extracted_data,
            "source_filename": file.filename
        }

    except json_mod.JSONDecodeError as e:
        logger.error(f"Erreur parsing JSON IA: {str(e)}")
        raise HTTPException(status_code=422, detail=f"L'IA n'a pas retourné un JSON valide: {str(e)}")
    except ImportError:
        raise HTTPException(status_code=500, detail="Module IA non disponible (emergentintegrations)")
    except Exception as e:
        logger.error(f"Erreur extraction IA fournisseur: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'analyse IA: {str(e)}")



# ==================== PURCHASE HISTORY ROUTES ====================
@router.get("/purchase-history/template", tags=["Historique Achats"])
async def download_purchase_history_template(
    format: str = "csv",
    current_user: dict = Depends(require_permission("purchaseHistory", "view"))
):
    """Telecharger le template CSV pour l'import d'historique d'achat"""
    import io
    import csv
    
    headers = [
        "fournisseur", "numeroCommande", "numeroReception", "dateCreation",
        "article", "description", "groupeStatistique", "quantite",
        "montantLigneHT", "quantiteRetournee", "site", "creationUser"
    ]
    
    example_row = [
        "Fournisseur ABC", "CMD-2026-001", "REC-2026-001", "2026-01-15",
        "Roulement SKF 6205", "Roulement a billes", "Pieces mecaniques", "10",
        "150.00", "0", "Site principal", "admin"
    ]
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(headers)
    writer.writerow(example_row)
    
    content = output.getvalue()
    output.close()
    
    from starlette.responses import Response
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_historique_achat.csv"}
    )

@router.get("/purchase-history/grouped", tags=["Historique Achats"])
async def get_purchase_history_grouped(current_user: dict = Depends(require_permission("purchaseHistory", "view"))):
    """Liste tous les achats groupés par N° Commande"""
    purchases = await db.purchase_history.find().sort("dateCreation", -1).to_list(5000)
    
    # Grouper par numeroCommande
    grouped = {}
    for p in purchases:
        num_cmd = p.get('numeroCommande')
        if not num_cmd:
            continue
            
        if num_cmd not in grouped:
            # Utiliser Fournisseur2 (colonne M) si disponible, sinon fournisseur
            fournisseur_display = p.get('Fournisseur2') or p.get('fournisseur', 'Inconnu')
            
            grouped[num_cmd] = {
                'numeroCommande': num_cmd,
                'fournisseur': fournisseur_display,
                'numeroReception': p.get('numeroReception'),  # Premier N° reception de la commande
                'dateCreation': p.get('dateCreation'),
                'site': p.get('site'),
                'items': [],
                'montantTotal': 0.0,
                'itemCount': 0
            }
        
        # Ajouter l'item au groupe
        item_data = {
            'article': p.get('article'),
            'description': p.get('description'),
            'quantite': p.get('quantite', 0.0),
            'montantLigneHT': p.get('montantLigneHT', 0.0),
            'numeroReception': p.get('numeroReception'),
            'groupeStatistique': p.get('groupeStatistique')
        }
        
        grouped[num_cmd]['items'].append(item_data)
        grouped[num_cmd]['montantTotal'] += item_data['montantLigneHT']
        grouped[num_cmd]['itemCount'] += 1
    
    # Convertir en liste
    result = list(grouped.values())
    return result


@router.delete("/purchase-history/all", tags=["Historique Achats"])
async def delete_all_purchase_history(current_user: dict = Depends(get_current_admin_user)):
    """Supprimer tout l'historique d'achat (admin uniquement)"""
    result = await db.purchase_history.delete_many({})
    return {
        "message": f"{result.deleted_count} achats supprimés",
        "deleted_count": result.deleted_count
    }


@router.get("/purchase-history", response_model=List[PurchaseHistory], tags=["Historique Achats"])
async def get_purchase_history(current_user: dict = Depends(require_permission("purchaseHistory", "view"))):
    """Liste tous les achats"""
    purchases = await db.purchase_history.find().sort("dateCreation", -1).to_list(5000)
    
    # Filtrer pour ne garder que les champs du modèle PurchaseHistory
    allowed_fields = {
        '_id', 'id', 'fournisseur', 'numeroCommande', 'numeroReception', 
        'dateCreation', 'article', 'description', 'groupeStatistique',
        'quantite', 'montantLigneHT', 'quantiteRetournee', 'site', 
        'creationUser', 'dateEnregistrement'
    }
    
    result = []
    for p in purchases:
        # Ne garder que les champs autorisés
        filtered_doc = {k: v for k, v in p.items() if k in allowed_fields}
        
        # S'assurer que les champs obligatoires existent avec des valeurs par défaut
        if 'montantLigneHT' not in filtered_doc or filtered_doc['montantLigneHT'] is None:
            filtered_doc['montantLigneHT'] = 0.0
        if 'quantite' not in filtered_doc or filtered_doc['quantite'] is None:
            filtered_doc['quantite'] = 0.0
        if 'quantiteRetournee' not in filtered_doc:
            filtered_doc['quantiteRetournee'] = 0.0
        
        try:
            result.append(PurchaseHistory(**serialize_doc(filtered_doc)))
        except Exception as e:
            logger.error(f"Erreur serialization purchase {filtered_doc.get('numeroCommande')}: {e}")
            continue
    
    return result

@router.get("/purchase-history/stats", tags=["Historique Achats"])
async def get_purchase_stats(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(require_permission("purchaseHistory", "view"))
):
    """Statistiques complètes des achats"""
    
    # Filtres de date
    match_filter = {}
    if start_date:
        match_filter["dateCreation"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "dateCreation" in match_filter:
            match_filter["dateCreation"]["$lte"] = datetime.fromisoformat(end_date)
        else:
            match_filter["dateCreation"] = {"$lte": datetime.fromisoformat(end_date)}
    
    # Total des achats
    all_purchases = await db.purchase_history.find(match_filter).to_list(10000)
    
    if not all_purchases:
        return {
            "totalAchats": 0,
            "montantTotal": 0,
            "commandesTotales": 0,
            "parFournisseur": [],
            "parMois": [],
            "parSite": [],
            "parGroupeStatistique": [],
            "articlesTop": [],
            "par_utilisateur": [],
            "par_mois": [],
            "par_mois_categories": []
        }
    
    total_achats = len(all_purchases)
    montant_total = sum(p.get("montantLigneHT", 0) for p in all_purchases)
    
    # Compter les commandes uniques (pas les lignes)
    commandes_uniques = set()
    for p in all_purchases:
        num_cmd = p.get("numeroCommande")
        if num_cmd:
            commandes_uniques.add(num_cmd)
    
    commandes_totales = len(commandes_uniques)
    
    # NOUVELLES STATS - Par utilisateur (créateur colonne L)
    user_stats = {}
    for purchase in all_purchases:
        user = purchase.get('creationUser', 'Inconnu')
        num_commande = purchase.get('numeroCommande')
        montant = purchase.get('montantLigneHT', 0)
        
        if user not in user_stats:
            user_stats[user] = {
                'utilisateur': user,
                'commandes': set(),
                'montant_total': 0,
                'nb_lignes': 0
            }
        
        if num_commande:
            user_stats[user]['commandes'].add(num_commande)
        user_stats[user]['montant_total'] += montant
        user_stats[user]['nb_lignes'] += 1
    
    # Convertir en liste
    users_list = []
    for user, data in user_stats.items():
        nb_commandes = len(data['commandes'])
        montant = data['montant_total']
        pourcentage = (montant / montant_total * 100) if montant_total > 0 else 0
        
        users_list.append({
            'utilisateur': user,
            'nb_commandes': nb_commandes,
            'nb_lignes': data['nb_lignes'],
            'montant_total': round(montant, 2),
            'pourcentage': round(pourcentage, 2)
        })
    
    users_list.sort(key=lambda x: x['montant_total'], reverse=True)
    
    # NOUVELLES STATS - Par mois avec catégorisation PAR (ARTICLE + DM6)
    monthly_stats = {}
    monthly_article_dm6_stats = {}  # Structure: {mois: {(article, dm6): {category, montant, etc.}}}
    
    for purchase in all_purchases:
        date = purchase.get('dateCreation')
        if date:
            if isinstance(date, str):
                date = datetime.fromisoformat(date.replace('Z', '+00:00'))
            month_key = date.strftime('%Y-%m')
            num_commande = purchase.get('numeroCommande')
            montant = purchase.get('montantLigneHT', 0)
            article = purchase.get('article', '')
            dm6 = purchase.get('DM6', 'Non défini')
            
            # Stats globales par mois
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {
                    'mois': month_key,
                    'commandes': set(),
                    'montant': 0,
                    'nb_lignes': 0
                }
            
            if num_commande:
                monthly_stats[month_key]['commandes'].add(num_commande)
            monthly_stats[month_key]['montant'] += montant
            monthly_stats[month_key]['nb_lignes'] += 1
            
            # Stats par (ARTICLE, DM6) - CHAQUE COMBINAISON est unique!
            try:
                category = get_category_from_article_dm6(article, dm6)
            except Exception:
                category = "Non catégorisé"
            
            if month_key not in monthly_article_dm6_stats:
                monthly_article_dm6_stats[month_key] = {}
            
            # Clé unique = (ARTICLE, DM6) - PAS juste DM6!
            unique_key = (article, dm6)
            
            if unique_key not in monthly_article_dm6_stats[month_key]:
                monthly_article_dm6_stats[month_key][unique_key] = {
                    'article': article,
                    'dm6': dm6,
                    'categorie': category,
                    'montant': 0,
                    'nb_lignes': 0,
                    'commandes': set()
                }
            
            monthly_article_dm6_stats[month_key][unique_key]['montant'] += montant
            monthly_article_dm6_stats[month_key][unique_key]['nb_lignes'] += 1
            if num_commande:
                monthly_article_dm6_stats[month_key][unique_key]['commandes'].add(num_commande)
    
    # Créer la liste mensuelle globale
    monthly_list = []
    for month, data in monthly_stats.items():
        monthly_list.append({
            'mois': month,
            'nb_commandes': len(data['commandes']),
            'nb_lignes': data['nb_lignes'],
            'montant_total': round(data['montant'], 2)
        })
    monthly_list.sort(key=lambda x: x['mois'])
    
    # Créer la liste mensuelle par (ARTICLE, DM6)
    monthly_category_list = []
    for month in sorted(monthly_article_dm6_stats.keys()):
        month_data = {
            'mois': month,
            'categories': []  # Garder 'categories' pour compatibilité frontend
        }
        
        for (article, dm6), data in monthly_article_dm6_stats[month].items():
            month_data['categories'].append({
                'article': article,
                'dm6': dm6,
                'nom': data['categorie'],
                'montant': round(data['montant'], 2),
                'nb_lignes': data['nb_lignes'],
                'nb_commandes': len(data['commandes'])
            })
        
        # Trier par montant décroissant
        month_data['categories'].sort(key=lambda x: x['montant'], reverse=True)
        monthly_category_list.append(month_data)
    
    # Par fournisseur (ancienne stat - gardée)
    fournisseurs = {}
    for p in all_purchases:
        fournisseur = p.get("Fournisseur2") or p.get("fournisseur", "Inconnu")
        if fournisseur not in fournisseurs:
            fournisseurs[fournisseur] = {"montant": 0, "quantite": 0, "count": 0}
        fournisseurs[fournisseur]["montant"] += p.get("montantLigneHT", 0)
        fournisseurs[fournisseur]["quantite"] += p.get("quantite", 0)
        fournisseurs[fournisseur]["count"] += 1
    
    par_fournisseur = [
        {
            "fournisseur": k,
            "montant": v["montant"],
            "quantite": v["quantite"],
            "count": v["count"],
            "pourcentage": round((v["montant"] / montant_total * 100) if montant_total > 0 else 0, 2)
        }
        for k, v in sorted(fournisseurs.items(), key=lambda x: x[1]["montant"], reverse=True)
    ]
    
    # Par mois (ancien format - gardé pour compatibilité)
    mois_dict = {}
    for p in all_purchases:
        date_creation = p.get("dateCreation")
        if date_creation:
            if isinstance(date_creation, str):
                date_creation = datetime.fromisoformat(date_creation.replace('Z', '+00:00'))
            mois_annee = date_creation.strftime("%Y-%m")
            if mois_annee not in mois_dict:
                mois_dict[mois_annee] = {"montant": 0, "quantite": 0, "count": 0}
            mois_dict[mois_annee]["montant"] += p.get("montantLigneHT", 0)
            mois_dict[mois_annee]["quantite"] += p.get("quantite", 0)
            mois_dict[mois_annee]["count"] += 1
    
    par_mois = [
        {"mois": k, "montant": v["montant"], "quantite": v["quantite"], "count": v["count"]}
        for k, v in sorted(mois_dict.items())
    ]
    
    # Par site
    sites = {}
    for p in all_purchases:
        site = p.get("site", "Non défini")
        if site not in sites:
            sites[site] = {"montant": 0, "quantite": 0, "count": 0}
        sites[site]["montant"] += p.get("montantLigneHT", 0)
        sites[site]["quantite"] += p.get("quantite", 0)
        sites[site]["count"] += 1
    
    par_site = [
        {"site": k, "montant": v["montant"], "quantite": v["quantite"], "count": v["count"]}
        for k, v in sorted(sites.items(), key=lambda x: x[1]["montant"], reverse=True)
    ]
    
    # Par groupe statistique
    groupes = {}
    for p in all_purchases:
        groupe = p.get("groupeStatistique", "Non défini")
        if groupe not in groupes:
            groupes[groupe] = {"montant": 0, "quantite": 0, "count": 0}
        groupes[groupe]["montant"] += p.get("montantLigneHT", 0)
        groupes[groupe]["quantite"] += p.get("quantite", 0)
        groupes[groupe]["count"] += 1
    
    par_groupe = [
        {"groupe": k, "montant": v["montant"], "quantite": v["quantite"], "count": v["count"]}
        for k, v in sorted(groupes.items(), key=lambda x: x[1]["montant"], reverse=True)
    ]
    
    # Articles top
    articles = {}
    for p in all_purchases:
        article = p.get("article", "Inconnu")
        if article not in articles:
            articles[article] = {"montant": 0, "quantite": 0, "count": 0, "description": p.get("description", "")}
        articles[article]["montant"] += p.get("montantLigneHT", 0)
        articles[article]["quantite"] += p.get("quantite", 0)
        articles[article]["count"] += 1
    
    articles_top = [
        {"article": k, **v}
        for k, v in sorted(articles.items(), key=lambda x: x[1]["montant"], reverse=True)[:20]
    ]
    
    return {
        "totalAchats": total_achats,
        "montantTotal": round(montant_total, 2),
        "commandesTotales": commandes_totales,
        "parFournisseur": par_fournisseur,
        "parMois": par_mois,
        "parSite": par_site,
        "parGroupeStatistique": par_groupe,
        "articlesTop": articles_top,
        "par_utilisateur": users_list,  # NOUVELLES STATS
        "par_mois": monthly_list,  # NOUVELLES STATS (format différent)
        "par_mois_categories": monthly_category_list  # NOUVELLES STATS - Catégorisation mensuelle
    }

@router.post("/purchase-history", response_model=PurchaseHistory, tags=["Historique Achats"])
async def create_purchase(purchase: PurchaseHistoryCreate, current_user: dict = Depends(require_permission("purchaseHistory", "edit"))):
    """Créer un nouvel achat"""
    purchase_dict = purchase.model_dump()
    
    # Convertir datetime en ISO string si nécessaire
    if isinstance(purchase_dict.get("dateCreation"), datetime):
        purchase_dict["dateCreation"] = purchase_dict["dateCreation"].isoformat()
    
    purchase_dict["dateEnregistrement"] = datetime.utcnow()
    purchase_dict["_id"] = ObjectId()
    
    # Ajouter l'utilisateur créateur si non fourni
    if not purchase_dict.get("creationUser"):
        purchase_dict["creationUser"] = current_user.get("email")
    
    await db.purchase_history.insert_one(purchase_dict)
    
    return PurchaseHistory(**serialize_doc(purchase_dict))

@router.put("/purchase-history/{purchase_id}", response_model=PurchaseHistory, tags=["Historique Achats"])
async def update_purchase(purchase_id: str, purchase_update: PurchaseHistoryUpdate, current_user: dict = Depends(require_permission("purchaseHistory", "edit"))):
    """Modifier un achat"""
    try:
        update_data = {k: v for k, v in purchase_update.model_dump().items() if v is not None}
        
        # Convertir datetime en ISO string si nécessaire
        if "dateCreation" in update_data and isinstance(update_data["dateCreation"], datetime):
            update_data["dateCreation"] = update_data["dateCreation"].isoformat()
        
        await db.purchase_history.update_one(
            {"_id": ObjectId(purchase_id)},
            {"$set": update_data}
        )
        
        purchase = await db.purchase_history.find_one({"_id": ObjectId(purchase_id)})
        return PurchaseHistory(**serialize_doc(purchase))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/purchase-history/{purchase_id}", tags=["Historique Achats"])
async def delete_purchase(purchase_id: str, current_user: dict = Depends(require_permission("purchaseHistory", "delete"))):
    """Supprimer un achat"""
    try:
        result = await db.purchase_history.delete_one({"_id": ObjectId(purchase_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Achat non trouvé")
        return {"message": "Achat supprimé"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


