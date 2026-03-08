"""
Routes API pour les Widgets Personnalisés
"""
from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks, Path, UploadFile, File
from pathlib import Path as FilePath
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
import uuid
import logging
import asyncio

from custom_widgets_models import (
    CustomWidget, CustomWidgetCreate, CustomWidgetUpdate,
    DataSource, DataSourceType, WidgetType,
    ServiceDashboardConfig
)
from excel_smb_service import read_excel_from_smb, test_smb_connection, get_excel_preview, configure_smb
from gmao_data_service import get_gmao_data, get_available_gmao_data_types, init_gmao_data_service
from formula_engine import evaluate_formula, validate_formula

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/custom-widgets", tags=["Custom Widgets"])

# Référence à la base de données (initialisée dans init_custom_widgets_routes)
db = None
audit_service = None


def init_custom_widgets_routes(database, audit_svc=None):
    """Initialise les routes avec la base de données"""
    global db, audit_service
    db = database
    audit_service = audit_svc
    init_gmao_data_service(database)
    logger.info("Routes Custom Widgets initialisées")


# === Dépendances ===

from dependencies import get_current_user


# === Routes CRUD Widgets ===

@router.get("", response_model=List[Dict])
async def get_widgets(
    service: Optional[str] = None,
    shared_only: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """
    Récupère les widgets accessibles par l'utilisateur
    
    - Ses propres widgets
    - Les widgets partagés avec son rôle
    - Les widgets de son service
    """
    user_id = current_user.get("id")
    user_role = current_user.get("role")
    user_service = current_user.get("service")
    
    # Si un filtre service est specifie, montrer les widgets de ce service
    if service:
        query = {"service": service}
    else:
        query = {"$or": [
            {"created_by": user_id},
            {"is_shared": True, "shared_with_roles": user_role},
        ]}
        if user_service:
            query["$or"].append({"service": user_service, "is_shared": True})
    
    # Filtre partagés uniquement
    if shared_only:
        query["is_shared"] = True
    
    widgets = await db.custom_widgets.find(query, {"_id": 0}).to_list(length=None)
    
    # Trier par position puis par date de création
    widgets.sort(key=lambda w: (w.get("position", 999), w.get("created_at", "")))
    
    return widgets


@router.get("/my-widgets", response_model=List[Dict])
async def get_my_widgets(current_user: dict = Depends(get_current_user)):
    """Récupère uniquement les widgets créés par l'utilisateur"""
    user_id = current_user.get("id")
    widgets = await db.custom_widgets.find(
        {"created_by": user_id},
        {"_id": 0}
    ).to_list(length=None)
    return widgets


@router.get("/{widget_id}", response_model=Dict)
async def get_widget(widget_id: str = Path(..., regex=r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$"), current_user: dict = Depends(get_current_user)):
    """Récupère un widget par son ID"""
    widget = await db.custom_widgets.find_one({"id": widget_id}, {"_id": 0})
    if not widget:
        raise HTTPException(status_code=404, detail="Widget non trouvé")
    
    # Vérifier l'accès
    user_id = current_user.get("id")
    user_role = current_user.get("role")
    
    can_access = (
        widget.get("created_by") == user_id or
        (widget.get("is_shared") and user_role in widget.get("shared_with_roles", []))
    )
    
    if not can_access:
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    return widget


@router.post("", response_model=Dict)
async def create_widget(
    widget_data: CustomWidgetCreate,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Crée un nouveau widget personnalisé"""
    user_id = current_user.get("id")
    user_name = f"{current_user.get('prenom', '')} {current_user.get('nom', '')}".strip()
    user_service = current_user.get("service")
    
    widget = {
        "id": str(uuid.uuid4()),
        "name": widget_data.name,
        "description": widget_data.description,
        "data_sources": [ds.model_dump() for ds in widget_data.data_sources],
        "primary_source_id": widget_data.primary_source_id,
        "visualization": widget_data.visualization.model_dump(),
        "refresh_interval": widget_data.refresh_interval or 5,
        "is_shared": widget_data.is_shared,
        "shared_with_roles": widget_data.shared_with_roles,
        "service": widget_data.service or user_service,
        "created_by": user_id,
        "created_by_name": user_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "position": 0,
        "is_active": True
    }
    
    await db.custom_widgets.insert_one(widget)
    widget.pop("_id", None)
    
    # Rafraîchir les données en arrière-plan
    background_tasks.add_task(refresh_widget_data, widget["id"])
    
    logger.info(f"Widget créé: {widget['name']} par {user_name}")
    
    # L'audit est facultatif - ne pas bloquer si le service n'est pas disponible
    # ou si l'appel échoue
    
    return widget


@router.put("/{widget_id}", response_model=Dict)
async def update_widget(
    widget_id: str,
    widget_data: CustomWidgetUpdate,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Met à jour un widget existant"""
    widget = await db.custom_widgets.find_one({"id": widget_id})
    if not widget:
        raise HTTPException(status_code=404, detail="Widget non trouvé")
    
    # Vérifier que l'utilisateur est le créateur
    if widget.get("created_by") != current_user.get("id"):
        raise HTTPException(status_code=403, detail="Seul le créateur peut modifier ce widget")
    
    update_data = {}
    if widget_data.name is not None:
        update_data["name"] = widget_data.name
    if widget_data.description is not None:
        update_data["description"] = widget_data.description
    if widget_data.data_sources is not None:
        update_data["data_sources"] = [ds.model_dump() for ds in widget_data.data_sources]
    if widget_data.primary_source_id is not None:
        update_data["primary_source_id"] = widget_data.primary_source_id
    if widget_data.visualization is not None:
        update_data["visualization"] = widget_data.visualization.model_dump()
    if widget_data.refresh_interval is not None:
        update_data["refresh_interval"] = widget_data.refresh_interval
    if widget_data.is_shared is not None:
        update_data["is_shared"] = widget_data.is_shared
    if widget_data.shared_with_roles is not None:
        update_data["shared_with_roles"] = widget_data.shared_with_roles
    if widget_data.service is not None:
        update_data["service"] = widget_data.service
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.custom_widgets.update_one({"id": widget_id}, {"$set": update_data})
    
    # Rafraîchir les données en arrière-plan
    background_tasks.add_task(refresh_widget_data, widget_id)
    
    updated = await db.custom_widgets.find_one({"id": widget_id}, {"_id": 0})
    return updated


@router.delete("/{widget_id}")
async def delete_widget(widget_id: str, current_user: dict = Depends(get_current_user)):
    """Supprime un widget"""
    widget = await db.custom_widgets.find_one({"id": widget_id})
    if not widget:
        raise HTTPException(status_code=404, detail="Widget non trouvé")
    
    # Vérifier que l'utilisateur est le créateur ou admin
    if widget.get("created_by") != current_user.get("id") and current_user.get("role") != "ADMIN":
        raise HTTPException(status_code=403, detail="Non autorisé à supprimer ce widget")
    
    await db.custom_widgets.delete_one({"id": widget_id})
    
    logger.info(f"Widget supprimé: {widget_id}")
    
    return {"message": "Widget supprimé", "id": widget_id}


@router.put("/{widget_id}/position")
async def update_widget_position(
    widget_id: str,
    position: int,
    current_user: dict = Depends(get_current_user)
):
    """Met à jour la position d'un widget sur le dashboard"""
    result = await db.custom_widgets.update_one(
        {"id": widget_id, "created_by": current_user.get("id")},
        {"$set": {"position": position}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Widget non trouvé")
    
    return {"message": "Position mise à jour"}


# === Routes de rafraîchissement des données ===

@router.post("/{widget_id}/refresh")
async def refresh_widget(widget_id: str, current_user: dict = Depends(get_current_user)):
    """Force le rafraîchissement des données d'un widget"""
    widget = await db.custom_widgets.find_one({"id": widget_id})
    if not widget:
        raise HTTPException(status_code=404, detail="Widget non trouvé")
    
    try:
        await refresh_widget_data(widget_id)
        updated = await db.custom_widgets.find_one({"id": widget_id}, {"_id": 0})
        return {"message": "Widget rafraîchi", "widget": updated}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur de rafraîchissement: {str(e)}")


async def refresh_widget_data(widget_id: str):
    """Rafraîchit les données d'un widget (appelé en arrière-plan)"""
    widget = await db.custom_widgets.find_one({"id": widget_id})
    if not widget:
        return
    
    try:
        sources_values = {}
        errors = []
        
        # Rafraîchir chaque source de données
        for source in widget.get("data_sources", []):
            source_id = source.get("id")
            source_type = source.get("type")
            
            try:
                value = None
                
                if source_type == "manual":
                    value = source.get("manual_value")
                
                elif source_type == "excel":
                    excel_config = source.get("excel_config", {})
                    value = read_excel_from_smb(
                        smb_path=excel_config.get("smb_path"),
                        sheet_name=excel_config.get("sheet_name"),
                        cell_reference=excel_config.get("cell_reference"),
                        column_name=excel_config.get("column_name"),
                        row_filter=excel_config.get("row_filter"),
                        aggregation=excel_config.get("aggregation"),
                        username=excel_config.get("smb_username"),
                        password=excel_config.get("smb_password")
                    )
                
                elif source_type == "gmao":
                    gmao_config = source.get("gmao_config", {})
                    value = await get_gmao_data(
                        data_type=gmao_config.get("data_type"),
                        service_filter=gmao_config.get("service_filter"),
                        status_filter=gmao_config.get("status_filter"),
                        date_from=gmao_config.get("date_from"),
                        date_to=gmao_config.get("date_to"),
                        group_by=gmao_config.get("group_by"),
                        sensor_id=gmao_config.get("sensor_id"),
                        meter_id=gmao_config.get("meter_id")
                    )
                
                sources_values[source.get("name")] = value
                
                # Mettre à jour la valeur en cache
                source["cached_value"] = value
                source["last_updated"] = datetime.now(timezone.utc).isoformat()
                source["error_message"] = None
                
            except Exception as e:
                logger.error(f"Erreur source {source_id}: {e}")
                source["error_message"] = str(e)
                errors.append(f"Source '{source.get('name')}': {str(e)}")
        
        # Évaluer les formules
        for source in widget.get("data_sources", []):
            if source.get("type") == "formula" and source.get("formula"):
                try:
                    value = evaluate_formula(source["formula"], sources_values)
                    sources_values[source.get("name")] = value
                    source["cached_value"] = value
                    source["last_updated"] = datetime.now(timezone.utc).isoformat()
                    source["error_message"] = None
                except Exception as e:
                    logger.error(f"Erreur formule: {e}")
                    source["error_message"] = str(e)
                    errors.append(f"Formule '{source.get('name')}': {str(e)}")
        
        # Mettre à jour le widget
        update_data = {
            "data_sources": widget["data_sources"],
            "last_refresh": datetime.now(timezone.utc).isoformat(),
            "refresh_error": "; ".join(errors) if errors else None
        }
        
        await db.custom_widgets.update_one({"id": widget_id}, {"$set": update_data})
        
    except Exception as e:
        logger.error(f"Erreur rafraîchissement widget {widget_id}: {e}")
        await db.custom_widgets.update_one(
            {"id": widget_id},
            {"$set": {"refresh_error": str(e), "last_refresh": datetime.now(timezone.utc).isoformat()}}
        )


# === Routes utilitaires ===

@router.get("/data-types/gmao")
async def get_gmao_data_types(current_user: dict = Depends(get_current_user)):
    """Retourne la liste des types de données FSAO disponibles"""
    return get_available_gmao_data_types()


# === Routes pour lister les sources de données disponibles ===

@router.get("/data-sources/sensors")
async def get_available_sensors(current_user: dict = Depends(get_current_user)):
    """Retourne la liste des capteurs MQTT disponibles"""
    sensors = await db.sensors.find(
        {},
        {"_id": 0, "id": 1, "name": 1, "type": 1, "unit": 1, "location": 1, "current_value": 1, "status": 1}
    ).to_list(length=500)
    
    return [
        {
            "id": s.get("id"),
            "name": s.get("name", f"Capteur {s.get('id', 'N/A')[:8]}"),
            "type": s.get("type", "Inconnu"),
            "unit": s.get("unit", ""),
            "location": s.get("location", ""),
            "current_value": s.get("current_value"),
            "status": s.get("status", "unknown"),
            "label": f"{s.get('name', 'Capteur')} ({s.get('type', '')} - {s.get('location', 'Sans emplacement')})"
        }
        for s in sensors
    ]


@router.get("/data-sources/meters")
async def get_available_meters(current_user: dict = Depends(get_current_user)):
    """Retourne la liste des compteurs disponibles"""
    meters = await db.meters.find(
        {},
        {"_id": 0, "id": 1, "name": 1, "type": 1, "unit": 1, "location": 1, "current_value": 1, "last_reading": 1}
    ).to_list(length=500)
    
    return [
        {
            "id": m.get("id"),
            "name": m.get("name", f"Compteur {m.get('id', 'N/A')[:8]}"),
            "type": m.get("type", "Inconnu"),
            "unit": m.get("unit", ""),
            "location": m.get("location", ""),
            "current_value": m.get("current_value") or m.get("last_reading"),
            "label": f"{m.get('name', 'Compteur')} ({m.get('type', '')} - {m.get('unit', '')})"
        }
        for m in meters
    ]


@router.get("/data-sources/equipments")
async def get_available_equipments(current_user: dict = Depends(get_current_user)):
    """Retourne la liste des équipements disponibles"""
    equipments = await db.equipments.find(
        {},
        {"_id": 0, "id": 1, "name": 1, "type": 1, "status": 1, "location": 1, "service": 1}
    ).to_list(length=500)
    
    return [
        {
            "id": e.get("id"),
            "name": e.get("name", f"Équipement {e.get('id', 'N/A')[:8]}"),
            "type": e.get("type", "Inconnu"),
            "status": e.get("status", ""),
            "location": e.get("location", ""),
            "service": e.get("service", ""),
            "label": f"{e.get('name', 'Équipement')} ({e.get('type', '')})"
        }
        for e in equipments
    ]


@router.get("/data-sources/services")
async def get_available_services(current_user: dict = Depends(get_current_user)):
    """Retourne la liste des services disponibles"""
    # Récupérer les services distincts depuis les utilisateurs et équipements
    user_services = await db.users.distinct("service")
    equipment_services = await db.equipments.distinct("service")
    
    all_services = list(set([s for s in user_services + equipment_services if s]))
    all_services.sort()
    
    return [
        {"id": s, "name": s, "label": s}
        for s in all_services
    ]


# === Templates de widgets prédéfinis ===

@router.get("/tpl/list")
async def get_widget_templates(current_user: dict = Depends(get_current_user)):
    """Retourne la liste des templates de widgets prédéfinis"""
    templates = [
        # === Ordres de Travail ===
        {
            "id": "tpl_ot_completion_today",
            "name": "Taux de complétion OT",
            "description": "Pourcentage d'ordres de travail terminés sur la période",
            "category": "Ordres de Travail",
            "icon": "CheckCircle",
            "preview_value": "78%",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Taux complétion",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "work_orders_completion_rate",
                            "date_from": "-30d"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Complétion OT",
                    "subtitle": "30 derniers jours",
                    "type": "gauge",
                    "min_value": 0,
                    "max_value": 100,
                    "unit": "%",
                    "size": "medium",
                    "color_scheme": "green"
                },
                "refresh_interval": 5
            }
        },
        {
            "id": "tpl_ot_pending",
            "name": "OT en attente",
            "description": "Nombre d'ordres de travail en attente de traitement",
            "category": "Ordres de Travail",
            "icon": "Clock",
            "preview_value": "12",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "OT en attente",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "work_orders_count",
                            "status_filter": "EN_ATTENTE"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "OT en attente",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "orange"
                },
                "refresh_interval": 5
            }
        },
        {
            "id": "tpl_ot_in_progress",
            "name": "OT en cours",
            "description": "Nombre d'ordres de travail actuellement en cours",
            "category": "Ordres de Travail",
            "icon": "Wrench",
            "preview_value": "8",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "OT en cours",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "work_orders_count",
                            "status_filter": "EN_COURS"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "OT en cours",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "blue"
                },
                "refresh_interval": 5
            }
        },
        
        # === Équipements ===
        {
            "id": "tpl_equipment_failures",
            "name": "Équipements en panne",
            "description": "Nombre d'équipements actuellement en panne",
            "category": "Équipements",
            "icon": "AlertTriangle",
            "preview_value": "3",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "En panne",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "assets_count",
                            "status_filter": "EN_PANNE"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Équipements en panne",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "red"
                },
                "refresh_interval": 5
            }
        },
        {
            "id": "tpl_equipment_availability",
            "name": "Disponibilité équipements",
            "description": "Taux de disponibilité des équipements",
            "category": "Équipements",
            "icon": "Activity",
            "preview_value": "94%",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Disponibilité",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "assets_availability_rate"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Disponibilité",
                    "subtitle": "Équipements",
                    "type": "gauge",
                    "min_value": 0,
                    "max_value": 100,
                    "unit": "%",
                    "size": "medium",
                    "color_scheme": "green"
                },
                "refresh_interval": 15
            }
        },
        
        # === Inventaire ===
        {
            "id": "tpl_stock_critical",
            "name": "Stock critique",
            "description": "Nombre d'articles sous le seuil de stock minimum",
            "category": "Inventaire",
            "icon": "Package",
            "preview_value": "5",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Stock bas",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "inventory_low_stock"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Stock critique",
                    "subtitle": "Articles sous seuil",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "red"
                },
                "refresh_interval": 30
            }
        },
        {
            "id": "tpl_stock_value",
            "name": "Valeur du stock",
            "description": "Valeur totale du stock en inventaire",
            "category": "Inventaire",
            "icon": "Euro",
            "preview_value": "45 230 €",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Valeur",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "inventory_value"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Valeur stock",
                    "type": "value",
                    "suffix": " €",
                    "size": "medium",
                    "color_scheme": "blue"
                },
                "refresh_interval": 60
            }
        },
        
        # === Maintenance Préventive ===
        {
            "id": "tpl_mprev_completion",
            "name": "Réalisation M.Prev",
            "description": "Taux de réalisation de la maintenance préventive",
            "category": "Maintenance Préventive",
            "icon": "Calendar",
            "preview_value": "85%",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Taux réalisation",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "preventive_completion_rate",
                            "date_from": "-30d"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Réalisation M.Prev",
                    "subtitle": "30 derniers jours",
                    "type": "gauge",
                    "min_value": 0,
                    "max_value": 100,
                    "unit": "%",
                    "size": "medium",
                    "color_scheme": "purple"
                },
                "refresh_interval": 15
            }
        },
        {
            "id": "tpl_mprev_overdue",
            "name": "M.Prev en retard",
            "description": "Nombre de maintenances préventives en retard",
            "category": "Maintenance Préventive",
            "icon": "AlertCircle",
            "preview_value": "2",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "En retard",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "preventive_overdue_count"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "M.Prev en retard",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "red"
                },
                "refresh_interval": 15
            }
        },
        
        # === IoT / Capteurs ===
        {
            "id": "tpl_sensor_value",
            "name": "Valeur capteur",
            "description": "Affiche la valeur actuelle d'un capteur MQTT spécifique",
            "category": "IoT",
            "icon": "Activity",
            "preview_value": "23.5°C",
            "requires_selection": "sensor",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Valeur capteur",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "sensor_value",
                            "sensor_id": None
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Capteur",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "cyan"
                },
                "refresh_interval": 1
            }
        },
        {
            "id": "tpl_meter_value",
            "name": "Relevé compteur",
            "description": "Affiche le dernier relevé d'un compteur spécifique",
            "category": "IoT",
            "icon": "Gauge",
            "preview_value": "1 234 kWh",
            "requires_selection": "meter",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Valeur compteur",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "meter_value",
                            "meter_id": None
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Compteur",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "teal"
                },
                "refresh_interval": 5
            }
        },
        
        # === Demandes ===
        {
            "id": "tpl_intervention_requests",
            "name": "Demandes d'intervention",
            "description": "Nombre de demandes d'intervention en attente",
            "category": "Demandes",
            "icon": "MessageSquare",
            "preview_value": "7",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Demandes",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "intervention_requests_count",
                            "status_filter": "EN_ATTENTE"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Demandes d'inter.",
                    "subtitle": "En attente",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "orange"
                },
                "refresh_interval": 5
            }
        },
        {
            "id": "tpl_purchase_requests",
            "name": "Demandes d'achat",
            "description": "Nombre de demandes d'achat en attente de validation",
            "category": "Demandes",
            "icon": "ShoppingCart",
            "preview_value": "4",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Demandes achat",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "purchase_requests_count",
                            "status_filter": "EN_ATTENTE"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Demandes d'achat",
                    "subtitle": "En attente",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "indigo"
                },
                "refresh_interval": 10
            }
        },
        
        # === Sécurité ===
        {
            "id": "tpl_near_miss_month",
            "name": "Presqu'accidents du mois",
            "description": "Nombre de presqu'accidents signalés ce mois",
            "category": "Sécurité",
            "icon": "AlertTriangle",
            "preview_value": "2",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Presqu'accidents",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "near_miss_count",
                            "date_from": "-1m"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Presqu'accidents",
                    "subtitle": "Ce mois",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "yellow"
                },
                "refresh_interval": 30
            }
        },

        # === KPIs Maintenance ===
        {
            "id": "tpl_mttr",
            "name": "MTTR (Temps moyen de reparation)",
            "description": "Temps moyen entre l'ouverture et la cloture d'un OT, en heures",
            "category": "KPIs Maintenance",
            "icon": "Clock",
            "preview_value": "4.5h",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "MTTR",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "mttr_hours",
                            "date_from": "-30d"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "MTTR",
                    "subtitle": "30 derniers jours",
                    "type": "value",
                    "suffix": " h",
                    "size": "small",
                    "color_scheme": "cyan"
                },
                "refresh_interval": 60
            }
        },
        {
            "id": "tpl_upcoming_maintenance",
            "name": "Maintenances a venir (7j)",
            "description": "Nombre de maintenances preventives planifiees dans les 7 prochains jours",
            "category": "KPIs Maintenance",
            "icon": "Calendar",
            "preview_value": "3",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "A venir",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "upcoming_maintenance_7d"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Maintenances a venir",
                    "subtitle": "7 prochains jours",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "blue"
                },
                "refresh_interval": 30
            }
        },
        {
            "id": "tpl_team_workload",
            "name": "Charge de travail equipe",
            "description": "Nombre d'OT en cours et en attente pour l'equipe",
            "category": "KPIs Maintenance",
            "icon": "Users",
            "preview_value": "15",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Charge",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "team_workload"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Charge equipe",
                    "subtitle": "OT actifs",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "purple"
                },
                "refresh_interval": 5
            }
        },
        {
            "id": "tpl_equipment_status_changes",
            "name": "Changements de statut recents",
            "description": "Nombre de changements de statut d'equipements dans les 7 derniers jours",
            "category": "Equipements",
            "icon": "Activity",
            "preview_value": "8",
            "config": {
                "data_sources": [
                    {
                        "id": "src_main",
                        "name": "Changements",
                        "type": "gmao",
                        "gmao_config": {
                            "data_type": "status_changes_7d"
                        }
                    }
                ],
                "primary_source_id": "src_main",
                "visualization": {
                    "title": "Changements statut",
                    "subtitle": "7 derniers jours",
                    "type": "value",
                    "size": "small",
                    "color_scheme": "orange"
                },
                "refresh_interval": 15
            }
        }
    ]
    
    return templates


@router.post("/tpl/{template_id}/create")
async def create_widget_from_template(
    template_id: str,
    sensor_id: Optional[str] = None,
    meter_id: Optional[str] = None,
    equipment_id: Optional[str] = None,
    custom_name: Optional[str] = None,
    service: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Crée un widget à partir d'un template prédéfini"""
    # Récupérer les templates
    templates = await get_widget_templates(current_user)
    template = next((t for t in templates if t["id"] == template_id), None)
    
    if not template:
        raise HTTPException(status_code=404, detail="Template non trouvé")
    
    # Copier la configuration
    import copy
    config = copy.deepcopy(template["config"])
    
    # Appliquer les sélections spécifiques (capteur, compteur, etc.)
    for source in config.get("data_sources", []):
        gmao_config = source.get("gmao_config", {})
        
        if sensor_id and gmao_config.get("data_type") == "sensor_value":
            gmao_config["sensor_id"] = sensor_id
            # Récupérer le nom du capteur pour le titre
            sensor = await db.sensors.find_one({"id": sensor_id}, {"_id": 0, "name": 1, "unit": 1})
            if sensor:
                config["visualization"]["title"] = sensor.get("name", "Capteur")
                config["visualization"]["unit"] = sensor.get("unit", "")
        
        if meter_id and gmao_config.get("data_type") == "meter_value":
            gmao_config["meter_id"] = meter_id
            # Récupérer le nom du compteur
            meter = await db.meters.find_one({"id": meter_id}, {"_id": 0, "name": 1, "unit": 1})
            if meter:
                config["visualization"]["title"] = meter.get("name", "Compteur")
                config["visualization"]["unit"] = meter.get("unit", "")
    
    # Créer le widget
    user_id = current_user.get("id")
    user_name = current_user.get("name", current_user.get("email", "Inconnu"))
    
    widget = {
        "id": str(uuid.uuid4()),
        "name": custom_name or template["name"],
        "description": template["description"],
        "created_by": user_id,
        "created_by_name": user_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "from_template": template_id,
        "service": service or current_user.get("service"),
        **config,
        "is_shared": True,
        "shared_with_roles": []
    }
    
    await db.custom_widgets.insert_one(widget)
    
    # Rafraîchir immédiatement les données
    try:
        await refresh_widget_data(widget["id"])
        widget = await db.custom_widgets.find_one({"id": widget["id"]}, {"_id": 0})
    except Exception as e:
        logger.warning(f"Erreur rafraîchissement initial: {e}")
    
    return widget


@router.post("/test/excel-connection")
async def test_excel_connection(
    smb_path: str,
    username: Optional[str] = None,
    password: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Teste la connexion à un fichier Excel via SMB"""
    result = test_smb_connection(smb_path, username, password)
    return result


@router.post("/upload/excel")
async def upload_excel_file(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload un fichier Excel depuis l'ordinateur de l'utilisateur"""
    import openpyxl
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Format non supporte. Utilisez .xlsx, .xls ou .csv")
    
    try:
        # Sauvegarder le fichier
        upload_dir = FilePath("/app/backend/uploads/excel")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        file_id = str(uuid.uuid4())[:8]
        safe_name = f"{file_id}_{file.filename.replace(' ', '_')}"
        file_path = upload_dir / safe_name
        
        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)
        
        # Lire les infos du fichier
        sheets = []
        preview = []
        if file.filename.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            sheets = wb.sheetnames
            # Preview de la premiere feuille
            ws = wb[sheets[0]]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 10:
                    break
                preview.append([str(c) if c is not None else '' for c in row])
            wb.close()
        elif file.filename.endswith('.csv'):
            import csv
            with open(file_path, 'r', encoding='utf-8', errors='replace') as cf:
                reader = csv.reader(cf)
                for i, row in enumerate(reader):
                    if i >= 10:
                        break
                    preview.append(row)
            sheets = ['CSV']
        
        # Sauvegarder la reference en base
        file_record = {
            "id": file_id,
            "filename": file.filename,
            "stored_path": str(file_path),
            "uploaded_by": current_user.get("id"),
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "sheets": sheets,
            "size_bytes": len(content)
        }
        await db.uploaded_excel_files.insert_one(file_record)
        
        return {
            "success": True,
            "file_id": file_id,
            "filename": file.filename,
            "sheets": sheets,
            "preview": preview,
            "stored_path": str(file_path)
        }
    except Exception as e:
        logger.error(f"Erreur upload Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/preview/excel-local/{file_id}")
async def preview_local_excel(
    file_id: str,
    sheet_name: Optional[str] = None,
    max_rows: int = 10,
    current_user: dict = Depends(get_current_user)
):
    """Preview un fichier Excel uploade localement"""
    import openpyxl
    
    file_record = await db.uploaded_excel_files.find_one({"id": file_id}, {"_id": 0})
    if not file_record:
        raise HTTPException(status_code=404, detail="Fichier non trouve")
    
    file_path = file_record.get("stored_path")
    if not file_path or not FilePath(file_path).exists():
        raise HTTPException(status_code=404, detail="Fichier supprime du serveur")
    
    try:
        data = []
        columns = []
        if file_path.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                row_data = [str(c) if c is not None else '' for c in row]
                if i == 0:
                    columns = row_data
                else:
                    data.append(row_data)
                if i >= max_rows:
                    break
            wb.close()
        elif file_path.endswith('.csv'):
            import csv
            with open(file_path, 'r', encoding='utf-8', errors='replace') as cf:
                reader = csv.reader(cf)
                for i, row in enumerate(reader):
                    if i == 0:
                        columns = row
                    else:
                        data.append(row)
                    if i >= max_rows:
                        break
        
        return {
            "success": True,
            "columns": columns,
            "data": data,
            "sheets": file_record.get("sheets", []),
            "total_rows": len(data)
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.post("/preview/excel")
async def preview_excel_file(
    smb_path: str,
    sheet_name: Optional[str] = None,
    max_rows: int = 10,
    username: Optional[str] = None,
    password: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Prévisualise le contenu d'un fichier Excel"""
    result = get_excel_preview(smb_path, sheet_name, max_rows, username, password)
    return result


@router.post("/validate/formula")
async def validate_formula_endpoint(
    formula: str,
    source_names: List[str],
    current_user: dict = Depends(get_current_user)
):
    """Valide une formule et retourne les erreurs éventuelles"""
    return validate_formula(formula, source_names)


@router.post("/test/formula")
async def test_formula(
    formula: str,
    test_values: Dict[str, Any],
    current_user: dict = Depends(get_current_user)
):
    """Teste une formule avec des valeurs de test"""
    try:
        result = evaluate_formula(formula, test_values)
        return {"success": True, "result": result}
    except Exception as e:
        return {"success": False, "error": str(e)}


# === Routes Dashboard Responsable de Service ===

@router.get("/dashboard/config")
async def get_dashboard_config(current_user: dict = Depends(get_current_user)):
    """Récupère la configuration du dashboard de l'utilisateur"""
    user_id = current_user.get("id")
    
    config = await db.service_dashboard_configs.find_one(
        {"user_id": user_id},
        {"_id": 0}
    )
    
    if not config:
        # Créer une configuration par défaut
        config = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "service": current_user.get("service"),
            "widget_ids": [],
            "layout": None,
            "auto_refresh": True,
            "theme": "light",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.service_dashboard_configs.insert_one(config)
        config.pop("_id", None)
    
    return config


@router.put("/dashboard/config")
async def update_dashboard_config(
    config_data: Dict[str, Any],
    current_user: dict = Depends(get_current_user)
):
    """Met à jour la configuration du dashboard"""
    user_id = current_user.get("id")
    
    config_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.service_dashboard_configs.update_one(
        {"user_id": user_id},
        {"$set": config_data},
        upsert=True
    )
    
    return {"message": "Configuration mise à jour"}


# === Configuration SMB globale ===

@router.post("/config/smb")
async def configure_smb_credentials(
    username: str,
    password: str,
    domain: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Configure les credentials SMB pour l'accès aux fichiers Excel"""
    # Seuls les admins peuvent configurer
    if current_user.get("role") != "ADMIN":
        raise HTTPException(status_code=403, detail="Réservé aux administrateurs")
    
    # Stocker de manière sécurisée (crypté)
    import os
    os.environ["SMB_USERNAME"] = username
    os.environ["SMB_PASSWORD"] = password
    if domain:
        os.environ["SMB_DOMAIN"] = domain
    
    configure_smb(username, password, domain)
    
    # Stocker en DB aussi (crypté)
    await db.system_config.update_one(
        {"key": "smb_config"},
        {"$set": {
            "key": "smb_config",
            "username": username,
            "domain": domain,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.get("id")
        }},
        upsert=True
    )
    
    logger.info(f"Configuration SMB mise à jour par {current_user.get('email')}")
    
    return {"message": "Configuration SMB enregistrée"}
